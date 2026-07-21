from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
import io
from datetime import date, timedelta, datetime
from urllib.parse import quote
from openpyxl import Workbook
from app.core.database import get_db
from app.core.log_helper import write_log
from app.core.query_utils import filter_active_employees, get_pending_resign_status_id
from app.models.models import Employee, EmployeeSalary, SalaryCalculation, AttendanceRecord, SysDictBase, ExportTemplate, PerformanceScore, SocialInsurance, SalaryPeriodStep
from app.api.auth import get_current_user, UserInfo, require_permission
from sqlalchemy import func

router = APIRouter()


def _filter_active_without_pending_resign(query, db: Session):
    """过滤活跃员工，排除待离职状态"""
    query = filter_active_employees(query, db)
    pending_resign_ids = get_pending_resign_status_id(db)
    if pending_resign_ids:
        query = query.filter(Employee.status_id.notin_(pending_resign_ids))
    return query


STEP_DEFINITIONS = [
    {"key": "employee", "title": "员工档案", "description": "确认人员信息", "route": "/employees"},
    {"key": "attendance", "title": "考勤数据", "description": "确认考勤数据", "route": "/attendance"},
    {"key": "performance", "title": "绩效评分", "description": "录入绩效系数", "route": "/performance"},
    {"key": "insurance", "title": "社保数据", "description": "确认社保公积金", "route": "/insurance"},
    {"key": "tax", "title": "个税申报", "description": "导出报税并导入个税结果", "route": "/salary"},
    {"key": "salary", "title": "薪资计算", "description": "计算应发与实发工资", "route": "/salary"},
    {"key": "payment", "title": "工资发放", "description": "审核通过后导出报表完成", "route": "/reports"},
]


@router.get("/stats", dependencies=[Depends(require_permission("dashboard:view"))])
def get_stats(
    period: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    query = db.query(Employee)
    active_employees = _filter_active_without_pending_resign(query, db).all()
    total_employees = len(active_employees)
    active_employee_ids = [emp.id for emp in active_employees]

    status_count = {}
    status_id_counts = {}
    for emp in active_employees:
        status_id_counts[emp.status_id] = status_id_counts.get(emp.status_id, 0) + 1
    for status_id, count in status_id_counts.items():
        dict_item = db.query(SysDictBase).filter(SysDictBase.id == status_id).first()
        status_count[dict_item.name if dict_item else "未知"] = count

    latest_period = period
    if not latest_period:
        latest_period_row = db.query(SalaryCalculation.period).order_by(SalaryCalculation.period.desc()).first()
        latest_period = latest_period_row[0] if latest_period_row else None

    salary_stats = {}
    att_stats = {}
    perf_count = 0
    si_count = 0
    att_count = 0
    att_missing = total_employees
    perf_missing = total_employees
    si_missing = total_employees
    salary_completed_count = 0
    salary_pending_count = total_employees
    tax_imported_count = 0
    tax_missing_count = total_employees
    review_passed_count = 0

    if latest_period:
        calcs = db.query(SalaryCalculation).filter(
            SalaryCalculation.period == latest_period,
            SalaryCalculation.employee_id.in_(active_employee_ids)
        ).all()
        gross_completed_calcs = [c for c in calcs if c.calculation_status in ("应发已核算", "实发已核算")]
        salary_completed_count = len(gross_completed_calcs)
        salary_pending_count = total_employees - salary_completed_count
        tax_imported_count = sum(1 for c in calcs if c.tax_deduction is not None)
        tax_missing_count = total_employees - tax_imported_count
        review_passed_count = sum(1 for c in calcs if c.review_status == "审核通过")
        gross_values = [float(c.gross_salary) for c in calcs if c.gross_salary is not None]
        net_values = [float(c.net_salary) for c in calcs if c.net_salary is not None]
        salary_stats = {
            "period": latest_period,
            "total": total_employees,
            "completed": salary_completed_count,
            "pending": salary_pending_count,
            "tax_imported": tax_imported_count,
            "tax_missing": tax_missing_count,
            "review_passed": review_passed_count,
            "review_rejected": sum(1 for c in calcs if c.review_status == "审核驳回"),
            "avg_gross_salary": round(sum(gross_values) / len(gross_values), 2) if gross_values else 0,
            "avg_net_salary": round(sum(net_values) / len(net_values), 2) if net_values else 0,
        }

        atts = db.query(AttendanceRecord).filter(
            AttendanceRecord.period == latest_period,
            AttendanceRecord.employee_id.in_(active_employee_ids)
        ).all()
        att_count = len(atts)
        att_missing = total_employees - att_count
        if atts:
            valid_att_rates = [float(a.attendance_rate) for a in atts if a.attendance_rate is not None]
            att_stats = {
                "period": latest_period,
                "total": att_count,
                "avg_rate": round(sum(valid_att_rates) / len(valid_att_rates) * 100, 1) if valid_att_rates else 0,
                "total_late": sum(a.late_count or 0 for a in atts),
                "total_leave": sum((a.sick_leave_days or 0) + (a.personal_leave_days or 0) for a in atts),
            }

        perfs = db.query(PerformanceScore).filter(
            PerformanceScore.period == latest_period,
            PerformanceScore.employee_id.in_(active_employee_ids),
            PerformanceScore.final_score.isnot(None)
        ).all()
        perf_count = len(perfs)
        perf_missing = total_employees - perf_count

        sis = db.query(SocialInsurance).filter(
            SocialInsurance.period == latest_period,
            SocialInsurance.employee_id.in_(active_employee_ids)
        ).all()
        valid_si_employee_ids = set()
        for si in sis:
            personal_total = (
                float(si.pension_personal or 0) +
                float(si.unemployment_personal or 0) +
                float(si.medical_personal or 0) +
                float(si.hf_personal or 0)
            )
            if personal_total > 0:
                valid_si_employee_ids.add(si.employee_id)
        si_count = len(valid_si_employee_ids)
        si_missing = total_employees - si_count

    confirmed_steps = {}
    if latest_period:
        step_records = db.query(SalaryPeriodStep).filter(
            SalaryPeriodStep.period == latest_period
        ).all()
        for sr in step_records:
            confirmed_steps[sr.step_key] = {
                "is_confirmed": sr.is_confirmed,
                "is_force_confirmed": sr.is_force_confirmed,
                "confirmed_at": str(sr.confirmed_at) if sr.confirmed_at else None,
            }

    steps_detail = []
    prerequisite_keys_for_salary = ["employee", "attendance", "performance", "insurance", "tax"]
    prerequisites_confirmed_for_salary = all(confirmed_steps.get(pk, {}).get("is_confirmed", False) for pk in prerequisite_keys_for_salary)
    prerequisite_keys_for_tax = ["employee", "attendance", "performance", "insurance"]
    prerequisites_confirmed_for_tax = all(confirmed_steps.get(pk, {}).get("is_confirmed", False) for pk in prerequisite_keys_for_tax)
    prerequisite_keys_for_payment = ["employee", "attendance", "performance", "insurance", "tax", "salary"]
    prerequisites_confirmed_for_payment = all(confirmed_steps.get(pk, {}).get("is_confirmed", False) for pk in prerequisite_keys_for_payment)
    for step_def in STEP_DEFINITIONS:
        key = step_def["key"]
        data_ready = False
        missing_count = 0
        can_confirm = total_employees > 0

        if key == "employee":
            data_ready = total_employees > 0
            missing_count = 0 if data_ready else total_employees
        elif key == "attendance":
            data_ready = att_count >= total_employees and total_employees > 0
            missing_count = att_missing
        elif key == "performance":
            data_ready = perf_count >= total_employees and total_employees > 0
            missing_count = perf_missing
        elif key == "insurance":
            data_ready = si_count >= total_employees and total_employees > 0
            missing_count = si_missing
        elif key == "tax":
            if not prerequisites_confirmed_for_tax:
                data_ready = False
                missing_count = 0
                can_confirm = False
            else:
                data_ready = True
                missing_count = tax_missing_count
        elif key == "salary":
            if not prerequisites_confirmed_for_salary:
                data_ready = False
                missing_count = 0
                can_confirm = False
            else:
                data_ready = True
                missing_count = 0
        elif key == "payment":
            if not prerequisites_confirmed_for_payment:
                data_ready = False
                missing_count = 0
                can_confirm = False
            else:
                data_ready = review_passed_count >= total_employees and total_employees > 0
                missing_count = total_employees - review_passed_count if review_passed_count < total_employees else 0

        confirmed_info = confirmed_steps.get(key, {})
        is_confirmed = confirmed_info.get("is_confirmed", False)
        is_force_confirmed = confirmed_info.get("is_force_confirmed", False)

        steps_detail.append({
            "key": key,
            "title": step_def["title"],
            "description": step_def["description"],
            "route": step_def["route"],
            "data_ready": data_ready,
            "missing_count": missing_count,
            "can_confirm": can_confirm,
            "is_confirmed": is_confirmed,
            "is_force_confirmed": is_force_confirmed,
            "confirmed_at": confirmed_info.get("confirmed_at"),
            "prerequisites_met": prerequisites_confirmed_for_salary if key == "salary" else (prerequisites_confirmed_for_tax if key == "tax" else (prerequisites_confirmed_for_payment if key == "payment" else None)),
        })

    return {
        "total_employees": total_employees,
        "status_breakdown": status_count,
        "salary_stats": salary_stats,
        "attendance_stats": att_stats,
        "steps": steps_detail,
        "period": latest_period
    }


class StepConfirmRequest(BaseModel):
    period: str
    step_key: str
    is_confirmed: bool
    remark: Optional[str] = None


@router.post("/steps/confirm", dependencies=[Depends(require_permission("salary:step_confirm"))])
def confirm_step(
    data: StepConfirmRequest,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """确认或取消确认某个算薪步骤"""
    valid_keys = [s["key"] for s in STEP_DEFINITIONS]
    if data.step_key not in valid_keys:
        raise HTTPException(status_code=400, detail=f"无效的步骤标识: {data.step_key}")

    step_def = next(s for s in STEP_DEFINITIONS if s["key"] == data.step_key)
    is_force = False

    if data.is_confirmed:
        query = db.query(Employee)
        active_employees = _filter_active_without_pending_resign(query, db).all()
        total_employees = len(active_employees)
        active_employee_ids = [emp.id for emp in active_employees]

        data_ready = True
        if data.step_key == "employee":
            data_ready = total_employees > 0
        elif data.step_key == "attendance":
            att_count = db.query(AttendanceRecord).filter(
                AttendanceRecord.period == data.period,
                AttendanceRecord.employee_id.in_(active_employee_ids)
            ).count()
            data_ready = att_count >= total_employees and total_employees > 0
        elif data.step_key == "performance":
            perf_count = db.query(PerformanceScore).filter(
                PerformanceScore.period == data.period,
                PerformanceScore.employee_id.in_(active_employee_ids),
                PerformanceScore.final_score.isnot(None)
            ).count()
            data_ready = perf_count >= total_employees and total_employees > 0
        elif data.step_key == "insurance":
            sis = db.query(SocialInsurance).filter(
                SocialInsurance.period == data.period,
                SocialInsurance.employee_id.in_(active_employee_ids)
            ).all()
            valid_si_count = 0
            for si in sis:
                personal_total = (
                    float(si.pension_personal or 0) +
                    float(si.unemployment_personal or 0) +
                    float(si.medical_personal or 0) +
                    float(si.hf_personal or 0)
                )
                if personal_total > 0:
                    valid_si_count += 1
            data_ready = valid_si_count >= total_employees and total_employees > 0
        elif data.step_key == "tax":
            prerequisite_keys = ["employee", "attendance", "performance", "insurance"]
            confirmed_prereqs = db.query(SalaryPeriodStep).filter(
                SalaryPeriodStep.period == data.period,
                SalaryPeriodStep.step_key.in_(prerequisite_keys),
                SalaryPeriodStep.is_confirmed == True
            ).count()
            if confirmed_prereqs < len(prerequisite_keys):
                unconfirmed = []
                for pk in prerequisite_keys:
                    rec = db.query(SalaryPeriodStep).filter(
                        SalaryPeriodStep.period == data.period,
                        SalaryPeriodStep.step_key == pk,
                        SalaryPeriodStep.is_confirmed == True
                    ).first()
                    if not rec:
                        pd = next(s for s in STEP_DEFINITIONS if s["key"] == pk)
                        unconfirmed.append(pd["title"])
                raise HTTPException(status_code=400, detail=f"请先确认以下步骤后再确认个税申报：{', '.join(unconfirmed)}")
            data_ready = True
        elif data.step_key == "salary":
            prerequisite_keys = ["employee", "attendance", "performance", "insurance", "tax"]
            confirmed_prereqs = db.query(SalaryPeriodStep).filter(
                SalaryPeriodStep.period == data.period,
                SalaryPeriodStep.step_key.in_(prerequisite_keys),
                SalaryPeriodStep.is_confirmed == True
            ).count()
            if confirmed_prereqs < len(prerequisite_keys):
                unconfirmed = []
                for pk in prerequisite_keys:
                    rec = db.query(SalaryPeriodStep).filter(
                        SalaryPeriodStep.period == data.period,
                        SalaryPeriodStep.step_key == pk,
                        SalaryPeriodStep.is_confirmed == True
                    ).first()
                    if not rec:
                        pd = next(s for s in STEP_DEFINITIONS if s["key"] == pk)
                        unconfirmed.append(pd["title"])
                raise HTTPException(status_code=400, detail=f"请先确认以下步骤后再确认薪资计算：{', '.join(unconfirmed)}")
            data_ready = True
        elif data.step_key == "payment":
            prerequisite_keys = ["employee", "attendance", "performance", "insurance", "tax", "salary"]
            confirmed_prereqs = db.query(SalaryPeriodStep).filter(
                SalaryPeriodStep.period == data.period,
                SalaryPeriodStep.step_key.in_(prerequisite_keys),
                SalaryPeriodStep.is_confirmed == True
            ).count()
            if confirmed_prereqs < len(prerequisite_keys):
                unconfirmed = []
                for pk in prerequisite_keys:
                    rec = db.query(SalaryPeriodStep).filter(
                        SalaryPeriodStep.period == data.period,
                        SalaryPeriodStep.step_key == pk,
                        SalaryPeriodStep.is_confirmed == True
                    ).first()
                    if not rec:
                        pd = next(s for s in STEP_DEFINITIONS if s["key"] == pk)
                        unconfirmed.append(pd["title"])
                raise HTTPException(status_code=400, detail=f"请先确认以下步骤后再确认工资发放：{', '.join(unconfirmed)}")
            salary_calcs = db.query(SalaryCalculation).filter(
                SalaryCalculation.period == data.period,
                SalaryCalculation.employee_id.in_(active_employee_ids)
            ).all()
            passed_count = sum(1 for c in salary_calcs if c.review_status == "审核通过")
            data_ready = passed_count >= total_employees and total_employees > 0

        is_force = not data_ready

    step_record = db.query(SalaryPeriodStep).filter(
        SalaryPeriodStep.period == data.period,
        SalaryPeriodStep.step_key == data.step_key
    ).first()

    if data.is_confirmed:
        if not step_record:
            step_record = SalaryPeriodStep(
                period=data.period,
                step_key=data.step_key,
            )
            db.add(step_record)

        step_record.is_confirmed = True
        step_record.confirmed_by = current_user.id
        step_record.confirmed_at = datetime.now()
        step_record.is_force_confirmed = is_force
        step_record.remark = data.remark
        action_desc = f"确认{step_def['title']}" + ("（数据不全，强制确认）" if is_force else "")
    else:
        if step_record:
            step_record.is_confirmed = False
            step_record.is_force_confirmed = False
            step_record.confirmed_by = None
            step_record.confirmed_at = None
            step_record.remark = None
        action_desc = f"取消确认{step_def['title']}"

    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "report", "edit", f"{action_desc} - {data.period}")

    return {"message": "操作成功", "is_confirmed": data.is_confirmed, "is_force": is_force}


@router.get("/roster", dependencies=[Depends(require_permission("report:export")), Depends(require_permission("employee:export"))])
def export_roster(db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    query = db.query(Employee)
    employees = _filter_active_without_pending_resign(query, db).order_by(Employee.employee_no).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "花名册"

    headers = [
        "序号", "姓名", "员工编号", "证件号码", "性别", "联系方式", "邮箱", "办公地点",
        "合同公司", "1级部门", "2级部门", "3级部门", "4级部门", "5级部门", "部门",
        "职位", "员工类型", "岗位职级", "岗位级别",
        "用工状态", "成本归属", "直属主管",
        "入职时间", "计划转正日期", "转正时间", "离职日期",
        "出生日期", "年龄", "民族", "婚姻状况", "子女情况", "政治面貌",
        "籍贯", "户籍类型", "户籍地址", "家庭住址",
        "首次参加工作时间", "司龄", "工龄",
        "学历", "毕业院校", "毕业时间", "所学专业",
        "资格证/职称证1", "资格证/职称证2",
        "紧急联系人姓名", "联系人关系", "联系人电话",
        "现合同起始日", "现合同到期日", "合同类型",
        "五险一金起购日期", "社保公积金购买地",
        "招聘渠道", "特长爱好", "商业保险类型", "外包类型",
        "银行卡号", "开户行", "开户行支行",
        "基本工资", "绩效奖金标准", "餐补", "交通补贴", "通讯补贴", "电脑补贴", "住房补贴",
        "月薪标准", "薪资生效日期", "备注"
    ]
    ws.append(headers)

    today = date.today()

    for idx, emp in enumerate(employees, 1):
        salary = db.query(EmployeeSalary).filter(
            EmployeeSalary.employee_id == emp.id
        ).order_by(EmployeeSalary.effective_date.desc()).first()

        dict_ids = [emp.contract_company_id, emp.department_id, emp.position_id, emp.status_id]
        dict_items = db.query(SysDictBase).filter(SysDictBase.id.in_(dict_ids)).all()
        name_map = {d.id: d.name for d in dict_items}

        age_val = ""
        if emp.birth_date:
            age_val = today.year - emp.birth_date.year - ((today.month, today.day) < (emp.birth_date.month, emp.birth_date.day))

        def calc_years_months(start_date):
            if not start_date:
                return ""
            years = today.year - start_date.year
            months = today.month - start_date.month
            if today.day < start_date.day:
                months -= 1
            if months < 0:
                years -= 1
                months += 12
            return f"{years}年{months}月"

        company_tenure = calc_years_months(emp.entry_date)
        work_tenure = calc_years_months(emp.first_work_date)

        monthly_standard = 0
        if salary:
            monthly_standard = round(
                float(salary.base_salary) + float(salary.performance_standard) +
                float(salary.meal_allowance or 0) + float(salary.transport_allowance or 0) +
                float(salary.communication_allowance or 0) + float(salary.computer_allowance or 0) +
                float(salary.housing_allowance or 0), 2
            )

        ws.append([
            idx, emp.name, emp.employee_no, emp.id_card or "", emp.gender or "",
            emp.phone or "", emp.email or "", emp.work_place or "",
            name_map.get(emp.contract_company_id, ""),
            emp.dept_level1 or "", emp.dept_level2 or "", emp.dept_level3 or "",
            emp.dept_level4 or "", emp.dept_level5 or "", name_map.get(emp.department_id, ""),
            name_map.get(emp.position_id, ""),
            emp.employee_type or "", emp.position_level or "", emp.job_level or "",
            name_map.get(emp.status_id, ""),
            emp.cost_owner or "", emp.report_manager or "",
            str(emp.entry_date) if emp.entry_date else "",
            str(emp.regular_date) if emp.regular_date else "",
            str(emp.regular_date) if emp.regular_date else "",
            str(emp.resign_date) if emp.resign_date else "",
            str(emp.birth_date) if emp.birth_date else "", age_val,
            emp.nation or "", emp.marital_status or "", emp.children_status or "",
            emp.political_status or "",
            emp.native_place or "", emp.residence_type or "",
            emp.census_address or "", emp.home_address or "",
            str(emp.first_work_date) if emp.first_work_date else "",
            company_tenure, work_tenure,
            emp.education or "", emp.graduate_school or "",
            str(emp.graduate_date) if emp.graduate_date else "", emp.major or "",
            emp.cert1 or "", emp.cert2 or "",
            emp.emergency_contact_name or "", emp.emergency_contact_relation or "",
            emp.emergency_contact_phone or "",
            str(emp.contract_start_date) if emp.contract_start_date else "",
            str(emp.contract_end_date) if emp.contract_end_date else "",
            emp.contract_type or "",
            str(emp.insurance_start_date) if emp.insurance_start_date else "",
            emp.insurance_location or "",
            emp.recruitment_channel or "", emp.hobby or "",
            emp.commercial_insurance_type or "", "",
            emp.bank_card or "", emp.bank_branch or "", emp.bank_branch_detail or "",
            float(salary.base_salary) if salary else "",
            float(salary.performance_standard) if salary else "",
            float(salary.meal_allowance) if salary else "",
            float(salary.transport_allowance) if salary else "",
            float(salary.communication_allowance) if salary else "",
            float(salary.computer_allowance) if salary else "",
            float(salary.housing_allowance) if salary else "",
            monthly_standard if salary else "",
            str(salary.effective_date) if salary and salary.effective_date else "",
            emp.remark or ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=roster.xlsx"}
    )


@router.get("/salary/{period}", dependencies=[Depends(require_permission("salary:export"))])
def export_salary(period: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    query = db.query(Employee)
    active_employees = _filter_active_without_pending_resign(query, db).order_by(Employee.employee_no).all()
    active_emp_map = {emp.id: emp for emp in active_employees}

    calcs = db.query(SalaryCalculation).filter(
        SalaryCalculation.period == period,
        SalaryCalculation.employee_id.in_(list(active_emp_map.keys()))
    ).order_by(SalaryCalculation.employee_id, SalaryCalculation.record_type).all()

    dict_name_map = _get_emp_dict_maps(active_employees, db)

    wb = Workbook()
    wb.remove(wb.active)
    headers = [
        "员工ID", "姓名", "合同公司", "发放公司", "记录类型", "部门", "职务", "成本归属", "状态",
        "基本工资", "绩效标准", "绩效系数", "实际绩效",
        "餐补", "交通补", "通讯补", "电脑补", "住房补", "补贴合计",
        "提成/奖金", "税前调整", "税后调整",
        "当月总计薪天数", "实际计薪天数", "出勤率",
        "总应发工资", "养老个人", "失业个人", "医疗个人", "社保个人合计", "公积金个人", "社保公积金合计",
        "上月未报税", "差旅未报税",
        "本月应扣个税", "实发工资", "本月实际工资报税金额",
        "实发离职补偿金", "未报税补偿金", "未报税年终奖", "实发年终奖",
        "核算状态", "数据完整性", "审核状态", "备注"
    ]

    company_groups = {}
    for c in calcs:
        emp = active_emp_map.get(c.employee_id)
        if not emp:
            continue
        company_name = c.pay_company_name or c.contract_company or "全部"
        if company_name not in company_groups:
            company_groups[company_name] = []
        company_groups[company_name].append((emp, c))

    if not company_groups:
        ws = wb.create_sheet(f"薪资核算表_{period}")
        ws.append(headers)
        for emp in active_employees:
            ws.append([
                emp.id, emp.name,
                dict_name_map.get(emp.contract_company_id, ""), "", "",
                dict_name_map.get(emp.department_id, ""),
                dict_name_map.get(emp.position_id, ""),
                emp.cost_owner or "",
                dict_name_map.get(emp.status_id, ""),
                0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                "", "", "", ""
            ])
    else:
        sorted_companies = sorted(company_groups.keys(), key=lambda x: (0 if "易玖" in x else 1, x))
        for company_name in sorted_companies:
            sheet_name = company_name[:28]
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
            records = company_groups[company_name]
            for emp, c in records:
                record_type_label = {"single": "单条", "contract": "合同公司", "payroll": "发放公司"}.get(c.record_type or "single", c.record_type)
                ws.append([
                    c.employee_id, emp.name,
                    c.contract_company or "", c.pay_company_name or "", record_type_label,
                    c.department or "", c.position or "", c.cost_owner or "", c.status or "",
                    float(c.base_salary or 0), float(c.performance_standard or 0),
                    float(c.performance_coefficient or 0), float(c.actual_performance or 0),
                    float(c.meal_allowance or 0), float(c.transport_allowance or 0),
                    float(c.communication_allowance or 0), float(c.computer_allowance or 0),
                    float(c.housing_allowance or 0), float(c.allowance_total or 0),
                    float(c.commission_bonus or 0), float(c.pretax_adjustment or 0),
                    float(c.posttax_adjustment or 0),
                    float(c.total_work_days or 0), float(c.actual_work_days or 0),
                    float(c.attendance_rate or 0),
                    float(c.gross_salary or 0),
                    float(c.pension_personal or 0), float(c.unemployment_personal or 0),
                    float(c.medical_personal or 0), float(c.social_insurance_personal or 0),
                    float(c.housing_fund_personal or 0), float(c.si_hf_total or 0),
                    float(c.last_month_untaxed or 0), float(c.travel_untaxed or 0),
                    float(c.tax_deduction) if c.tax_deduction is not None else "", float(c.net_salary or 0),
                    float(c.actual_taxable or 0),
                    float(c.severance_pay or 0), float(c.compensation_tax or 0),
                    float(c.year_end_bonus_untaxed or 0), float(c.year_end_bonus_net or 0),
                    c.calculation_status or "", c.data_completeness or "", c.review_status or "",
                    c.remark or ""
                ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_{period}.xlsx"}
    )


@router.get("/attendance/{period}", dependencies=[Depends(require_permission("report:export")), Depends(require_permission("attendance:export"))])
def export_attendance(period: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    # 获取所有在职员工
    query = db.query(Employee)
    active_employees = _filter_active_without_pending_resign(query, db).order_by(Employee.employee_no).all()
    active_employee_ids = [emp.id for emp in active_employees]

    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.period == period,
        AttendanceRecord.employee_id.in_(active_employee_ids)
    ).all()
    
    record_map = {r.employee_id: r for r in records}
    dict_name_map = _get_emp_dict_maps(active_employees, db)

    wb = Workbook()
    ws = wb.active
    ws.title = f"考勤表_{period}"

    headers = [
        "员工ID", "姓名", "部门", "职务",
        "当月总计薪天数", "实际计薪天数", "出勤率",
        "迟到次数", "早退次数", "缺卡次数",
        "病假天数", "事假天数", "年假天数", "其他假天数",
        "是否在家打卡", "需核实", "核实状态", "备注"
    ]
    ws.append(headers)

    for emp in active_employees:
        r = record_map.get(emp.id)
        if r:
            ws.append([
                r.employee_id, emp.name,
                dict_name_map.get(emp.department_id, ""),
                dict_name_map.get(emp.position_id, ""),
                float(r.total_work_days or 0), float(r.actual_work_days or 0),
                float(r.attendance_rate or 0),
                r.late_count or 0, r.early_leave_count or 0, r.missed_punch_count or 0,
                float(r.sick_leave_days or 0), float(r.personal_leave_days or 0),
                float(r.annual_leave_days or 0), float(r.other_leave_days or 0),
                "是" if r.is_home_checkin else "否",
                "是" if r.need_verify else "否",
                r.verify_status or "", r.remark or ""
            ])
        else:
            ws.append([
                emp.id, emp.name,
                dict_name_map.get(emp.department_id, ""),
                dict_name_map.get(emp.position_id, ""),
                0, 0, 0,
                0, 0, 0, 0, 0, 0, 0,
                "否", "否", "", ""
            ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_{period}.xlsx"}
    )


# ============================================================
# 导出表配置相关
# ============================================================

SALARY_FIELDS = [
    {"key": "employee_no", "label": "员工编号", "width": 90},
    {"key": "contract_company", "label": "合同公司", "width": 120},
    {"key": "employee_name", "label": "姓名", "width": 70},
    {"key": "department", "label": "部门", "width": 90},
    {"key": "position", "label": "职务", "width": 90},
    {"key": "cost_owner", "label": "费用负责人", "width": 90},
    {"key": "status", "label": "状态", "width": 80},
    {"key": "entry_date", "label": "入职时间", "width": 100},
    {"key": "total_work_days", "label": "当月总计薪天数", "width": 105},
    {"key": "actual_work_days", "label": "实际计薪天数", "width": 95},
    {"key": "attendance_rate", "label": "出勤率", "width": 75},
    {"key": "base_salary", "label": "基本工资", "width": 95},
    {"key": "commission_bonus", "label": "提成/项目奖金/补发", "width": 95},
    {"key": "meal_allowance", "label": "餐补", "width": 70},
    {"key": "transport_allowance", "label": "交通补", "width": 75},
    {"key": "communication_allowance", "label": "通讯补", "width": 75},
    {"key": "computer_allowance", "label": "电脑补贴（非固定收入）", "width": 95},
    {"key": "housing_allowance", "label": "住房补（非固定收入）", "width": 95},
    {"key": "allowance_total", "label": "补贴合计", "width": 85},
    {"key": "performance_standard", "label": "绩效奖金标准", "width": 95},
    {"key": "performance_coefficient", "label": "实发绩效奖金系数", "width": 95},
    {"key": "actual_performance", "label": "实发绩效奖金标准", "width": 95},
    {"key": "effective_performance", "label": "实发绩效奖金", "width": 95},
    {"key": "monthly_standard", "label": "月薪标准", "width": 90},
    {"key": "gross_salary", "label": "总应发工资", "width": 100},
    {"key": "pretax_adjustment", "label": "税前调整金额", "width": 90},
    {"key": "pretax_adjustment_reason", "label": "税前调整原因", "width": 120},
    {"key": "pension_personal", "label": "养老保险（个人）", "width": 95},
    {"key": "unemployment_personal", "label": "失业保险（个人）", "width": 95},
    {"key": "medical_personal", "label": "医疗保险（个人）", "width": 95},
    {"key": "housing_fund_personal", "label": "公积金（个人）", "width": 90},
    {"key": "si_hf_total", "label": "社保、公积金（个人）合计", "width": 100},
    {"key": "salary_after_si_hf", "label": "扣掉社保公积金工资", "width": 110},
    {"key": "tax_deduction", "label": "本月应扣个税额", "width": 100},
    {"key": "posttax_adjustment", "label": "税后调整金额", "width": 90},
    {"key": "posttax_adjustment_reason", "label": "税后调整原因", "width": 120},
    {"key": "net_salary", "label": "实发工资", "width": 100},
    {"key": "last_month_untaxed", "label": "上月未报税金额", "width": 100},
    {"key": "travel_untaxed", "label": "临时性差旅补贴未报税费用", "width": 100},
    {"key": "actual_taxable", "label": "本月实际工资报税金额", "width": 120},
    {"key": "severance_pay", "label": "实发离职补偿金", "width": 100},
    {"key": "compensation_tax", "label": "未报税补偿金", "width": 100},
    {"key": "year_end_bonus_untaxed", "label": "未报税年终奖", "width": 100},
    {"key": "year_end_bonus_net", "label": "实发年终奖", "width": 100},
    {"key": "remark", "label": "备注", "width": 200},
]

ROSTER_FIELDS = [
    {"key": "idx", "label": "序号", "width": 60},
    {"key": "name", "label": "姓名", "width": 70},
    {"key": "employee_no", "label": "员工编号", "width": 90},
    {"key": "id_card", "label": "证件号码", "width": 180},
    {"key": "gender", "label": "性别", "width": 50},
    {"key": "phone", "label": "联系方式", "width": 110},
    {"key": "email", "label": "邮箱", "width": 150},
    {"key": "work_place", "label": "办公地点", "width": 100},
    {"key": "contract_company", "label": "合同公司", "width": 120},
    {"key": "dept_level1", "label": "1级部门", "width": 100},
    {"key": "dept_level2", "label": "2级部门", "width": 100},
    {"key": "dept_level3", "label": "3级部门", "width": 100},
    {"key": "dept_level4", "label": "4级部门", "width": 100},
    {"key": "dept_level5", "label": "5级部门", "width": 100},
    {"key": "department", "label": "部门", "width": 90},
    {"key": "position", "label": "职位", "width": 90},
    {"key": "employee_type", "label": "员工类型", "width": 80},
    {"key": "position_level", "label": "岗位职级", "width": 80},
    {"key": "job_level", "label": "岗位级别", "width": 80},
    {"key": "status", "label": "用工状态", "width": 80},
    {"key": "cost_owner", "label": "成本归属", "width": 90},
    {"key": "report_manager", "label": "直属主管", "width": 80},
    {"key": "entry_date", "label": "入职时间", "width": 100},
    {"key": "regular_date", "label": "计划转正日期", "width": 110},
    {"key": "regular_confirm_date", "label": "转正时间", "width": 100},
    {"key": "resign_date", "label": "离职日期", "width": 100},
    {"key": "birth_date", "label": "出生日期", "width": 100},
    {"key": "age", "label": "年龄", "width": 50},
    {"key": "nation", "label": "民族", "width": 60},
    {"key": "marital_status", "label": "婚姻状况", "width": 70},
    {"key": "children_status", "label": "子女情况", "width": 80},
    {"key": "political_status", "label": "政治面貌", "width": 80},
    {"key": "native_place", "label": "籍贯", "width": 100},
    {"key": "residence_type", "label": "户籍类型", "width": 80},
    {"key": "census_address", "label": "户籍地址", "width": 200},
    {"key": "home_address", "label": "家庭住址", "width": 200},
    {"key": "first_work_date", "label": "首次参加工作时间", "width": 120},
    {"key": "company_tenure", "label": "司龄", "width": 70},
    {"key": "work_tenure", "label": "工龄", "width": 70},
    {"key": "education", "label": "学历", "width": 60},
    {"key": "graduate_school", "label": "毕业院校", "width": 150},
    {"key": "graduate_date", "label": "毕业时间", "width": 100},
    {"key": "major", "label": "所学专业", "width": 100},
    {"key": "cert1", "label": "资格证/职称证1", "width": 150},
    {"key": "cert2", "label": "资格证/职称证2", "width": 150},
    {"key": "emergency_contact_name", "label": "紧急联系人姓名", "width": 120},
    {"key": "emergency_contact_relation", "label": "联系人关系", "width": 90},
    {"key": "emergency_contact_phone", "label": "联系人电话", "width": 110},
    {"key": "contract_start_date", "label": "现合同起始日", "width": 110},
    {"key": "contract_end_date", "label": "现合同到期日", "width": 110},
    {"key": "contract_type", "label": "合同类型", "width": 80},
    {"key": "insurance_start_date", "label": "五险一金起购日期", "width": 130},
    {"key": "insurance_location", "label": "社保公积金购买地", "width": 130},
    {"key": "recruitment_channel", "label": "招聘渠道", "width": 100},
    {"key": "hobby", "label": "特长爱好", "width": 100},
    {"key": "commercial_insurance_type", "label": "商业保险类型", "width": 110},
    {"key": "bank_card", "label": "银行卡号", "width": 180},
    {"key": "bank_branch", "label": "开户行", "width": 120},
    {"key": "bank_branch_detail", "label": "开户行支行", "width": 150},
    {"key": "base_salary", "label": "基本工资", "width": 95},
    {"key": "performance_standard", "label": "绩效奖金标准", "width": 110},
    {"key": "meal_allowance", "label": "餐补", "width": 70},
    {"key": "transport_allowance", "label": "交通补贴", "width": 80},
    {"key": "communication_allowance", "label": "通讯补贴", "width": 80},
    {"key": "computer_allowance", "label": "电脑补贴", "width": 80},
    {"key": "housing_allowance", "label": "住房补贴", "width": 80},
    {"key": "monthly_standard", "label": "月薪标准", "width": 90},
    {"key": "salary_effective_date", "label": "薪资生效日期", "width": 110},
    {"key": "remark", "label": "备注", "width": 200},
]

ATTENDANCE_FIELDS = [
    {"key": "employee_no", "label": "员工编号", "width": 90},
    {"key": "name", "label": "姓名", "width": 70},
    {"key": "department", "label": "部门", "width": 90},
    {"key": "position", "label": "职务", "width": 90},
    {"key": "standard_days", "label": "应出勤天数", "width": 100},
    {"key": "actual_days", "label": "实际出勤天数", "width": 110},
    {"key": "attendance_rate", "label": "出勤率", "width": 75},
    {"key": "late_count", "label": "迟到次数", "width": 80},
    {"key": "early_leave_count", "label": "早退次数", "width": 80},
    {"key": "absent_days", "label": "旷工天数", "width": 80},
    {"key": "personal_leave_days", "label": "事假天数", "width": 80},
    {"key": "sick_leave_days", "label": "病假天数", "width": 80},
    {"key": "annual_leave_days", "label": "年假天数", "width": 80},
    {"key": "overtime_hours", "label": "加班小时数", "width": 90},
    {"key": "travel_days", "label": "出差天数", "width": 80},
    {"key": "remark", "label": "备注", "width": 200},
]

SOCIAL_INSURANCE_FIELDS = [
    {"key": "employee_no", "label": "员工编号", "width": 90},
    {"key": "name", "label": "姓名", "width": 70},
    {"key": "contract_company", "label": "合同公司", "width": 120},
    {"key": "department", "label": "部门", "width": 90},
    {"key": "position", "label": "职务", "width": 90},
    {"key": "si_base", "label": "社保基数", "width": 90},
    {"key": "hf_base", "label": "公积金基数", "width": 100},
    {"key": "pension_personal", "label": "养老（个人）", "width": 100},
    {"key": "pension_company", "label": "养老（公司）", "width": 100},
    {"key": "medical_personal", "label": "医疗（个人）", "width": 100},
    {"key": "medical_company", "label": "医疗（公司）", "width": 100},
    {"key": "unemployment_personal", "label": "失业（个人）", "width": 100},
    {"key": "unemployment_company", "label": "失业（公司）", "width": 100},
    {"key": "injury_company", "label": "工伤（公司）", "width": 100},
    {"key": "maternity_company", "label": "生育（公司）", "width": 100},
    {"key": "housing_fund_personal", "label": "公积金（个人）", "width": 110},
    {"key": "housing_fund_company", "label": "公积金（公司）", "width": 110},
    {"key": "personal_total", "label": "个人合计", "width": 90},
    {"key": "company_total", "label": "公司合计", "width": 90},
    {"key": "total_amount", "label": "五险一金合计", "width": 110},
]


class ExportTemplateIn(BaseModel):
    name: str
    template_type: str
    description: Optional[str] = None
    fields: List[dict]
    is_default: bool = False
    is_enabled: bool = True


class ExportTemplateOut(BaseModel):
    id: int
    name: str
    template_type: str
    description: Optional[str] = None
    fields: List[dict]
    is_default: bool
    is_enabled: bool
    created_at: str

    class Config:
        from_attributes = True


@router.get("/export/available-fields", dependencies=[Depends(require_permission("report:view"))])
def get_available_fields(current_user: UserInfo = Depends(get_current_user)):
    """获取可配置的导出字段列表"""
    return {
        "salary": SALARY_FIELDS,
        "roster": ROSTER_FIELDS,
        "attendance": ATTENDANCE_FIELDS,
        "social_insurance": SOCIAL_INSURANCE_FIELDS,
    }


@router.get("/export/templates", response_model=List[ExportTemplateOut], dependencies=[Depends(require_permission("report:view"))])
def get_export_templates(
    template_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """获取导出模板列表"""
    query = db.query(ExportTemplate)
    if template_type:
        query = query.filter(ExportTemplate.template_type == template_type)
    templates = query.order_by(ExportTemplate.is_default.desc(), ExportTemplate.created_at.desc()).all()
    return [
        ExportTemplateOut(
            id=t.id, name=t.name, template_type=t.template_type,
            description=t.description, fields=t.fields,
            is_default=t.is_default, is_enabled=t.is_enabled,
            created_at=str(t.created_at)
        ) for t in templates
    ]


@router.post("/export/templates", response_model=ExportTemplateOut, dependencies=[Depends(require_permission("report:export"))])
def create_export_template(
    data: ExportTemplateIn,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """创建导出模板"""
    if data.is_default:
        db.query(ExportTemplate).filter(
            ExportTemplate.template_type == data.template_type,
            ExportTemplate.is_default == True
        ).update({"is_default": False})
    
    template = ExportTemplate(
        name=data.name,
        template_type=data.template_type,
        description=data.description,
        fields=data.fields,
        is_default=data.is_default,
        is_enabled=data.is_enabled,
        created_by=current_user.id
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    write_log(db, "data_change", current_user.id, current_user.username, "report", "create", f"创建导出模板 {data.name}")
    return ExportTemplateOut(
        id=template.id, name=template.name, template_type=template.template_type,
        description=template.description, fields=template.fields,
        is_default=template.is_default, is_enabled=template.is_enabled,
        created_at=str(template.created_at)
    )


@router.put("/export/templates/{tpl_id}", response_model=ExportTemplateOut, dependencies=[Depends(require_permission("report:export"))])
def update_export_template(
    tpl_id: int,
    data: ExportTemplateIn,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """更新导出模板"""
    template = db.query(ExportTemplate).filter(ExportTemplate.id == tpl_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    
    if data.is_default and not template.is_default:
        db.query(ExportTemplate).filter(
            ExportTemplate.template_type == data.template_type,
            ExportTemplate.is_default == True,
            ExportTemplate.id != tpl_id
        ).update({"is_default": False})
    
    template.name = data.name
    template.template_type = data.template_type
    template.description = data.description
    template.fields = data.fields
    template.is_default = data.is_default
    template.is_enabled = data.is_enabled
    db.commit()
    db.refresh(template)
    write_log(db, "data_change", current_user.id, current_user.username, "report", "edit", f"更新导出模板 {data.name}")
    return ExportTemplateOut(
        id=template.id, name=template.name, template_type=template.template_type,
        description=template.description, fields=template.fields,
        is_default=template.is_default, is_enabled=template.is_enabled,
        created_at=str(template.created_at)
    )


@router.delete("/export/templates/{tpl_id}", dependencies=[Depends(require_permission("report:export"))])
def delete_export_template(
    tpl_id: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """删除导出模板"""
    template = db.query(ExportTemplate).filter(ExportTemplate.id == tpl_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    if template.is_default:
        raise HTTPException(status_code=400, detail="默认模板不能删除，请先设置其他模板为默认")
    name = template.name
    db.delete(template)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "report", "delete", f"删除导出模板 {name}")
    return {"message": "模板已删除"}


def _get_emp_dict_maps(emps, db):
    """批量获取员工的字典映射（合同公司、部门、职位、状态）"""
    dict_ids = set()
    for emp in emps:
        if emp.contract_company_id:
            dict_ids.add(emp.contract_company_id)
        if emp.department_id:
            dict_ids.add(emp.department_id)
        if emp.position_id:
            dict_ids.add(emp.position_id)
        if emp.status_id:
            dict_ids.add(emp.status_id)
    
    name_map = {}
    if dict_ids:
        dict_items = db.query(SysDictBase).filter(SysDictBase.id.in_(list(dict_ids))).all()
        name_map = {d.id: d.name for d in dict_items}
    return name_map


def _safe_float(val, default=0.0):
    """安全转换为浮点数"""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _build_salary_row_data(emp, calc, dict_name_map):
    """构建单条薪资导出行数据，兼容无核算记录的情况"""
    row_data = {}
    row_data["employee_no"] = getattr(emp, 'employee_no', "") if emp else ""
    row_data["employee_name"] = getattr(emp, 'name', "") if emp else ""
    
    if calc:
        row_data["contract_company"] = getattr(calc, 'contract_company', "") or ""
        row_data["department"] = getattr(calc, 'department', "") or ""
        row_data["position"] = getattr(calc, 'position', "") or ""
        row_data["cost_owner"] = getattr(calc, 'cost_owner', "") or ""
        row_data["status"] = getattr(calc, 'status', "") or ""
        entry_date = getattr(calc, 'entry_date', None)
        row_data["entry_date"] = str(entry_date) if entry_date else ""
        row_data["total_work_days"] = _safe_float(getattr(calc, 'total_work_days', 0))
        row_data["actual_work_days"] = _safe_float(getattr(calc, 'actual_work_days', 0))
        row_data["attendance_rate"] = _safe_float(getattr(calc, 'attendance_rate', 0))
        row_data["base_salary"] = _safe_float(getattr(calc, 'base_salary', 0))
        row_data["commission_bonus"] = _safe_float(getattr(calc, 'commission_bonus', 0))
        row_data["meal_allowance"] = _safe_float(getattr(calc, 'meal_allowance', 0))
        row_data["transport_allowance"] = _safe_float(getattr(calc, 'transport_allowance', 0))
        row_data["communication_allowance"] = _safe_float(getattr(calc, 'communication_allowance', 0))
        row_data["computer_allowance"] = _safe_float(getattr(calc, 'computer_allowance', 0))
        row_data["housing_allowance"] = _safe_float(getattr(calc, 'housing_allowance', 0))
        row_data["allowance_total"] = _safe_float(getattr(calc, 'allowance_total', 0))
        row_data["performance_standard"] = _safe_float(getattr(calc, 'performance_standard', 0))
        row_data["performance_coefficient"] = _safe_float(getattr(calc, 'performance_coefficient', 0))
        row_data["actual_performance"] = _safe_float(getattr(calc, 'actual_performance', 0))
        row_data["effective_performance"] = _safe_float(getattr(calc, 'effective_performance', 0))
        row_data["monthly_standard"] = _safe_float(getattr(calc, 'monthly_standard', 0))
        row_data["gross_salary"] = _safe_float(getattr(calc, 'gross_salary', 0))
        row_data["pretax_adjustment"] = _safe_float(getattr(calc, 'pretax_adjustment', 0))
        row_data["pretax_adjustment_reason"] = getattr(calc, 'pretax_adjustment_reason', "") or ""
        row_data["pension_personal"] = _safe_float(getattr(calc, 'pension_personal', 0))
        row_data["unemployment_personal"] = _safe_float(getattr(calc, 'unemployment_personal', 0))
        row_data["medical_personal"] = _safe_float(getattr(calc, 'medical_personal', 0))
        row_data["housing_fund_personal"] = _safe_float(getattr(calc, 'housing_fund_personal', 0))
        row_data["social_insurance_personal"] = _safe_float(getattr(calc, 'social_insurance_personal', 0))
        row_data["si_hf_total"] = _safe_float(getattr(calc, 'si_hf_total', 0))
        salary_after_si_hf = (
            _safe_float(getattr(calc, 'gross_salary', 0)) + 
            _safe_float(getattr(calc, 'pretax_adjustment', 0)) - 
            _safe_float(getattr(calc, 'si_hf_total', 0))
        )
        row_data["salary_after_si_hf"] = round(salary_after_si_hf, 2)
        tax_val = getattr(calc, 'tax_deduction', None)
        row_data["tax_deduction"] = _safe_float(tax_val, None) if tax_val is not None else None
        row_data["posttax_adjustment"] = _safe_float(getattr(calc, 'posttax_adjustment', 0))
        row_data["posttax_adjustment_reason"] = getattr(calc, 'posttax_adjustment_reason', "") or ""
        row_data["severance_pay"] = _safe_float(getattr(calc, 'severance_pay', 0))
        row_data["year_end_bonus_untaxed"] = _safe_float(getattr(calc, 'year_end_bonus_untaxed', 0))
        row_data["year_end_bonus_net"] = _safe_float(getattr(calc, 'year_end_bonus_net', 0))
        row_data["net_salary"] = _safe_float(getattr(calc, 'net_salary', 0))
        row_data["last_month_untaxed"] = _safe_float(getattr(calc, 'last_month_untaxed', 0))
        row_data["travel_untaxed"] = _safe_float(getattr(calc, 'travel_untaxed', 0))
        row_data["compensation_tax"] = _safe_float(getattr(calc, 'compensation_tax', 0))
        row_data["actual_taxable"] = _safe_float(getattr(calc, 'actual_taxable', 0))
        row_data["data_completeness"] = getattr(calc, 'data_completeness', "") or ""
        row_data["calculation_status"] = getattr(calc, 'calculation_status', "") or ""
        row_data["review_status"] = getattr(calc, 'review_status', "") or ""
        row_data["remark"] = getattr(calc, 'remark', "") or ""
    else:
        row_data["contract_company"] = dict_name_map.get(getattr(emp, 'contract_company_id', None), "") if emp else ""
        row_data["department"] = dict_name_map.get(getattr(emp, 'department_id', None), "") if emp else ""
        row_data["position"] = dict_name_map.get(getattr(emp, 'position_id', None), "") if emp else ""
        row_data["cost_owner"] = getattr(emp, 'cost_owner', "") if emp else ""
        row_data["status"] = dict_name_map.get(getattr(emp, 'status_id', None), "") if emp else ""
        emp_entry = getattr(emp, 'entry_date', None) if emp else None
        row_data["entry_date"] = str(emp_entry) if emp_entry else ""
        row_data["total_work_days"] = 0
        row_data["actual_work_days"] = 0
        row_data["attendance_rate"] = 0
        row_data["base_salary"] = 0
        row_data["commission_bonus"] = 0
        row_data["meal_allowance"] = 0
        row_data["transport_allowance"] = 0
        row_data["communication_allowance"] = 0
        row_data["computer_allowance"] = 0
        row_data["housing_allowance"] = 0
        row_data["allowance_total"] = 0
        row_data["performance_standard"] = 0
        row_data["performance_coefficient"] = 0
        row_data["actual_performance"] = 0
        row_data["effective_performance"] = 0
        row_data["monthly_standard"] = 0
        row_data["gross_salary"] = 0
        row_data["pretax_adjustment"] = 0
        row_data["pretax_adjustment_reason"] = ""
        row_data["pension_personal"] = 0
        row_data["unemployment_personal"] = 0
        row_data["medical_personal"] = 0
        row_data["housing_fund_personal"] = 0
        row_data["social_insurance_personal"] = 0
        row_data["si_hf_total"] = 0
        row_data["salary_after_si_hf"] = 0
        row_data["tax_deduction"] = None
        row_data["posttax_adjustment"] = 0
        row_data["posttax_adjustment_reason"] = ""
        row_data["severance_pay"] = 0
        row_data["year_end_bonus_untaxed"] = 0
        row_data["year_end_bonus_net"] = 0
        row_data["net_salary"] = 0
        row_data["last_month_untaxed"] = 0
        row_data["travel_untaxed"] = 0
        row_data["compensation_tax"] = 0
        row_data["actual_taxable"] = 0
        row_data["data_completeness"] = ""
        row_data["calculation_status"] = ""
        row_data["review_status"] = ""
        row_data["remark"] = ""
    
    return row_data


@router.get("/salary-by-template/{period}", dependencies=[Depends(require_permission("report:export")), Depends(require_permission("salary:export"))])
def export_salary_by_template(
    period: str,
    template_id: Optional[int] = Query(None, description="模板ID，不传则使用默认模板"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """按自定义模板导出薪资表（包含所有在职员工，无核算记录则字段留空）"""
    query = db.query(Employee)
    active_employees = _filter_active_without_pending_resign(query, db).order_by(Employee.employee_no).all()
    active_employee_ids = [emp.id for emp in active_employees]

    calcs = db.query(SalaryCalculation).filter(
        SalaryCalculation.period == period,
        SalaryCalculation.employee_id.in_(active_employee_ids)
    ).all()
    
    calc_map = {c.employee_id: c for c in calcs}
    dict_name_map = _get_emp_dict_maps(active_employees, db)
    
    if template_id:
        template = db.query(ExportTemplate).filter(ExportTemplate.id == template_id).first()
    else:
        template = db.query(ExportTemplate).filter(
            ExportTemplate.template_type.in_(["salary_finance", "salary_slip", "tax", "custom"]),
            ExportTemplate.is_default == True,
            ExportTemplate.is_enabled == True
        ).first()
    
    field_list = SALARY_FIELDS
    if template:
        field_list = template.fields
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"薪资表_{period}"
    
    headers = [f["label"] for f in field_list]
    ws.append(headers)
    
    field_keys = [f["key"] for f in field_list]
    
    for emp in active_employees:
        calc = calc_map.get(emp.id)
        row_data = _build_salary_row_data(emp, calc, dict_name_map)
        ws.append([row_data.get(k, "") for k in field_keys])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"薪资表_{period}.xlsx"
    if template:
        filename = f"{template.name}_{period}.xlsx"
    
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/contract-expiry-warning", dependencies=[Depends(require_permission("report:contract_warning_view"))])
def get_contract_expiry_warning(
    days_ahead: int = Query(30, description="提前预警天数，默认30天"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """获取合同到期预警员工列表"""
    today = date.today()
    warning_date = today + timedelta(days=days_ahead)
    
    query = db.query(Employee)
    active_employees = _filter_active_without_pending_resign(query, db).all()
    
    dept_ids = [emp.department_id for emp in active_employees if emp.department_id]
    dept_map = {}
    if dept_ids:
        dept_items = db.query(SysDictBase).filter(SysDictBase.id.in_(dept_ids)).all()
        dept_map = {d.id: d.name for d in dept_items}
    
    warning_list = []
    for emp in active_employees:
        if emp.contract_end_date and emp.contract_end_date <= warning_date:
            days_remaining = (emp.contract_end_date - today).days
            dept_name = dept_map.get(emp.department_id, "")
            display_text = f"{emp.name}-{dept_name}-{emp.contract_end_date}"
            warning_list.append({
                "id": emp.id,
                "name": emp.name,
                "employee_no": emp.employee_no,
                "department": dept_name,
                "contract_end_date": str(emp.contract_end_date),
                "days_remaining": days_remaining,
                "display_text": display_text,
                "is_expired": days_remaining < 0
            })
    
    warning_list.sort(key=lambda x: x["contract_end_date"])
    
    return {
        "days_ahead": days_ahead,
        "today": str(today),
        "warning_date": str(warning_date),
        "total_count": len(warning_list),
        "expired_count": sum(1 for w in warning_list if w["is_expired"]),
        "list": warning_list
    }


@router.get("/social-insurance/{period}", dependencies=[Depends(require_permission("report:export")), Depends(require_permission("insurance:export"))])
def export_social_insurance_report(
    period: str,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """报表中心专用：社保公积金导出（需同时拥有报表导出和社保导出权限）"""
    query = db.query(Employee)
    employees = _filter_active_without_pending_resign(query, db).order_by(Employee.employee_no).all()
    si_map = {}
    records = db.query(SocialInsurance).filter(SocialInsurance.period == period).all()
    for r in records:
        si_map[r.employee_id] = r

    wb = Workbook()
    ws = wb.active
    ws.title = f"社保公积金_{period}"
    headers = [
        "员工编号", "姓名",
        "养老保险基数(个人)", "养老保险基数(单位)", "养老保险个人", "养老保险公司",
        "失业保险基数(个人)", "失业保险基数(单位)", "失业保险个人", "失业保险公司",
        "医疗保险基数(个人)", "医疗保险基数(单位)", "医疗保险个人", "医疗保险公司",
        "工伤保险基数(单位)", "工伤保险公司",
        "社保个人合计", "社保公司合计",
        "公积金基数", "公积金个人", "公积金公司",
    ]
    ws.append(headers)

    for emp in employees:
        si = si_map.get(emp.id)
        ws.append([
            emp.employee_no, emp.name,
            float(si.pension_personal_base or 0) if si else "",
            float(si.pension_company_base or 0) if si else "",
            float(si.pension_personal or 0) if si else "",
            float(si.pension_company or 0) if si else "",
            float(si.unemployment_personal_base or 0) if si else "",
            float(si.unemployment_company_base or 0) if si else "",
            float(si.unemployment_personal or 0) if si else "",
            float(si.unemployment_company or 0) if si else "",
            float(si.medical_personal_base or 0) if si else "",
            float(si.medical_company_base or 0) if si else "",
            float(si.medical_personal or 0) if si else "",
            float(si.medical_company or 0) if si else "",
            float(si.injury_company_base or 0) if si else "",
            float(si.injury_company or 0) if si else "",
            float(si.si_personal) if si else "",
            float(si.si_company) if si else "",
            float(si.hf_base) if si else "",
            float(si.hf_personal) if si else "",
            float(si.hf_company) if si else "",
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"社保公积金_{period}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )