from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date, timedelta
from io import BytesIO
from openpyxl import Workbook, load_workbook
from urllib.parse import quote
import xlrd
import uuid
from app.core.database import get_db
from app.core.log_helper import write_log
from app.core.query_utils import filter_active_employees, apply_data_scope
from app.services.work_calendar import get_month_salary_days
from app.models.models import (
    Employee, EmployeeSalary, EmployeeSalaryAdjustment, AttendanceRecord, PerformanceScore,
    SocialInsurance, LegacyAdjustment, TravelReimbursement, LaborCompensation,
    SalaryCalculation, CalculationLog, SysDictBase
)
from app.api.auth import get_current_user, UserInfo, require_permission

router = APIRouter()


def _get_period_end(period: str) -> date:
    if '-' in period:
        year, month = map(int, period.split('-'))
    else:
        year = int(period[:4])
        month = int(period[4:6])
    if month == 12:
        return date(year + 1, 1, 1) - timedelta(days=1)
    return date(year, month + 1, 1) - timedelta(days=1)


def _ensure_period_calculations(db: Session, period: str, hide_status_id: Optional[int] = None):
    """[已废弃] 不再自动创建记录，仅返回当前已有记录数。正式算薪请使用 calculate_salary API。"""
    existing_calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    return len(existing_calcs)


def _safe_float(val, default=None):
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _get_salary_split_config(db, dict_items):
    """获取工资拆分配置
    返回: (yijiu_company_dict, split_config_map)
    - yijiu_company_dict: 北京易玖智能科技有限公司的字典项
    - split_config_map: {company_id: target_company_dict} 配置了拆分的公司映射
    默认规则：未配置的情况下，北京易玖和上海瑞方不拆分，其他公司默认拆分到北京易玖
    """
    yijiu_company = None
    for d in dict_items:
        if '北京易玖智能科技' in d.name and d.category == 'contract_company':
            yijiu_company = d
            break

    split_config = {}

    for d in dict_items:
        if d.category != 'contract_company' or not d.is_enabled:
            continue
        extra = d.extra or {}
        split_target = extra.get('salary_split_target')

        if split_target == 'none':
            continue
        elif split_target and split_target != 'none':
            try:
                target_id = int(split_target)
                target = db.query(SysDictBase).filter(SysDictBase.id == target_id).first()
                if target:
                    split_config[d.id] = target
                continue
            except (ValueError, TypeError):
                pass

        if '北京易玖智能科技' in d.name or '上海瑞方' in d.name:
            continue

        if yijiu_company:
            split_config[d.id] = yijiu_company

    return yijiu_company, split_config


class DataSourceStatus(BaseModel):
    source_name: str
    source_key: str
    status: str
    count: int
    missing_employees: List[str] = []


class DataCompletenessOut(BaseModel):
    period: str
    total_employees: int
    complete_count: int
    missing_count: int
    optional_missing_count: int
    sources: List[DataSourceStatus] = []


class SalaryCalcOut(BaseModel):
    id: Optional[int] = None
    period: str
    employee_id: int
    employee_no: str = ""
    employee_name: str = ""
    contract_company: str = ""
    pay_company_id: Optional[int] = None
    pay_company_name: str = ""
    record_type: str = "single"
    department: str = ""
    position: str = ""
    cost_owner: str = ""
    status: str = ""
    entry_date: Optional[str] = None
    base_salary: Optional[float] = None
    base_salary_prorated: Optional[float] = None
    performance_standard_prorated: Optional[float] = None
    adjustment_id: Optional[int] = None
    monthly_standard: Optional[float] = None
    performance_standard: Optional[float] = None
    performance_coefficient: Optional[float] = None
    actual_performance: Optional[float] = None
    effective_performance: Optional[float] = None
    meal_allowance: Optional[float] = None
    transport_allowance: Optional[float] = None
    communication_allowance: Optional[float] = None
    computer_allowance: Optional[float] = None
    housing_allowance: Optional[float] = None
    allowance_total: Optional[float] = None
    commission_bonus: Optional[float] = None
    pretax_adjustment: Optional[float] = None
    pretax_adjustment_reason: Optional[str] = None
    posttax_adjustment: Optional[float] = None
    posttax_adjustment_reason: Optional[str] = None
    total_work_days: Optional[float] = None
    actual_work_days: Optional[float] = None
    attendance_rate: Optional[float] = None
    gross_salary: Optional[float] = None
    pension_personal: Optional[float] = None
    unemployment_personal: Optional[float] = None
    medical_personal: Optional[float] = None
    social_insurance_personal: Optional[float] = None
    housing_fund_personal: Optional[float] = None
    si_hf_total: Optional[float] = None
    tax_deduction: Optional[float] = None
    net_salary: Optional[float] = None
    last_month_untaxed: Optional[float] = None
    travel_untaxed: Optional[float] = None
    compensation_tax: Optional[float] = None
    severance_pay: Optional[float] = None
    year_end_bonus_untaxed: Optional[float] = None
    year_end_bonus_net: Optional[float] = None
    actual_taxable: Optional[float] = None
    special_deduction: Optional[float] = None
    salary_after_si_hf: Optional[float] = None
    remark: Optional[str] = None
    review_status: str = ""
    calculation_status: str = ""
    data_completeness: str = ""

    class Config:
        from_attributes = True


class SalaryCalcUpdate(BaseModel):
    commission_bonus: Optional[float] = None
    pretax_adjustment: Optional[float] = None
    pretax_adjustment_reason: Optional[str] = None
    posttax_adjustment: Optional[float] = None
    posttax_adjustment_reason: Optional[str] = None
    performance_coefficient: Optional[float] = None
    tax_deduction: Optional[float] = None
    special_deduction: Optional[float] = None
    last_month_untaxed: Optional[float] = None
    travel_untaxed: Optional[float] = None
    compensation_tax: Optional[float] = None
    severance_pay: Optional[float] = None
    year_end_bonus_untaxed: Optional[float] = None
    year_end_bonus_net: Optional[float] = None
    remark: Optional[str] = None
    review_status: Optional[str] = None


class CalculationSummary(BaseModel):
    period: str
    total_employees: int
    success_count: int
    failed_count: int
    total_gross_salary: float
    avg_gross_salary: float
    total_net_salary: float
    avg_net_salary: float
    batch_no: str


@router.get("/periods", dependencies=[Depends(require_permission("salary:view"))])
def get_available_periods(db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    periods = db.query(SalaryCalculation.period).distinct().order_by(SalaryCalculation.period.desc()).all()
    return [p[0] for p in periods]


@router.get("/check-completeness/{period}", response_model=DataCompletenessOut, dependencies=[Depends(require_permission("salary:view"))])
def check_data_completeness(period: str, hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"), db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    query = db.query(Employee)
    active_employees = filter_active_employees(query, db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()

    sources = []
    missing_map = {}

    period_end = _get_period_end(period)
    salaries = {s.employee_id: s for s in db.query(EmployeeSalary).filter(
        EmployeeSalary.employee_id.in_([e.id for e in active_employees]),
        EmployeeSalary.effective_date <= period_end
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).all()}

    attendance = {a.employee_id: a for a in db.query(AttendanceRecord).filter(
        AttendanceRecord.period == period,
        AttendanceRecord.employee_id.in_([e.id for e in active_employees])
    ).all()}

    performances = {p.employee_id: p for p in db.query(PerformanceScore).filter(
        PerformanceScore.period == period,
        PerformanceScore.employee_id.in_([e.id for e in active_employees])
    ).all()}

    social_ins = {s.employee_id: s for s in db.query(SocialInsurance).filter(
        SocialInsurance.period == period,
        SocialInsurance.employee_id.in_([e.id for e in active_employees])
    ).all()}

    for emp in active_employees:
        missing = []
        if emp.id not in salaries:
            missing.append("薪资档案")
        if emp.id not in attendance:
            missing.append("考勤数据")
        if emp.id not in social_ins:
            missing.append("社保公积金")
        if missing:
            missing_map[emp.id] = missing

    sources.append(DataSourceStatus(
        source_name="员工档案", source_key="employee",
        status="完整", count=len(active_employees)
    ))
    sources.append(DataSourceStatus(
        source_name="薪资档案", source_key="salary",
        status="完整" if len(salaries) == len(active_employees) else "部分缺失",
        count=len(salaries),
        missing_employees=[e.name for e in active_employees if e.id not in salaries]
    ))
    sources.append(DataSourceStatus(
        source_name="考勤数据", source_key="attendance",
        status="完整" if len(attendance) == len(active_employees) else "部分缺失",
        count=len(attendance),
        missing_employees=[e.name for e in active_employees if e.id not in attendance]
    ))
    sources.append(DataSourceStatus(
        source_name="绩效系数", source_key="performance",
        status="完整" if len(performances) == len(active_employees) else "部分缺失",
        count=len(performances),
        missing_employees=[e.name for e in active_employees if e.id not in performances]
    ))
    sources.append(DataSourceStatus(
        source_name="社保公积金", source_key="social_insurance",
        status="完整" if len(social_ins) == len(active_employees) else "部分缺失",
        count=len(social_ins),
        missing_employees=[e.name for e in active_employees if e.id not in social_ins]
    ))

    travel_reimbs = {t.employee_id: t for t in db.query(TravelReimbursement).filter(
        TravelReimbursement.period == period
    ).all()}
    sources.append(DataSourceStatus(
        source_name="临时差旅报销", source_key="travel_reimbursement",
        status="可选数据",
        count=len(travel_reimbs),
        missing_employees=[]
    ))

    labor_comps = {l.employee_id: l for l in db.query(LaborCompensation).filter(
        LaborCompensation.period == period
    ).all()}
    sources.append(DataSourceStatus(
        source_name="劳动补偿金", source_key="labor_compensation",
        status="可选数据",
        count=len(labor_comps),
        missing_employees=[]
    ))

    legacy_adjs = {l.employee_id: l for l in db.query(LegacyAdjustment).filter(
        LegacyAdjustment.period == period
    ).all()}
    sources.append(DataSourceStatus(
        source_name="遗留金额调整", source_key="legacy_adjustment",
        status="可选数据",
        count=len(legacy_adjs),
        missing_employees=[]
    ))

    complete_count = sum(1 for e in active_employees if e.id not in missing_map)
    missing_count = len(missing_map)

    return DataCompletenessOut(
        period=period,
        total_employees=len(active_employees),
        complete_count=complete_count,
        missing_count=missing_count,
        optional_missing_count=0,
        sources=sources
    )


@router.post("/ensure-records/{period}", dependencies=[Depends(require_permission("salary:create"))])
def ensure_salary_records(
    period: str,
    hide_status_id: Optional[int] = Query(None),
    employee_ids: Optional[str] = Query(None, description="逗号分隔的员工ID列表，为空则为所有在职员工创建"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    existing_count = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).count()
    if existing_count == 0:
        result = _perform_gross_calculation(db, period, hide_status_id)
        db.commit()
        write_log(db, "data_change", current_user.id, current_user.username, "salary", "create", f"自动创建薪资记录 (period={period}, count={result['success_count']})")
        return {"created_count": result["success_count"], "total": result["total_employees"]}

    dict_items = db.query(SysDictBase).all()
    name_map = {d.id: d.name for d in dict_items}
    query = db.query(Employee)
    active_employees = filter_active_employees(query, db, hide_status_id=hide_status_id).all()
    existing = {c.employee_id: c for c in db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()}
    updated_count = 0

    for emp in active_employees:
        contract_company_name = name_map.get(emp.contract_company_id, '')
        if emp.id not in existing:
            calc = SalaryCalculation(
                period=period,
                employee_id=emp.id,
                contract_company=contract_company_name,
                pay_company_id=emp.contract_company_id,
                pay_company_name=contract_company_name,
                record_type='single',
                department=name_map.get(emp.department_id, ''),
                position=name_map.get(emp.position_id, ''),
                cost_owner=emp.cost_owner or '',
                status=name_map.get(emp.status_id, ''),
                entry_date=emp.entry_date,
                calculation_status="待核算",
                review_status=""
            )
            db.add(calc)
            updated_count += 1
        else:
            existing_calc = existing[emp.id]
            need_update = False
            if not existing_calc.department:
                existing_calc.department = name_map.get(emp.department_id, '')
                need_update = True
            if not existing_calc.position:
                existing_calc.position = name_map.get(emp.position_id, '')
                need_update = True
            if not existing_calc.status:
                existing_calc.status = name_map.get(emp.status_id, '')
                need_update = True
            if not existing_calc.entry_date:
                existing_calc.entry_date = emp.entry_date
                need_update = True
            if not existing_calc.contract_company:
                existing_calc.contract_company = contract_company_name
                need_update = True
            if existing_calc.record_type is None:
                existing_calc.record_type = 'single'
                need_update = True
            if existing_calc.pay_company_id is None:
                existing_calc.pay_company_id = emp.contract_company_id
                existing_calc.pay_company_name = contract_company_name
                need_update = True
            if need_update:
                updated_count += 1

    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "create", f"确保薪资记录存在 (period={period}, updated={updated_count})")
    return {"created_count": updated_count, "total": len(active_employees)}


def _perform_gross_calculation(db, period, hide_status_id, current_user=None, batch_no=None):
    is_batch_calc = batch_no is not None
    is_admin = current_user and current_user.is_admin

    query = db.query(Employee)
    if current_user and not is_admin:
        query = apply_data_scope(query, db, current_user.data_scope, current_user.id)
    active_employees = filter_active_employees(query, db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()

    period_end = _get_period_end(period)
    salaries = {}
    for s in db.query(EmployeeSalary).filter(
        EmployeeSalary.effective_date <= period_end
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).all():
        if s.employee_id not in salaries:
            salaries[s.employee_id] = s

    attendance = {a.employee_id: a for a in db.query(AttendanceRecord).filter(
        AttendanceRecord.period == period
    ).all()}

    performances = {p.employee_id: p for p in db.query(PerformanceScore).filter(
        PerformanceScore.period == period
    ).all()}

    social_ins = {s.employee_id: s for s in db.query(SocialInsurance).filter(
        SocialInsurance.period == period
    ).all()}

    travel_reimbs = {}
    for tr in db.query(TravelReimbursement).filter(TravelReimbursement.period == period).all():
        travel_reimbs[tr.employee_id] = travel_reimbs.get(tr.employee_id, 0) + float(tr.amount)

    labor_comps = {}
    for lc in db.query(LaborCompensation).filter(LaborCompensation.period == period).all():
        labor_comps[lc.employee_id] = labor_comps.get(lc.employee_id, 0) + float(lc.amount)

    legacy_adjs = {}
    for la in db.query(LegacyAdjustment).filter(LegacyAdjustment.period == period).all():
        if la.employee_id not in legacy_adjs:
            legacy_adjs[la.employee_id] = {"pretax": None, "posttax": None}
        val = float(la.amount)
        if la.is_pretax:
            legacy_adjs[la.employee_id]["pretax"] = (legacy_adjs[la.employee_id]["pretax"] or 0) + val
        else:
            legacy_adjs[la.employee_id]["posttax"] = (legacy_adjs[la.employee_id]["posttax"] or 0) + val

    adjustments = {a.employee_id: a for a in db.query(EmployeeSalaryAdjustment).filter(
        EmployeeSalaryAdjustment.period == period
    ).all()}

    all_existing = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    existing_by_emp = {}
    for c in all_existing:
        key = (c.employee_id, c.pay_company_id)
        existing_by_emp[key] = c
    existing_emp_ids = {c.employee_id for c in all_existing}

    success_count = 0
    failed_count = 0
    total_gross = 0
    total_net = 0

    dict_items = db.query(SysDictBase).all()
    name_map = {d.id: d.name for d in dict_items}
    id_map = {d.name: d.id for d in dict_items}
    standard_salary_days = get_month_salary_days(db, period)

    yijiu_company, split_config = _get_salary_split_config(db, dict_items)

    for emp in active_employees:
        try:
            sal = salaries.get(emp.id)
            att = attendance.get(emp.id)
            perf = performances.get(emp.id)
            si = social_ins.get(emp.id)
            contract_company_name = name_map.get(emp.contract_company_id, '')
            split_target = split_config.get(emp.contract_company_id)

            has_salary_data = sal is not None
            has_si_data = si is not None
            has_any_data = has_salary_data or has_si_data

            db.query(SalaryCalculation).filter(
                SalaryCalculation.period == period,
                SalaryCalculation.employee_id == emp.id
            ).delete(synchronize_session=False)

            records_to_create = []

            if not has_any_data:
                empty_record = SalaryCalculation(
                    period=period,
                    employee_id=emp.id,
                    contract_company=contract_company_name,
                    pay_company_id=emp.contract_company_id,
                    pay_company_name=contract_company_name,
                    record_type='single',
                    department=name_map.get(emp.department_id, ''),
                    position=name_map.get(emp.position_id, ''),
                    cost_owner=emp.cost_owner or '',
                    status=name_map.get(emp.status_id, ''),
                    entry_date=emp.entry_date,
                    calculation_status="待核算",
                    review_status=""
                )
                records_to_create.append(empty_record)
                for rec in records_to_create:
                    db.add(rec)
                success_count += 1
                continue

            if has_salary_data:
                base_salary = float(sal.base_salary)
                perf_std = float(sal.performance_standard)
                meal = float(sal.meal_allowance or 0)
                transport = float(sal.transport_allowance or 0)
                comm = float(sal.communication_allowance or 0)
                computer = float(sal.computer_allowance or 0)
                housing = float(sal.housing_allowance or 0)
                allowance_total = meal + transport + comm + computer + housing
            else:
                base_salary = None
                perf_std = None
                meal = None
                transport = None
                comm = None
                computer = None
                housing = None
                allowance_total = None

            if att:
                att_rate = float(att.attendance_rate) if att.attendance_rate else 1.00
                total_work_days = float(att.adjusted_salary_days) if att.adjusted_salary_days else standard_salary_days
                actual_work_days = float(att.actual_salary_days) if att.actual_salary_days else total_work_days
            else:
                att_rate = 1.00
                total_work_days = standard_salary_days
                actual_work_days = standard_salary_days

            adj = adjustments.get(emp.id)
            if adj and has_salary_data:
                base_salary_prorated = float(adj.base_salary_prorated or 0)
                perf_std_prorated = float(adj.performance_standard_prorated or 0)
                commission_prorated = float(adj.commission_prorated) if adj.commission_prorated is not None else None
                adjustment_id = adj.id
                monthly_standard = round(base_salary_prorated + perf_std_prorated + allowance_total, 2)
            elif has_salary_data:
                base_salary_prorated = base_salary
                perf_std_prorated = perf_std
                commission_prorated = None
                adjustment_id = None
                monthly_standard = round(base_salary + perf_std + (allowance_total or 0), 2)
            else:
                base_salary_prorated = None
                perf_std_prorated = None
                commission_prorated = None
                adjustment_id = None
                monthly_standard = None

            perf_coef = float(perf.final_score) if perf and perf.final_score is not None else None
            actual_perf = round(perf_std_prorated * perf_coef, 2) if perf_coef is not None and perf_std_prorated is not None else None
            effective_performance = round(actual_perf * att_rate, 2) if actual_perf is not None else None

            travel_untaxed = travel_reimbs.get(emp.id) if emp.id in travel_reimbs else None
            compensation_tax_val = labor_comps.get(emp.id) if emp.id in labor_comps else None
            legacy = legacy_adjs.get(emp.id)
            pretax_adj = float(legacy["pretax"]) if legacy and legacy["pretax"] is not None else None
            posttax_adj = float(legacy["posttax"]) if legacy and legacy["posttax"] is not None else None

            if has_salary_data:
                commission_calc = _safe_float(commission_prorated) or 0
                gross_salary = round((base_salary_prorated + (allowance_total or 0) + commission_calc) * att_rate + (effective_performance or 0), 2)
            else:
                gross_salary = None

            pension_personal = float(si.pension_personal) if si and si.pension_personal is not None else None
            unemployment_personal = float(si.unemployment_personal) if si and si.unemployment_personal is not None else None
            medical_personal = float(si.medical_personal) if si and si.medical_personal is not None else None
            hf_personal = float(si.hf_personal) if si and si.hf_personal is not None else None
            
            si_personal = None
            si_hf_total = None
            if has_si_data:
                p = pension_personal or 0
                u = unemployment_personal or 0
                m = medical_personal or 0
                h = hf_personal or 0
                si_personal = round(p + u + m, 2) if any(v is not None for v in [pension_personal, unemployment_personal, medical_personal]) else None
                si_hf_total = round(p + u + m + h, 2) if any(v is not None for v in [pension_personal, unemployment_personal, medical_personal, hf_personal]) else None

            should_split = split_target is not None and has_salary_data and has_si_data and si_hf_total is not None and si_hf_total > 0

            if should_split:
                contract_existing = existing_by_emp.get((emp.id, emp.contract_company_id))
                yijiu_existing_keys = [k for k in existing_by_emp.keys() if k[0] == emp.id and k[1] != emp.contract_company_id]
                payroll_existing = existing_by_emp.get(yijiu_existing_keys[0]) if yijiu_existing_keys else None

                single_existing = None
                if not contract_existing and not payroll_existing:
                    for k, v in existing_by_emp.items():
                        if k[0] == emp.id:
                            single_existing = v
                            break

                contract_tax = _safe_float(contract_existing.tax_deduction) if contract_existing else None
                contract_pretax = _safe_float(contract_existing.pretax_adjustment) if contract_existing else None
                contract_posttax = _safe_float(contract_existing.posttax_adjustment) if contract_existing else None
                contract_remark = contract_existing.remark if contract_existing else None

                if contract_tax is None and single_existing:
                    pass

                if contract_existing:
                    contract_bonus_final = _merge_val(contract_existing, contract_existing.commission_bonus, None)
                    contract_pretax_final = _merge_val(contract_existing, contract_existing.pretax_adjustment, None)
                    contract_posttax_final = _merge_val(contract_existing, contract_existing.posttax_adjustment, None)
                else:
                    contract_bonus_final = None
                    contract_pretax_final = None
                    contract_posttax_final = None

                contract_gross = round(si_hf_total, 2) if si_hf_total is not None else None
                contract_si = si_hf_total
                contract_sh = si_hf_total if si_hf_total is not None else 0
                contract_pretax_calc = _safe_float(contract_pretax_final) or 0
                contract_posttax_calc = _safe_float(contract_posttax_final) or 0
                
                if contract_gross is not None:
                    contract_salary_after_si = round(contract_gross + contract_pretax_calc - contract_sh, 2)
                    contract_actual_taxable = round(contract_gross + contract_pretax_calc, 2)
                    if contract_tax is not None:
                        contract_net = round(contract_gross + contract_pretax_calc - contract_sh - contract_tax + contract_posttax_calc, 2)
                        contract_status = "实发已核算"
                    else:
                        contract_net = round(contract_gross + contract_pretax_calc - contract_sh + contract_posttax_calc, 2)
                        contract_status = "应发已核算"
                else:
                    contract_salary_after_si = None
                    contract_actual_taxable = None
                    contract_net = None
                    contract_status = "待核算"

                contract_record = SalaryCalculation(
                    period=period,
                    employee_id=emp.id,
                    contract_company=contract_company_name,
                    pay_company_id=emp.contract_company_id,
                    pay_company_name=contract_company_name,
                    record_type='contract',
                    department=name_map.get(emp.department_id, ''),
                    position=name_map.get(emp.position_id, ''),
                    cost_owner=emp.cost_owner or '',
                    status=name_map.get(emp.status_id, ''),
                    entry_date=emp.entry_date,
                    base_salary=None,
                    base_salary_prorated=None,
                    performance_standard_prorated=None,
                    adjustment_id=None,
                    monthly_standard=None,
                    performance_standard=None,
                    performance_coefficient=None,
                    actual_performance=None,
                    effective_performance=None,
                    meal_allowance=None,
                    transport_allowance=None,
                    communication_allowance=None,
                    computer_allowance=None,
                    housing_allowance=None,
                    allowance_total=None,
                    commission_bonus=contract_bonus_final,
                    pretax_adjustment=contract_pretax_final,
                    pretax_adjustment_reason=contract_existing.pretax_adjustment_reason if contract_existing else None,
                    posttax_adjustment=contract_posttax_final,
                    posttax_adjustment_reason=contract_existing.posttax_adjustment_reason if contract_existing else None,
                    total_work_days=total_work_days,
                    actual_work_days=actual_work_days,
                    attendance_rate=att_rate,
                    gross_salary=contract_gross,
                    pension_personal=pension_personal,
                    unemployment_personal=unemployment_personal,
                    medical_personal=medical_personal,
                    social_insurance_personal=si_personal,
                    housing_fund_personal=hf_personal,
                    si_hf_total=contract_si,
                    salary_after_si_hf=contract_salary_after_si,
                    tax_deduction=contract_tax,
                    net_salary=contract_net,
                    last_month_untaxed=None,
                    travel_untaxed=None,
                    compensation_tax=None,
                    severance_pay=None,
                    year_end_bonus_untaxed=None,
                    year_end_bonus_net=None,
                    actual_taxable=contract_actual_taxable,
                    special_deduction=None,
                    remark=contract_remark,
                    review_status=contract_existing.review_status if contract_existing else "",
                    calculation_status=contract_status,
                    data_completeness="完整" if has_si_data else "部分缺失"
                )
                records_to_create.append(contract_record)
                if contract_gross is not None:
                    total_gross += contract_gross
                if contract_net is not None:
                    total_net += contract_net

                payroll_tax = _safe_float(payroll_existing.tax_deduction) if payroll_existing else None
                payroll_last_month = _safe_float(payroll_existing.last_month_untaxed) if payroll_existing else None
                payroll_year_end = _safe_float(payroll_existing.year_end_bonus_untaxed) if payroll_existing else None
                payroll_year_end_net = _safe_float(payroll_existing.year_end_bonus_net) if payroll_existing else None
                payroll_severance = _safe_float(payroll_existing.severance_pay) if payroll_existing else None
                payroll_special = _safe_float(payroll_existing.special_deduction) if payroll_existing else None
                payroll_remark = payroll_existing.remark if payroll_existing else None

                if payroll_tax is None and single_existing:
                    payroll_tax = _safe_float(single_existing.tax_deduction)
                if payroll_last_month is None and single_existing:
                    payroll_last_month = _safe_float(single_existing.last_month_untaxed)
                if payroll_year_end is None and single_existing:
                    payroll_year_end = _safe_float(single_existing.year_end_bonus_untaxed)
                if payroll_year_end_net is None and single_existing:
                    payroll_year_end_net = _safe_float(single_existing.year_end_bonus_net)
                if payroll_severance is None and single_existing:
                    payroll_severance = _safe_float(single_existing.severance_pay)
                if payroll_special is None and single_existing:
                    payroll_special = _safe_float(single_existing.special_deduction)

                if payroll_existing:
                    payroll_bonus_final = _merge_val(payroll_existing, payroll_existing.commission_bonus, commission_prorated)
                    payroll_pretax_final = _merge_val(payroll_existing, payroll_existing.pretax_adjustment, pretax_adj)
                    payroll_posttax_final = _merge_val(payroll_existing, payroll_existing.posttax_adjustment, posttax_adj)
                    payroll_travel_final = _merge_val(payroll_existing, payroll_existing.travel_untaxed, travel_untaxed)
                    payroll_comp_final = _merge_val(payroll_existing, payroll_existing.compensation_tax, compensation_tax_val)
                    payroll_review = payroll_existing.review_status
                    payroll_pretax_reason = payroll_existing.pretax_adjustment_reason if pretax_adj is None else None
                    payroll_posttax_reason = payroll_existing.posttax_adjustment_reason if posttax_adj is None else None
                elif single_existing:
                    payroll_bonus_final = _merge_val(single_existing, single_existing.commission_bonus, commission_prorated)
                    payroll_pretax_final = _merge_val(single_existing, single_existing.pretax_adjustment, pretax_adj)
                    payroll_posttax_final = _merge_val(single_existing, single_existing.posttax_adjustment, posttax_adj)
                    payroll_travel_final = _merge_val(single_existing, single_existing.travel_untaxed, travel_untaxed)
                    payroll_comp_final = _merge_val(single_existing, single_existing.compensation_tax, compensation_tax_val)
                    payroll_review = single_existing.review_status
                    payroll_pretax_reason = single_existing.pretax_adjustment_reason if pretax_adj is None else None
                    payroll_posttax_reason = single_existing.posttax_adjustment_reason if posttax_adj is None else None
                else:
                    payroll_bonus_final = commission_prorated
                    payroll_pretax_final = pretax_adj
                    payroll_posttax_final = posttax_adj
                    payroll_travel_final = travel_untaxed
                    payroll_comp_final = compensation_tax_val
                    payroll_review = ""
                    payroll_pretax_reason = None
                    payroll_posttax_reason = None

                if gross_salary is not None:
                    si_hf_deduct = si_hf_total if si_hf_total is not None else 0
                    payroll_gross = round(max(0, gross_salary - si_hf_deduct), 2)
                else:
                    payroll_gross = None
                payroll_si = 0

                payroll_pretax_calc = _safe_float(payroll_pretax_final) or 0
                payroll_posttax_calc = _safe_float(payroll_posttax_final) or 0
                payroll_last_month_calc = _safe_float(payroll_last_month) or 0
                payroll_travel_calc = _safe_float(payroll_travel_final) or 0
                payroll_comp_calc = _safe_float(payroll_comp_final) or 0
                payroll_severance_calc = _safe_float(payroll_severance) or 0
                payroll_year_end_calc = _safe_float(payroll_year_end) or 0
                payroll_special_calc = _safe_float(payroll_special) or 0

                if monthly_standard is not None:
                    si_hf_deduct = si_hf_total if si_hf_total is not None else 0
                    payroll_monthly_standard = round(monthly_standard - si_hf_deduct, 2)
                else:
                    payroll_monthly_standard = None

                if payroll_gross is not None:
                    payroll_actual_taxable = round(payroll_gross + payroll_pretax_calc + payroll_last_month_calc + payroll_travel_calc + payroll_comp_calc + payroll_severance_calc + payroll_year_end_calc - payroll_special_calc, 2)
                    payroll_salary_after_si = round(payroll_gross + payroll_pretax_calc - payroll_si, 2)
                    if payroll_tax is not None:
                        payroll_net = round(payroll_gross + payroll_pretax_calc - payroll_si - payroll_tax + payroll_posttax_calc, 2)
                        payroll_status = "实发已核算"
                    else:
                        payroll_net = round(payroll_gross + payroll_pretax_calc - payroll_si + payroll_posttax_calc, 2)
                        payroll_status = "应发已核算"
                else:
                    payroll_actual_taxable = None
                    payroll_salary_after_si = None
                    payroll_net = None
                    payroll_status = "待核算"

                payroll_record = SalaryCalculation(
                    period=period,
                    employee_id=emp.id,
                    contract_company=split_target.name,
                    pay_company_id=split_target.id,
                    pay_company_name=split_target.name,
                    record_type='payroll',
                    department=name_map.get(emp.department_id, ''),
                    position=name_map.get(emp.position_id, ''),
                    cost_owner=emp.cost_owner or '',
                    status=name_map.get(emp.status_id, ''),
                    entry_date=emp.entry_date,
                    base_salary=base_salary,
                    base_salary_prorated=base_salary_prorated,
                    performance_standard_prorated=perf_std_prorated,
                    adjustment_id=adjustment_id,
                    monthly_standard=payroll_monthly_standard,
                    performance_standard=perf_std,
                    performance_coefficient=perf_coef,
                    actual_performance=actual_perf,
                    effective_performance=effective_performance,
                    meal_allowance=meal,
                    transport_allowance=transport,
                    communication_allowance=comm,
                    computer_allowance=computer,
                    housing_allowance=housing,
                    allowance_total=allowance_total,
                    commission_bonus=payroll_bonus_final,
                    pretax_adjustment=payroll_pretax_final,
                    pretax_adjustment_reason=payroll_pretax_reason,
                    posttax_adjustment=payroll_posttax_final,
                    posttax_adjustment_reason=payroll_posttax_reason,
                    total_work_days=total_work_days,
                    actual_work_days=actual_work_days,
                    attendance_rate=att_rate,
                    gross_salary=payroll_gross,
                    pension_personal=None,
                    unemployment_personal=None,
                    medical_personal=None,
                    social_insurance_personal=None,
                    housing_fund_personal=None,
                    si_hf_total=None,
                    salary_after_si_hf=payroll_salary_after_si,
                    tax_deduction=payroll_tax,
                    net_salary=payroll_net,
                    last_month_untaxed=payroll_last_month,
                    travel_untaxed=payroll_travel_final,
                    compensation_tax=payroll_comp_final,
                    severance_pay=payroll_severance,
                    year_end_bonus_untaxed=payroll_year_end,
                    year_end_bonus_net=payroll_year_end_net,
                    actual_taxable=payroll_actual_taxable,
                    special_deduction=payroll_special,
                    remark=payroll_remark,
                    review_status=payroll_review,
                    calculation_status=payroll_status,
                    data_completeness="完整" if has_salary_data else "部分缺失"
                )
                records_to_create.append(payroll_record)
                if payroll_gross is not None:
                    total_gross += payroll_gross
                if payroll_net is not None:
                    total_net += payroll_net
            else:
                existing_single = None
                for k, v in existing_by_emp.items():
                    if k[0] == emp.id:
                        existing_single = v
                        break

                tax_deduction_val = _safe_float(existing_single.tax_deduction) if existing_single else None
                last_month_untaxed_val = _safe_float(existing_single.last_month_untaxed) if existing_single else None
                year_end_bonus_untaxed_val = _safe_float(existing_single.year_end_bonus_untaxed) if existing_single else None
                year_end_bonus_net_val = _safe_float(existing_single.year_end_bonus_net) if existing_single else None
                severance_pay_val = _safe_float(existing_single.severance_pay) if existing_single else None
                special_deduction_val = _safe_float(existing_single.special_deduction) if existing_single else None

                if existing_single:
                    commission_bonus_final = _merge_val(existing_single, existing_single.commission_bonus, commission_prorated)
                    pretax_adj_final = _merge_val(existing_single, existing_single.pretax_adjustment, pretax_adj)
                    posttax_adj_final = _merge_val(existing_single, existing_single.posttax_adjustment, posttax_adj)
                    travel_untaxed_final = _merge_val(existing_single, existing_single.travel_untaxed, travel_untaxed)
                    compensation_tax_final = _merge_val(existing_single, existing_single.compensation_tax, compensation_tax_val)
                    pretax_adjustment_reason_final = existing_single.pretax_adjustment_reason if pretax_adj is None else None
                    posttax_adjustment_reason_final = existing_single.posttax_adjustment_reason if posttax_adj is None else None
                    review_status_final = existing_single.review_status
                    remark_final = existing_single.remark
                else:
                    commission_bonus_final = commission_prorated
                    pretax_adj_final = pretax_adj
                    posttax_adj_final = posttax_adj
                    travel_untaxed_final = travel_untaxed
                    compensation_tax_final = compensation_tax_val
                    pretax_adjustment_reason_final = None
                    posttax_adjustment_reason_final = None
                    review_status_final = ""
                    remark_final = None

                pretax_calc = _safe_float(pretax_adj_final) or 0
                posttax_calc = _safe_float(posttax_adj_final) or 0
                last_month_calc = _safe_float(last_month_untaxed_val) or 0
                travel_calc = _safe_float(travel_untaxed_final) or 0

                sh = si_hf_total if si_hf_total is not None else 0
                if gross_salary is not None:
                    actual_taxable = round(gross_salary + pretax_calc + last_month_calc + travel_calc, 2)
                    salary_after_si_hf_val = round(gross_salary + pretax_calc - sh, 2)
                    if tax_deduction_val is not None:
                        net_salary = round(gross_salary + pretax_calc - sh - tax_deduction_val + posttax_calc, 2)
                        calc_status = "实发已核算"
                    else:
                        net_salary = round(gross_salary + pretax_calc - sh + posttax_calc, 2)
                        calc_status = "应发已核算"
                else:
                    actual_taxable = None
                    salary_after_si_hf_val = None
                    net_salary = None
                    calc_status = "待核算"

                if split_target is not None:
                    if has_salary_data and not has_si_data:
                        pay_company_id = split_target.id
                        pay_company_name_val = split_target.name
                    else:
                        pay_company_id = emp.contract_company_id
                        pay_company_name_val = contract_company_name
                else:
                    pay_company_id = emp.contract_company_id
                    pay_company_name_val = contract_company_name
                record_type_val = 'single'

                single_record = SalaryCalculation(
                    period=period,
                    employee_id=emp.id,
                    contract_company=contract_company_name,
                    pay_company_id=pay_company_id,
                    pay_company_name=pay_company_name_val,
                    record_type=record_type_val,
                    department=name_map.get(emp.department_id, ''),
                    position=name_map.get(emp.position_id, ''),
                    cost_owner=emp.cost_owner or '',
                    status=name_map.get(emp.status_id, ''),
                    entry_date=emp.entry_date,
                    base_salary=base_salary,
                    base_salary_prorated=base_salary_prorated,
                    performance_standard_prorated=perf_std_prorated,
                    adjustment_id=adjustment_id,
                    monthly_standard=monthly_standard,
                    performance_standard=perf_std,
                    performance_coefficient=perf_coef,
                    actual_performance=actual_perf,
                    effective_performance=effective_performance,
                    meal_allowance=meal,
                    transport_allowance=transport,
                    communication_allowance=comm,
                    computer_allowance=computer,
                    housing_allowance=housing,
                    allowance_total=allowance_total,
                    commission_bonus=commission_bonus_final,
                    pretax_adjustment=pretax_adj_final,
                    pretax_adjustment_reason=pretax_adjustment_reason_final,
                    posttax_adjustment=posttax_adj_final,
                    posttax_adjustment_reason=posttax_adjustment_reason_final,
                    travel_untaxed=travel_untaxed_final,
                    compensation_tax=compensation_tax_final,
                    severance_pay=severance_pay_val,
                    year_end_bonus_untaxed=year_end_bonus_untaxed_val,
                    year_end_bonus_net=year_end_bonus_net_val,
                    total_work_days=total_work_days,
                    actual_work_days=actual_work_days,
                    attendance_rate=att_rate,
                    gross_salary=gross_salary,
                    pension_personal=pension_personal,
                    unemployment_personal=unemployment_personal,
                    medical_personal=medical_personal,
                    social_insurance_personal=si_personal,
                    housing_fund_personal=hf_personal,
                    si_hf_total=si_hf_total,
                    salary_after_si_hf=salary_after_si_hf_val,
                    tax_deduction=tax_deduction_val,
                    net_salary=net_salary,
                    last_month_untaxed=last_month_untaxed_val,
                    actual_taxable=actual_taxable,
                    special_deduction=special_deduction_val,
                    remark=remark_final,
                    review_status=review_status_final,
                    calculation_status=calc_status
                )
                records_to_create.append(single_record)
                if gross_salary is not None:
                    total_gross += gross_salary
                if net_salary is not None:
                    total_net += net_salary

            for rec in records_to_create:
                db.add(rec)
            success_count += 1

        except Exception as e:
            import traceback
            traceback.print_exc()
            failed_count += 1

    active_emp_ids = {e.id for e in active_employees}
    if is_batch_calc and is_admin:
        for old_calc in all_existing:
            if old_calc.employee_id not in active_emp_ids:
                db.delete(old_calc)

    if current_user and batch_no:
        start_time = datetime.now()
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        log = CalculationLog(
            batch_no=batch_no,
            period=period,
            calculation_type="应发工资核算",
            status="成功" if failed_count == 0 else "部分成功",
            start_time=start_time,
            end_time=end_time,
            duration_seconds=duration,
            total_employees=len(active_employees),
            success_count=success_count,
            failed_count=failed_count,
            operator_id=current_user.id
        )
        db.add(log)
        write_log(db, "operation", current_user.id, current_user.username, "salary", "calculate", f"应发工资核算 (period={period}, batch={batch_no})", {"success": success_count, "failed": failed_count})

    return {
        "period": period,
        "total_employees": len(active_employees),
        "success_count": success_count,
        "failed_count": failed_count,
        "total_gross_salary": round(total_gross, 2),
        "avg_gross_salary": round(total_gross / success_count, 2) if success_count > 0 else 0,
        "total_net_salary": round(total_net, 2),
        "avg_net_salary": round(total_net / success_count, 2) if success_count > 0 else 0,
        "batch_no": batch_no
    }


@router.post("/calculate/{period}", response_model=CalculationSummary, dependencies=[Depends(require_permission("salary:calculate"))])
def calculate_salary(period: str, hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"), db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    batch_no = f"CALC-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"
    result = _perform_gross_calculation(db, period, hide_status_id, current_user, batch_no)
    db.commit()
    return CalculationSummary(**result)


def _safe_float(val):
    if val is None:
        return None
    return float(val)


def _merge_val(existing, existing_val, new_val):
    if new_val is not None:
        return new_val
    if existing is not None and existing_val is not None:
        return float(existing_val)
    return None


def _build_employee_salary_data(
    emp, sal, att, perf, si, adj,
    travel_untaxed_val, compensation_tax_val,
    pretax_adj_val, posttax_adj_val,
    last_month_untaxed_val, severance_pay_val, year_end_bonus_untaxed_val,
    year_end_bonus_net_val,
    tax_deduction_val, special_deduction_val,
    commission_bonus_val,
    name_map, standard_salary_days
):
    contract_company_name = name_map.get(emp.contract_company_id, '')
    department_name = name_map.get(emp.department_id, '')
    position_name = name_map.get(emp.position_id, '')
    status_name = name_map.get(emp.status_id, '')

    base_salary = _safe_float(sal.base_salary) if sal else None
    perf_std = _safe_float(sal.performance_standard) if sal else None
    meal = _safe_float(sal.meal_allowance) if sal else None
    transport = _safe_float(sal.transport_allowance) if sal else None
    comm = _safe_float(sal.communication_allowance) if sal else None
    computer = _safe_float(sal.computer_allowance) if sal else None
    housing = _safe_float(sal.housing_allowance) if sal else None

    allowance_total = None
    if sal:
        total = 0
        has_any = False
        for v in [meal, transport, comm, computer, housing]:
            if v is not None:
                total += v
                has_any = True
        allowance_total = round(total, 2) if has_any else None

    if att:
        total_work_days = _safe_float(att.adjusted_salary_days) or standard_salary_days
        actual_work_days = _safe_float(att.actual_salary_days) or total_work_days
        attendance_rate = _safe_float(att.attendance_rate)
        if attendance_rate is None and total_work_days and total_work_days > 0 and actual_work_days is not None:
            attendance_rate = round(actual_work_days / total_work_days, 4)
    else:
        total_work_days = standard_salary_days
        actual_work_days = standard_salary_days
        attendance_rate = 1.0

    if adj:
        base_salary_prorated = _safe_float(adj.base_salary_prorated)
        perf_std_prorated = _safe_float(adj.performance_standard_prorated)
        commission_from_adj = _safe_float(adj.commission_prorated)
        adjustment_id = adj.id
    else:
        base_salary_prorated = base_salary
        perf_std_prorated = perf_std
        commission_from_adj = None
        adjustment_id = None

    perf_coef = _safe_float(perf.final_score) if perf else None

    actual_perf = None
    if perf_coef is not None and perf_std_prorated is not None:
        actual_perf = round(perf_std_prorated * perf_coef, 2)

    effective_perf = None
    if actual_perf is not None and attendance_rate is not None:
        effective_perf = round(actual_perf * attendance_rate, 2)

    monthly_standard = None
    if base_salary_prorated is not None or perf_std_prorated is not None or allowance_total is not None:
        bs = base_salary_prorated or 0
        ps = perf_std_prorated or 0
        al = allowance_total or 0
        monthly_standard = round(bs + ps + al, 2)

    commission = commission_bonus_val if commission_bonus_val is not None else commission_from_adj
    pension_personal = _safe_float(si.pension_personal) if si else None
    unemployment_personal = _safe_float(si.unemployment_personal) if si else None
    medical_personal = _safe_float(si.medical_personal) if si else None
    hf_personal = _safe_float(si.hf_personal) if si else None

    si_personal = None
    if si:
        si_sum = 0
        has_si = False
        for v in [pension_personal, unemployment_personal, medical_personal]:
            if v is not None:
                si_sum += v
                has_si = True
        si_personal = round(si_sum, 2) if has_si else None

    si_hf_total_val = None
    if si:
        total = 0
        has_total = False
        for v in [si_personal, hf_personal]:
            if v is not None:
                total += v
                has_total = True
        si_hf_total_val = round(total, 2) if has_total else None

    gross_salary_val = None
    can_calc_gross = all(v is not None for v in [base_salary_prorated, allowance_total, attendance_rate])
    if can_calc_gross:
        bs = base_salary_prorated or 0
        al = allowance_total or 0
        cm = commission or 0
        ar = attendance_rate or 1.0
        ep = effective_perf or 0
        gross_salary_val = round((bs + al + cm) * ar + ep, 2)

    salary_after_si_hf_val = None
    net_salary_val = None
    actual_taxable_val = None

    if gross_salary_val is not None:
        pa = pretax_adj_val or 0
        sh = si_hf_total_val or 0
        pt = posttax_adj_val or 0
        salary_after_si_hf_val = round(gross_salary_val + pa - sh, 2)

        if tax_deduction_val is not None:
            net_salary_val = round(gross_salary_val + pa - sh - tax_deduction_val + pt, 2)
        else:
            net_salary_val = round(gross_salary_val + pa - sh + pt, 2)

        lm = last_month_untaxed_val or 0
        tu = travel_untaxed_val or 0
        actual_taxable_val = round(gross_salary_val + pa + lm + tu, 2)

    return {
        'contract_company': contract_company_name,
        'department': department_name,
        'position': position_name,
        'cost_owner': emp.cost_owner or '',
        'status': status_name,
        'entry_date': emp.entry_date.isoformat() if emp.entry_date else None,
        'base_salary': base_salary,
        'base_salary_prorated': base_salary_prorated,
        'performance_standard_prorated': perf_std_prorated,
        'adjustment_id': adjustment_id,
        'monthly_standard': monthly_standard,
        'performance_standard': perf_std,
        'performance_coefficient': perf_coef,
        'actual_performance': actual_perf,
        'effective_performance': effective_perf,
        'meal_allowance': meal,
        'transport_allowance': transport,
        'communication_allowance': comm,
        'computer_allowance': computer,
        'housing_allowance': housing,
        'allowance_total': allowance_total,
        'commission_bonus': commission,
        'pretax_adjustment': pretax_adj_val,
        'posttax_adjustment': posttax_adj_val,
        'travel_untaxed': travel_untaxed_val,
        'compensation_tax': compensation_tax_val,
        'total_work_days': total_work_days,
        'actual_work_days': actual_work_days,
        'attendance_rate': attendance_rate,
        'gross_salary': gross_salary_val,
        'pension_personal': pension_personal,
        'unemployment_personal': unemployment_personal,
        'medical_personal': medical_personal,
        'social_insurance_personal': si_personal,
        'housing_fund_personal': hf_personal,
        'si_hf_total': si_hf_total_val,
        'salary_after_si_hf': salary_after_si_hf_val,
        'tax_deduction': tax_deduction_val,
        'net_salary': net_salary_val,
        'last_month_untaxed': last_month_untaxed_val,
        'compensation_tax': compensation_tax_val,
        'severance_pay': severance_pay_val,
        'year_end_bonus_untaxed': year_end_bonus_untaxed_val,
        'year_end_bonus_net': year_end_bonus_net_val,
        'actual_taxable': actual_taxable_val,
        'special_deduction': special_deduction_val,
    }


@router.get("/results/{period}", response_model=List[SalaryCalcOut], dependencies=[Depends(require_permission("salary:view"))])
def get_calculation_results(
    period: str,
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    pay_company_id: Optional[int] = Query(None, description="按发放公司筛选"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    _perform_gross_calculation(db, period, hide_status_id, current_user)
    db.commit()

    query = db.query(Employee)
    if not current_user.is_admin:
        query = apply_data_scope(query, db, current_user.data_scope, current_user.id)
    active_employees = filter_active_employees(query, db, hide_status_id=hide_status_id).all()
    active_emp_ids = {e.id: e for e in active_employees}

    dict_items = db.query(SysDictBase).all()
    name_map = {d.id: d.name for d in dict_items}

    standard_salary_days = get_month_salary_days(db, period)
    attendance_map = {a.employee_id: a for a in db.query(AttendanceRecord).filter(
        AttendanceRecord.period == period
    ).all()}

    calcs_query = db.query(SalaryCalculation).filter(SalaryCalculation.period == period)
    if not current_user.is_admin:
        calcs_query = apply_data_scope(calcs_query, db, current_user.data_scope, current_user.id, employee_id_field=SalaryCalculation.employee_id)
    if pay_company_id:
        calcs_query = calcs_query.filter(SalaryCalculation.pay_company_id == pay_company_id)
    calcs = calcs_query.order_by(SalaryCalculation.employee_id, SalaryCalculation.record_type).all()

    results = []
    for c in calcs:
        if c.employee_id not in active_emp_ids:
            continue
        emp = active_emp_ids[c.employee_id]

        att = attendance_map.get(emp.id)
        if att:
            att_rate = float(att.attendance_rate) if att.attendance_rate else 1.00
            total_work_days = float(att.adjusted_salary_days) if att.adjusted_salary_days else standard_salary_days
            actual_work_days = float(att.actual_salary_days) if att.actual_salary_days else total_work_days
        else:
            att_rate = 1.00
            total_work_days = standard_salary_days
            actual_work_days = standard_salary_days

        department_val = c.department or name_map.get(emp.department_id, '')
        position_val = c.position or name_map.get(emp.position_id, '')
        status_val = c.status or name_map.get(emp.status_id, '')
        entry_date_val = c.entry_date or emp.entry_date
        total_work_days_val = _safe_float(c.total_work_days) or total_work_days
        actual_work_days_val = _safe_float(c.actual_work_days) or actual_work_days
        attendance_rate_val = _safe_float(c.attendance_rate) or att_rate

        contract_company_val = c.contract_company or name_map.get(emp.contract_company_id, '')

        result = SalaryCalcOut(
            id=c.id,
            period=c.period,
            employee_id=c.employee_id,
            employee_no=emp.employee_no,
            employee_name=emp.name,
            contract_company=contract_company_val,
            pay_company_id=c.pay_company_id,
            pay_company_name=c.pay_company_name or "",
            record_type=c.record_type or "single",
            department=department_val,
            position=position_val,
            status=status_val,
            entry_date=entry_date_val.isoformat() if entry_date_val else None,
            cost_owner=c.cost_owner or emp.cost_owner or '',
            base_salary=_safe_float(c.base_salary),
            base_salary_prorated=_safe_float(c.base_salary_prorated),
            performance_standard_prorated=_safe_float(c.performance_standard_prorated),
            adjustment_id=c.adjustment_id,
            monthly_standard=_safe_float(c.monthly_standard),
            performance_standard=_safe_float(c.performance_standard),
            performance_coefficient=_safe_float(c.performance_coefficient),
            actual_performance=_safe_float(c.actual_performance),
            effective_performance=_safe_float(c.effective_performance),
            meal_allowance=_safe_float(c.meal_allowance),
            transport_allowance=_safe_float(c.transport_allowance),
            communication_allowance=_safe_float(c.communication_allowance),
            computer_allowance=_safe_float(c.computer_allowance),
            housing_allowance=_safe_float(c.housing_allowance),
            allowance_total=_safe_float(c.allowance_total),
            commission_bonus=_safe_float(c.commission_bonus),
            pretax_adjustment=_safe_float(c.pretax_adjustment),
            pretax_adjustment_reason=c.pretax_adjustment_reason,
            posttax_adjustment=_safe_float(c.posttax_adjustment),
            posttax_adjustment_reason=c.posttax_adjustment_reason,
            total_work_days=total_work_days_val,
            actual_work_days=actual_work_days_val,
            attendance_rate=attendance_rate_val,
            gross_salary=_safe_float(c.gross_salary),
            pension_personal=_safe_float(c.pension_personal),
            unemployment_personal=_safe_float(c.unemployment_personal),
            medical_personal=_safe_float(c.medical_personal),
            social_insurance_personal=_safe_float(c.social_insurance_personal),
            housing_fund_personal=_safe_float(c.housing_fund_personal),
            si_hf_total=_safe_float(c.si_hf_total),
            salary_after_si_hf=_safe_float(c.salary_after_si_hf),
            tax_deduction=_safe_float(c.tax_deduction),
            net_salary=_safe_float(c.net_salary),
            last_month_untaxed=_safe_float(c.last_month_untaxed),
            travel_untaxed=_safe_float(c.travel_untaxed),
            compensation_tax=_safe_float(c.compensation_tax),
            severance_pay=_safe_float(c.severance_pay),
            year_end_bonus_untaxed=_safe_float(c.year_end_bonus_untaxed),
            year_end_bonus_net=_safe_float(c.year_end_bonus_net),
            actual_taxable=_safe_float(c.actual_taxable),
            special_deduction=_safe_float(c.special_deduction),
            remark=c.remark,
            review_status=c.review_status or "",
            calculation_status=c.calculation_status or "",
            data_completeness=c.data_completeness or "",
        )
        results.append(result)

    return results


@router.get("/pay-companies/{period}", dependencies=[Depends(require_permission("salary:view"))])
def get_pay_companies(
    period: str,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    calcs_query = db.query(SalaryCalculation).filter(SalaryCalculation.period == period)
    if not current_user.is_admin:
        calcs_query = apply_data_scope(calcs_query, db, current_user.data_scope, current_user.id, employee_id_field=SalaryCalculation.employee_id)
    calcs = calcs_query.all()
    companies = []
    seen = set()
    for c in calcs:
        key = (c.pay_company_id or -1, c.pay_company_name or c.contract_company or "未知公司")
        if key not in seen:
            seen.add(key)
            record_count = sum(1 for x in calcs if (x.pay_company_id or -1) == key[0] and (x.pay_company_name or x.contract_company or "未知公司") == key[1])
            companies.append({
                "pay_company_id": c.pay_company_id,
                "pay_company_name": c.pay_company_name or c.contract_company,
                "record_count": record_count,
                "record_types": list(set(x.record_type or "single" for x in calcs if (x.pay_company_id or -1) == key[0]))
            })
    companies.sort(key=lambda x: (0 if x["pay_company_name"] and "易玖" in x["pay_company_name"] else 1, x["pay_company_name"] or ""))
    return companies


@router.put("/results/{calc_id}", response_model=SalaryCalcOut, dependencies=[Depends(require_permission("salary:edit"))])
def update_calculation_result(
    calc_id: int,
    data: SalaryCalcUpdate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    calc = db.query(SalaryCalculation).filter(SalaryCalculation.id == calc_id).first()
    if not calc:
        raise HTTPException(status_code=404, detail="核算记录不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(calc, key, value)

    record_type = calc.record_type or "single"

    perf_std_used = _safe_float(calc.performance_standard_prorated)
    if perf_std_used is None:
        perf_std_used = _safe_float(calc.performance_standard)

    if "performance_coefficient" in update_data:
        if record_type == "contract":
            calc.actual_performance = None
        elif calc.performance_coefficient is not None and perf_std_used is not None:
            calc.actual_performance = round(perf_std_used * float(calc.performance_coefficient), 2)
        else:
            calc.actual_performance = None

    base = _safe_float(calc.base_salary_prorated)
    if base is None:
        base = _safe_float(calc.base_salary)
    allowance = _safe_float(calc.allowance_total)
    actual_perf = _safe_float(calc.actual_performance)
    commission = _safe_float(calc.commission_bonus)
    att_rate = _safe_float(calc.attendance_rate)

    if record_type != "contract":
        if actual_perf is not None and att_rate is not None:
            calc.effective_performance = round(actual_perf * att_rate, 2)
        else:
            calc.effective_performance = None

        if base is not None and perf_std_used is not None and allowance is not None:
            calc.monthly_standard = round(base + perf_std_used + allowance, 2)
        else:
            calc.monthly_standard = None

    si_hf = 0
    has_si = False
    pension = _safe_float(calc.pension_personal)
    unemployment = _safe_float(calc.unemployment_personal)
    medical = _safe_float(calc.medical_personal)
    hf = _safe_float(calc.housing_fund_personal)
    si_personal_sum = 0
    for v in [pension, unemployment, medical]:
        if v is not None:
            si_personal_sum += v
            has_si = True
    calc.social_insurance_personal = round(si_personal_sum, 2) if any(v is not None for v in [pension, unemployment, medical]) else None
    for v in [pension, unemployment, medical, hf]:
        if v is not None:
            si_hf += v
    calc.si_hf_total = round(si_hf, 2) if any(v is not None for v in [pension, unemployment, medical, hf]) else None

    if record_type == "contract":
        if calc.si_hf_total is not None:
            calc.gross_salary = calc.si_hf_total
    elif record_type == "payroll":
        calc.pension_personal = None
        calc.unemployment_personal = None
        calc.medical_personal = None
        calc.social_insurance_personal = None
        calc.housing_fund_personal = None
        calc.si_hf_total = None
    else:
        if base is not None and allowance is not None and att_rate is not None:
            c = commission if commission is not None else 0
            ep = _safe_float(calc.effective_performance) or 0
            calc.gross_salary = round((base + allowance + c) * att_rate + ep, 2)

    tax = _safe_float(calc.tax_deduction)
    pretax = _safe_float(calc.pretax_adjustment) or 0
    posttax = _safe_float(calc.posttax_adjustment) or 0
    last_month_untaxed = _safe_float(calc.last_month_untaxed) or 0
    travel_untaxed = _safe_float(calc.travel_untaxed) or 0
    special_deduction = _safe_float(calc.special_deduction) or 0
    compensation_tax = _safe_float(calc.compensation_tax) or 0
    severance_pay = _safe_float(calc.severance_pay) or 0
    year_end_bonus = _safe_float(calc.year_end_bonus_untaxed) or 0

    calc.salary_after_si_hf = None
    calc.net_salary = None
    calc.actual_taxable = None

    gross = _safe_float(calc.gross_salary)
    sh = _safe_float(calc.si_hf_total) or 0
    if gross is not None:
        calc.salary_after_si_hf = round(gross + pretax - sh, 2)
        if tax is not None:
            calc.net_salary = round(gross + pretax - sh - tax + posttax, 2)
            calc.calculation_status = "实发已核算"
        else:
            calc.net_salary = round(gross + pretax - sh + posttax, 2)
            calc.calculation_status = "应发已核算"
        calc.actual_taxable = round(gross + pretax + last_month_untaxed + travel_untaxed + compensation_tax + severance_pay + year_end_bonus - special_deduction, 2)

    net_val = _safe_float(calc.net_salary)
    if net_val is not None and net_val < 0:
        if record_type != "contract":
            raise HTTPException(status_code=400, detail="实发工资不能为负数，请检查各项数据是否正确")
    if gross is not None and has_si and record_type != "payroll" and si_hf > gross + pretax:
        raise HTTPException(status_code=400, detail="社保公积金个人合计不能超过总应发工资+税前调整，请检查社保公积金数据")

    db.commit()
    db.refresh(calc)
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "edit", f"编辑核算结果 (id={calc_id}, employee_id={calc.employee_id}, company={calc.pay_company_name})")

    emp = db.query(Employee).filter(Employee.id == calc.employee_id).first()
    return SalaryCalcOut(
        id=calc.id, period=calc.period, employee_id=calc.employee_id,
        employee_no=emp.employee_no if emp else "", employee_name=emp.name if emp else "",
        contract_company=calc.contract_company,
        pay_company_id=calc.pay_company_id,
        pay_company_name=calc.pay_company_name or "",
        record_type=calc.record_type or "single",
        department=calc.department,
        position=calc.position, status=calc.status,
        entry_date=calc.entry_date.isoformat() if calc.entry_date else None,
        base_salary=calc.base_salary,
        base_salary_prorated=calc.base_salary_prorated,
        performance_standard_prorated=calc.performance_standard_prorated,
        adjustment_id=calc.adjustment_id,
        monthly_standard=calc.monthly_standard,
        performance_standard=calc.performance_standard,
        performance_coefficient=calc.performance_coefficient,
        actual_performance=calc.actual_performance,
        effective_performance=calc.effective_performance,
        meal_allowance=calc.meal_allowance, transport_allowance=calc.transport_allowance,
        communication_allowance=calc.communication_allowance,
        computer_allowance=calc.computer_allowance,
        housing_allowance=calc.housing_allowance,
        allowance_total=calc.allowance_total,
        commission_bonus=calc.commission_bonus,
        pretax_adjustment=calc.pretax_adjustment,
        pretax_adjustment_reason=calc.pretax_adjustment_reason,
        posttax_adjustment=calc.posttax_adjustment,
        posttax_adjustment_reason=calc.posttax_adjustment_reason,
        total_work_days=calc.total_work_days,
        actual_work_days=calc.actual_work_days,
        attendance_rate=calc.attendance_rate,
        gross_salary=calc.gross_salary,
        pension_personal=calc.pension_personal,
        unemployment_personal=calc.unemployment_personal,
        medical_personal=calc.medical_personal,
        social_insurance_personal=calc.social_insurance_personal,
        housing_fund_personal=calc.housing_fund_personal,
        si_hf_total=calc.si_hf_total,
        salary_after_si_hf=calc.salary_after_si_hf,
        tax_deduction=calc.tax_deduction,
        net_salary=calc.net_salary,
        last_month_untaxed=calc.last_month_untaxed,
        travel_untaxed=calc.travel_untaxed,
        compensation_tax=calc.compensation_tax,
        severance_pay=calc.severance_pay,
        year_end_bonus_untaxed=calc.year_end_bonus_untaxed,
        year_end_bonus_net=calc.year_end_bonus_net,
        actual_taxable=calc.actual_taxable,
        special_deduction=calc.special_deduction,
        remark=calc.remark,
        review_status=calc.review_status or "",
        calculation_status=calc.calculation_status or "",
        data_completeness=calc.data_completeness or "",
    )


@router.post("/calculate-net/{period}", response_model=CalculationSummary, dependencies=[Depends(require_permission("salary:calculate", "salary:tax_import", "salary:travel_import"))])
def calculate_net_salary(period: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    if not calcs:
        _perform_gross_calculation(db, period, None)
        calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    if not calcs:
        raise HTTPException(status_code=400, detail="未找到员工薪资数据，请先导入员工信息和薪资标准")

    batch_no = f"CALC-NET-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6]}"
    start_time = datetime.now()

    total_gross = 0
    total_net = 0
    tax_imported_count = 0
    for calc in calcs:
        record_type = calc.record_type or "single"

        if record_type == "contract":
            if calc.si_hf_total is not None:
                calc.gross_salary = calc.si_hf_total
        elif record_type != "payroll":
            base = _safe_float(calc.base_salary_prorated)
            if base is None:
                base = _safe_float(calc.base_salary)
            allowance = _safe_float(calc.allowance_total)
            perf_std_used = _safe_float(calc.performance_standard_prorated)
            if perf_std_used is None:
                perf_std_used = _safe_float(calc.performance_standard)
            if base is not None and perf_std_used is not None and allowance is not None:
                calc.monthly_standard = round(base + perf_std_used + allowance, 2)
            else:
                calc.monthly_standard = None

            actual_perf = _safe_float(calc.actual_performance)
            att_rate = _safe_float(calc.attendance_rate)
            commission = _safe_float(calc.commission_bonus) or 0
            if actual_perf is not None and att_rate is not None:
                calc.effective_performance = round(actual_perf * att_rate, 2)
            else:
                calc.effective_performance = None

            if base is not None and allowance is not None and att_rate is not None:
                ep = _safe_float(calc.effective_performance) or 0
                calc.gross_salary = round((base + allowance + commission) * att_rate + ep, 2)

        si_hf = 0
        has_si = False
        pension = _safe_float(calc.pension_personal)
        unemployment = _safe_float(calc.unemployment_personal)
        medical = _safe_float(calc.medical_personal)
        hf = _safe_float(calc.housing_fund_personal)
        si_personal_sum = 0
        for v in [pension, unemployment, medical]:
            if v is not None:
                si_personal_sum += v
                has_si = True
        calc.social_insurance_personal = round(si_personal_sum, 2) if any(v is not None for v in [pension, unemployment, medical]) else None
        for v in [pension, unemployment, medical, hf]:
            if v is not None:
                si_hf += v
        if record_type == "payroll":
            calc.pension_personal = None
            calc.unemployment_personal = None
            calc.medical_personal = None
            calc.social_insurance_personal = None
            calc.housing_fund_personal = None
            calc.si_hf_total = None
            si_hf = 0
        else:
            calc.si_hf_total = round(si_hf, 2) if any(v is not None for v in [pension, unemployment, medical, hf]) else None

        tax = _safe_float(calc.tax_deduction)
        pretax = _safe_float(calc.pretax_adjustment) or 0
        posttax = _safe_float(calc.posttax_adjustment) or 0
        last_month_untaxed = _safe_float(calc.last_month_untaxed) or 0
        travel_untaxed = _safe_float(calc.travel_untaxed) or 0
        special_deduction = _safe_float(calc.special_deduction) or 0
        compensation_tax = _safe_float(calc.compensation_tax) or 0
        severance_pay = _safe_float(calc.severance_pay) or 0
        year_end_bonus = _safe_float(calc.year_end_bonus_untaxed) or 0
        gross = _safe_float(calc.gross_salary)

        calc.salary_after_si_hf = None
        calc.net_salary = None
        calc.actual_taxable = None

        sh = si_hf if has_si and record_type != "payroll" else 0
        if gross is not None:
            calc.salary_after_si_hf = round(gross + pretax - sh, 2)
            if tax is not None:
                calc.net_salary = round(gross + pretax - sh - tax + posttax, 2)
                tax_imported_count += 1
            else:
                calc.net_salary = round(gross + pretax - sh + posttax, 2)
            calc.actual_taxable = round(gross + pretax + last_month_untaxed + travel_untaxed + compensation_tax + severance_pay + year_end_bonus - special_deduction, 2)

        calc.calculation_status = "实发已核算" if tax is not None else "应发已核算"
        if gross is not None:
            total_gross += gross
        net_val = _safe_float(calc.net_salary)
        if net_val is not None:
            total_net += net_val

    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()

    log = CalculationLog(
        batch_no=batch_no,
        period=period,
        calculation_type="实发工资核算",
        status="成功",
        start_time=start_time,
        end_time=end_time,
        duration_seconds=duration,
        total_employees=len(calcs),
        success_count=len(calcs),
        failed_count=0,
        operator_id=current_user.id
    )
    db.add(log)
    db.commit()
    write_log(db, "operation", current_user.id, current_user.username, "salary", "calculate", f"实发工资核算 (period={period}, batch={batch_no})")

    return CalculationSummary(
        period=period,
        total_employees=len(calcs),
        success_count=len(calcs),
        failed_count=0,
        total_gross_salary=round(total_gross, 2),
        avg_gross_salary=round(total_gross / len(calcs), 2) if calcs else 0,
        total_net_salary=round(total_net, 2),
        avg_net_salary=round(total_net / len(calcs), 2) if calcs else 0,
        batch_no=batch_no
    )


@router.get("/logs/{period}", dependencies=[Depends(require_permission("salary:view"))])
def get_calculation_logs(period: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    logs = db.query(CalculationLog).filter(CalculationLog.period == period).order_by(CalculationLog.start_time.desc()).all()
    return [
        {
            "batch_no": l.batch_no,
            "calculation_type": l.calculation_type,
            "status": l.status,
            "start_time": str(l.start_time),
            "end_time": str(l.end_time) if l.end_time else None,
            "duration_seconds": float(l.duration_seconds) if l.duration_seconds else 0,
            "total_employees": l.total_employees,
            "success_count": l.success_count,
            "failed_count": l.failed_count
        }
        for l in logs
    ]


@router.get("/export/{period}", dependencies=[Depends(require_permission("salary:export"))])
def export_salary_results(
    period: str,
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    calc_map = {c.employee_id: c for c in calcs}
    employees = filter_active_employees(db.query(Employee), db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()

    missing = [emp.name for emp in employees if emp.id not in calc_map]
    if missing or not calcs:
        raise HTTPException(status_code=400, detail=f"请先完成薪资核算后再导出" + (f"，缺少：{', '.join(missing)}" if missing else ""))

    def _f(val):
        v = _safe_float(val)
        return v if v is not None else ""

    wb = Workbook()
    ws = wb.active
    ws.title = f"薪酬核算_{period}"
    headers = [
        "序号", "合同公司", "姓名", "部门", "职务", "费用负责人", "状态", "入职时间",
        "当月总计薪天数", "实际计薪天数", "基本工资", "提成/项目奖金/补发", "税前调整金额", "税前调整原因",
        "餐补", "交通补", "通讯补", "电脑补贴（非固定收入）", "住房补（非固定收入）", "补贴合计",
        "绩效奖金标准", "实发绩效奖金系数", "实发绩效奖金标准", "实发绩效奖金",
        "月薪标准", "总应发工资",
        "养老保险（个人）", "失业保险（个人）", "医疗保险（个人）", "公积金（个人）",
        "社保、公积金（个人）合计", "扣掉社保公积金工资",
        "本月应扣个税额", "税后调整金额", "税后调整原因", "实发工资", "实发离职补偿金",
        "上月未报税金额", "临时性差旅补贴未报税费用", "本月实际工资报税金额",
        "未报税补偿金", "未报税年终奖", "实发年终奖",
        "员工编号", "成本归属", "核算状态", "审核状态"
    ]
    ws.append(headers)

    for idx, emp in enumerate(employees, 1):
        c = calc_map.get(emp.id)
        if not c:
            continue
        ws.append([
            idx,
            c.contract_company or "", emp.name,
            c.department or "", c.position or "", c.cost_owner or "", c.status or "",
            c.entry_date.isoformat() if c.entry_date else "",
            _f(c.total_work_days), _f(c.actual_work_days),
            _f(c.base_salary),
            _f(c.commission_bonus),
            _f(c.pretax_adjustment),
            c.pretax_adjustment_reason or "",
            _f(c.meal_allowance), _f(c.transport_allowance),
            _f(c.communication_allowance), _f(c.computer_allowance),
            _f(c.housing_allowance), _f(c.allowance_total),
            _f(c.performance_standard),
            _f(c.performance_coefficient), _f(c.actual_performance),
            _f(c.effective_performance),
            _f(c.monthly_standard),
            _f(c.gross_salary),
            _f(c.pension_personal), _f(c.unemployment_personal),
            _f(c.medical_personal), _f(c.housing_fund_personal),
            _f(c.si_hf_total), _f(c.salary_after_si_hf),
            _f(c.tax_deduction), _f(c.posttax_adjustment),
            c.posttax_adjustment_reason or "",
            _f(c.net_salary),
            _f(c.severance_pay),
            _f(c.last_month_untaxed), _f(c.travel_untaxed),
            _f(c.actual_taxable),
            _f(c.compensation_tax), _f(c.year_end_bonus_untaxed), _f(c.year_end_bonus_net),
            emp.employee_no, c.cost_owner or "",
            c.calculation_status or "", c.review_status or ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_{period}.xlsx"}
    )


class TaxImportItem(BaseModel):
    employee_no: str
    last_month_untaxed: Optional[float] = 0
    travel_untaxed: Optional[float] = 0
    compensation_tax: Optional[float] = 0


@router.post("/import-tax/{period}", dependencies=[Depends(require_permission("salary:tax_import", "salary:travel_import"))])
def import_tax_data(
    period: str,
    items: List[TaxImportItem],
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    updated = 0
    created = 0
    for item in items:
        emp = db.query(Employee).filter(Employee.employee_no == item.employee_no).first()
        if not emp:
            continue
        calc = db.query(SalaryCalculation).filter(
            SalaryCalculation.period == period,
            SalaryCalculation.employee_id == emp.id
        ).first()
        if not calc:
            calc = SalaryCalculation(
                period=period,
                employee_id=emp.id,
                calculation_status="待核算"
            )
            db.add(calc)
            db.flush()
            created += 1

        if item.last_month_untaxed is not None:
            calc.last_month_untaxed = item.last_month_untaxed
        if item.travel_untaxed is not None:
            calc.travel_untaxed = item.travel_untaxed
        if item.compensation_tax is not None:
            calc.compensation_tax = item.compensation_tax

        gross = _safe_float(calc.gross_salary)
        si_hf = _safe_float(calc.si_hf_total)
        pretax = _safe_float(calc.pretax_adjustment) or 0
        posttax = _safe_float(calc.posttax_adjustment) or 0
        tax = _safe_float(calc.tax_deduction)

        if si_hf is None and gross is not None:
            pension = _safe_float(calc.pension_personal) or 0
            unemployment = _safe_float(calc.unemployment_personal) or 0
            medical = _safe_float(calc.medical_personal) or 0
            hf = _safe_float(calc.housing_fund_personal) or 0
            si_hf = pension + unemployment + medical + hf
            calc.si_hf_total = si_hf

        if gross is not None:
            sh = si_hf or 0
            calc.salary_after_si_hf = round(gross + pretax - sh, 2)
            if tax is not None:
                calc.net_salary = round(gross + pretax - sh - tax + posttax, 2)
                calc.calculation_status = "实发已核算"
            else:
                calc.net_salary = round(gross + pretax - sh + posttax, 2)
                calc.calculation_status = "应发已核算"
            lm = _safe_float(calc.last_month_untaxed) or 0
            tu = _safe_float(calc.travel_untaxed) or 0
            ct = _safe_float(calc.compensation_tax) or 0
            yeb = _safe_float(calc.year_end_bonus_untaxed) or 0
            calc.actual_taxable = round(gross + pretax + lm + tu + ct + yeb, 2)
        updated += 1
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "import", f"导入报税数据 (period={period}, updated={updated}, created={created})")
    msg = f"成功导入 {updated} 条报税数据"
    if created > 0:
        msg += f"（新建 {created} 条记录）"
    return {"message": msg, "updated": updated, "created": created}


@router.post("/batch-delete", dependencies=[Depends(require_permission("salary:delete"))])
def batch_delete_salary_calculations(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的核算记录ID列表")
    calcs = db.query(SalaryCalculation).filter(SalaryCalculation.id.in_(ids)).all()
    if not calcs:
        raise HTTPException(status_code=404, detail="未找到指定的核算记录")
    for c in calcs:
        db.delete(c)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "delete", f"批量删除 {len(calcs)} 条核算记录")
    return {"message": f"成功删除 {len(calcs)} 条核算记录", "deleted_count": len(calcs)}


# ============================================================
# 月中调薪（EmployeeSalaryAdjustment）API
# ============================================================

class SalaryAdjustmentIn(BaseModel):
    period: str
    employee_id: int
    base_salary_before: float
    bonus_before: float = 0
    performance_standard_before: float = 0
    base_salary_after: float
    bonus_after: float = 0
    performance_standard_after: float = 0
    month_start: date
    adjustment_date: date
    month_end: date
    total_days: float
    days_before: float = 0
    days_after: float = 0
    base_salary_ratio: float = 1.00
    adjustment_type: str = "转正调薪"


class SalaryAdjustmentOut(BaseModel):
    id: int
    period: str
    employee_id: int
    employee_name: str = ""
    base_salary_before: float
    bonus_before: float
    performance_standard_before: float
    total_before: float
    base_salary_after: float
    bonus_after: float
    performance_standard_after: float
    total_after: float
    base_salary_prorated: float
    commission_prorated: float
    performance_standard_prorated: float
    total_prorated: float
    month_start: str
    adjustment_date: str
    month_end: str
    days_before: float
    days_after: float
    total_days: float
    base_salary_ratio: float
    adjustment_type: str

    class Config:
        from_attributes = True


def _calc_prorated(adj: EmployeeSalaryAdjustment):
    """按 Excel 公式计算折算后工资"""
    total_days = float(adj.total_days)
    days_before = float(adj.days_before)
    days_after = float(adj.days_after)

    if total_days == 0:
        return

    base_before = float(adj.base_salary_before)
    base_after = float(adj.base_salary_after)
    perf_before = float(adj.performance_standard_before)
    perf_after = float(adj.performance_standard_after)

    # 折算后基本工资 = ROUND(调前天数 * 调前基本工资 / 总计薪天数 + 调后基本工资 * 调后天数 / 总计薪天数, 2)
    adj.base_salary_prorated = round(days_before * base_before / total_days + base_after * days_after / total_days, 2)
    # 折算后绩效奖金标准 = ROUND(调前天数 * 调前绩效标准 / 总计薪天数 + 调后绩效标准 * 调后天数 / 总计薪天数, 2)
    adj.performance_standard_prorated = round(days_before * perf_before / total_days + perf_after * days_after / total_days, 2)
    # 折算后工资标准合计 = ROUND(折算后基本工资 + 提成 + 折算后绩效标准, 0)
    adj.total_prorated = round(float(adj.base_salary_prorated) + float(adj.commission_prorated or 0) + float(adj.performance_standard_prorated), 0)

    # 调前/调后合计
    adj.total_before = round(base_before + float(adj.bonus_before or 0) + perf_before, 2)
    adj.total_after = round(base_after + float(adj.bonus_after or 0) + perf_after, 2)


@router.get("/adjustments/{period}", response_model=List[SalaryAdjustmentOut], dependencies=[Depends(require_permission("salary:view"))])
def get_adjustments(period: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    """获取指定月份的月中调薪记录"""
    adjs = db.query(EmployeeSalaryAdjustment).filter(
        EmployeeSalaryAdjustment.period == period
    ).order_by(EmployeeSalaryAdjustment.id).all()

    emp_map = {e.id: e.name for e in db.query(Employee).all()}
    return [
        SalaryAdjustmentOut(
            id=a.id, period=a.period, employee_id=a.employee_id,
            employee_name=emp_map.get(a.employee_id, ""),
            base_salary_before=float(a.base_salary_before),
            bonus_before=float(a.bonus_before or 0),
            performance_standard_before=float(a.performance_standard_before),
            total_before=float(a.total_before),
            base_salary_after=float(a.base_salary_after),
            bonus_after=float(a.bonus_after or 0),
            performance_standard_after=float(a.performance_standard_after),
            total_after=float(a.total_after),
            base_salary_prorated=float(a.base_salary_prorated or 0),
            commission_prorated=float(a.commission_prorated or 0),
            performance_standard_prorated=float(a.performance_standard_prorated or 0),
            total_prorated=float(a.total_prorated or 0),
            month_start=a.month_start.isoformat(),
            adjustment_date=a.adjustment_date.isoformat(),
            month_end=a.month_end.isoformat(),
            days_before=float(a.days_before),
            days_after=float(a.days_after),
            total_days=float(a.total_days),
            base_salary_ratio=float(a.base_salary_ratio),
            adjustment_type=a.adjustment_type,
        ) for a in adjs
    ]


@router.post("/adjustments", response_model=SalaryAdjustmentOut, dependencies=[Depends(require_permission("salary:edit"))])
def create_adjustment(data: SalaryAdjustmentIn, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    """创建月中调薪记录（自动计算折算后工资）"""
    emp = db.query(Employee).filter(Employee.id == data.employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="员工不存在")

    adj = EmployeeSalaryAdjustment(
        period=data.period,
        employee_id=data.employee_id,
        base_salary_before=data.base_salary_before,
        bonus_before=data.bonus_before,
        performance_standard_before=data.performance_standard_before,
        base_salary_after=data.base_salary_after,
        bonus_after=data.bonus_after,
        performance_standard_after=data.performance_standard_after,
        month_start=data.month_start,
        adjustment_date=data.adjustment_date,
        month_end=data.month_end,
        total_days=data.total_days,
        days_before=data.days_before,
        days_after=data.days_after,
        base_salary_ratio=data.base_salary_ratio,
        adjustment_type=data.adjustment_type,
    )
    _calc_prorated(adj)
    db.add(adj)
    db.commit()
    db.refresh(adj)
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "create", f"创建月中调薪记录 (period={data.period}, employee_id={data.employee_id})")

    return SalaryAdjustmentOut(
        id=adj.id, period=adj.period, employee_id=adj.employee_id,
        employee_name=emp.name,
        base_salary_before=float(adj.base_salary_before),
        bonus_before=float(adj.bonus_before or 0),
        performance_standard_before=float(adj.performance_standard_before),
        total_before=float(adj.total_before),
        base_salary_after=float(adj.base_salary_after),
        bonus_after=float(adj.bonus_after or 0),
        performance_standard_after=float(adj.performance_standard_after),
        total_after=float(adj.total_after),
        base_salary_prorated=float(adj.base_salary_prorated or 0),
        commission_prorated=float(adj.commission_prorated or 0),
        performance_standard_prorated=float(adj.performance_standard_prorated or 0),
        total_prorated=float(adj.total_prorated or 0),
        month_start=adj.month_start.isoformat(),
        adjustment_date=adj.adjustment_date.isoformat(),
        month_end=adj.month_end.isoformat(),
        days_before=float(adj.days_before),
        days_after=float(adj.days_after),
        total_days=float(adj.total_days),
        base_salary_ratio=float(adj.base_salary_ratio),
        adjustment_type=adj.adjustment_type,
    )


@router.put("/adjustments/{adj_id}", response_model=SalaryAdjustmentOut, dependencies=[Depends(require_permission("salary:edit"))])
def update_adjustment(adj_id: int, data: SalaryAdjustmentIn, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    """更新月中调薪记录（重新计算折算后工资）"""
    adj = db.query(EmployeeSalaryAdjustment).filter(EmployeeSalaryAdjustment.id == adj_id).first()
    if not adj:
        raise HTTPException(status_code=404, detail="调薪记录不存在")

    adj.period = data.period
    adj.employee_id = data.employee_id
    adj.base_salary_before = data.base_salary_before
    adj.bonus_before = data.bonus_before
    adj.performance_standard_before = data.performance_standard_before
    adj.base_salary_after = data.base_salary_after
    adj.bonus_after = data.bonus_after
    adj.performance_standard_after = data.performance_standard_after
    adj.month_start = data.month_start
    adj.adjustment_date = data.adjustment_date
    adj.month_end = data.month_end
    adj.total_days = data.total_days
    adj.days_before = data.days_before
    adj.days_after = data.days_after
    adj.base_salary_ratio = data.base_salary_ratio
    adj.adjustment_type = data.adjustment_type
    _calc_prorated(adj)

    db.commit()
    db.refresh(adj)
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "edit", f"更新月中调薪记录 (id={adj_id})")

    emp = db.query(Employee).filter(Employee.id == adj.employee_id).first()
    return SalaryAdjustmentOut(
        id=adj.id, period=adj.period, employee_id=adj.employee_id,
        employee_name=emp.name if emp else "",
        base_salary_before=float(adj.base_salary_before),
        bonus_before=float(adj.bonus_before or 0),
        performance_standard_before=float(adj.performance_standard_before),
        total_before=float(adj.total_before),
        base_salary_after=float(adj.base_salary_after),
        bonus_after=float(adj.bonus_after or 0),
        performance_standard_after=float(adj.performance_standard_after),
        total_after=float(adj.total_after),
        base_salary_prorated=float(adj.base_salary_prorated or 0),
        commission_prorated=float(adj.commission_prorated or 0),
        performance_standard_prorated=float(adj.performance_standard_prorated or 0),
        total_prorated=float(adj.total_prorated or 0),
        month_start=adj.month_start.isoformat(),
        adjustment_date=adj.adjustment_date.isoformat(),
        month_end=adj.month_end.isoformat(),
        days_before=float(adj.days_before),
        days_after=float(adj.days_after),
        total_days=float(adj.total_days),
        base_salary_ratio=float(adj.base_salary_ratio),
        adjustment_type=adj.adjustment_type,
    )


@router.delete("/adjustments/{adj_id}", dependencies=[Depends(require_permission("salary:edit"))])
def delete_adjustment(adj_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    """删除月中调薪记录"""
    adj = db.query(EmployeeSalaryAdjustment).filter(EmployeeSalaryAdjustment.id == adj_id).first()
    if not adj:
        raise HTTPException(status_code=404, detail="调薪记录不存在")
    db.delete(adj)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "delete", f"删除月中调薪记录 (id={adj_id})")
    return {"message": "调薪记录已删除"}


def _parse_tax_refund_excel(file_content: bytes) -> List[dict]:
    """解析个人所得税申报导出表，支持多Sheet（每个公司一个Sheet），提取姓名、应补/退税额和Sheet名称"""
    results = []

    def find_column_by_header(headers, keywords, row_idx_hint=None):
        for col_idx, header_val in enumerate(headers):
            if header_val is None:
                continue
            header_str = str(header_val).strip()
            for kw in keywords:
                if kw in header_str:
                    return col_idx
        return None

    def parse_sheet_rows(ws, is_xlrd=True):
        sheet_results = []
        name_col = 1
        tax_col = None

        if is_xlrd:
            for row_idx in range(min(10, ws.nrows)):
                row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
                for col_idx, val in enumerate(row):
                    val_str = str(val).strip() if val else ''
                    if '姓名' in val_str and '纳税人' not in val_str:
                        name_col = col_idx
                    if '应补/退税额' in val_str or '应补退税额' in val_str:
                        tax_col = col_idx
            if tax_col is None:
                tax_col = 49
            for row_idx in range(8, ws.nrows):
                name = ws.cell_value(row_idx, name_col)
                if not name or str(name).strip() in ('', '合    计', '合计', '谨声明', '扣缴义务人'):
                    continue
                if str(name).strip().startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.')):
                    continue
                tax_amount = ws.cell_value(row_idx, tax_col) if tax_col < ws.ncols else None
                try:
                    tax_val = float(tax_amount) if tax_amount not in ('', '-', None) else 0.0
                except (ValueError, TypeError):
                    tax_val = 0.0
                sheet_results.append({"name": str(name).strip(), "tax_amount": tax_val})
        else:
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                for col_idx, val in enumerate(row):
                    val_str = str(val).strip() if val else ''
                    if '姓名' in val_str and '纳税人' not in val_str:
                        name_col = col_idx
                    if '应补/退税额' in val_str or '应补退税额' in val_str:
                        tax_col = col_idx
            if tax_col is None:
                tax_col = 49
            for row_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
                name = row[name_col] if name_col < len(row) else None
                if not name or str(name).strip() in ('', '合    计', '合计', '谨声明', '扣缴义务人'):
                    continue
                if str(name).strip().startswith(('1.', '2.', '3.', '4.', '5.', '6.', '7.', '8.', '9.')):
                    continue
                tax_amount = row[tax_col] if tax_col < len(row) else None
                try:
                    tax_val = float(tax_amount) if tax_amount not in ('', '-', None) else 0.0
                except (ValueError, TypeError):
                    tax_val = 0.0
                sheet_results.append({"name": str(name).strip(), "tax_amount": tax_val})
        return sheet_results

    try:
        try:
            wb = xlrd.open_workbook(file_contents=file_content)
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                sheet_name = wb.sheet_names()[sheet_idx]
                sheet_results = parse_sheet_rows(ws, is_xlrd=True)
                for item in sheet_results:
                    item["sheet_name"] = sheet_name
                results.extend(sheet_results)
        except Exception:
            wb = load_workbook(BytesIO(file_content), data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if "数据缺失提示" in sheet_name:
                    continue
                sheet_results = parse_sheet_rows(ws, is_xlrd=False)
                for item in sheet_results:
                    item["sheet_name"] = sheet_name
                results.extend(sheet_results)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析Excel文件，请确认文件格式正确：{str(e)}")
    return results


@router.post("/upload-tax-excel/{period}", dependencies=[Depends(require_permission("salary:tax_import"))])
async def upload_tax_excel(
    period: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """上传个人所得税申报导出表Excel，自动提取姓名和应补/退税额并更新（支持多Sheet按公司导入）"""
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="仅支持 .xls 或 .xlsx 格式的文件")
    
    content = await file.read()
    tax_items = _parse_tax_refund_excel(content)
    
    if not tax_items:
        raise HTTPException(status_code=400, detail="未能从文件中解析到有效数据，请确认文件格式正确")

    calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    company_names_in_db = set()
    for c in calcs:
        cn = c.pay_company_name or c.contract_company or ""
        if cn:
            company_names_in_db.add(cn)

    updated = 0
    created = 0
    not_found = []
    unmatched_sheets = set()
    for item in tax_items:
        emp = db.query(Employee).filter(Employee.name == item["name"]).first()
        if not emp:
            not_found.append(item["name"])
            continue

        sheet_name = item.get("sheet_name", "")
        target_calc = None
        if sheet_name and "数据缺失提示" not in sheet_name:
            for c in calcs:
                if c.employee_id == emp.id:
                    cn = c.pay_company_name or c.contract_company or ""
                    if sheet_name in cn or cn in sheet_name:
                        target_calc = c
                        break
        
        if target_calc is None:
            for c in calcs:
                if c.employee_id == emp.id and c.record_type != "contract":
                    target_calc = c
                    break
            if target_calc is None:
                for c in calcs:
                    if c.employee_id == emp.id:
                        target_calc = c
                        break

        if not target_calc:
            target_calc = SalaryCalculation(
                period=period,
                employee_id=emp.id,
                calculation_status="待核算"
            )
            db.add(target_calc)
            db.flush()
            created += 1

        target_calc.tax_deduction = item["tax_amount"]

        record_type = target_calc.record_type or "single"
        gross = _safe_float(target_calc.gross_salary)
        pretax = _safe_float(target_calc.pretax_adjustment) or 0
        posttax = _safe_float(target_calc.posttax_adjustment) or 0
        tax = _safe_float(target_calc.tax_deduction)
        special_deduction = _safe_float(target_calc.special_deduction) or 0

        if record_type == "payroll":
            sh = 0
        else:
            sh = _safe_float(target_calc.si_hf_total)
            if sh is None:
                pension = _safe_float(target_calc.pension_personal) or 0
                unemployment = _safe_float(target_calc.unemployment_personal) or 0
                medical = _safe_float(target_calc.medical_personal) or 0
                hf = _safe_float(target_calc.housing_fund_personal) or 0
                sh = pension + unemployment + medical + hf
                target_calc.si_hf_total = sh if sh > 0 else None

        if gross is not None:
            target_calc.salary_after_si_hf = round(gross + pretax - sh, 2)
            if tax is not None:
                target_calc.net_salary = round(gross + pretax - sh - tax + posttax, 2)
                target_calc.calculation_status = "实发已核算"
            else:
                target_calc.net_salary = round(gross + pretax - sh + posttax, 2)
                target_calc.calculation_status = "应发已核算"
            lm = _safe_float(target_calc.last_month_untaxed) or 0
            tu = _safe_float(target_calc.travel_untaxed) or 0
            ct = _safe_float(target_calc.compensation_tax) or 0
            sp = _safe_float(target_calc.severance_pay) or 0
            yeb = _safe_float(target_calc.year_end_bonus_untaxed) or 0
            target_calc.actual_taxable = round(gross + pretax + lm + tu + ct + sp + yeb - special_deduction, 2)
        updated += 1
    
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "import", f"导入个税Excel (period={period}, updated={updated}, created={created})")
    
    message = f"成功导入 {updated} 条个税数据"
    if created > 0:
        message += f"（新建 {created} 条记录）"
    if not_found:
        message += f"，{len(not_found)}人未找到：{', '.join(not_found[:10])}"
        if len(not_found) > 10:
            message += f"等{len(not_found)}人"
    
    return {
        "message": message,
        "updated": updated,
        "created": created,
        "not_found": not_found,
        "parsed_count": len(tax_items)
    }


@router.get("/export-tax-template/{period}", dependencies=[Depends(require_permission("salary:tax_export"))])
def export_tax_template(
    period: str,
    type: str = Query("salary", description="导出类型: salary(正常工资薪金), bonus(全年一次性奖金), severance(解除劳动合同一次性补偿金)"),
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """导出报税系统模板格式的表格（单个Excel文件，按type参数区分）"""
    if type not in ("salary", "bonus", "severance"):
        raise HTTPException(status_code=400, detail="无效的导出类型，可选值: salary, bonus, severance")

    calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    calc_map = {c.employee_id: c for c in calcs}
    employees = filter_active_employees(db.query(Employee), db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()
    
    period_end = _get_period_end(period)
    salary_map = {}
    for s in db.query(EmployeeSalary).filter(
        EmployeeSalary.effective_date <= period_end
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).all():
        if s.employee_id not in salary_map:
            salary_map[s.employee_id] = s

    perf_map = {p.employee_id: p for p in db.query(PerformanceScore).filter(
        PerformanceScore.period == period
    ).all()}

    att_map = {a.employee_id: a for a in db.query(AttendanceRecord).filter(
        AttendanceRecord.period == period
    ).all()}

    si_map = {s.employee_id: s for s in db.query(SocialInsurance).filter(
        SocialInsurance.period == period
    ).all()}
    
    adjustments = {a.employee_id: a for a in db.query(EmployeeSalaryAdjustment).filter(
        EmployeeSalaryAdjustment.period == period
    ).all()}

    def _create_workbook_response(wb, filename):
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        encoded_filename = quote(filename)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    def _build_salary_workbook():
        wb = Workbook()
        wb.remove(wb.active)
        headers = [
            "工号", "*姓名", "*证件类型", "*证件号码", "本期收入", "本期免税收入",
            "基本养老保险费", "基本医疗保险费", "失业保险费", "住房公积金",
            "累计子女教育", "累计继续教育", "累计住房贷款利息", "累计住房租金",
            "累计赡养老人", "累计3岁以下婴幼儿照护", "累计个人养老金",
            "企业(职业)年金", "商业健康保险", "税延养老保险", "其他",
            "准予扣除的捐赠额", "减免税额", "备注"
        ]

        company_groups = {}
        for c in calcs:
            emp_id = c.employee_id
            emp = next((e for e in employees if e.id == emp_id), None)
            if not emp:
                continue
            company_name = c.pay_company_name or c.contract_company or "未知公司"
            if company_name not in company_groups:
                company_groups[company_name] = []
            company_groups[company_name].append((emp, c))

        all_warnings = []
        if not company_groups:
            ws = wb.create_sheet("正常工资薪金收入")
            ws.append(headers)
            for emp in employees:
                ws.append(["", emp.name, "居民身份证", emp.id_card or "", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, ""])
        else:
            sorted_companies = sorted(company_groups.keys(), key=lambda x: (0 if "易玖" in x else 1, x))
            for company_name in sorted_companies:
                sheet_name = company_name[:31]
                ws = wb.create_sheet(sheet_name)
                ws.append(headers)
                company_records = company_groups[company_name]
                for emp, c in company_records:
                    remarks = []
                    id_card = emp.id_card or ""
                    if not id_card:
                        remarks.append("证件号码缺失")

                    record_type = c.record_type or "single"
                    taxable_income = float(c.actual_taxable or 0)
                    pension = float(c.pension_personal or 0) if record_type != "payroll" else 0
                    medical = float(c.medical_personal or 0) if record_type != "payroll" else 0
                    unemployment = float(c.unemployment_personal or 0) if record_type != "payroll" else 0
                    housing = float(c.housing_fund_personal or 0) if record_type != "payroll" else 0

                    if taxable_income > 0 or pension > 0 or medical > 0 or unemployment > 0 or housing > 0:
                        if remarks:
                            all_warnings.append(f"[{company_name}] {emp.name}({emp.employee_no}): {', '.join(remarks)}")
                        ws.append([
                            "", emp.name, "居民身份证", id_card,
                            round(taxable_income, 2), 0,
                            round(pension, 2), round(medical, 2), round(unemployment, 2), round(housing, 2),
                            0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                            ""
                        ])

        if all_warnings:
            ws_warn = wb.create_sheet("数据缺失提示")
            ws_warn.append(["以下员工存在必填字段缺失或数据不完整，导入报税系统前请补充："])
            ws_warn.append([])
            for w in all_warnings:
                ws_warn.append([w])
        return _create_workbook_response(wb, f"正常工资薪金_{period}.xlsx")

    def _build_bonus_workbook():
        wb = Workbook()
        ws = wb.active
        ws.title = "全年一次性奖金收入"
        headers = [
            "工号", "*姓名", "*证件类型", "*证件号码", "*全年一次性奖金额",
            "免税收入", "其他", "准予扣除的捐赠额", "减免税额", "备注"
        ]
        ws.append(headers)
        warnings = []
        bonus_emps = []
        for emp in employees:
            c = calc_map.get(emp.id)
            amount = float(c.year_end_bonus_untaxed or 0) if c else 0
            if amount > 0:
                bonus_emps.append((emp, amount))
        if not bonus_emps:
            warnings.append(f"{period} 没有全年一次性奖金数据，无需导入")
        else:
            for emp, amount in bonus_emps:
                id_card = emp.id_card or ""
                remarks = []
                if not id_card:
                    remarks.append("证件号码缺失")
                if remarks:
                    warnings.append(f"{emp.name}({emp.employee_no}): {', '.join(remarks)}")
                ws.append([
                    "", emp.name, "居民身份证", id_card,
                    round(amount, 2), 0, 0, 0, 0, ""
                ])
        if warnings:
            ws2 = wb.create_sheet("数据缺失提示")
            ws2.append(["提示信息："])
            ws2.append([])
            for w in warnings:
                ws2.append([w])
        return _create_workbook_response(wb, f"全年一次性奖金_{period}.xlsx")

    def _build_severance_workbook():
        wb = Workbook()
        ws = wb.active
        ws.title = "解除劳动合同一次性补偿金"
        headers = [
            "工号", "*姓名", "*证件类型", "*证件号码", "*一次性补偿收入",
            "免税收入", "其他", "准予扣除的捐赠额", "减免税额", "备注"
        ]
        ws.append(headers)
        warnings = []
        comp_emps = []
        for emp in employees:
            c = calc_map.get(emp.id)
            amount = float(c.compensation_tax or 0) if c else 0
            if amount > 0:
                comp_emps.append((emp, amount))
        if not comp_emps:
            warnings.append(f"{period} 没有解除劳动合同一次性补偿金数据，无需导入")
        else:
            for emp, amount in comp_emps:
                id_card = emp.id_card or ""
                remarks = []
                if not id_card:
                    remarks.append("证件号码缺失")
                if remarks:
                    warnings.append(f"{emp.name}({emp.employee_no}): {', '.join(remarks)}")
                ws.append([
                    "", emp.name, "居民身份证", id_card,
                    round(amount, 2), 0, 0, 0, 0, ""
                ])
        if warnings:
            ws2 = wb.create_sheet("数据缺失提示")
            ws2.append(["提示信息："])
            ws2.append([])
            for w in warnings:
                ws2.append([w])
        return _create_workbook_response(wb, f"解除劳动合同一次性补偿金_{period}.xlsx")

    if type == "salary":
        return _build_salary_workbook()
    elif type == "bonus":
        return _build_bonus_workbook()
    elif type == "severance":
        return _build_severance_workbook()


def _parse_travel_untaxed_excel(file_content: bytes) -> dict:
    """解析临时性差旅补贴未报税Excel，兼容明细表和数据透视表，返回{员工姓名: 报税应纳税所得额合计}"""
    
    def _parse_xlrd_sheet(ws, is_pivot=False):
        """解析xlrd工作表，返回解析结果dict或None"""
        name_col = None
        amount_col = None
        data_start_row = None
        
        if is_pivot:
            for row_idx in range(min(15, ws.nrows)):
                row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
                for col_idx, val in enumerate(row):
                    val_str = str(val).strip() if val else ''
                    if val_str == '员工姓名':
                        name_col = col_idx
                    if '报税应纳税所得额' in val_str and '求和项' in val_str:
                        amount_col = col_idx
                if name_col is not None and amount_col is not None:
                    data_start_row = row_idx + 1
                    break
        else:
            for row_idx in range(min(15, ws.nrows)):
                row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
                for col_idx, val in enumerate(row):
                    val_str = str(val).strip() if val else ''
                    if val_str == '员工姓名':
                        name_col = col_idx
                    if val_str == '报税应纳税所得额':
                        amount_col = col_idx
                if name_col is not None and amount_col is not None:
                    data_start_row = row_idx + 1
                    break
        
        if name_col is None or amount_col is None or data_start_row is None:
            return None
        
        sheet_result = {}
        hit_total = False
        for row_idx in range(data_start_row, ws.nrows):
            if name_col >= ws.ncols or amount_col >= ws.ncols:
                continue
            name_val = ws.cell_value(row_idx, name_col)
            amount_val = ws.cell_value(row_idx, amount_col)
            
            if not name_val:
                if hit_total:
                    break
                continue
            name = str(name_val).strip()
            
            if name in ('总计', '合计', '合    计'):
                hit_total = True
                continue
            
            if name == '员工姓名' or name == '':
                if hit_total:
                    break
                continue
            
            try:
                amount = float(amount_val) if amount_val not in ('', '-', None) else 0.0
            except (ValueError, TypeError):
                amount = 0.0
            
            if name in sheet_result:
                sheet_result[name] += amount
            else:
                sheet_result[name] = amount
        
        return sheet_result if sheet_result else None
    
    def _parse_openpyxl_sheet(ws, is_pivot=False):
        """解析openpyxl工作表，返回解析结果dict或None"""
        name_col = None
        amount_col = None
        data_start_row = None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            for col_idx, val in enumerate(row):
                val_str = str(val).strip() if val else ''
                if val_str == '员工姓名':
                    name_col = col_idx
                if is_pivot:
                    if '报税应纳税所得额' in val_str and '求和项' in val_str:
                        amount_col = col_idx
                else:
                    if val_str == '报税应纳税所得额':
                        amount_col = col_idx
            if name_col is not None and amount_col is not None:
                data_start_row = row_idx + 1
                break
        
        if name_col is None or amount_col is None or data_start_row is None:
            return None
        
        sheet_result = {}
        hit_total = False
        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            if name_col >= len(row) or amount_col >= len(row):
                continue
            name_val = row[name_col]
            amount_val = row[amount_col]
            
            if not name_val:
                if hit_total:
                    break
                continue
            name = str(name_val).strip()
            
            if name in ('总计', '合计', '合    计'):
                hit_total = True
                continue
            
            if name == '员工姓名' or name == '':
                if hit_total:
                    break
                continue
            
            try:
                amount = float(amount_val) if amount_val not in ('', '-', None) else 0.0
            except (ValueError, TypeError):
                amount = 0.0
            
            if name in sheet_result:
                sheet_result[name] += amount
            else:
                sheet_result[name] = amount
        
        return sheet_result if sheet_result else None
    
    result = None
    
    try:
        wb = xlrd.open_workbook(file_contents=file_content)
        sheets = []
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheets.append((ws.name, ws))
        
        pivot_sheets = [(name, ws) for name, ws in sheets if 'Sheet2' in name or '透视' in name]
        other_sheets = [(name, ws) for name, ws in sheets if not ('Sheet2' in name or '透视' in name)]
        
        for name, ws in pivot_sheets + other_sheets:
            is_pivot = ('Sheet2' in name or '透视' in name)
            res = _parse_xlrd_sheet(ws, is_pivot=is_pivot)
            if res:
                result = res
                break
            if not is_pivot:
                res = _parse_xlrd_sheet(ws, is_pivot=True)
                if res:
                    result = res
                    break
    except Exception:
        pass
    
    if result is None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_content), data_only=True)
            sheets = []
            for ws in wb.worksheets:
                sheets.append((ws.title, ws))
            
            pivot_sheets = [(name, ws) for name, ws in sheets if 'Sheet2' in name or '透视' in name]
            other_sheets = [(name, ws) for name, ws in sheets if not ('Sheet2' in name or '透视' in name)]
            
            for name, ws in pivot_sheets + other_sheets:
                is_pivot = ('Sheet2' in name or '透视' in name)
                res = _parse_openpyxl_sheet(ws, is_pivot=is_pivot)
                if res:
                    result = res
                    break
                if not is_pivot:
                    res = _parse_openpyxl_sheet(ws, is_pivot=True)
                    if res:
                        result = res
                        break
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"无法解析Excel文件，请确认文件格式正确：{str(e)}")
    
    if result is None:
        return {}
    
    rounded = {}
    for name, amount in result.items():
        rounded[name] = round(amount, 2)
    return rounded


@router.post("/import-travel-untaxed/{period}", dependencies=[Depends(require_permission("salary:travel_import"))])
async def import_travel_untaxed(
    period: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """导入临时性差旅补贴未报税费用Excel，自动识别明细表/数据透视表，按员工汇总保存到TravelReimbursement表"""
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="仅支持 .xls 或 .xlsx 格式的文件")
    
    content = await file.read()
    travel_data = _parse_travel_untaxed_excel(content)
    
    if not travel_data:
        raise HTTPException(status_code=400, detail="未能从文件中解析到有效数据，请确认文件格式正确")

    emp_map = {e.name: e for e in db.query(Employee).all()}
    calcs = db.query(SalaryCalculation).filter(SalaryCalculation.period == period).all()
    calc_map = {c.employee_id: c for c in calcs}

    imported = 0
    not_found = []
    
    for name, amount in travel_data.items():
        emp = emp_map.get(name)
        if not emp:
            not_found.append(name)
            continue

        tr = TravelReimbursement(
            period=period,
            employee_id=emp.id,
            amount=round(amount, 2),
            description="临时性差旅补贴未报税费用导入"
        )
        db.add(tr)

        calc = calc_map.get(emp.id)
        if calc:
            existing_travel = _safe_float(calc.travel_untaxed) or 0
            calc.travel_untaxed = round(existing_travel + amount, 2)

            gross = _safe_float(calc.gross_salary)
            si_hf = _safe_float(calc.si_hf_total)
            pretax = _safe_float(calc.pretax_adjustment) or 0
            posttax = _safe_float(calc.posttax_adjustment) or 0
            tax = _safe_float(calc.tax_deduction)
            special_deduction = _safe_float(calc.special_deduction) or 0

            if si_hf is None and gross is not None:
                pension = _safe_float(calc.pension_personal) or 0
                unemployment = _safe_float(calc.unemployment_personal) or 0
                medical = _safe_float(calc.medical_personal) or 0
                hf = _safe_float(calc.housing_fund_personal) or 0
                si_hf = pension + unemployment + medical + hf
                calc.si_hf_total = si_hf if si_hf > 0 else None

            if gross is not None:
                sh = si_hf or 0
                calc.salary_after_si_hf = round(gross + pretax - sh, 2)
                if tax is not None:
                    calc.net_salary = round(gross + pretax - sh - tax + posttax, 2)
                    calc.calculation_status = "实发已核算"
                else:
                    calc.net_salary = round(gross + pretax - sh + posttax, 2)
                    calc.calculation_status = "应发已核算"
                lm = _safe_float(calc.last_month_untaxed) or 0
                tu = _safe_float(calc.travel_untaxed) or 0
                ct = _safe_float(calc.compensation_tax) or 0
                sp = _safe_float(calc.severance_pay) or 0
                yeb = _safe_float(calc.year_end_bonus_untaxed) or 0
                calc.actual_taxable = round(gross + pretax + lm + tu + ct + sp + yeb - special_deduction, 2)

        imported += 1
    
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "salary", "import", f"导入临时性差旅补贴 (period={period}, imported={imported})")
    
    message = f"成功导入 {imported} 条临时性差旅补贴数据"
    if not_found:
        message += f"，{len(not_found)}人未找到：{', '.join(not_found[:10])}"
        if len(not_found) > 10:
            message += f"等{len(not_found)}人"
    
    return {
        "message": message,
        "imported": imported,
        "not_found": not_found,
        "parsed_count": len(travel_data)
    }