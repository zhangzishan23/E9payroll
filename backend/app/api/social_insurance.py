from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from io import BytesIO
from urllib.parse import quote
import logging
from openpyxl import Workbook
from app.core.database import get_db
from app.core.log_helper import write_log
from app.core.query_utils import filter_active_employees, apply_data_scope
from app.models.models import SocialInsurance, Employee, SiImportTemplate, SiImportLog
from app.api.auth import get_current_user, UserInfo, require_permission
from app.services.si_import_engine import (
    run_smart_import, SI_FIELD_LABELS, auto_detect_template,
    save_batch_files, precheck_batch, auto_detect_from_batch,
    run_smart_import_from_batch, cleanup_batch, create_batch_from_unmatched,
)

logger = logging.getLogger(__name__)
router = APIRouter()


class SocialInsuranceOut(BaseModel):
    id: Optional[int] = None
    period: str
    employee_id: int
    employee_no: str = ""
    employee_name: str = ""
    employee_social_insurance_no: Optional[str] = None
    # 社保 — 各险种基数
    pension_personal_base: Optional[float] = None
    pension_company_base: Optional[float] = None
    unemployment_personal_base: Optional[float] = None
    unemployment_company_base: Optional[float] = None
    medical_personal_base: Optional[float] = None
    medical_company_base: Optional[float] = None
    injury_company_base: Optional[float] = None
    # 社保 — 各险种金额
    pension_personal: Optional[float] = None
    pension_company: Optional[float] = None
    unemployment_personal: Optional[float] = None
    unemployment_company: Optional[float] = None
    medical_personal: Optional[float] = None
    medical_company: Optional[float] = None
    injury_company: Optional[float] = None
    si_personal: Optional[float] = None
    si_company: Optional[float] = None
    # 社保 — 各险种比例
    pension_personal_rate: Optional[float] = None
    pension_company_rate: Optional[float] = None
    unemployment_personal_rate: Optional[float] = None
    unemployment_company_rate: Optional[float] = None
    medical_personal_rate: Optional[float] = None
    medical_company_rate: Optional[float] = None
    injury_company_rate: Optional[float] = None
    # 社保 — 合计
    pension_total: Optional[float] = None
    unemployment_total: Optional[float] = None
    medical_total: Optional[float] = None
    injury_total: Optional[float] = None
    si_grand_total: Optional[float] = None
    # 公积金
    hf_base: Optional[float] = None
    hf_personal: Optional[float] = None
    hf_company: Optional[float] = None
    hf_personal_rate: Optional[float] = None
    hf_company_rate: Optional[float] = None
    hf_total: Optional[float] = None
    # 总合计
    grand_total: Optional[float] = None

    class Config:
        from_attributes = True


class SocialInsuranceCreate(BaseModel):
    period: str
    employee_id: int
    employee_social_insurance_no: Optional[str] = None
    pension_personal_base: float = 0
    pension_company_base: float = 0
    unemployment_personal_base: float = 0
    unemployment_company_base: float = 0
    medical_personal_base: float = 0
    medical_company_base: float = 0
    injury_company_base: float = 0
    pension_personal: float = 0
    unemployment_personal: float = 0
    medical_personal: float = 0
    si_personal: float = 0
    pension_company: float = 0
    unemployment_company: float = 0
    medical_company: float = 0
    injury_company: float = 0
    si_company: float = 0
    pension_personal_rate: float = 0
    pension_company_rate: float = 0
    unemployment_personal_rate: float = 0
    unemployment_company_rate: float = 0
    medical_personal_rate: float = 0
    medical_company_rate: float = 0
    injury_company_rate: float = 0
    pension_total: float = 0
    unemployment_total: float = 0
    medical_total: float = 0
    injury_total: float = 0
    si_grand_total: float = 0
    hf_base: float = 0
    hf_personal: float = 0
    hf_company: float = 0
    hf_personal_rate: float = 0
    hf_company_rate: float = 0
    hf_total: float = 0
    grand_total: float = 0


class SocialInsuranceUpdate(BaseModel):
    employee_social_insurance_no: Optional[str] = None
    pension_personal_base: Optional[float] = None
    pension_company_base: Optional[float] = None
    unemployment_personal_base: Optional[float] = None
    unemployment_company_base: Optional[float] = None
    medical_personal_base: Optional[float] = None
    medical_company_base: Optional[float] = None
    injury_company_base: Optional[float] = None
    pension_personal: Optional[float] = None
    unemployment_personal: Optional[float] = None
    medical_personal: Optional[float] = None
    si_personal: Optional[float] = None
    pension_company: Optional[float] = None
    unemployment_company: Optional[float] = None
    medical_company: Optional[float] = None
    injury_company: Optional[float] = None
    si_company: Optional[float] = None
    pension_personal_rate: Optional[float] = None
    pension_company_rate: Optional[float] = None
    unemployment_personal_rate: Optional[float] = None
    unemployment_company_rate: Optional[float] = None
    medical_personal_rate: Optional[float] = None
    medical_company_rate: Optional[float] = None
    injury_company_rate: Optional[float] = None
    pension_total: Optional[float] = None
    unemployment_total: Optional[float] = None
    medical_total: Optional[float] = None
    injury_total: Optional[float] = None
    si_grand_total: Optional[float] = None
    hf_base: Optional[float] = None
    hf_personal: Optional[float] = None
    hf_company: Optional[float] = None
    hf_personal_rate: Optional[float] = None
    hf_company_rate: Optional[float] = None
    hf_total: Optional[float] = None
    grand_total: Optional[float] = None


class SiImportTemplateOut(BaseModel):
    id: int
    name: str
    source_category: str
    file_type: str
    city: Optional[str] = None
    description: Optional[str] = None
    file_pattern: Optional[str] = None
    file_keywords: Optional[list] = None
    sheet_pattern: Optional[str] = None
    header_rows: List[int]
    data_start_row: int
    skip_footer_rows: int = 0
    column_mappings: dict
    row_filters: Optional[dict] = None
    number_format: Optional[dict] = None
    default_rates: Optional[dict] = None
    is_active: bool = True
    sort_order: int = 0

    class Config:
        from_attributes = True


class SiImportTemplateCreate(BaseModel):
    name: str
    source_category: str
    file_type: str
    city: Optional[str] = None
    description: Optional[str] = None
    file_pattern: Optional[str] = None
    file_keywords: Optional[list] = None
    sheet_pattern: Optional[str] = None
    header_rows: List[int]
    data_start_row: int
    skip_footer_rows: int = 0
    column_mappings: dict
    row_filters: Optional[dict] = None
    number_format: Optional[dict] = None
    default_rates: Optional[dict] = None
    is_active: bool = True
    sort_order: int = 0


class SiImportTemplateUpdate(BaseModel):
    name: Optional[str] = None
    source_category: Optional[str] = None
    file_type: Optional[str] = None
    city: Optional[str] = None
    description: Optional[str] = None
    file_pattern: Optional[str] = None
    file_keywords: Optional[list] = None
    sheet_pattern: Optional[str] = None
    header_rows: Optional[List[int]] = None
    data_start_row: Optional[int] = None
    skip_footer_rows: Optional[int] = None
    column_mappings: Optional[dict] = None
    row_filters: Optional[dict] = None
    number_format: Optional[dict] = None
    default_rates: Optional[dict] = None
    is_active: Optional[bool] = None
    sort_order: Optional[int] = None


class SiImportLogOut(BaseModel):
    id: int
    period: str
    batch_id: str
    file_name: Optional[str] = None
    row_number: Optional[int] = None
    employee_name: Optional[str] = None
    error_level: str
    error_type: str
    error_message: str
    resolved: bool = False

    class Config:
        from_attributes = True


@router.get("/", response_model=List[SocialInsuranceOut], dependencies=[Depends(require_permission("insurance:view"))])
def get_social_insurance(
    period: Optional[str] = Query(None),
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if period:
        from app.services.si_import_engine import _calc_summaries

        query = db.query(Employee)
        if not current_user.is_admin:
            query = apply_data_scope(query, db, current_user.data_scope, current_user.id)
        employees = filter_active_employees(query, db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()
        si_map = {}
        records = db.query(SocialInsurance).filter(SocialInsurance.period == period).all()
        needs_commit = False
        for r in records:
            si_map[r.employee_id] = r
            # 自动修复缺失的合计（兜底：已有分项数据但合计为0时自动计算）
            if r.si_grand_total is None or r.si_grand_total == 0:
                if (r.pension_personal or 0) or (r.pension_company or 0) or (r.hf_personal or 0):
                    _calc_summaries(r)
                    needs_commit = True
        if needs_commit:
            db.commit()

        result = []
        for emp in employees:
            si = si_map.get(emp.id)
            if si:
                out = SocialInsuranceOut.model_validate(si)
                out.employee_no = emp.employee_no
                out.employee_name = emp.name
            else:
                out = SocialInsuranceOut(
                    id=None, period=period, employee_id=emp.id,
                    employee_no=emp.employee_no, employee_name=emp.name,
                )
            result.append(out)
        return result

    records = db.query(SocialInsurance).order_by(SocialInsurance.period.desc()).all()

    # 自动修复缺失的合计
    needs_commit = False
    from app.services.si_import_engine import _calc_summaries
    for r in records:
        if r.si_grand_total is None or r.si_grand_total == 0:
            if (r.pension_personal or 0) or (r.pension_company or 0) or (r.hf_personal or 0):
                _calc_summaries(r)
                needs_commit = True
    if needs_commit:
        db.commit()

    return [SocialInsuranceOut.model_validate(r) for r in records]


@router.post("/", response_model=SocialInsuranceOut, dependencies=[Depends(require_permission("insurance:create"))])
def create_social_insurance(
    data: SocialInsuranceCreate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    existing = db.query(SocialInsurance).filter(
        SocialInsurance.period == data.period,
        SocialInsurance.employee_id == data.employee_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该员工本月社保公积金记录已存在，请使用编辑功能")

    emp = db.query(Employee).filter(Employee.id == data.employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="员工不存在")

    from app.services.si_import_engine import _calc_summaries

    si = SocialInsurance(
        period=data.period,
        employee_id=data.employee_id,
        **data.model_dump(exclude={"period", "employee_id"}),
    )
    db.add(si)
    db.commit()
    db.refresh(si)
    _calc_summaries(si)
    db.commit()
    db.refresh(si)
    write_log(db, "data_change", current_user.id, current_user.username, "social_insurance", "create", f"新增社保公积金记录 (period={data.period}, employee={emp.name})")
    out = SocialInsuranceOut.model_validate(si)
    out.employee_no = emp.employee_no
    out.employee_name = emp.name
    return out


@router.put("/{record_id}", response_model=SocialInsuranceOut, dependencies=[Depends(require_permission("insurance:edit"))])
def update_social_insurance(
    record_id: int,
    data: SocialInsuranceUpdate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    si = db.query(SocialInsurance).filter(SocialInsurance.id == record_id).first()
    if not si:
        raise HTTPException(status_code=404, detail="社保公积金记录不存在")

    emp = db.query(Employee).filter(Employee.id == si.employee_id).first()
    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(si, key, value)

    db.commit()
    db.refresh(si)

    from app.services.si_import_engine import _calc_summaries
    _calc_summaries(si)
    db.commit()
    db.refresh(si)
    write_log(db, "data_change", current_user.id, current_user.username, "social_insurance", "edit", f"编辑社保公积金记录 (id={record_id})")
    out = SocialInsuranceOut.model_validate(si)
    out.employee_no = emp.employee_no if emp else ""
    out.employee_name = emp.name if emp else ""
    return out


@router.post("/batch-delete", dependencies=[Depends(require_permission("insurance:delete"))])
def batch_delete_social_insurance(
    ids: List[int],
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    valid_ids = [i for i in ids if i is not None and isinstance(i, int)]
    if not valid_ids:
        raise HTTPException(status_code=400, detail="请选择要删除的有效社保公积金记录")
    records = db.query(SocialInsurance).filter(SocialInsurance.id.in_(valid_ids)).all()
    if not records:
        raise HTTPException(status_code=404, detail="未找到指定的社保公积金记录")
    deleted_count = len(records)
    for r in records:
        db.delete(r)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "social_insurance", "batch_delete", f"批量删除 {deleted_count} 条社保公积金记录")
    return {"message": f"成功删除 {deleted_count} 条社保公积金记录", "deleted_count": deleted_count}


@router.delete("/{record_id}", dependencies=[Depends(require_permission("insurance:delete"))])
def delete_social_insurance(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    si = db.query(SocialInsurance).filter(SocialInsurance.id == record_id).first()
    if not si:
        raise HTTPException(status_code=404, detail="社保公积金记录不存在")
    db.delete(si)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "social_insurance", "delete", f"删除社保公积金记录 (id={record_id})")
    return {"message": "删除成功"}


def _safe_float(val):
    """安全将单元格值转为float，失败返回None"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ("-", "--", "——", "/", ""):
        return None
    s = s.replace(",", "").replace("，", "").replace("¥", "").replace("￥", "").replace("$", "")
    s = s.replace("元", "").replace("圆", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


@router.post("/import/{period}", dependencies=[Depends(require_permission("insurance:import"))])
def import_social_insurance(
    period: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file.file)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析Excel文件，请确认格式正确：{str(e)}")

    from app.services.si_import_engine import _calc_summaries

    if not all_rows or len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel文件内容为空或格式不正确")

    # ── 读取表头，建立动态列映射 ──
    header_row = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    data_rows = all_rows[1:]

    # 表头文本 → 数据库字段名映射
    HEADER_TO_FIELD = {
        "员工编号": "_emp_no",
        "工号": "_emp_no",
        "员工工号": "_emp_no",
        "姓名": "_emp_name",
        "员工姓名": "_emp_name",
        "养老保险基数(个人)": "pension_personal_base",
        "养老保险基数（个人）": "pension_personal_base",
        "养老基数(个人)": "pension_personal_base",
        "养老保险基数(单位)": "pension_company_base",
        "养老保险基数（单位）": "pension_company_base",
        "养老基数(单位)": "pension_company_base",
        "养老保险个人": "pension_personal",
        "养老保险公司": "pension_company",
        "失业保险基数(个人)": "unemployment_personal_base",
        "失业保险基数（个人）": "unemployment_personal_base",
        "失业基数(个人)": "unemployment_personal_base",
        "失业保险基数(单位)": "unemployment_company_base",
        "失业保险基数（单位）": "unemployment_company_base",
        "失业基数(单位)": "unemployment_company_base",
        "失业保险个人": "unemployment_personal",
        "失业保险公司": "unemployment_company",
        "医疗保险基数(个人)": "medical_personal_base",
        "医疗保险基数（个人）": "medical_personal_base",
        "医疗基数(个人)": "medical_personal_base",
        "医疗保险基数(单位)": "medical_company_base",
        "医疗保险基数（单位）": "medical_company_base",
        "医疗基数(单位)": "medical_company_base",
        "医疗保险个人": "medical_personal",
        "医疗保险公司": "medical_company",
        "工伤保险基数(单位)": "injury_company_base",
        "工伤保险基数（单位）": "injury_company_base",
        "工伤基数(单位)": "injury_company_base",
        "工伤保险公司": "injury_company",
        "社保个人合计": "si_personal",
        "社保公司合计": "si_company",
        "社保个人": "si_personal",
        "社保公司": "si_company",
        "公积金基数": "hf_base",
        "公积金缴存基数": "hf_base",
        "公积金个人": "hf_personal",
        "公积金公司": "hf_company",
        "公积金个人金额": "hf_personal",
        "公积金单位金额": "hf_company",
    }

    # 建立列索引映射：col_idx → field_name
    col_map = {}
    emp_no_col = None
    emp_name_col = None
    for idx, header_text in enumerate(header_row):
        if not header_text:
            continue
        field = HEADER_TO_FIELD.get(header_text)
        if field == "_emp_no":
            emp_no_col = idx
        elif field == "_emp_name":
            emp_name_col = idx
        elif field:
            col_map[idx] = field

    # 检测是否为新格式（有表头匹配），否则回退到旧硬编码格式
    use_new_format = emp_no_col is not None and len(col_map) > 0

    # 所有数值字段（值为0且旧值非0时保留旧值）
    numeric_fields = {
        "pension_personal_base", "pension_company_base",
        "unemployment_personal_base", "unemployment_company_base",
        "medical_personal_base", "medical_company_base",
        "injury_company_base",
        "pension_personal", "unemployment_personal", "medical_personal",
        "si_personal", "pension_company", "unemployment_company", "medical_company",
        "injury_company", "si_company",
        "hf_base", "hf_personal", "hf_company",
    }

    created = 0
    updated = 0
    skipped = 0
    emp_map = {e.employee_no: e for e in db.query(Employee).all()}
    emp_name_map = {}
    for e in db.query(Employee).all():
        emp_name_map.setdefault(e.name, []).append(e)

    for row in data_rows:
        if not row:
            continue

        emp = None
        si_data = {}

        if use_new_format:
            # ── 新格式：通过表头动态匹配 ──
            # 获取员工编号
            emp_no = None
            if emp_no_col is not None and emp_no_col < len(row) and row[emp_no_col]:
                emp_no = str(row[emp_no_col]).strip()
                emp = emp_map.get(emp_no)

            # 如果工号没匹配到，尝试用姓名匹配
            if not emp and emp_name_col is not None and emp_name_col < len(row) and row[emp_name_col]:
                emp_name = str(row[emp_name_col]).strip()
                name_matches = emp_name_map.get(emp_name, [])
                if name_matches:
                    emp = name_matches[0]

            if not emp:
                skipped += 1
                continue

            # 提取各字段数据
            for col_idx, field_name in col_map.items():
                if col_idx >= len(row):
                    continue
                val = _safe_float(row[col_idx])
                if val is not None:
                    si_data[field_name] = val
        else:
            # ── 旧格式兜底：硬编码列索引（向后兼容） ──
            if not row[0]:
                continue
            emp_no = str(row[0]).strip()
            emp = emp_map.get(emp_no)
            if not emp:
                skipped += 1
                continue

            if len(row) > 1 and row[1]:
                v = _safe_float(row[1])
                if v is not None:
                    si_data["pension_personal"] = v
            if len(row) > 2 and row[2]:
                v = _safe_float(row[2])
                if v is not None:
                    si_data["unemployment_personal"] = v
            if len(row) > 3 and row[3]:
                v = _safe_float(row[3])
                if v is not None:
                    si_data["medical_personal"] = v
            if len(row) > 4 and row[4]:
                v = _safe_float(row[4])
                if v is not None:
                    si_data["si_personal"] = v
            if len(row) > 5 and row[5]:
                v = _safe_float(row[5])
                if v is not None:
                    si_data["pension_company"] = v
            if len(row) > 6 and row[6]:
                v = _safe_float(row[6])
                if v is not None:
                    si_data["unemployment_company"] = v
            if len(row) > 7 and row[7]:
                v = _safe_float(row[7])
                if v is not None:
                    si_data["medical_company"] = v
            if len(row) > 8 and row[8]:
                v = _safe_float(row[8])
                if v is not None:
                    si_data["si_company"] = v
            if len(row) > 9 and row[9]:
                v = _safe_float(row[9])
                if v is not None:
                    si_data["hf_base"] = v
            if len(row) > 10 and row[10]:
                v = _safe_float(row[10])
                if v is not None:
                    si_data["hf_personal"] = v
            if len(row) > 11 and row[11]:
                v = _safe_float(row[11])
                if v is not None:
                    si_data["hf_company"] = v

        # 如果该行没有任何有效数据（只有工号/姓名，无社保/公积金数值），跳过
        if not si_data:
            skipped += 1
            continue

        existing = db.query(SocialInsurance).filter(
            SocialInsurance.period == period,
            SocialInsurance.employee_id == emp.id
        ).first()

        if existing:
            for key, value in si_data.items():
                # 防止不同文件互相覆盖：新值为0且旧值非0时保留旧值
                if key in numeric_fields and value == 0:
                    old_val = getattr(existing, key, None)
                    if old_val is not None and old_val != 0:
                        continue
                setattr(existing, key, value)
            updated += 1
        else:
            si = SocialInsurance(period=period, employee_id=emp.id, **si_data)
            db.add(si)
            created += 1

    db.commit()

    # 自动计算合计字段
    all_records = db.query(SocialInsurance).filter(
        SocialInsurance.period == period
    ).all()
    for rec in all_records:
        _calc_summaries(rec)
    db.commit()

    write_log(db, "data_change", current_user.id, current_user.username, "social_insurance", "import", f"批量导入社保公积金 (period={period}, 新增{created}条, 更新{updated}条, 跳过{skipped}条)")
    return {"message": f"导入完成：新增 {created} 条，更新 {updated} 条，跳过 {skipped} 条不匹配的员工记录", "created": created, "updated": updated, "skipped": skipped}


@router.get("/export/{period}", dependencies=[Depends(require_permission("insurance:export"))])
def export_social_insurance(
    period: str,
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    query = db.query(Employee)
    employees = filter_active_employees(query, db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()
    si_map = {}
    records = db.query(SocialInsurance).filter(SocialInsurance.period == period).all()
    for r in records:
        si_map[r.employee_id] = r

    wb = Workbook()
    ws = wb.active
    ws.title = f"社保公积金_{period}"
    headers = [
        "员工编号", "姓名",
        "养老保险基数(个人)", "养老保险基数(单位)", "养老保险个人", "养老保险公司",
        "失业保险基数(个人)", "失业保险基数(单位)", "失业保险个人", "失业保险公司",
        "医疗保险基数(个人)", "医疗保险基数(单位)", "医疗保险个人", "医疗保险公司",
        "工伤保险基数(单位)", "工伤保险公司",
        "社保个人合计", "社保公司合计",
        "公积金基数", "公积金个人", "公积金公司",
    ]
    ws.append(headers)

    for emp in employees:
        si = si_map.get(emp.id)
        ws.append([
            emp.employee_no, emp.name,
            float(si.pension_personal_base or 0) if si else "",
            float(si.pension_company_base or 0) if si else "",
            float(si.pension_personal or 0) if si else "",
            float(si.pension_company or 0) if si else "",
            float(si.unemployment_personal_base or 0) if si else "",
            float(si.unemployment_company_base or 0) if si else "",
            float(si.unemployment_personal or 0) if si else "",
            float(si.unemployment_company or 0) if si else "",
            float(si.medical_personal_base or 0) if si else "",
            float(si.medical_company_base or 0) if si else "",
            float(si.medical_personal or 0) if si else "",
            float(si.medical_company or 0) if si else "",
            float(si.injury_company_base or 0) if si else "",
            float(si.injury_company or 0) if si else "",
            float(si.si_personal) if si else "",
            float(si.si_company) if si else "",
            float(si.hf_base) if si else "",
            float(si.hf_personal) if si else "",
            float(si.hf_company) if si else "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"社保公积金_{period}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


# ── 智能导入（多文件批量上传）──────────────────────────────────────────
@router.post("/smart-import/{period}", dependencies=[Depends(require_permission("insurance:import"))])
def smart_import_social_insurance(
    period: str,
    files: List[UploadFile] = File(...),
    file_paths: Optional[str] = Form(None, description="每个文件的相对路径JSON数组，用于文件夹导入时提供上下文关键词"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """智能批量导入：支持多文件、Excel+PDF、自动模板匹配、姓名匹配，支持文件夹导入"""
    import json as _json
    paths_list = None
    if file_paths:
        try:
            paths_list = _json.loads(file_paths)
        except Exception:
            paths_list = None
    result = run_smart_import(db, period, files, current_user.id, current_user.username, file_paths=paths_list)
    write_log(
        db, "data_change", current_user.id, current_user.username,
        "social_insurance", "smart_import",
        f"智能导入 (period={period}, 文件{result.total_files}个, "
        f"解析成功{result.parsed_files}个, 新增{result.created}条, 更新{result.updated}条)"
    )
    return {
        "total_files": result.total_files,
        "parsed_files": result.parsed_files,
        "failed_files": result.failed_files,
        "total_rows": result.total_rows,
        "created": result.created,
        "updated": result.updated,
        "errors": result.errors,
        "warnings": result.warnings,
    }


# ── 模板自动识别（上传样本文件，自动生成模板配置）───────────────────
@router.post("/templates/auto-detect", dependencies=[Depends(require_permission("insurance:template"))])
def auto_detect_template_config(
    file: UploadFile = File(...),
    current_user: UserInfo = Depends(get_current_user)
):
    """上传一个样本文件，自动识别表头、字段映射等模板配置"""
    try:
        file_bytes = file.file.read()
        result = auto_detect_template(file_bytes, file.filename or "unknown")
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"自动识别失败：{str(e)}")


# ── 字段标签列表（供前端下拉菜单使用）─────────────────────────────────
@router.get("/field-labels", dependencies=[Depends(require_permission("insurance:view"))])
def get_field_labels(
    current_user: UserInfo = Depends(get_current_user)
):
    """返回所有可映射字段及其中文标签"""
    return SI_FIELD_LABELS


# ── 导入模板 CRUD ─────────────────────────────────────────────────────
@router.get("/templates", response_model=List[SiImportTemplateOut], dependencies=[Depends(require_permission("insurance:template"))])
def list_templates(
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    return db.query(SiImportTemplate).order_by(SiImportTemplate.sort_order).all()


DEFAULT_NUMBER_FORMAT = {"remove_chars": [",", "，"], "decimal_separator": "."}


@router.post("/templates", response_model=SiImportTemplateOut, dependencies=[Depends(require_permission("insurance:template"))])
def create_template(
    data: SiImportTemplateCreate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    tpl_data = data.model_dump()
    if not tpl_data.get("number_format"):
        tpl_data["number_format"] = DEFAULT_NUMBER_FORMAT.copy()
    else:
        nf = tpl_data["number_format"]
        if "remove_chars" not in nf:
            nf["remove_chars"] = [",", "，"]
        else:
            existing = set(nf["remove_chars"])
            existing.update([",", "，"])
            nf["remove_chars"] = list(existing)
    tpl = SiImportTemplate(**tpl_data)
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    write_log(db, "data_change", current_user.id, current_user.username, "si_template", "create", f"新增导入模板「{tpl.name}」")
    return tpl


@router.put("/templates/{template_id}", response_model=SiImportTemplateOut, dependencies=[Depends(require_permission("insurance:template"))])
def update_template(
    template_id: int,
    data: SiImportTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    tpl = db.query(SiImportTemplate).filter(SiImportTemplate.id == template_id).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="模板不存在")
    update_data = data.model_dump(exclude_unset=True)
    if "number_format" in update_data and update_data["number_format"] is not None:
        nf = update_data["number_format"]
        if "remove_chars" not in nf:
            nf["remove_chars"] = [",", "，"]
        else:
            existing = set(nf["remove_chars"])
            existing.update([",", "，"])
            nf["remove_chars"] = list(existing)
    for key, value in update_data.items():
        setattr(tpl, key, value)
    db.commit()
    db.refresh(tpl)
    write_log(db, "data_change", current_user.id, current_user.username, "si_template", "edit", f"编辑导入模板「{tpl.name}」")
    return tpl


@router.delete("/templates/{template_id}")
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    tpl = db.query(SiImportTemplate).filter(SiImportTemplate.id == template_id).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="模板不存在")
    db.delete(tpl)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "si_template", "delete", f"删除导入模板「{tpl.name}」")
    return {"message": "删除成功"}


# ── 导入日志查询 ──────────────────────────────────────────────────────
@router.get("/import-logs/{period}", response_model=List[SiImportLogOut])
def get_import_logs(
    period: str,
    batch_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    q = db.query(SiImportLog).filter(SiImportLog.period == period)
    if batch_id:
        q = q.filter(SiImportLog.batch_id == batch_id)
    return q.order_by(SiImportLog.id.desc()).all()


# ── 批量模板配置向导 ──────────────────────────────────────────────────
@router.post("/smart-import-prepare/{period}")
def smart_import_prepare(
    period: str,
    files: List[UploadFile] = File(...),
    file_paths: Optional[str] = Form(None, description="每个文件的相对路径JSON数组，用于文件夹导入时提供上下文关键词"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    第一步：上传文件并预检查。
    保存文件到临时目录，返回batch_id和模板匹配结果。
    如果有文件未匹配模板，前端引导用户进入批量模板配置流程。
    支持文件夹导入（通过file_paths传递相对路径）。
    """
    try:
        import json as _json
        paths_list = None
        if file_paths:
            try:
                paths_list = _json.loads(file_paths)
            except Exception:
                paths_list = None
        batch_meta = save_batch_files(period, files, file_paths=paths_list)
        batch_id = batch_meta["batch_id"]
        precheck_result = precheck_batch(db, batch_id)
        return precheck_result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("批量导入准备失败")
        raise HTTPException(status_code=500, detail=f"准备失败：{str(e)}")


@router.get("/smart-import-batch/{batch_id}")
def get_smart_import_batch(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """获取批次预检查结果"""
    try:
        return precheck_batch(db, batch_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/templates/auto-detect-batch/{batch_id}/{file_index}")
def auto_detect_template_from_batch(
    batch_id: str,
    file_index: int,
    sheet_name: Optional[str] = Query(None, description="Excel工作表名称，指定解析哪个工作表"),
    current_user: UserInfo = Depends(get_current_user)
):
    """对批次中指定索引的文件（可指定工作表）执行自动模板识别"""
    try:
        result = auto_detect_from_batch(batch_id, file_index, sheet_name=sheet_name)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("批量自动识别失败")
        raise HTTPException(status_code=500, detail=f"自动识别失败：{str(e)}")


@router.post("/smart-import-batch/{batch_id}/execute")
def execute_smart_import_batch(
    batch_id: str,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """执行批次导入（所有模板配置完成后调用）"""
    try:
        # 执行前再次预检查，确保所有文件都能匹配到模板
        precheck = precheck_batch(db, batch_id)
        if precheck["has_unmatched"]:
            # 仍有未匹配文件，返回需要配置模板的信息，不删除临时文件
            return {
                "needs_config": True,
                "partial_success": False,
                "config_batch_id": batch_id,
                **precheck
            }

        result = run_smart_import_from_batch(db, batch_id, current_user.id, current_user.username)
        write_log(
            db, "data_change", current_user.id, current_user.username,
            "social_insurance", "smart_import_batch",
            f"批量导入 (batch={batch_id}, 文件{result.total_files}个, "
            f"解析成功{result.parsed_files}个, 新增{result.created}条, 更新{result.updated}条)"
        )

        # 执行后检查是否有未匹配模板的文件（预检查通过但实际提取失败）
        no_template_files = getattr(result, "no_template_filenames", [])
        if no_template_files:
            try:
                new_batch = create_batch_from_unmatched(batch_id, no_template_files)
                cleanup_batch(batch_id)
                # 对新批次做预检查，获取未匹配文件列表
                new_precheck = precheck_batch(db, new_batch["batch_id"])
                return {
                    "needs_config": True,
                    "partial_success": True,
                    "config_batch_id": new_batch["batch_id"],
                    "total_files": result.total_files,
                    "parsed_files": result.parsed_files,
                    "failed_files": result.failed_files,
                    "total_rows": result.total_rows,
                    "created": result.created,
                    "updated": result.updated,
                    "errors": result.errors,
                    "warnings": result.warnings,
                    "unmatched_count": new_precheck["unmatched_count"],
                    "unmatched_files": new_precheck["unmatched_files"],
                    "has_unmatched": True,
                }
            except ValueError:
                pass

        cleanup_batch(batch_id)
        return {
            "needs_config": False,
            "total_files": result.total_files,
            "parsed_files": result.parsed_files,
            "failed_files": result.failed_files,
            "total_rows": result.total_rows,
            "created": result.created,
            "updated": result.updated,
            "errors": result.errors,
            "warnings": result.warnings,
        }
    except ValueError as e:
        cleanup_batch(batch_id)
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("批量导入执行失败")
        cleanup_batch(batch_id)
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router.post("/smart-import-batch/{batch_id}/cancel")
def cancel_smart_import_batch(
    batch_id: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """取消批量导入，清理临时文件"""
    cleanup_batch(batch_id)
    return {"message": "已取消"}


def _get_prev_period(period: str) -> str:
    """根据当前月份计算上月月份，如 202607 -> 202606, 202601 -> 202512"""
    year = int(period[:4])
    month = int(period[4:6])
    if month == 1:
        prev_year = year - 1
        prev_month = 12
    else:
        prev_year = year
        prev_month = month - 1
    return f"{prev_year}{prev_month:02d}"


@router.post("/copy-from-prev/{period}")
def copy_from_prev_month(
    period: str,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    用上月数据预填本月社保公积金：
    - 只对本月无记录但上月有记录的员工复制数据
    - 本月已有记录的员工完全跳过，不覆盖
    """
    prev_period = _get_prev_period(period)

    prev_records = db.query(SocialInsurance).filter(
        SocialInsurance.period == prev_period
    ).all()

    if not prev_records:
        return {
            "message": f"{prev_period[:4]}年{prev_period[4:6]}月没有社保公积金数据，无法预填",
            "copied": 0,
            "skipped": 0,
            "prev_period": prev_period
        }

    existing_ids = {r.employee_id for r in db.query(SocialInsurance).filter(
        SocialInsurance.period == period
    ).all()}

    copied = 0
    skipped = 0

    for prev in prev_records:
        if prev.employee_id in existing_ids:
            skipped += 1
            continue

        new_si = SocialInsurance(
            period=period,
            employee_id=prev.employee_id,
            employee_social_insurance_no=prev.employee_social_insurance_no,
            pension_personal_base=prev.pension_personal_base,
            pension_company_base=prev.pension_company_base,
            unemployment_personal_base=prev.unemployment_personal_base,
            unemployment_company_base=prev.unemployment_company_base,
            medical_personal_base=prev.medical_personal_base,
            medical_company_base=prev.medical_company_base,
            injury_company_base=prev.injury_company_base,
            pension_personal=prev.pension_personal,
            pension_company=prev.pension_company,
            unemployment_personal=prev.unemployment_personal,
            unemployment_company=prev.unemployment_company,
            medical_personal=prev.medical_personal,
            medical_company=prev.medical_company,
            injury_company=prev.injury_company,
            si_personal=prev.si_personal,
            si_company=prev.si_company,
            pension_personal_rate=prev.pension_personal_rate,
            pension_company_rate=prev.pension_company_rate,
            unemployment_personal_rate=prev.unemployment_personal_rate,
            unemployment_company_rate=prev.unemployment_company_rate,
            medical_personal_rate=prev.medical_personal_rate,
            medical_company_rate=prev.medical_company_rate,
            injury_company_rate=prev.injury_company_rate,
            pension_total=prev.pension_total,
            unemployment_total=prev.unemployment_total,
            medical_total=prev.medical_total,
            injury_total=prev.injury_total,
            si_grand_total=prev.si_grand_total,
            hf_base=prev.hf_base,
            hf_personal=prev.hf_personal,
            hf_company=prev.hf_company,
            hf_personal_rate=prev.hf_personal_rate,
            hf_company_rate=prev.hf_company_rate,
            hf_total=prev.hf_total,
            grand_total=prev.grand_total,
        )
        db.add(new_si)
        copied += 1

    db.commit()

    write_log(
        db, "data_change", current_user.id, current_user.username,
        "social_insurance", "copy_from_prev",
        f"从上月({prev_period})预填社保公积金到{period}: 新增{copied}条, 跳过已有{skipped}条"
    )

    return {
        "message": f"预填完成：从{prev_period[:4]}年{prev_period[4:6]}月复制了{copied}条数据，{skipped}条已有数据被跳过",
        "copied": copied,
        "skipped": skipped,
        "prev_period": prev_period
    }


@router.post("/ensure-records/{period}")
def ensure_si_records(
    period: str,
    hide_status_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """确保所有在职员工都有社保公积金记录（空记录），便于批量编辑"""
    query = db.query(Employee)
    employees = filter_active_employees(query, db, hide_status_id=hide_status_id).all()
    existing = {r.employee_id: r for r in db.query(SocialInsurance).filter(SocialInsurance.period == period).all()}
    created_count = 0

    for emp in employees:
        if emp.id not in existing:
            si = SocialInsurance(
                period=period,
                employee_id=emp.id,
                pension_personal_base=0, pension_company_base=0,
                unemployment_personal_base=0, unemployment_company_base=0,
                medical_personal_base=0, medical_company_base=0,
                injury_company_base=0,
                pension_personal=0, unemployment_personal=0, medical_personal=0,
                si_personal=0, pension_company=0, unemployment_company=0, medical_company=0,
                injury_company=0, si_company=0,
                pension_personal_rate=0, pension_company_rate=0,
                unemployment_personal_rate=0, unemployment_company_rate=0,
                medical_personal_rate=0, medical_company_rate=0, injury_company_rate=0,
                pension_total=0, unemployment_total=0, medical_total=0, injury_total=0,
                si_grand_total=0,
                hf_base=0, hf_personal=0, hf_company=0,
                hf_personal_rate=0, hf_company_rate=0, hf_total=0,
                grand_total=0
            )
            db.add(si)
            created_count += 1

    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "social_insurance", "create", f"确保社保记录存在 (period={period}, 新增{created_count}条)")
    return {"created_count": created_count, "total": len(employees)}


@router.get("/prev-month-data/{period}/{employee_id}")
def get_prev_month_data(
    period: str,
    employee_id: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """获取指定员工上月的社保数据，用于单条录入时预填表单"""
    prev_period = _get_prev_period(period)
    prev = db.query(SocialInsurance).filter(
        SocialInsurance.period == prev_period,
        SocialInsurance.employee_id == employee_id
    ).first()

    if not prev:
        return {"has_data": False, "prev_period": prev_period}

    return {
        "has_data": True,
        "prev_period": prev_period,
        "data": SocialInsuranceOut.model_validate(prev).model_dump()
    }