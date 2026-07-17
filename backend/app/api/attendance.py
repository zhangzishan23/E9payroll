from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from io import BytesIO
from openpyxl import Workbook
from datetime import date
import calendar
from app.core.database import get_db
from app.core.log_helper import write_log
from app.models.models import AttendanceRecord, Employee, SysDictBase
from app.api.auth import get_current_user, UserInfo, require_permission
from app.core.query_utils import filter_active_employees, apply_data_scope
from app.services import work_calendar as work_cal

router = APIRouter()

DAILY_WORK_HOURS = 7


def _get_dict_name_cached(dict_id: Optional[int], name_map: dict) -> str:
    if not dict_id:
        return ""
    return name_map.get(dict_id, "")


def _batch_load_dict_names(db: Session) -> dict:
    dict_items = db.query(SysDictBase).all()
    return {d.id: d.name for d in dict_items}


def _calc_salary_dates(period: str) -> tuple:
    year = int(period[:4])
    month = int(period[4:6])
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    return first_day, last_day


def _get_adjusted_salary_days(db: Session, period: str) -> float:
    return work_cal.get_month_salary_days(db, period)


def _recalculate_attendance_fields(db: Session, att: AttendanceRecord) -> None:
    """根据当前工作日历和请假数据，重算应计薪天数、计薪天数、出勤率"""
    adjusted = _get_adjusted_salary_days(db, att.period)
    late_leave = float(att.late_to_personal_leave_days) if att.late_to_personal_leave_days else _calc_late_to_personal_leave(
        att.late_count or 0, att.late_duration or 0
    )
    leave_total = _calc_leave_total(
        late_leave,
        float(att.personal_leave_days or 0),
        float(att.full_pay_sick_days or 0),
        float(att.reduced_pay_sick_days or 0),
        float(att.statutory_sick_days or 0),
        float(att.compensatory_leave_days or 0),
        float(att.annual_leave_days or 0),
        float(att.prenatal_checkup_days or 0),
        float(att.maternity_leave_days or 0),
        float(att.paternity_leave_days or 0),
        float(att.marriage_leave_days or 0),
        float(att.funeral_leave_days or 0),
        float(att.engineering_compensatory_days or 0),
    )
    actual_salary = _calc_actual_salary_days(adjusted, leave_total)
    att_rate = round(actual_salary / adjusted, 4) if adjusted > 0 else 0

    att.adjusted_salary_days = adjusted
    att.actual_salary_days = actual_salary
    att.actual_work_days = actual_salary
    att.late_to_personal_leave_days = late_leave
    att.leave_total_days = leave_total
    att.attendance_rate = att_rate


def _recalculate_month_attendance(db: Session, period: str) -> int:
    """重算指定月份所有考勤记录的应计薪天数、计薪天数、出勤率"""
    records = db.query(AttendanceRecord).filter(AttendanceRecord.period == period).all()
    updated_count = 0
    for att in records:
        _recalculate_attendance_fields(db, att)
        updated_count += 1
    if updated_count > 0:
        db.commit()
    return updated_count


def _calc_late_to_personal_leave(late_count: int, late_duration: int) -> float:
    if late_count <= 3 or late_duration <= 0:
        return 0.0
    return round(late_duration / 60 / DAILY_WORK_HOURS, 2)


def _calc_actual_salary_days(adjusted_days: float, leave_total: float) -> float:
    return max(round(adjusted_days - leave_total, 2), 0)


def _calc_leave_total(
    late_to_personal: float,
    personal_leave: float, full_pay_sick: float, reduced_pay_sick: float,
    statutory_sick: float, compensatory: float, annual: float,
    prenatal: float, maternity: float, paternity: float,
    marriage: float, funeral: float, engineering: float
) -> float:
    return round(
        late_to_personal + personal_leave + reduced_pay_sick +
        statutory_sick + prenatal + maternity +
        paternity + marriage + funeral + engineering, 2
    )


# ==================== Schemas ====================

class AttendanceOut(BaseModel):
    id: Optional[int] = None
    period: str
    employee_id: int
    employee_no: str = ""
    employee_name: str = ""
    contract_company: str = ""
    department: str = ""
    salary_start_date: Optional[str] = None
    salary_end_date: Optional[str] = None
    total_work_days: Optional[float] = None
    adjusted_salary_days: Optional[float] = None
    actual_salary_days: Optional[float] = None
    attendance_rate: Optional[float] = None
    half_day_missed_punch: Optional[int] = None
    absenteeism_days: Optional[float] = None
    late_count: Optional[int] = None
    late_duration: Optional[int] = None
    severe_late_count: Optional[int] = None
    severe_late_duration: Optional[int] = None
    early_count: Optional[int] = None
    early_duration: Optional[int] = None
    total_overtime: Optional[float] = None
    late_to_personal_leave_days: Optional[float] = None
    personal_leave_days: Optional[float] = None
    full_pay_sick_days: Optional[float] = None
    reduced_pay_sick_days: Optional[float] = None
    statutory_sick_days: Optional[float] = None
    compensatory_leave_days: Optional[float] = None
    annual_leave_days: Optional[float] = None
    prenatal_checkup_days: Optional[float] = None
    maternity_leave_days: Optional[float] = None
    paternity_leave_days: Optional[float] = None
    marriage_leave_days: Optional[float] = None
    funeral_leave_days: Optional[float] = None
    engineering_compensatory_days: Optional[float] = None
    leave_total_days: Optional[float] = None
    remark: Optional[str] = None
    is_row_locked: Optional[bool] = False
    locked_fields: Optional[dict] = None
    special_apply_ids: Optional[list] = None

    class Config:
        from_attributes = True

    @classmethod
    def from_record(cls, att, emp, db: Session = None, fallback_period: str = "",
                    name_map: dict = None, adjusted_from_calendar: float = None) -> "AttendanceOut":
        period = att.period if att else fallback_period
        start_date = None
        end_date = None

        if period:
            start_date, end_date = _calc_salary_dates(period)

        if db is not None and name_map is None:
            name_map = _batch_load_dict_names(db)
        if name_map is None:
            name_map = {}

        if db is not None and adjusted_from_calendar is None and period:
            adjusted_from_calendar = _get_adjusted_salary_days(db, period)

        if att:
            contract_company = ""
            department = ""
            if emp:
                contract_company = _get_dict_name_cached(emp.contract_company_id, name_map)
                department = _get_dict_name_cached(emp.department_id, name_map)

            leave_total = float(att.leave_total_days) if att.leave_total_days else 0
            total_work = float(att.total_work_days) if att.total_work_days is not None else None
            if adjusted_from_calendar is not None:
                adjusted = adjusted_from_calendar
                actual = max(round(adjusted - leave_total, 2), 0)
                att_rate = round(actual / adjusted, 4) if adjusted > 0 else 0
            else:
                adjusted = float(att.adjusted_salary_days) if att.adjusted_salary_days else total_work
                actual = float(att.actual_salary_days) if att.actual_salary_days else None
                att_rate = float(att.attendance_rate)

            start_str = start_date.isoformat() if start_date else (att.salary_start_date.isoformat() if att.salary_start_date else None)
            end_str = end_date.isoformat() if end_date else (att.salary_end_date.isoformat() if att.salary_end_date else None)

            return cls(
                id=att.id, period=att.period, employee_id=att.employee_id,
                employee_no=emp.employee_no if emp else "",
                employee_name=emp.name if emp else "",
                contract_company=contract_company,
                department=department,
                salary_start_date=start_str,
                salary_end_date=end_str,
                total_work_days=total_work,
                adjusted_salary_days=adjusted,
                actual_salary_days=actual,
                attendance_rate=att_rate,
                half_day_missed_punch=att.half_day_missed_punch,
                absenteeism_days=float(att.absenteeism_days),
                late_count=att.late_count,
                late_duration=att.late_duration,
                severe_late_count=att.severe_late_count,
                severe_late_duration=att.severe_late_duration,
                early_count=att.early_count,
                early_duration=att.early_duration,
                total_overtime=float(att.total_overtime),
                late_to_personal_leave_days=float(att.late_to_personal_leave_days) if att.late_to_personal_leave_days else 0,
                personal_leave_days=float(att.personal_leave_days),
                full_pay_sick_days=float(att.full_pay_sick_days),
                reduced_pay_sick_days=float(att.reduced_pay_sick_days),
                statutory_sick_days=float(att.statutory_sick_days),
                compensatory_leave_days=float(att.compensatory_leave_days),
                annual_leave_days=float(att.annual_leave_days),
                prenatal_checkup_days=float(att.prenatal_checkup_days),
                maternity_leave_days=float(att.maternity_leave_days),
                paternity_leave_days=float(att.paternity_leave_days),
                marriage_leave_days=float(att.marriage_leave_days),
                funeral_leave_days=float(att.funeral_leave_days),
                engineering_compensatory_days=float(att.engineering_compensatory_days),
                leave_total_days=leave_total,
                remark=att.remark,
                is_row_locked=att.is_row_locked or False,
                locked_fields=att.locked_fields or {},
                special_apply_ids=att.special_apply_ids or [],
            )
        else:
            contract_company = ""
            department = ""
            if emp:
                contract_company = _get_dict_name_cached(emp.contract_company_id, name_map)
                department = _get_dict_name_cached(emp.department_id, name_map)

            if adjusted_from_calendar is not None:
                total_work = None
                adjusted = adjusted_from_calendar
                actual = adjusted_from_calendar
                att_rate = 1.0 if adjusted_from_calendar > 0 else 0
            else:
                total_work = None
                adjusted = None
                actual = None
                att_rate = None

            return cls(
                period=fallback_period,
                employee_id=emp.id if emp else 0,
                employee_no=emp.employee_no if emp else "",
                employee_name=emp.name if emp else "",
                contract_company=contract_company,
                department=department,
                salary_start_date=start_date.isoformat() if start_date else None,
                salary_end_date=end_date.isoformat() if end_date else None,
                total_work_days=total_work,
                adjusted_salary_days=adjusted,
                actual_salary_days=actual,
                attendance_rate=att_rate,
            )


class AttendanceCreate(BaseModel):
    period: str
    employee_id: int
    total_work_days: float
    adjusted_salary_days: float = 0
    actual_salary_days: float = 0
    late_count: int = 0
    late_duration: int = 0
    severe_late_count: int = 0
    severe_late_duration: int = 0
    early_count: int = 0
    early_duration: int = 0
    half_day_missed_punch: int = 0
    absenteeism_days: float = 0
    total_overtime: float = 0
    late_to_personal_leave_days: float = 0
    personal_leave_days: float = 0
    full_pay_sick_days: float = 0
    reduced_pay_sick_days: float = 0
    statutory_sick_days: float = 0
    compensatory_leave_days: float = 0
    annual_leave_days: float = 0
    prenatal_checkup_days: float = 0
    maternity_leave_days: float = 0
    paternity_leave_days: float = 0
    marriage_leave_days: float = 0
    funeral_leave_days: float = 0
    engineering_compensatory_days: float = 0
    leave_total_days: float = 0
    remark: Optional[str] = None


class AttendanceUpdate(BaseModel):
    total_work_days: Optional[float] = None
    late_count: Optional[int] = None
    late_duration: Optional[int] = None
    severe_late_count: Optional[int] = None
    severe_late_duration: Optional[int] = None
    early_count: Optional[int] = None
    early_duration: Optional[int] = None
    half_day_missed_punch: Optional[int] = None
    absenteeism_days: Optional[float] = None
    total_overtime: Optional[float] = None
    late_to_personal_leave_days: Optional[float] = None
    personal_leave_days: Optional[float] = None
    full_pay_sick_days: Optional[float] = None
    reduced_pay_sick_days: Optional[float] = None
    statutory_sick_days: Optional[float] = None
    compensatory_leave_days: Optional[float] = None
    annual_leave_days: Optional[float] = None
    prenatal_checkup_days: Optional[float] = None
    maternity_leave_days: Optional[float] = None
    paternity_leave_days: Optional[float] = None
    marriage_leave_days: Optional[float] = None
    funeral_leave_days: Optional[float] = None
    engineering_compensatory_days: Optional[float] = None
    remark: Optional[str] = None


# ==================== 考勤查询 ====================

@router.get("/", response_model=List[AttendanceOut], dependencies=[Depends(require_permission("attendance:view"))])
def get_attendance(
    period: Optional[str] = Query(None),
    employee_id: Optional[int] = Query(None),
    filter_field: Optional[str] = Query(None),
    filter_value: Optional[str] = Query(None),
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if period:
        employee_query = db.query(Employee)
        if not current_user.is_admin:
            employee_query = apply_data_scope(employee_query, db, current_user.data_scope, current_user.id)
        employee_query = filter_active_employees(employee_query, db, hide_status_id=hide_status_id)
        if filter_field and filter_value:
            if filter_field == 'employee_no':
                employee_query = employee_query.filter(Employee.employee_no.ilike(f'%{filter_value}%'))
            elif filter_field == 'employee_name':
                employee_query = employee_query.filter(Employee.name.ilike(f'%{filter_value}%'))
        employees = employee_query.order_by(Employee.employee_no).all()

        attendance_map = {}
        # 使用子查询获取每个员工该周期最新的记录ID，避免重复记录问题
        from sqlalchemy import func
        latest_ids = db.query(func.max(AttendanceRecord.id).label('max_id')).filter(
            AttendanceRecord.period == period
        ).group_by(AttendanceRecord.employee_id).subquery()
        records = db.query(AttendanceRecord).join(
            latest_ids, AttendanceRecord.id == latest_ids.c.max_id
        ).all()
        for r in records:
            attendance_map[r.employee_id] = r

        name_map = _batch_load_dict_names(db)
        adjusted_days = _get_adjusted_salary_days(db, period)

        result = []
        for emp in employees:
            att = attendance_map.get(emp.id)
            result.append(AttendanceOut.from_record(att, emp, db=db, fallback_period=period,
                                                   name_map=name_map, adjusted_from_calendar=adjusted_days))
        return result

    query = db.query(AttendanceRecord)
    if not current_user.is_admin:
        query = apply_data_scope(query, db, current_user.data_scope, current_user.id,
                                 employee_model=AttendanceRecord,
                                 employee_id_field=AttendanceRecord.employee_id)
    if employee_id:
        query = query.filter(AttendanceRecord.employee_id == employee_id)
    records = query.order_by(AttendanceRecord.period.desc(), AttendanceRecord.employee_id).all()
    result = []
    name_map = _batch_load_dict_names(db)
    for r in records:
        emp = db.query(Employee).filter(Employee.id == r.employee_id).first()
        result.append(AttendanceOut.from_record(r, emp, db=db, name_map=name_map))
    return result


# ==================== 导出 ====================

@router.get("/export", dependencies=[Depends(require_permission("attendance:export"))])
def export_attendance(
    period: Optional[str] = Query(None),
    employee_id: Optional[int] = Query(None),
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if period:
        employee_query = db.query(Employee)
        if not current_user.is_admin:
            employee_query = apply_data_scope(employee_query, db, current_user.data_scope, current_user.id)
        employee_query = filter_active_employees(employee_query, db, hide_status_id=hide_status_id)
        employees = employee_query.order_by(Employee.employee_no).all()

        attendance_map = {}
        # 使用子查询获取每个员工该周期最新的记录ID，避免重复记录问题
        from sqlalchemy import func
        latest_ids = db.query(func.max(AttendanceRecord.id).label('max_id')).filter(
            AttendanceRecord.period == period
        ).group_by(AttendanceRecord.employee_id).subquery()
        records = db.query(AttendanceRecord).join(
            latest_ids, AttendanceRecord.id == latest_ids.c.max_id
        ).all()
        for r in records:
            attendance_map[r.employee_id] = r
        data = []
        name_map = _batch_load_dict_names(db)
        adjusted_days = _get_adjusted_salary_days(db, period)
        for emp in employees:
            att = attendance_map.get(emp.id)
            data.append((emp, att, name_map, adjusted_days))
    else:
        query = db.query(AttendanceRecord)
        if not current_user.is_admin:
            query = apply_data_scope(query, db, current_user.data_scope, current_user.id,
                                     employee_model=AttendanceRecord,
                                     employee_id_field=AttendanceRecord.employee_id)
        if employee_id:
            query = query.filter(AttendanceRecord.employee_id == employee_id)
        records = query.order_by(AttendanceRecord.period.desc(), AttendanceRecord.employee_id).all()
        data = []
        name_map = _batch_load_dict_names(db)
        for r in records:
            emp = db.query(Employee).filter(Employee.id == r.employee_id).first()
            data.append((emp, r, name_map, None))

    wb = Workbook()
    ws = wb.active
    ws.title = "考勤数据"
    headers = [
        "核算周期", "员工编号", "员工姓名", "合同主体", "部门",
        "计薪开始日期", "计薪截至日期",
        "当月总计薪天数", "应计薪天数", "计薪天数", "出勤率",
        "半天缺卡(次数)", "全天缺卡(天数)",
        "迟到次数", "迟到时长(分钟)", "严重迟到次数", "严重迟到时长(分钟)",
        "早退次数", "早退时长(分钟)", "加班(小时)", "迟到转事假",
        "事假", "全薪病假", "减薪病假", "法定病假",
        "调休", "年假", "产检假", "产假", "陪产假", "婚假", "丧假",
        "调休-工程交付(天)", "合计", "备注",
    ]
    ws.append(headers)
    for emp, r, nm, adj_days in data:
        out = AttendanceOut.from_record(r, emp, db=db, fallback_period=period if period else (r.period if r else ""),
                                        name_map=nm, adjusted_from_calendar=adj_days)
        ws.append([
            out.period, out.employee_no, out.employee_name,
            out.contract_company or "", out.department or "",
            out.salary_start_date or "", out.salary_end_date or "",
            out.total_work_days if out.total_work_days is not None else "",
            out.adjusted_salary_days if out.adjusted_salary_days is not None else "",
            out.actual_salary_days if out.actual_salary_days is not None else "",
            out.attendance_rate if out.attendance_rate is not None else "",
            out.half_day_missed_punch or 0, out.absenteeism_days or 0,
            out.late_count or 0, out.late_duration or 0,
            out.severe_late_count or 0, out.severe_late_duration or 0,
            out.early_count or 0, out.early_duration or 0,
            out.total_overtime or 0,
            out.late_to_personal_leave_days or 0,
            out.personal_leave_days or 0, out.full_pay_sick_days or 0,
            out.reduced_pay_sick_days or 0, out.statutory_sick_days or 0,
            out.compensatory_leave_days or 0, out.annual_leave_days or 0,
            out.prenatal_checkup_days or 0, out.maternity_leave_days or 0,
            out.paternity_leave_days or 0, out.marriage_leave_days or 0,
            out.funeral_leave_days or 0, out.engineering_compensatory_days or 0,
            out.leave_total_days or 0,
            out.remark or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=attendance_export.xlsx"}
    )


# ==================== 新增 ====================

@router.post("/", response_model=AttendanceOut, dependencies=[Depends(require_permission("attendance:create"))])
def create_attendance(att: AttendanceCreate, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    existing = db.query(AttendanceRecord).filter(
        AttendanceRecord.period == att.period,
        AttendanceRecord.employee_id == att.employee_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"员工 [{att.employee_id}] 在周期 [{att.period}] 的考勤记录已存在")

    start_date, end_date = _calc_salary_dates(att.period)
    adjusted = _get_adjusted_salary_days(db, att.period)

    late_leave = att.late_to_personal_leave_days if att.late_to_personal_leave_days > 0 else _calc_late_to_personal_leave(att.late_count, att.late_duration)

    leave_total = att.leave_total_days if att.leave_total_days > 0 else _calc_leave_total(
        late_leave, att.personal_leave_days, att.full_pay_sick_days,
        att.reduced_pay_sick_days, att.statutory_sick_days,
        att.compensatory_leave_days, att.annual_leave_days,
        att.prenatal_checkup_days, att.maternity_leave_days,
        att.paternity_leave_days, att.marriage_leave_days,
        att.funeral_leave_days, att.engineering_compensatory_days
    )

    actual_salary = _calc_actual_salary_days(adjusted, leave_total)
    attendance_rate = round(actual_salary / adjusted, 4) if adjusted > 0 else 0

    db_att = AttendanceRecord(
        period=att.period, employee_id=att.employee_id,
        total_work_days=att.total_work_days,
        actual_work_days=actual_salary,
        attendance_rate=attendance_rate,
        salary_start_date=start_date, salary_end_date=end_date,
        adjusted_salary_days=adjusted,
        actual_salary_days=actual_salary,
        late_to_personal_leave_days=late_leave,
        leave_total_days=leave_total,
        late_count=att.late_count, late_duration=att.late_duration,
        severe_late_count=att.severe_late_count, severe_late_duration=att.severe_late_duration,
        early_count=att.early_count, early_duration=att.early_duration,
        half_day_missed_punch=att.half_day_missed_punch,
        absenteeism_days=att.absenteeism_days,
        total_overtime=att.total_overtime,
        personal_leave_days=att.personal_leave_days,
        full_pay_sick_days=att.full_pay_sick_days,
        reduced_pay_sick_days=att.reduced_pay_sick_days,
        statutory_sick_days=att.statutory_sick_days,
        compensatory_leave_days=att.compensatory_leave_days,
        annual_leave_days=att.annual_leave_days,
        prenatal_checkup_days=att.prenatal_checkup_days,
        maternity_leave_days=att.maternity_leave_days,
        paternity_leave_days=att.paternity_leave_days,
        marriage_leave_days=att.marriage_leave_days,
        funeral_leave_days=att.funeral_leave_days,
        engineering_compensatory_days=att.engineering_compensatory_days,
        remark=att.remark,
    )
    db.add(db_att)
    db.commit()
    db.refresh(db_att)
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "create", f"新增考勤记录 (employee_id={att.employee_id}, period={att.period})")
    emp = db.query(Employee).filter(Employee.id == att.employee_id).first()
    name_map = _batch_load_dict_names(db)
    return AttendanceOut.from_record(db_att, emp, db=db, name_map=name_map)


# ==================== 特殊申请 API ====================

@router.get("/special-applies", dependencies=[Depends(require_permission("attendance:view", "attendance:writeoff"))])
def get_special_applies(
    period: str = Query(..., description="核算周期 YYYYMM"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """获取指定月份的特殊申请审批单（从钉钉拉取）"""
    try:
        from app.services import dingtalk_service
        applies = dingtalk_service.get_month_special_applies(period)
        
        # 补充申请人姓名
        emp_by_dingtalk = {}
        for emp in db.query(Employee).filter(Employee.dingtalk_user_id.isnot(None)).all():
            emp_by_dingtalk[emp.dingtalk_user_id] = emp
        
        result = []
        for apply in applies:
            user_id = apply["originator_user_id"]
            emp = emp_by_dingtalk.get(user_id)
            result.append({
                **apply,
                "applicant_name": emp.name if emp else user_id,
                "employee_id": emp.id if emp else None,
            })
        
        return {"period": period, "applies": result, "total": len(result)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"获取特殊申请失败: {str(e)}")


@router.post("/apply-special-applies", dependencies=[Depends(require_permission("attendance:writeoff"))])
def apply_special_applies(
    period: str = Form(...),
    apply_ids: Optional[str] = Form(None, description="逗号分隔的申请单ID，为空则应用所有"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """将特殊申请应用到考勤记录"""
    try:
        from app.services import dingtalk_service
        applies = dingtalk_service.get_month_special_applies(period)
        
        if apply_ids:
            target_ids = set(apply_ids.split(","))
            applies = [a for a in applies if a["instance_id"] in target_ids]
        
        stats = dingtalk_service.apply_special_applies_to_attendance(db, period, applies)
        write_log(db, "data_change", current_user.id, current_user.username, "attendance", "apply_apply", f"应用特殊申请到考勤：成功{stats['applied']}个，更新{stats['updated']}人 (period={period})")
        return {
            "message": f"应用完成：成功处理 {stats['applied']} 个申请单，更新 {stats['updated']} 名员工考勤",
            **stats
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"应用特殊申请失败: {str(e)}")


# ==================== 缺卡自动核销 API ====================

@router.get("/missed-punch-check", dependencies=[Depends(require_permission("attendance:writeoff"))])
def check_missed_punch(
    period: str = Query(..., description="核算周期 YYYYMM"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    检查指定月份有缺卡记录的员工是否有审批通过的考勤特殊申请
    返回比对结果（匹配、不匹配、无申请），供用户预览确认
    """
    try:
        from app.services import dingtalk_service
        result = dingtalk_service.check_missed_punch_applies(db, period)
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"检查缺卡申请失败: {str(e)}")


@router.post("/missed-punch-write-off", dependencies=[Depends(require_permission("attendance:writeoff"))])
def write_off_missed_punch(
    period: str = Form(..., description="核算周期 YYYYMM"),
    apply_all: bool = Form(False, description="是否核销所有匹配的员工"),
    employee_ids: Optional[str] = Form(None, description="逗号分隔的员工ID列表，指定核销部分员工"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    执行缺卡核销：将匹配员工的半天缺卡和全天缺卡清零
    """
    try:
        from app.services import dingtalk_service
        
        emp_ids = None
        if employee_ids:
            emp_ids = [int(x.strip()) for x in employee_ids.split(",") if x.strip()]
        
        result = dingtalk_service.apply_missed_punch_write_off(
            db, period,
            employee_ids=emp_ids,
            apply_all_matched=apply_all
        )
        
        write_log(
            db, "data_change", current_user.id, current_user.username,
            "attendance", "missed_punch_write_off",
            f"缺卡核销：更新{result['updated_count']}人 (period={period})"
        )
        
        return {
            "message": f"核销完成，共更新 {result['updated_count']} 名员工的缺卡记录",
            **result
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"缺卡核销失败: {str(e)}")


# ==================== 批量删除 ====================

@router.post("/batch-delete", dependencies=[Depends(require_permission("attendance:delete"))])
def batch_delete_attendance(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的考勤记录ID列表")
    records = db.query(AttendanceRecord).filter(AttendanceRecord.id.in_(ids)).all()
    if not records:
        raise HTTPException(status_code=404, detail="未找到指定的考勤记录")
    for r in records:
        db.delete(r)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "delete", f"批量删除 {len(records)} 条考勤记录")
    return {"message": f"成功删除 {len(records)} 条考勤记录", "deleted_count": len(records)}


# ==================== 编辑 ====================

@router.put("/{record_id}", response_model=AttendanceOut, dependencies=[Depends(require_permission("attendance:edit"))])
def update_attendance(record_id: int, data: AttendanceUpdate, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    att = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="考勤记录不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(att, key, value)

    _recalculate_attendance_fields(db, att)

    db.commit()
    db.refresh(att)
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "edit", f"编辑考勤记录 (id={record_id})")

    emp = db.query(Employee).filter(Employee.id == att.employee_id).first()
    name_map = _batch_load_dict_names(db)
    return AttendanceOut.from_record(att, emp, db=db, name_map=name_map)


# ==================== 导入 ====================

@router.post("/import", dependencies=[Depends(require_permission("attendance:import"))])
async def import_attendance(
    file: UploadFile = File(...),
    period: str = Form(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头，没有数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    header_map = {}
    for i, h in enumerate(headers):
        if "员工编号" in h or "编号" in h:
            header_map["employee_no"] = i
        elif "总计薪" in h or "应出勤" in h:
            header_map["total_work_days"] = i
        elif "迟到" in h and "转事假" in h:
            header_map["late_to_personal_leave"] = i
        elif "迟到" in h:
            header_map["late_count"] = i
        elif "早退" in h:
            header_map["early_count"] = i
        elif "半天缺卡" in h or "缺卡" in h:
            header_map["half_day_missed_punch"] = i
        elif "全天缺卡" in h:
            header_map["absenteeism_days"] = i
        elif "加班" in h:
            header_map["total_overtime"] = i
        elif "事假" in h:
            header_map["personal_leave_days"] = i
        elif "全薪病假" in h:
            header_map["full_pay_sick_days"] = i
        elif "减薪病假" in h:
            header_map["reduced_pay_sick_days"] = i
        elif "法定病假" in h:
            header_map["statutory_sick_days"] = i
        elif "调休" in h and "工程" in h:
            header_map["engineering_compensatory_days"] = i
        elif "调休" in h:
            header_map["compensatory_leave_days"] = i
        elif "年假" in h:
            header_map["annual_leave_days"] = i
        elif "产检假" in h:
            header_map["prenatal_checkup_days"] = i
        elif "产假" in h:
            header_map["maternity_leave_days"] = i
        elif "陪产假" in h:
            header_map["paternity_leave_days"] = i
        elif "婚假" in h:
            header_map["marriage_leave_days"] = i
        elif "丧假" in h:
            header_map["funeral_leave_days"] = i
        elif "合计" in h:
            header_map["leave_total_days"] = i
        elif "备注" in h:
            header_map["remark"] = i

    if "employee_no" not in header_map:
        raise HTTPException(status_code=400, detail="Excel 表头缺少「员工编号」列")

    emp_map = {e.employee_no: e for e in db.query(Employee).all()}
    start_date, end_date = _calc_salary_dates(period)
    calendar_adjusted = _get_adjusted_salary_days(db, period)

    created = 0
    updated = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            emp_no = str(row[header_map["employee_no"]]).strip() if row[header_map["employee_no"]] else ""
            emp = emp_map.get(emp_no)
            if not emp:
                errors.append(f"第{row_idx}行：员工编号「{emp_no}」不存在")
                continue

            def get_float(key, default=0):
                if key in header_map and row[header_map[key]] is not None:
                    try:
                        return float(row[header_map[key]])
                    except (ValueError, TypeError):
                        return default
                return default

            def get_int(key, default=0):
                if key in header_map and row[header_map[key]] is not None:
                    try:
                        return int(float(row[header_map[key]]))
                    except (ValueError, TypeError):
                        return default
                return default

            def get_str(key, default=None):
                if key in header_map and row[header_map[key]] is not None:
                    return str(row[header_map[key]]).strip()
                return default

            total_work = get_float("total_work_days", 0)
            late_count_val = get_int("late_count")
            late_duration_val = get_int("late_duration", 0)
            late_leave = get_float("late_to_personal_leave", 0) or _calc_late_to_personal_leave(late_count_val, late_duration_val)

            personal = get_float("personal_leave_days")
            full_sick = get_float("full_pay_sick_days")
            reduced_sick = get_float("reduced_pay_sick_days")
            stat_sick = get_float("statutory_sick_days")
            comp = get_float("compensatory_leave_days")
            annual = get_float("annual_leave_days")
            prenatal = get_float("prenatal_checkup_days")
            maternity = get_float("maternity_leave_days")
            paternity = get_float("paternity_leave_days")
            marriage = get_float("marriage_leave_days")
            funeral = get_float("funeral_leave_days")
            eng = get_float("engineering_compensatory_days")

            leave_total = get_float("leave_total_days", 0) or _calc_leave_total(
                late_leave, personal, full_sick, reduced_sick, stat_sick,
                comp, annual, prenatal, maternity, paternity, marriage, funeral, eng
            )
            adjusted = calendar_adjusted
            actual_salary = _calc_actual_salary_days(adjusted, leave_total)
            att_rate = round(actual_salary / adjusted, 4) if adjusted > 0 else 0

            existing_records = db.query(AttendanceRecord).filter(
                AttendanceRecord.period == period,
                AttendanceRecord.employee_id == emp.id
            ).order_by(AttendanceRecord.id.desc()).all()
            
            existing = existing_records[0] if existing_records else None
            
            # 如果有重复记录，删除旧的（保留最新的一条）
            if len(existing_records) > 1:
                for old_rec in existing_records[1:]:
                    db.delete(old_rec)

            if existing:
                # 跳过整行锁定的记录
                if existing.is_row_locked:
                    continue
                
                locked_fields = existing.locked_fields or {}
                
                # 定义字段导入映射（字段名 -> 值）
                import_data = {
                    "total_work_days": total_work,
                    "adjusted_salary_days": adjusted,
                    "actual_salary_days": actual_salary,
                    "attendance_rate": att_rate,
                    "actual_work_days": actual_salary,
                    "late_count": late_count_val,
                    "late_duration": late_duration_val,
                    "severe_late_count": get_int("severe_late_count"),
                    "severe_late_duration": get_int("severe_late_duration"),
                    "early_count": get_int("early_count"),
                    "early_duration": get_int("early_duration"),
                    "half_day_missed_punch": get_int("half_day_missed_punch"),
                    "absenteeism_days": get_float("absenteeism_days"),
                    "total_overtime": get_float("total_overtime"),
                    "late_to_personal_leave_days": late_leave,
                    "personal_leave_days": personal,
                    "full_pay_sick_days": full_sick,
                    "reduced_pay_sick_days": reduced_sick,
                    "statutory_sick_days": stat_sick,
                    "compensatory_leave_days": comp,
                    "annual_leave_days": annual,
                    "prenatal_checkup_days": prenatal,
                    "maternity_leave_days": maternity,
                    "paternity_leave_days": paternity,
                    "marriage_leave_days": marriage,
                    "funeral_leave_days": funeral,
                    "engineering_compensatory_days": eng,
                    "leave_total_days": leave_total,
                    "salary_start_date": start_date,
                    "salary_end_date": end_date,
                }
                remark_val = get_str("remark")
                if remark_val is not None:
                    import_data["remark"] = remark_val
                
                # 只更新未锁定的字段
                for field, value in import_data.items():
                    if not locked_fields.get(field):
                        setattr(existing, field, value)
                
                updated += 1
            else:
                new_att = AttendanceRecord(
                    period=period, employee_id=emp.id,
                    total_work_days=total_work,
                    actual_work_days=actual_salary,
                    attendance_rate=att_rate,
                    salary_start_date=start_date, salary_end_date=end_date,
                    adjusted_salary_days=adjusted,
                    actual_salary_days=actual_salary,
                    late_to_personal_leave_days=late_leave,
                    leave_total_days=leave_total,
                    late_count=late_count_val,
                    late_duration=late_duration_val,
                    severe_late_count=get_int("severe_late_count"),
                    severe_late_duration=get_int("severe_late_duration"),
                    early_count=get_int("early_count"),
                    early_duration=get_int("early_duration"),
                    half_day_missed_punch=get_int("half_day_missed_punch"),
                    absenteeism_days=get_float("absenteeism_days"),
                    total_overtime=get_float("total_overtime"),
                    personal_leave_days=personal,
                    full_pay_sick_days=full_sick,
                    reduced_pay_sick_days=reduced_sick,
                    statutory_sick_days=stat_sick,
                    compensatory_leave_days=comp,
                    annual_leave_days=annual,
                    prenatal_checkup_days=prenatal,
                    maternity_leave_days=maternity,
                    paternity_leave_days=paternity,
                    marriage_leave_days=marriage,
                    funeral_leave_days=funeral,
                    engineering_compensatory_days=eng,
                    remark=get_str("remark"),
                )
                db.add(new_att)
                created += 1
        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")

    db.commit()
    wb.close()
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "import", f"批量导入考勤：新增{created}条，更新{updated}条 (period={period})")

    return {
        "message": f"导入完成：新增 {created} 条，更新 {updated} 条",
        "created": created,
        "updated": updated,
        "errors": errors
    }


# ==================== 年度工作日历 API ====================

class WorkCalendarDayUpdate(BaseModel):
    date: str
    day_type: Optional[str] = None
    is_salary_day: Optional[bool] = None
    remark: Optional[str] = None


class WorkCalendarBatchUpdate(BaseModel):
    updates: List[WorkCalendarDayUpdate]


@router.get("/work-calendar/{year}", dependencies=[Depends(require_permission("attendance:view"))])
def get_work_calendar(
    year: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """获取指定年份的工作日历数据"""
    days = work_cal.get_year_calendar(db, year)
    summary = work_cal.get_calendar_status_summary(db, year)
    return {
        "year": year,
        "days": days,
        "summary": summary,
    }


@router.put("/work-calendar/{year}", dependencies=[Depends(require_permission("attendance:edit"))])
def update_work_calendar(
    year: int,
    data: WorkCalendarBatchUpdate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """批量更新工作日历"""
    updates = [u.model_dump(exclude_none=True) for u in data.updates]
    updated = work_cal.update_calendar_days(db, updates)
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "calendar_edit", f"更新工作日历 {year}年，共{updated}天")
    summary = work_cal.get_calendar_status_summary(db, year)
    return {
        "message": f"已更新 {updated} 天",
        "updated_count": updated,
        "summary": summary,
    }


@router.post("/work-calendar/{year}/toggle-day", dependencies=[Depends(require_permission("attendance:edit"))])
def toggle_work_calendar_day(
    year: int,
    date_str: str = Form(..., alias="date"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """点击切换单日状态：工作日→节假日→调休补班→工作日"""
    from datetime import datetime as dt
    cal_date = dt.strptime(date_str, "%Y-%m-%d").date()
    month_period = f"{year}{cal_date.month:02d}"

    work_cal.init_year_calendar(db, year)
    from app.models.models import WorkCalendar
    record = db.query(WorkCalendar).filter(WorkCalendar.cal_date == cal_date).first()
    if not record:
        raise HTTPException(status_code=404, detail="日期记录不存在")

    weekday = record.weekday
    is_weekend = weekday >= 5

    if record.day_type == "workday":
        record.day_type = "holiday"
        record.is_salary_day = False
        record.remark = "节假日"
    elif record.day_type == "holiday":
        record.day_type = "makeup_work"
        record.is_salary_day = True
        record.remark = "调休补班"
    elif record.day_type == "makeup_work":
        record.day_type = "weekend" if is_weekend else "workday"
        record.is_salary_day = not is_weekend
        record.remark = None
    else:
        record.day_type = "workday"
        record.is_salary_day = True
        record.remark = None

    record.is_ai_generated = False
    db.commit()

    summary = work_cal.get_calendar_status_summary(db, year)
    return {
        "message": "切换成功",
        "date": date_str,
        "day_type": record.day_type,
        "is_salary_day": record.is_salary_day,
        "summary": summary,
        "period": month_period,
    }


@router.post("/work-calendar/{year}/ai-generate", dependencies=[Depends(require_permission("attendance:edit"))])
def ai_generate_work_calendar(
    year: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """调用AI生成指定年份的法定节假日和调休补班安排"""
    try:
        result = work_cal.generate_holidays_by_ai(year)
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI生成失败：{str(e)}")

    holiday_count, workday_count = work_cal.apply_ai_holidays(
        db, year,
        result.get("holidays", []),
        result.get("workdays", [])
    )

    summary = work_cal.get_calendar_status_summary(db, year)
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "ai_calendar", f"AI预填{year}年工作日历：节假日{holiday_count}天，补班{workday_count}天")
    return {
        "message": f"AI预填完成：标记节假日 {holiday_count} 天，调休补班 {workday_count} 天",
        "holiday_count": holiday_count,
        "workday_count": workday_count,
        "summary": summary,
    }


@router.post("/work-calendar/recalculate", dependencies=[Depends(require_permission("attendance:edit"))])
def recalculate_attendance_days(
    period: Optional[str] = Form(None, description="指定月份重算，支持逗号分隔多个月份(如202606,202607)，为空则重算所有月份"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """根据当前工作日历重算考勤记录的应计薪天数、计薪天数、出勤率"""
    periods = []
    if period:
        periods = [p.strip() for p in period.split(',') if p.strip()]

    updated_count = 0
    updated_periods = []

    if periods:
        for p in periods:
            count = _recalculate_month_attendance(db, p)
            if count > 0:
                updated_count += count
                updated_periods.append(p)
        period_desc = ','.join(updated_periods) if updated_periods else '无'
    else:
        records = db.query(AttendanceRecord).all()
        for att in records:
            _recalculate_attendance_fields(db, att)
            updated_count += 1
        db.commit()
        period_desc = '全部'

    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "recalc", f"根据工作日历重算考勤：{updated_count}条记录 (period={period_desc})")
    return {"message": f"重算完成，共更新 {updated_count} 条考勤记录（月份：{period_desc}）", "updated_count": updated_count}


# ==================== 兼容旧版月度计薪日历API ====================

@router.get("/salary-calendar", dependencies=[Depends(require_permission("attendance:view"))])
def get_salary_calendar(
    period: str = Query(..., description="核算周期 YYYYMM"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """兼容旧版API：获取指定月份的计薪日历（从年度工作日历读取）"""
    year = int(period[:4])
    month = int(period[4:6])
    _, last_day = _calc_salary_dates(period)

    work_cal.init_year_calendar(db, year)
    from app.models.models import WorkCalendar
    records = db.query(WorkCalendar).filter(
        WorkCalendar.year == year,
        WorkCalendar.month == month
    ).order_by(WorkCalendar.day).all()

    days = []
    for r in records:
        is_workday = r.weekday < 5
        is_overridden = r.day_type not in ("workday", "weekend") or (is_workday and not r.is_salary_day) or (not is_workday and r.is_salary_day)
        days.append({
            "date": r.cal_date.isoformat(),
            "day": r.day,
            "weekday": r.weekday,
            "is_workday": is_workday,
            "is_salary_day": r.is_salary_day,
            "is_overridden": is_overridden,
            "day_type": r.day_type,
            "remark": r.remark,
        })

    total_salary_days = sum(1 for d in days if d["is_salary_day"])

    return {
        "period": period,
        "days": days,
        "total_salary_days": total_salary_days,
        "override_count": sum(1 for d in days if d["is_overridden"]),
    }


# ==================== 数据锁定 API ====================

@router.post("/{record_id}/lock-row", dependencies=[Depends(require_permission("attendance:edit"))])
def lock_row(
    record_id: int,
    locked: bool = Form(True),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """锁定/解锁整行考勤记录"""
    att = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="考勤记录不存在")
    
    att.is_row_locked = locked
    if locked:
        att.locked_fields = {}
    db.commit()
    
    action = "锁定" if locked else "解锁"
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "lock_row", f"{action}考勤行 (id={record_id})")
    
    emp = db.query(Employee).filter(Employee.id == att.employee_id).first()
    name_map = _batch_load_dict_names(db)
    return AttendanceOut.from_record(att, emp, db=db, name_map=name_map)


@router.post("/{record_id}/lock-field", dependencies=[Depends(require_permission("attendance:edit"))])
def lock_field(
    record_id: int,
    field: str = Form(...),
    locked: bool = Form(True),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    """锁定/解锁单个单元格（字段）"""
    att = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="考勤记录不存在")
    
    if att.is_row_locked:
        raise HTTPException(status_code=400, detail="整行已锁定，无需单独锁定字段")
    
    locked_fields = att.locked_fields or {}
    locked_fields[field] = locked
    locked_fields = {k: v for k, v in locked_fields.items() if v}
    att.locked_fields = locked_fields
    db.commit()
    
    action = "锁定" if locked else "解锁"
    write_log(db, "data_change", current_user.id, current_user.username, "attendance", "lock_field", f"{action}考勤字段 {field} (id={record_id})")
    
    emp = db.query(Employee).filter(Employee.id == att.employee_id).first()
    name_map = _batch_load_dict_names(db)
    return AttendanceOut.from_record(att, emp, db=db, name_map=name_map)
