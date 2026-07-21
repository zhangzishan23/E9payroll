from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, field_validator
from typing import Optional, List
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from app.core.database import get_db
from app.core.log_helper import write_log
from app.models.models import Employee, EmployeeSalary, SysDictBase, SysUser
from app.api.auth import get_current_user, UserInfo, require_permission
from app.core.query_utils import filter_active_employees, apply_data_scope

router = APIRouter()


class EmployeeOut(BaseModel):
    id: int
    employee_no: str
    dingtalk_user_id: Optional[str] = None
    name: str
    gender: Optional[str] = None
    id_card: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    work_place: Optional[str] = None
    contract_company_id: int
    contract_company_name: Optional[str] = None
    department_id: int
    department_name: Optional[str] = None
    position_id: int
    position_name: Optional[str] = None
    status_id: int
    status_name: Optional[str] = None
    position_level: Optional[str] = None
    employee_type: Optional[str] = None
    job_level: Optional[str] = None
    cost_owner: Optional[str] = None
    report_manager: Optional[str] = None
    entry_date: date
    regular_date: Optional[date] = None
    resign_date: Optional[date] = None
    birth_date: Optional[date] = None
    nation: Optional[str] = None
    marital_status: Optional[str] = None
    children_status: Optional[str] = None
    political_status: Optional[str] = None
    native_place: Optional[str] = None
    residence_type: Optional[str] = None
    census_address: Optional[str] = None
    first_work_date: Optional[date] = None
    education: Optional[str] = None
    graduate_school: Optional[str] = None
    graduate_date: Optional[date] = None
    major: Optional[str] = None
    cert1: Optional[str] = None
    cert2: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_relation: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    contract_start_date: Optional[date] = None
    contract_end_date: Optional[date] = None
    contract_type: Optional[str] = None
    insurance_start_date: Optional[date] = None
    insurance_location: Optional[str] = None
    recruitment_channel: Optional[str] = None
    hobby: Optional[str] = None
    commercial_insurance_type: Optional[str] = None
    remark: Optional[str] = None
    bank_card: Optional[str] = None
    bank_branch: Optional[str] = None
    bank_branch_detail: Optional[str] = None
    home_address: Optional[str] = None
    dept_path: Optional[str] = None
    dept_level1: Optional[str] = None
    dept_level2: Optional[str] = None
    dept_level3: Optional[str] = None
    dept_level4: Optional[str] = None
    dept_level5: Optional[str] = None
    base_salary: Optional[float] = None
    performance_standard: Optional[float] = None
    meal_allowance: Optional[float] = None
    transport_allowance: Optional[float] = None
    communication_allowance: Optional[float] = None
    computer_allowance: Optional[float] = None
    housing_allowance: Optional[float] = None
    salary_effective_date: Optional[str] = None

    class Config:
        from_attributes = True


class EmployeeCreate(BaseModel):
    name: str
    gender: Optional[str] = None
    id_card: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    work_place: Optional[str] = None
    contract_company_id: int
    department_id: int
    position_id: int
    status_id: int
    position_level: Optional[str] = None
    employee_type: Optional[str] = None
    job_level: Optional[str] = None
    cost_owner: Optional[str] = None
    report_manager: Optional[str] = None
    entry_date: date
    regular_date: Optional[date] = None
    birth_date: Optional[date] = None
    nation: Optional[str] = None
    marital_status: Optional[str] = None
    children_status: Optional[str] = None
    political_status: Optional[str] = None
    native_place: Optional[str] = None
    residence_type: Optional[str] = None
    census_address: Optional[str] = None
    first_work_date: Optional[date] = None
    education: Optional[str] = None
    graduate_school: Optional[str] = None
    graduate_date: Optional[date] = None
    major: Optional[str] = None
    cert1: Optional[str] = None
    cert2: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_relation: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    contract_start_date: Optional[date] = None
    contract_end_date: Optional[date] = None
    contract_type: Optional[str] = None
    insurance_start_date: Optional[date] = None
    insurance_location: Optional[str] = None
    recruitment_channel: Optional[str] = None
    hobby: Optional[str] = None
    commercial_insurance_type: Optional[str] = None
    remark: Optional[str] = None
    bank_card: Optional[str] = None
    bank_branch: Optional[str] = None
    bank_branch_detail: Optional[str] = None
    home_address: Optional[str] = None


class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    gender: Optional[str] = None
    id_card: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    work_place: Optional[str] = None
    contract_company_id: Optional[int] = None
    department_id: Optional[int] = None
    position_id: Optional[int] = None
    status_id: Optional[int] = None
    position_level: Optional[str] = None
    employee_type: Optional[str] = None
    job_level: Optional[str] = None
    cost_owner: Optional[str] = None
    report_manager: Optional[str] = None
    entry_date: Optional[date] = None
    regular_date: Optional[date] = None
    resign_date: Optional[date] = None
    birth_date: Optional[date] = None
    nation: Optional[str] = None
    marital_status: Optional[str] = None
    children_status: Optional[str] = None
    political_status: Optional[str] = None
    native_place: Optional[str] = None
    residence_type: Optional[str] = None
    census_address: Optional[str] = None
    first_work_date: Optional[date] = None
    education: Optional[str] = None
    graduate_school: Optional[str] = None
    graduate_date: Optional[date] = None
    major: Optional[str] = None
    cert1: Optional[str] = None
    cert2: Optional[str] = None
    emergency_contact_name: Optional[str] = None
    emergency_contact_relation: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    contract_start_date: Optional[date] = None
    contract_end_date: Optional[date] = None
    contract_type: Optional[str] = None
    insurance_start_date: Optional[date] = None
    insurance_location: Optional[str] = None
    recruitment_channel: Optional[str] = None
    hobby: Optional[str] = None
    commercial_insurance_type: Optional[str] = None
    remark: Optional[str] = None
    bank_card: Optional[str] = None
    bank_branch: Optional[str] = None
    bank_branch_detail: Optional[str] = None
    home_address: Optional[str] = None


class SalaryOut(BaseModel):
    id: int
    employee_id: int
    base_salary: float
    performance_standard: float
    meal_allowance: float
    transport_allowance: float
    communication_allowance: float
    computer_allowance: float
    housing_allowance: float
    effective_date: str
    change_reason: Optional[str] = None
    created_at: Optional[str] = None
    operator_name: Optional[str] = None

    @field_validator('effective_date', mode='before')
    @classmethod
    def convert_effective_date(cls, v):
        if isinstance(v, date):
            return v.isoformat()
        return str(v) if v else ''

    @field_validator('created_at', mode='before')
    @classmethod
    def convert_created_at(cls, v):
        if v:
            if hasattr(v, 'strftime'):
                return v.strftime('%Y-%m-%d %H:%M')
            return str(v)
        return None

    class Config:
        from_attributes = True


class SalaryCreate(BaseModel):
    employee_id: int
    base_salary: float
    performance_standard: float
    meal_allowance: float = 0
    transport_allowance: float = 0
    communication_allowance: float = 0
    computer_allowance: float = 0
    housing_allowance: float = 0
    effective_date: date
    change_reason: Optional[str] = None


def _enrich_employee(emp: Employee, db: Session = None, name_map: dict = None, salary_map: dict = None) -> dict:
    if name_map is None and db is not None:
        dict_items = db.query(SysDictBase).all()
        name_map = {d.id: d.name for d in dict_items}
    if name_map is None:
        name_map = {}

    def _get_name(dict_id):
        if dict_id is None:
            return None
        return name_map.get(dict_id)

    if salary_map is None and db is not None:
        latest_salaries = db.query(EmployeeSalary).order_by(
            EmployeeSalary.employee_id, EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()
        ).all()
        salary_map = {}
        for s in latest_salaries:
            if s.employee_id not in salary_map:
                salary_map[s.employee_id] = s
    if salary_map is None:
        salary_map = {}

    latest_salary = salary_map.get(emp.id)

    result = {
        "id": emp.id, "employee_no": emp.employee_no, "dingtalk_user_id": emp.dingtalk_user_id,
        "name": emp.name,
        "gender": emp.gender, "id_card": emp.id_card, "phone": emp.phone,
        "email": emp.email, "work_place": emp.work_place,
        "contract_company_id": emp.contract_company_id,
        "contract_company_name": _get_name(emp.contract_company_id),
        "department_id": emp.department_id,
        "department_name": _get_name(emp.department_id),
        "position_id": emp.position_id,
        "position_name": _get_name(emp.position_id),
        "status_id": emp.status_id,
        "status_name": _get_name(emp.status_id),
        "position_level": emp.position_level, "employee_type": emp.employee_type,
        "job_level": emp.job_level, "cost_owner": emp.cost_owner,
        "report_manager": emp.report_manager,
        "entry_date": emp.entry_date, "regular_date": emp.regular_date,
        "resign_date": emp.resign_date, "birth_date": emp.birth_date,
        "nation": emp.nation, "marital_status": emp.marital_status,
        "children_status": emp.children_status, "political_status": emp.political_status,
        "native_place": emp.native_place, "residence_type": emp.residence_type,
        "census_address": emp.census_address, "first_work_date": emp.first_work_date,
        "education": emp.education, "graduate_school": emp.graduate_school,
        "graduate_date": emp.graduate_date, "major": emp.major,
        "cert1": emp.cert1, "cert2": emp.cert2,
        "emergency_contact_name": emp.emergency_contact_name,
        "emergency_contact_relation": emp.emergency_contact_relation,
        "emergency_contact_phone": emp.emergency_contact_phone,
        "contract_start_date": emp.contract_start_date,
        "contract_end_date": emp.contract_end_date,
        "contract_type": emp.contract_type,
        "insurance_start_date": emp.insurance_start_date,
        "insurance_location": emp.insurance_location,
        "recruitment_channel": emp.recruitment_channel,
        "hobby": emp.hobby,
        "commercial_insurance_type": emp.commercial_insurance_type,
        "remark": emp.remark,
        "bank_card": emp.bank_card, "bank_branch": emp.bank_branch,
        "bank_branch_detail": emp.bank_branch_detail,
        "home_address": emp.home_address,
        "dept_path": emp.dept_path,
        "dept_level1": emp.dept_level1, "dept_level2": emp.dept_level2,
        "dept_level3": emp.dept_level3, "dept_level4": emp.dept_level4,
        "dept_level5": emp.dept_level5,
        "base_salary": float(latest_salary.base_salary) if latest_salary else None,
        "performance_standard": float(latest_salary.performance_standard) if latest_salary else None,
        "meal_allowance": float(latest_salary.meal_allowance) if latest_salary else None,
        "transport_allowance": float(latest_salary.transport_allowance) if latest_salary else None,
        "communication_allowance": float(latest_salary.communication_allowance) if latest_salary else None,
        "computer_allowance": float(latest_salary.computer_allowance) if latest_salary else None,
        "housing_allowance": float(latest_salary.housing_allowance) if latest_salary else None,
        "salary_effective_date": str(latest_salary.effective_date) if latest_salary and latest_salary.effective_date else None,
    }
    return result


def _batch_load_enrich_data(db: Session, employee_ids: list = None) -> tuple:
    dict_items = db.query(SysDictBase).all()
    name_map = {d.id: d.name for d in dict_items}

    salary_query = db.query(EmployeeSalary)
    if employee_ids:
        salary_query = salary_query.filter(EmployeeSalary.employee_id.in_(employee_ids))
    latest_salaries = salary_query.order_by(
        EmployeeSalary.employee_id, EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()
    ).all()
    salary_map = {}
    for s in latest_salaries:
        if s.employee_id not in salary_map:
            salary_map[s.employee_id] = s

    return name_map, salary_map


@router.get("/", response_model=List[EmployeeOut], dependencies=[Depends(require_permission("employee:view"))])
def get_employees(
    status_id: Optional[int] = Query(None),
    department_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    filter_field: Optional[str] = Query(None),
    filter_value: Optional[str] = Query(None),
    include_disabled_dept: bool = Query(False, description="是否包含已禁用部门的员工"),
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    query = db.query(Employee)
    if not current_user.is_admin:
        query = apply_data_scope(query, db, current_user.data_scope, current_user.id)
    if status_id:
        query = query.filter(Employee.status_id == status_id)
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if filter_field and filter_value:
        if filter_field == 'name':
            query = query.filter(Employee.name.contains(filter_value))
        elif filter_field == 'no':
            query = query.filter(Employee.employee_no.contains(filter_value))
        elif filter_field == 'department':
            dept_ids = db.query(SysDictBase.id).filter(
                SysDictBase.category == 'department',
                SysDictBase.name.contains(filter_value)
            ).all()
            if dept_ids:
                query = query.filter(Employee.department_id.in_([d[0] for d in dept_ids]))
        elif filter_field == 'company':
            company_ids = db.query(SysDictBase.id).filter(
                SysDictBase.category == 'contract_company',
                SysDictBase.name.contains(filter_value)
            ).all()
            if company_ids:
                query = query.filter(Employee.contract_company_id.in_([c[0] for c in company_ids]))
        elif filter_field == 'status':
            status_ids = db.query(SysDictBase.id).filter(
                SysDictBase.category == 'employee_status',
                SysDictBase.name.contains(filter_value)
            ).all()
            if status_ids:
                query = query.filter(Employee.status_id.in_([s[0] for s in status_ids]))
    elif keyword:
        query = query.filter(
            Employee.name.contains(keyword) | Employee.employee_no.contains(keyword)
        )
    query = filter_active_employees(query, db, hide_status_id=hide_status_id)
    employees = query.order_by(Employee.employee_no).all()
    name_map, salary_map = _batch_load_enrich_data(db, [e.id for e in employees])
    return [_enrich_employee(e, name_map=name_map, salary_map=salary_map) for e in employees]


@router.get("/export", dependencies=[Depends(require_permission("employee:export"))])
def export_employees(
    filter_field: Optional[str] = Query(None),
    filter_value: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    query = db.query(Employee)
    if not current_user.is_admin:
        query = apply_data_scope(query, db, current_user.data_scope, current_user.id)
    query = filter_active_employees(query, db)
    if filter_field and filter_value:
        if filter_field == 'name':
            query = query.filter(Employee.name.contains(filter_value))
        elif filter_field == 'no':
            query = query.filter(Employee.employee_no.contains(filter_value))
        elif filter_field == 'department':
            dept_ids = db.query(SysDictBase.id).filter(
                SysDictBase.category == 'department',
                SysDictBase.name.contains(filter_value)
            ).all()
            if dept_ids:
                query = query.filter(Employee.department_id.in_([d[0] for d in dept_ids]))
        elif filter_field == 'company':
            company_ids = db.query(SysDictBase.id).filter(
                SysDictBase.category == 'contract_company',
                SysDictBase.name.contains(filter_value)
            ).all()
            if company_ids:
                query = query.filter(Employee.contract_company_id.in_([c[0] for c in company_ids]))
        elif filter_field == 'status':
            status_ids = db.query(SysDictBase.id).filter(
                SysDictBase.category == 'employee_status',
                SysDictBase.name.contains(filter_value)
            ).all()
            if status_ids:
                query = query.filter(Employee.status_id.in_([s[0] for s in status_ids]))
    employees = query.order_by(Employee.employee_no).all()
    name_map, salary_map = _batch_load_enrich_data(db, [e.id for e in employees])
    enriched = [_enrich_employee(e, name_map=name_map, salary_map=salary_map) for e in employees]

    wb = Workbook()
    ws = wb.active
    ws.title = "员工数据"
    headers = ["编号", "姓名", "性别", "身份证号", "手机号", "合同公司", "部门", "岗位", "状态", "成本归属", "入职日期", "转正日期", "离职日期", "银行卡号", "开户行", "家庭地址", "基本工资", "绩效标准", "餐补", "交通补贴", "通讯补贴", "电脑补贴", "住房补贴", "薪资生效日期"]
    ws.append(headers)
    for e in enriched:
        ws.append([
            e["employee_no"], e["name"], e["gender"], e["id_card"], e.get("phone") or "",
            e.get("contract_company_name") or "", e.get("department_name") or "",
            e.get("position_name") or "", e.get("status_name") or "",
            e.get("cost_owner") or "", str(e["entry_date"]),
            str(e.get("regular_date") or ""), str(e.get("resign_date") or ""),
            e.get("bank_card") or "", e.get("bank_branch") or "", e.get("home_address") or "",
            e.get("base_salary") or "", e.get("performance_standard") or "",
            e.get("meal_allowance") or "", e.get("transport_allowance") or "",
            e.get("communication_allowance") or "", e.get("computer_allowance") or "",
            e.get("housing_allowance") or "", e.get("salary_effective_date") or ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_export.xlsx"}
    )


@router.get("/import-template", dependencies=[Depends(require_permission("employee:import"))])
def download_import_template(
    current_user: UserInfo = Depends(get_current_user)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "员工档案导入模板"

    headers = [
        "姓名*", "费用负责人",
        "基本工资", "绩效标准", "餐补", "交通补贴", "通讯补贴", "电脑补贴", "住房补贴", "薪资生效日期"
    ]

    ws.append(headers)

    examples = [
        "张三", "研发中心",
        "8000", "2000", "300", "200", "100", "100", "500", "2024-01-15"
    ]
    ws.append(examples)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    example_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    note_font = Font(color="6B7280", italic=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 16

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = example_fill

    note_start_row = 4
    notes = [
        "⚠️ 填写说明（导入前请将此区域所有说明行删除）：",
        "",
        "填写说明：",
        "1. 本模板仅用于补充钉钉同步不到的信息，员工必须先从钉钉同步后再导入",
        "2. 带 * 号的为必填项：姓名用于匹配已有员工",
        "3. 姓名、性别、手机号、部门、职位、入职时间、家庭住址、银行卡等所有基本信息",
        "   均会在「同步钉钉」时自动获取，无需在此模板填写",
        "4. 工资金额请填写数字，不要包含货币符号或千分位",
        "5. 填写了薪资数据时，「薪资生效日期」为必填；仅填费用负责人则无需填写",
        "6. 日期格式统一为 YYYY-MM-DD，如 2024-01-15",
        "7. 系统根据姓名匹配已有员工，不会创建新员工"
    ]
    from openpyxl.styles import Font as StFont
    warning_font = StFont(color="FF0000", bold=True, size=12)
    for i, note in enumerate(notes):
        cell = ws.cell(row=note_start_row + i, column=1, value=note)
        if i == 0:
            cell.font = warning_font
        else:
            cell.font = note_font

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from urllib.parse import quote
    filename = quote("员工档案导入模板.xlsx")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.post("/import", dependencies=[Depends(require_permission("employee:import"))])
async def import_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头，没有数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    header_map = {}
    for i, h in enumerate(headers):
        if "姓名" in h:
            header_map["name"] = i
        elif "成本" in h or "费用负责人" in h:
            header_map["cost_owner"] = i
        elif "基本工资" in h:
            header_map["base_salary"] = i
        elif "绩效标准" in h or "绩效奖金" in h:
            header_map["performance_standard"] = i
        elif "餐补" in h:
            header_map["meal_allowance"] = i
        elif "交通补贴" in h:
            header_map["transport_allowance"] = i
        elif "通讯补贴" in h:
            header_map["communication_allowance"] = i
        elif "电脑补贴" in h:
            header_map["computer_allowance"] = i
        elif "住房补贴" in h:
            header_map["housing_allowance"] = i
        elif "薪资生效" in h or "生效日期" in h:
            header_map["salary_effective_date"] = i

    required = ["name"]
    missing = [k for k in required if k not in header_map]
    if missing:
        name_map = {"name": "姓名"}
        missing_names = [name_map.get(k, k) for k in missing]
        raise HTTPException(status_code=400, detail=f"Excel 表头缺少必要列：{', '.join(missing_names)}")

    updated = 0
    salary_created = 0
    errors = []

    def _parse_float(val):
        if val is None or val == "":
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            name = str(row[header_map["name"]]).strip() if row[header_map["name"]] else ""
            
            if not name:
                errors.append(f"第{row_idx}行：姓名不能为空")
                continue

            cost_owner = None
            if header_map.get("cost_owner") is not None and row[header_map["cost_owner"]]:
                cost_owner = str(row[header_map["cost_owner"]]).strip()

            base_salary = _parse_float(row[header_map["base_salary"]]) if header_map.get("base_salary") is not None else None
            performance_standard = _parse_float(row[header_map["performance_standard"]]) if header_map.get("performance_standard") is not None else None
            meal_allowance = _parse_float(row[header_map["meal_allowance"]]) if header_map.get("meal_allowance") is not None else None
            transport_allowance = _parse_float(row[header_map["transport_allowance"]]) if header_map.get("transport_allowance") is not None else None
            communication_allowance = _parse_float(row[header_map["communication_allowance"]]) if header_map.get("communication_allowance") is not None else None
            computer_allowance = _parse_float(row[header_map["computer_allowance"]]) if header_map.get("computer_allowance") is not None else None
            housing_allowance = _parse_float(row[header_map["housing_allowance"]]) if header_map.get("housing_allowance") is not None else None
            salary_effective_str = str(row[header_map["salary_effective_date"]]).strip() if header_map.get("salary_effective_date") is not None and row[header_map["salary_effective_date"]] else None

            has_salary = any(v is not None and v > 0 for v in [
                base_salary, performance_standard, meal_allowance, transport_allowance,
                communication_allowance, computer_allowance, housing_allowance
            ])

            salary_effective_date = None
            if has_salary:
                if not salary_effective_str:
                    errors.append(f"第{row_idx}行：填写了薪资数据但未填写薪资生效日期")
                    continue
                try:
                    salary_effective_date = date.fromisoformat(salary_effective_str[:10])
                except ValueError:
                    errors.append(f"第{row_idx}行：薪资生效日期格式错误「{salary_effective_str}」，应为 YYYY-MM-DD")
                    continue

            existing = db.query(Employee).filter(Employee.name == name).first()
            if not existing:
                errors.append(f"第{row_idx}行：员工「{name}」在系统中不存在，请先从钉钉同步员工后再导入")
                continue

            if cost_owner:
                existing.cost_owner = cost_owner
            updated += 1
            emp_id = existing.id

            if has_salary:
                new_salary = EmployeeSalary(
                    employee_id=emp_id,
                    base_salary=base_salary or 0,
                    performance_standard=performance_standard or 0,
                    meal_allowance=meal_allowance or 0,
                    transport_allowance=transport_allowance or 0,
                    communication_allowance=communication_allowance or 0,
                    computer_allowance=computer_allowance or 0,
                    housing_allowance=housing_allowance or 0,
                    effective_date=salary_effective_date,
                    operator_id=current_user.id,
                    change_reason="批量导入"
                )
                db.add(new_salary)
                salary_created += 1
        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")

    db.commit()
    wb.close()
    write_log(db, "data_change", current_user.id, current_user.username, "employee", "import", 
              f"批量导入员工补充信息：更新{updated}人，薪资记录{salary_created}条")

    msg = f"导入完成：更新 {updated} 人"
    if salary_created > 0:
        msg += f"，新增薪资记录 {salary_created} 条"
    return {
        "message": msg,
        "created": 0,
        "updated": updated,
        "salary_created": salary_created,
        "errors": errors
    }


@router.post("/batch-delete", dependencies=[Depends(require_permission("employee:delete"))])
def batch_delete_employees(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的员工ID列表")
    employees = db.query(Employee).filter(Employee.id.in_(ids)).all()
    if not employees:
        raise HTTPException(status_code=404, detail="未找到指定的员工")
    for emp in employees:
        db.query(EmployeeSalary).filter(EmployeeSalary.employee_id == emp.id).delete()
        db.delete(emp)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "employee", "delete", f"批量删除 {len(employees)} 名员工")
    return {"message": f"成功删除 {len(employees)} 名员工", "deleted_count": len(employees)}


@router.post("/export-selected", dependencies=[Depends(require_permission("employee:export"))])
def export_selected_employees(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要导出的员工ID列表")
    employees = db.query(Employee).filter(Employee.id.in_(ids)).order_by(Employee.employee_no).all()
    name_map, salary_map = _batch_load_enrich_data(db, [e.id for e in employees])
    enriched = [_enrich_employee(e, name_map=name_map, salary_map=salary_map) for e in employees]

    wb = Workbook()
    ws = wb.active
    ws.title = "员工数据"
    headers = ["编号", "姓名", "性别", "身份证号", "手机号", "合同公司", "部门", "岗位", "状态", "成本归属", "入职日期", "转正日期", "离职日期", "银行卡号", "开户行", "家庭地址", "基本工资", "绩效标准", "餐补", "交通补贴", "通讯补贴", "电脑补贴", "住房补贴", "薪资生效日期"]
    ws.append(headers)
    for e in enriched:
        ws.append([
            e["employee_no"], e["name"], e["gender"], e["id_card"], e.get("phone") or "",
            e.get("contract_company_name") or "", e.get("department_name") or "",
            e.get("position_name") or "", e.get("status_name") or "",
            e.get("cost_owner") or "", str(e["entry_date"]),
            str(e.get("regular_date") or ""), str(e.get("resign_date") or ""),
            e.get("bank_card") or "", e.get("bank_branch") or "", e.get("home_address") or "",
            e.get("base_salary") or "", e.get("performance_standard") or "",
            e.get("meal_allowance") or "", e.get("transport_allowance") or "",
            e.get("communication_allowance") or "", e.get("computer_allowance") or "",
            e.get("housing_allowance") or "", e.get("salary_effective_date") or ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees_selected_export.xlsx"}
    )


@router.post("/salaries", response_model=SalaryOut, dependencies=[Depends(require_permission("employee:edit"))])
def create_employee_salary(salary: SalaryCreate, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    employee = db.query(Employee).filter(Employee.id == salary.employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="员工不存在，无法创建薪资记录")
    try:
        data = salary.model_dump()
        data["operator_id"] = current_user.id
        db_salary = EmployeeSalary(**data)
        db.add(db_salary)
        db.commit()
        db.refresh(db_salary)
        write_log(db, "data_change", current_user.id, current_user.username, "employee", "edit", f"新增员工薪资记录 (employee_id={salary.employee_id}, effective_date={salary.effective_date})")
        return db_salary
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"薪资记录创建失败：{str(e)}")


@router.get("/{employee_id}/salaries", response_model=List[SalaryOut], dependencies=[Depends(require_permission("employee:view"))])
def get_employee_salaries(employee_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    salaries = db.query(EmployeeSalary).filter(
        EmployeeSalary.employee_id == employee_id
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).all()
    result = []
    for s in salaries:
        item = {
            "id": s.id, "employee_id": s.employee_id,
            "base_salary": float(s.base_salary), "performance_standard": float(s.performance_standard),
            "meal_allowance": float(s.meal_allowance or 0), "transport_allowance": float(s.transport_allowance or 0),
            "communication_allowance": float(s.communication_allowance or 0),
            "computer_allowance": float(s.computer_allowance or 0), "housing_allowance": float(s.housing_allowance or 0),
            "effective_date": s.effective_date.isoformat() if s.effective_date else "",
            "change_reason": s.change_reason, "created_at": s.created_at, "operator_name": None
        }
        if s.operator_id:
            op = db.query(SysUser).filter(SysUser.id == s.operator_id).first()
            item["operator_name"] = op.username if op else None
        result.append(item)
    return result


@router.delete("/salaries/{salary_id}", dependencies=[Depends(require_permission("employee:edit"))])
def delete_employee_salary(salary_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    sal = db.query(EmployeeSalary).filter(EmployeeSalary.id == salary_id).first()
    if not sal:
        raise HTTPException(status_code=404, detail="薪资记录不存在")
    emp_id = sal.employee_id
    eff_date = sal.effective_date
    db.delete(sal)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "employee", "delete", f"删除员工薪资记录 (employee_id={emp_id}, effective_date={eff_date})")
    return {"message": "删除成功"}


@router.get("/{employee_id}", response_model=EmployeeOut, dependencies=[Depends(require_permission("employee:view"))])
def get_employee(employee_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="员工不存在")
    if not current_user.is_admin:
        if current_user.data_scope == "self":
            from app.core.query_utils import get_user_employee_id
            user_emp_id = get_user_employee_id(db, current_user.id)
            if user_emp_id != emp.id:
                raise HTTPException(status_code=403, detail="您没有权限查看该员工信息")
        elif current_user.data_scope == "dept":
            from app.core.query_utils import get_user_department_id
            user_dept_id = get_user_department_id(db, current_user.id)
            if user_dept_id != emp.department_id:
                raise HTTPException(status_code=403, detail="您没有权限查看其他部门员工信息")
    return _enrich_employee(emp, db)


@router.post("/", response_model=EmployeeOut, dependencies=[Depends(require_permission("employee:create"))])
def create_employee(emp: EmployeeCreate, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    existing = db.query(Employee).filter(Employee.id_card == emp.id_card).first()
    if existing:
        raise HTTPException(status_code=400, detail="该身份证号已存在")

    if emp.department_id:
        dept = db.query(SysDictBase).filter(
            SysDictBase.id == emp.department_id,
            SysDictBase.category == 'department'
        ).first()
        if dept and dept.is_enabled == False:
            raise HTTPException(status_code=400, detail=f"部门「{dept.name}」已被禁用，无法将员工分配到该部门")

    count = db.query(Employee).count()
    employee_no = f"E{count + 1:04d}"
    db_emp = Employee(employee_no=employee_no, **emp.model_dump())
    db.add(db_emp)
    db.commit()
    db.refresh(db_emp)
    write_log(db, "data_change", current_user.id, current_user.username, "employee", "create", f"新增员工 {db_emp.name}({db_emp.employee_no})")
    return _enrich_employee(db_emp, db)


@router.put("/{employee_id}", response_model=EmployeeOut, dependencies=[Depends(require_permission("employee:edit"))])
def update_employee(employee_id: int, emp: EmployeeUpdate, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    db_emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not db_emp:
        raise HTTPException(status_code=404, detail="员工不存在")

    dept_id = emp.department_id if emp.department_id is not None else db_emp.department_id
    if dept_id:
        dept = db.query(SysDictBase).filter(
            SysDictBase.id == dept_id,
            SysDictBase.category == 'department'
        ).first()
        if dept and dept.is_enabled == False:
            raise HTTPException(status_code=400, detail=f"部门「{dept.name}」已被禁用，无法将员工分配到该部门")

    update_data = emp.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_emp, key, value)
    db.commit()
    db.refresh(db_emp)
    write_log(db, "data_change", current_user.id, current_user.username, "employee", "edit", f"编辑员工 {db_emp.name}({db_emp.employee_no})")
    return _enrich_employee(db_emp, db)


@router.delete("/{employee_id}", dependencies=[Depends(require_permission("employee:delete"))])
def delete_employee(employee_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    db_emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not db_emp:
        raise HTTPException(status_code=404, detail="员工不存在")
    db.delete(db_emp)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "employee", "delete", f"删除员工 {db_emp.name}({db_emp.employee_no})")
    return {"message": "删除成功"}
