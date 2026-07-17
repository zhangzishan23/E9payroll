from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from app.core.database import get_db
from app.core.log_helper import write_log
from app.core.query_utils import filter_active_employees
from app.services.work_calendar import get_month_salary_days
from app.models.models import PerformanceScore, Employee, EmployeeSalary, AttendanceRecord
from app.api.auth import get_current_user, UserInfo, require_permission

router = APIRouter()


class PerformanceOut(BaseModel):
    id: Optional[int] = None
    period: str
    employee_id: int
    employee_no: str = ""
    employee_name: str = ""
    contract_company: str = ""
    department: str = ""
    position: str = ""
    performance_standard: float = 0
    initial_score: Optional[float] = None
    final_score: Optional[float] = None
    coefficient: Optional[float] = None
    evaluated_performance: Optional[float] = None
    performance_diff: Optional[float] = None
    score_diff: Optional[float] = None
    performance_category: Optional[str] = None
    score_reason: Optional[str] = None
    review_note: Optional[str] = None
    total_work_days: Optional[float] = None
    actual_work_days: Optional[float] = None
    attendance_rate: Optional[float] = None
    actual_paid_performance: Optional[float] = None
    reviewer_id: Optional[int] = None

    class Config:
        from_attributes = True


class PerformanceCreate(BaseModel):
    period: str
    employee_id: int
    initial_score: Optional[float] = None
    final_score: Optional[float] = None
    performance_category: Optional[str] = None
    score_reason: Optional[str] = None
    review_note: Optional[str] = None


class PerformanceUpdate(BaseModel):
    initial_score: Optional[float] = None
    final_score: Optional[float] = None
    performance_category: Optional[str] = None
    score_reason: Optional[str] = None
    review_note: Optional[str] = None


class PerformanceImportItem(BaseModel):
    employee_no: str
    initial_score: Optional[float] = None
    final_score: Optional[float] = None
    performance_category: Optional[str] = None
    score_reason: Optional[str] = None
    review_note: Optional[str] = None


def _build_performance_out(p: Optional[PerformanceScore], emp: Employee, perf_std: float,
                           total_work_days: Optional[float], actual_work_days: Optional[float],
                           period: str) -> PerformanceOut:
    initial = float(p.initial_score) if p and p.initial_score is not None else None
    final = float(p.final_score) if p and p.final_score is not None else None
    coef = final
    evaluated = round(perf_std * coef, 2) if coef is not None else None
    diff = round(evaluated - perf_std, 2) if evaluated is not None else None
    score_diff = round(final - initial, 2) if (initial is not None and final is not None) else None

    att_rate = None
    actual_paid = None
    if total_work_days and actual_work_days is not None and total_work_days > 0:
        att_rate = round(actual_work_days / total_work_days, 4)
        actual_paid = round(evaluated * att_rate, 2) if evaluated is not None else None

    return PerformanceOut(
        id=p.id if p else None,
        period=period,
        employee_id=emp.id,
        employee_no=emp.employee_no,
        employee_name=emp.name,
        contract_company="",
        department="",
        position="",
        performance_standard=perf_std,
        initial_score=initial,
        final_score=final,
        coefficient=coef,
        evaluated_performance=evaluated,
        performance_diff=diff,
        score_diff=score_diff,
        performance_category=p.performance_category if p else None,
        score_reason=p.score_reason if p else None,
        review_note=p.review_note if p else None,
        total_work_days=total_work_days,
        actual_work_days=actual_work_days,
        attendance_rate=att_rate,
        actual_paid_performance=actual_paid,
        reviewer_id=p.reviewer_id if p else None,
    )


def _get_dict_name(db: Session, dict_id: Optional[int], name_map: dict = None) -> str:
    if not dict_id:
        return ""
    if name_map is not None:
        return name_map.get(dict_id, "")
    from app.models.models import SysDictBase
    d = db.query(SysDictBase).filter(SysDictBase.id == dict_id).first()
    return d.name if d else ""


def _batch_load_dict_names(db: Session) -> dict:
    from app.models.models import SysDictBase
    dict_items = db.query(SysDictBase).all()
    return {d.id: d.name for d in dict_items}


@router.get("/", response_model=List[PerformanceOut], dependencies=[Depends(require_permission("performance:view"))])
def get_performances(
    period: Optional[str] = Query(None),
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if period:
        from datetime import date
        query = db.query(Employee)
        employees = filter_active_employees(query, db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()
        perf_map = {}
        records = db.query(PerformanceScore).filter(PerformanceScore.period == period).all()
        for r in records:
            perf_map[r.employee_id] = r

        year = int(period[:4])
        month = int(period[4:])
        if month == 12:
            period_end = date(year + 1, 1, 1)
        else:
            period_end = date(year, month + 1, 1)

        sal_map = {}
        for sal in db.query(EmployeeSalary).filter(
            EmployeeSalary.effective_date < period_end
        ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).all():
            if sal.employee_id not in sal_map:
                sal_map[sal.employee_id] = sal

        att_map = {}
        for att in db.query(AttendanceRecord).filter(AttendanceRecord.period == period).all():
            att_map[att.employee_id] = att

        name_map = _batch_load_dict_names(db)
        standard_salary_days = get_month_salary_days(db, period)

        result = []
        for emp in employees:
            p = perf_map.get(emp.id)
            sal = sal_map.get(emp.id)
            att = att_map.get(emp.id)
            perf_std = float(sal.performance_standard) if sal else 0
            
            if att:
                total_days = float(att.adjusted_salary_days) if att.adjusted_salary_days else standard_salary_days
                actual_days = float(att.actual_salary_days)
            else:
                total_days = standard_salary_days
                actual_days = standard_salary_days

            out = _build_performance_out(p, emp, perf_std, total_days, actual_days, period)
            out.contract_company = _get_dict_name(db, emp.contract_company_id, name_map)
            out.department = _get_dict_name(db, emp.department_id, name_map)
            out.position = _get_dict_name(db, emp.position_id, name_map)
            result.append(out)
        return result

    records = db.query(PerformanceScore).order_by(
        PerformanceScore.period.desc(), PerformanceScore.employee_id
    ).all()
    result = []
    sal_map = {}
    for sal in db.query(EmployeeSalary).order_by(
        EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()
    ).all():
        if sal.employee_id not in sal_map:
            sal_map[sal.employee_id] = sal

    emp_ids = list(set(r.employee_id for r in records))
    emp_map = {e.id: e for e in db.query(Employee).filter(Employee.id.in_(emp_ids)).all()}
    name_map = _batch_load_dict_names(db)

    for r in records:
        emp = emp_map.get(r.employee_id)
        if not emp:
            continue
        sal = sal_map.get(emp.id)
        perf_std = float(sal.performance_standard) if sal else 0
        out = _build_performance_out(r, emp, perf_std, None, None, r.period)
        out.contract_company = _get_dict_name(db, emp.contract_company_id, name_map)
        out.department = _get_dict_name(db, emp.department_id, name_map)
        out.position = _get_dict_name(db, emp.position_id, name_map)
        result.append(out)
    return result


@router.post("/", response_model=PerformanceOut, dependencies=[Depends(require_permission("performance:create"))])
def create_performance(
    data: PerformanceCreate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    existing = db.query(PerformanceScore).filter(
        PerformanceScore.period == data.period,
        PerformanceScore.employee_id == data.employee_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"员工 [{data.employee_id}] 在周期 [{data.period}] 的绩效数据已存在")

    perf = PerformanceScore(
        period=data.period,
        employee_id=data.employee_id,
        initial_score=data.initial_score,
        final_score=data.final_score,
        performance_category=data.performance_category,
        score_reason=data.score_reason,
        review_note=data.review_note,
        reviewer_id=current_user.id
    )
    db.add(perf)
    db.commit()
    db.refresh(perf)
    write_log(db, "data_change", current_user.id, current_user.username, "performance", "create", f"新增绩效数据 (employee_id={data.employee_id}, period={data.period})")

    emp = db.query(Employee).filter(Employee.id == perf.employee_id).first()
    sal = db.query(EmployeeSalary).filter(
        EmployeeSalary.employee_id == perf.employee_id
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).first()
    att = db.query(AttendanceRecord).filter(
        AttendanceRecord.period == perf.period,
        AttendanceRecord.employee_id == perf.employee_id
    ).first()
    perf_std = float(sal.performance_standard) if sal else 0
    standard_salary_days = get_month_salary_days(db, perf.period)
    if att:
        total_days = float(att.adjusted_salary_days) if att.adjusted_salary_days else standard_salary_days
        actual_days = float(att.actual_salary_days)
    else:
        total_days = standard_salary_days
        actual_days = standard_salary_days
    out = _build_performance_out(perf, emp, perf_std, total_days, actual_days, perf.period)
    out.contract_company = _get_dict_name(db, emp.contract_company_id)
    out.department = _get_dict_name(db, emp.department_id)
    out.position = _get_dict_name(db, emp.position_id)
    return out


@router.put("/{perf_id}", response_model=PerformanceOut, dependencies=[Depends(require_permission("performance:edit"))])
def update_performance(
    perf_id: int,
    data: PerformanceUpdate,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    perf = db.query(PerformanceScore).filter(PerformanceScore.id == perf_id).first()
    if not perf:
        raise HTTPException(status_code=404, detail="绩效记录不存在")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(perf, key, value)
    perf.reviewer_id = current_user.id

    db.commit()
    db.refresh(perf)
    write_log(db, "data_change", current_user.id, current_user.username, "performance", "edit", f"编辑绩效数据 (id={perf_id})")

    emp = db.query(Employee).filter(Employee.id == perf.employee_id).first()
    sal = db.query(EmployeeSalary).filter(
        EmployeeSalary.employee_id == perf.employee_id
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).first()
    att = db.query(AttendanceRecord).filter(
        AttendanceRecord.period == perf.period,
        AttendanceRecord.employee_id == perf.employee_id
    ).first()
    perf_std = float(sal.performance_standard) if sal else 0
    standard_salary_days = get_month_salary_days(db, perf.period)
    if att:
        total_days = float(att.adjusted_salary_days) if att.adjusted_salary_days else standard_salary_days
        actual_days = float(att.actual_salary_days)
    else:
        total_days = standard_salary_days
        actual_days = standard_salary_days
    out = _build_performance_out(perf, emp, perf_std, total_days, actual_days, perf.period)
    out.contract_company = _get_dict_name(db, emp.contract_company_id)
    out.department = _get_dict_name(db, emp.department_id)
    out.position = _get_dict_name(db, emp.position_id)
    return out


@router.delete("/{perf_id}", dependencies=[Depends(require_permission("performance:delete"))])
def delete_performance(
    perf_id: int,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    perf = db.query(PerformanceScore).filter(PerformanceScore.id == perf_id).first()
    if not perf:
        raise HTTPException(status_code=404, detail="绩效记录不存在")
    db.delete(perf)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "performance", "delete", f"删除绩效记录 (id={perf_id})")
    return {"message": "删除成功"}


@router.post("/import/{period}", dependencies=[Depends(require_permission("performance:import"))])
def import_performances(
    period: str,
    items: List[PerformanceImportItem],
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    emp_map = {e.employee_no: e for e in db.query(Employee).all()}
    created = 0
    updated = 0
    errors = []

    for item in items:
        emp = emp_map.get(item.employee_no)
        if not emp:
            errors.append(f"员工编号 [{item.employee_no}] 不存在")
            continue

        existing = db.query(PerformanceScore).filter(
            PerformanceScore.period == period,
            PerformanceScore.employee_id == emp.id
        ).first()

        if existing:
            if item.initial_score is not None:
                existing.initial_score = item.initial_score
            if item.final_score is not None:
                existing.final_score = item.final_score
            if item.performance_category is not None:
                existing.performance_category = item.performance_category
            if item.score_reason is not None:
                existing.score_reason = item.score_reason
            if item.review_note is not None:
                existing.review_note = item.review_note
            existing.reviewer_id = current_user.id
            updated += 1
        else:
            perf = PerformanceScore(
                period=period,
                employee_id=emp.id,
                initial_score=item.initial_score,
                final_score=item.final_score,
                performance_category=item.performance_category,
                score_reason=item.score_reason,
                review_note=item.review_note,
                reviewer_id=current_user.id
            )
            db.add(perf)
            created += 1

    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "performance", "import", f"批量导入绩效：新增{created}条，更新{updated}条 (period={period})")
    return {
        "message": f"导入完成：新增 {created} 条，更新 {updated} 条",
        "created": created,
        "updated": updated,
        "errors": errors
    }


@router.get("/export/{period}", dependencies=[Depends(require_permission("performance:export"))])
def export_performances(
    period: str,
    hide_status_id: Optional[int] = Query(None, description="要隐藏的员工状态ID"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    query = db.query(Employee)
    employees = filter_active_employees(query, db, hide_status_id=hide_status_id).order_by(Employee.employee_no).all()
    perf_map = {}
    records = db.query(PerformanceScore).filter(PerformanceScore.period == period).all()
    for r in records:
        perf_map[r.employee_id] = r

    sal_map = {}
    year = int(period[:4])
    month = int(period[4:])
    if month == 12:
        export_period_end = __import__('datetime').date(year + 1, 1, 1)
    else:
        export_period_end = __import__('datetime').date(year, month + 1, 1)
    for sal in db.query(EmployeeSalary).filter(
        EmployeeSalary.effective_date < export_period_end
    ).order_by(EmployeeSalary.effective_date.desc(), EmployeeSalary.id.desc()).all():
        if sal.employee_id not in sal_map:
            sal_map[sal.employee_id] = sal

    att_map = {}
    for att in db.query(AttendanceRecord).filter(AttendanceRecord.period == period).all():
        att_map[att.employee_id] = att

    name_map = _batch_load_dict_names(db)
    standard_salary_days = get_month_salary_days(db, period)

    wb = Workbook()
    ws = wb.active
    ws.title = f"绩效数据_{period}"
    headers = [
        "员工编号", "员工姓名", "合同公司", "部门", "职务",
        "应计薪天数", "实际计薪天数", "出勤率",
        "绩效奖金标准", "绩效类别",
        "初评", "复评", "评价后绩效标准", "差额",
        "实发绩效金额",
        "调整差异", "评分理由", "分管领导审核后调整"
    ]
    ws.append(headers)

    for emp in employees:
        p = perf_map.get(emp.id)
        sal = sal_map.get(emp.id)
        att = att_map.get(emp.id)
        perf_std = float(sal.performance_standard) if sal else 0
        initial = float(p.initial_score) if p and p.initial_score is not None else None
        final = float(p.final_score) if p and p.final_score is not None else None
        coef = final
        evaluated = round(perf_std * coef, 2) if coef is not None else None
        diff = round(evaluated - perf_std, 2) if evaluated is not None else None
        score_diff = round(final - initial, 2) if (initial is not None and final is not None) else None
        
        if att:
            total_days = float(att.adjusted_salary_days) if att.adjusted_salary_days else standard_salary_days
            actual_days = float(att.actual_salary_days)
        else:
            total_days = standard_salary_days
            actual_days = standard_salary_days
            
        att_rate_val = round(actual_days / total_days, 4) if total_days > 0 else 0
        actual_paid_val = round(evaluated * att_rate_val, 2) if evaluated is not None else None

        ws.append([
            emp.employee_no,
            emp.name,
            _get_dict_name(db, emp.contract_company_id, name_map),
            _get_dict_name(db, emp.department_id, name_map),
            _get_dict_name(db, emp.position_id, name_map),
            total_days,
            actual_days if att else "",
            f"{att_rate_val * 100:.1f}%" if att else "100.0%",
            perf_std,
            p.performance_category if p else "",
            initial if initial is not None else "",
            final if final is not None else "",
            evaluated if evaluated is not None else "",
            diff if diff is not None else "",
            actual_paid_val if actual_paid_val is not None else "",
            score_diff if score_diff is not None else "",
            p.score_reason if p else "",
            p.review_note if p else ""
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=performance_{period}.xlsx"}
    )


@router.post("/batch-delete", dependencies=[Depends(require_permission("performance:delete"))])
def batch_delete_performances(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的绩效记录ID列表")
    records = db.query(PerformanceScore).filter(PerformanceScore.id.in_(ids)).all()
    if not records:
        raise HTTPException(status_code=404, detail="未找到指定的绩效记录")
    for r in records:
        db.delete(r)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "performance", "delete", f"批量删除 {len(records)} 条绩效记录")
    return {"message": f"成功删除 {len(records)} 条绩效记录", "deleted_count": len(records)}


@router.post("/import-excel", dependencies=[Depends(require_permission("performance:import"))])
async def import_performances_excel(
    file: UploadFile = File(...),
    period: str = Form(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(get_current_user)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头，没有数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    header_map = {}
    for i, h in enumerate(headers):
        if "姓名" in h or "名字" in h:
            header_map["employee_name"] = i
        elif "初评" in h:
            header_map["initial_score"] = i
        elif "复评" in h and "审核" not in h and "调整" not in h:
            header_map["final_score"] = i
        elif "绩效类别" in h or "类别" in h:
            header_map["performance_category"] = i
        elif "评分理由" in h or "理由" in h:
            header_map["score_reason"] = i
        elif "分管领导" in h or "审核后调整" in h or "审核备注" in h:
            header_map["review_note"] = i

    if "employee_name" not in header_map:
        raise HTTPException(status_code=400, detail="Excel 表头缺少必需列：姓名")

    all_employees = db.query(Employee).all()
    name_count = {}
    emp_name_map = {}
    for emp in all_employees:
        name = emp.name.strip() if emp.name else ""
        if not name:
            continue
        name_count[name] = name_count.get(name, 0) + 1
        emp_name_map[name] = emp

    duplicate_names = [n for n, c in name_count.items() if c > 1]

    created = 0
    updated = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            emp_name = str(row[header_map["employee_name"]]).strip() if row[header_map["employee_name"]] else ""
            if not emp_name:
                errors.append(f"第{row_idx}行：姓名为空")
                continue

            if emp_name in duplicate_names:
                errors.append(f"第{row_idx}行：姓名「{emp_name}」存在重名，请使用员工编号或手动录入")
                continue

            emp = emp_name_map.get(emp_name)
            if not emp:
                errors.append(f"第{row_idx}行：员工姓名「{emp_name}」不存在")
                continue

            def get_float(key):
                if key in header_map and row[header_map[key]] is not None:
                    try:
                        val = row[header_map[key]]
                        if isinstance(val, str):
                            val = val.replace('%', '').strip()
                            if val.endswith('%'):
                                return float(val[:-1]) / 100
                        return float(val)
                    except (ValueError, TypeError):
                        return None
                return None

            def get_str(key):
                if key in header_map and row[header_map[key]] is not None:
                    return str(row[header_map[key]]).strip()
                return None

            initial = get_float("initial_score")
            final = get_float("final_score")
            category = get_str("performance_category")
            reason = get_str("score_reason")
            note = get_str("review_note")

            existing = db.query(PerformanceScore).filter(
                PerformanceScore.period == period,
                PerformanceScore.employee_id == emp.id
            ).first()

            if existing:
                if initial is not None:
                    existing.initial_score = initial
                if final is not None:
                    existing.final_score = final
                if category is not None:
                    existing.performance_category = category
                if reason is not None:
                    existing.score_reason = reason
                if note is not None:
                    existing.review_note = note
                existing.reviewer_id = current_user.id
                updated += 1
            else:
                perf = PerformanceScore(
                    period=period, employee_id=emp.id,
                    initial_score=initial, final_score=final,
                    performance_category=category,
                    score_reason=reason, review_note=note,
                    reviewer_id=current_user.id
                )
                db.add(perf)
                created += 1
        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")

    db.commit()
    wb.close()
    write_log(db, "data_change", current_user.id, current_user.username, "performance", "import", f"Excel批量导入绩效：新增{created}条，更新{updated}条 (period={period})")

    return {
        "message": f"导入完成：新增 {created} 条，更新 {updated} 条",
        "created": created,
        "updated": updated,
        "errors": errors
    }
