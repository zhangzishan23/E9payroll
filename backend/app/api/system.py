from datetime import datetime
import shutil
import subprocess
from pathlib import Path
from urllib.parse import urlparse
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, field_validator
from typing import Optional, List
from io import BytesIO
from openpyxl import Workbook
from app.core.database import get_db
from app.core.log_helper import write_log
from app.core.config import DATABASE_URL
from app.core.security import get_password_hash
from app.models.models import (
    SysUser, SysRole, SysUserRole, SysPermission, SysDictBase, SysLog, Employee,
    EmployeeSalary, PerformanceScore, CalculationLog, ApprovalRecord,
    ExportTemplate, SalaryPeriodStep
)
from app.api.auth import get_current_user, UserInfo, require_permission


PERMISSION_MODULES = [
    {
        "key": "dashboard",
        "label": "工作台",
        "actions": [
            {"key": "view", "label": "查看工作台"},
            {"key": "work_view", "label": "工作视角/智能工作台"},
            {"key": "leader_view", "label": "管理视角/数据看板"}
        ]
    },
    {
        "key": "employee",
        "label": "人事档案",
        "actions": [
            {"key": "view", "label": "查看档案"},
            {"key": "create", "label": "新增员工"},
            {"key": "edit", "label": "编辑档案"},
            {"key": "delete", "label": "删除员工"},
            {"key": "export", "label": "导出花名册"},
            {"key": "import", "label": "批量导入"},
            {"key": "sync", "label": "同步钉钉"},
        ]
    },
    {
        "key": "attendance",
        "label": "考勤管理",
        "actions": [
            {"key": "view", "label": "查看考勤"},
            {"key": "create", "label": "新增记录"},
            {"key": "edit", "label": "编辑考勤"},
            {"key": "delete", "label": "删除记录"},
            {"key": "export", "label": "导出考勤"},
            {"key": "import", "label": "导入考勤"},
            {"key": "sync", "label": "同步钉钉"},
            {"key": "writeoff", "label": "缺卡核销"},
        ]
    },
    {
        "key": "performance",
        "label": "绩效评分",
        "actions": [
            {"key": "view", "label": "查看绩效"},
            {"key": "create", "label": "新增评分"},
            {"key": "edit", "label": "编辑评分"},
            {"key": "delete", "label": "删除评分"},
            {"key": "export", "label": "导出绩效"},
            {"key": "import", "label": "导入绩效"},
        ]
    },
    {
        "key": "insurance",
        "label": "社保公积金",
        "actions": [
            {"key": "view", "label": "查看社保"},
            {"key": "create", "label": "新增记录"},
            {"key": "edit", "label": "编辑社保"},
            {"key": "delete", "label": "删除记录"},
            {"key": "export", "label": "导出社保"},
            {"key": "import", "label": "导入社保"},
            {"key": "template", "label": "管理导入模板"},
        ]
    },
    {
        "key": "salary",
        "label": "薪资计算",
        "actions": [
            {"key": "view", "label": "查看薪资"},
            {"key": "create", "label": "新增记录"},
            {"key": "edit", "label": "编辑薪资"},
            {"key": "delete", "label": "删除薪资"},
            {"key": "check", "label": "数据检查"},
            {"key": "step_confirm", "label": "步骤确认"},
            {"key": "tax_export", "label": "导出报税模板"},
            {"key": "tax_import", "label": "导入个税申报结果"},
            {"key": "travel_import", "label": "导入临时性差旅补贴"},
            {"key": "export", "label": "导出薪资"},
        ]
    },
    {
        "key": "approval",
        "label": "审批流程",
        "actions": [
            {"key": "view", "label": "查看审批"},
            {"key": "submit", "label": "提交审批"},
            {"key": "approve", "label": "审核操作"},
        ]
    },
    {
        "key": "report",
        "label": "报表导出",
        "actions": [
            {"key": "view", "label": "查看报表"},
            {"key": "export", "label": "导出报表"},
            {"key": "view_my_slip", "label": "查看个人工资条"},
        ]
    },
    {
        "key": "system",
        "label": "系统管理",
        "actions": [
            {"key": "view", "label": "查看系统设置"},
            {"key": "user", "label": "用户管理"},
            {"key": "role", "label": "角色权限"},
            {"key": "dict", "label": "数据字典"},
            {"key": "log", "label": "操作日志"},
            {"key": "backup", "label": "数据备份"},
        ]
    },
    {
        "key": "profile",
        "label": "个人中心",
        "actions": [
            {"key": "view", "label": "查看个人信息"},
            {"key": "edit", "label": "修改个人信息"},
        ]
    },
]

router = APIRouter()


class DictItemOut(BaseModel):
    id: int
    category: str
    code: str
    name: str
    parent_id: Optional[int] = None
    sort_order: int
    is_enabled: bool
    extra: Optional[dict] = None

    class Config:
        from_attributes = True


class DictItemCreate(BaseModel):
    category: str
    code: str
    name: str
    parent_id: Optional[int] = None
    sort_order: int = 0
    is_enabled: bool = True
    extra: Optional[dict] = None


class DictItemUpdate(BaseModel):
    name: Optional[str] = None
    parent_id: Optional[int] = None
    sort_order: Optional[int] = None
    is_enabled: Optional[bool] = None
    extra: Optional[dict] = None


class RoleOut(BaseModel):
    id: int
    name: str
    description: Optional[str] = None
    is_preset: bool
    data_scope: str

    class Config:
        from_attributes = True


class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = None
    data_scope: str = "all"


class RoleUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    data_scope: Optional[str] = None


class PermissionAssign(BaseModel):
    permissions: List[dict]


class UserRoleOut(BaseModel):
    id: int
    name: str

    class Config:
        from_attributes = True


class UserOut(BaseModel):
    id: int
    username: str
    display_name: str
    is_admin: bool
    is_active: bool
    role_ids: List[int] = []
    role_names: List[str] = []
    created_at: Optional[str] = None

    @field_validator('created_at', mode='before')
    @classmethod
    def convert_datetime(cls, v):
        if isinstance(v, datetime):
            return v.isoformat()
        return v

    class Config:
        from_attributes = True


class UserCreate(BaseModel):
    username: str
    password: str
    display_name: str
    is_admin: bool = False
    is_active: bool = True
    role_ids: List[int] = []


class UserUpdate(BaseModel):
    display_name: Optional[str] = None
    is_active: Optional[bool] = None
    is_admin: Optional[bool] = None
    role_ids: Optional[List[int]] = None


class PasswordReset(BaseModel):
    new_password: str = "123456"


class LogOut(BaseModel):
    id: int
    log_type: str
    username: Optional[str] = None
    module: Optional[str] = None
    action: Optional[str] = None
    target: Optional[str] = None
    detail: Optional[dict] = None
    ip_address: Optional[str] = None
    result: Optional[str] = None
    created_at: Optional[str] = None

    @field_validator('created_at', mode='before')
    @classmethod
    def convert_datetime(cls, v):
        if isinstance(v, datetime):
            return v.isoformat()
        return v

    class Config:
        from_attributes = True


class DictTreeNode(BaseModel):
    id: int
    category: str
    code: str
    name: str
    parent_id: Optional[int] = None
    sort_order: int
    is_enabled: bool
    extra: Optional[dict] = None
    children: List["DictTreeNode"] = []

    class Config:
        from_attributes = True


DictTreeNode.model_rebuild()


@router.get("/dict/{category}", response_model=List[DictItemOut])
def get_dict(category: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    items = db.query(SysDictBase).filter(
        SysDictBase.category == category,
    ).order_by(SysDictBase.sort_order).all()
    return items


@router.get("/dict/{category}/tree", response_model=List[DictTreeNode])
def get_dict_tree(category: str, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    items = db.query(SysDictBase).filter(
        SysDictBase.category == category,
    ).order_by(SysDictBase.sort_order).all()

    node_map = {}
    for item in items:
        node_map[item.id] = {
            "id": item.id,
            "category": item.category,
            "code": item.code,
            "name": item.name,
            "parent_id": item.parent_id,
            "sort_order": item.sort_order,
            "is_enabled": item.is_enabled,
            "extra": item.extra,
            "children": []
        }

    roots = []
    for item in items:
        node = node_map[item.id]
        if item.parent_id and item.parent_id in node_map:
            node_map[item.parent_id]["children"].append(node)
        else:
            roots.append(node)

    def clean_children(nodes):
        for node in nodes:
            if not node["children"]:
                node["children"] = []
            else:
                clean_children(node["children"])

    clean_children(roots)
    return roots


@router.get("/dict", response_model=List[DictItemOut])
def get_all_dicts(db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    items = db.query(SysDictBase).order_by(SysDictBase.category, SysDictBase.sort_order).all()
    return items


@router.post("/dict", response_model=DictItemOut)
def create_dict(item: DictItemCreate, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:dict"))):
    existing = db.query(SysDictBase).filter(
        SysDictBase.category == item.category,
        SysDictBase.code == item.code
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"字典项 [{item.code}] 已存在")
    db_item = SysDictBase(**item.model_dump())
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "create", f"新增字典项 {item.category}/{item.code}")
    return db_item


@router.put("/dict/{dict_id}/toggle")
def toggle_dict(dict_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:dict"))):
    db_item = db.query(SysDictBase).filter(SysDictBase.id == dict_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="字典项不存在")
    db_item.is_enabled = not db_item.is_enabled
    db.commit()
    db.refresh(db_item)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "toggle",
              f"{'启用' if db_item.is_enabled else '禁用'}字典项 {db_item.name} (id={dict_id})")
    return {"id": db_item.id, "is_enabled": db_item.is_enabled}


@router.put("/dict/{dict_id}", response_model=DictItemOut)
def update_dict(dict_id: int, item: DictItemUpdate, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:dict"))):
    db_item = db.query(SysDictBase).filter(SysDictBase.id == dict_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="字典项不存在")
    update_data = item.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_item, key, value)
    db.commit()
    db.refresh(db_item)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "edit", f"编辑字典项 (id={dict_id})")
    return db_item


@router.delete("/dict/{dict_id}")
def delete_dict(dict_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:dict"))):
    db_item = db.query(SysDictBase).filter(SysDictBase.id == dict_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="字典项不存在")
    db.delete(db_item)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "delete", f"删除字典项 (id={dict_id})")
    return {"message": "删除成功"}


@router.post("/dict/batch-delete")
def batch_delete_dict(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:dict"))):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的字典项ID列表")
    items = db.query(SysDictBase).filter(SysDictBase.id.in_(ids)).all()
    if not items:
        raise HTTPException(status_code=404, detail="未找到指定的字典项")
    for item in items:
        db.delete(item)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "delete", f"批量删除 {len(items)} 个字典项")
    return {"message": f"成功删除 {len(items)} 个字典项", "deleted_count": len(items)}


@router.post("/dict/dedup")
def dedup_dict(
    category: Optional[str] = Query(None, description="指定分类去重，为空则处理所有分类"),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(require_permission("system:dict")),
):
    q = db.query(SysDictBase)
    if category:
        q = q.filter(SysDictBase.category == category)
    items = q.order_by(SysDictBase.id).all()

    seen = {}
    to_delete = []
    kept_ids = set()
    for item in items:
        key = (item.category, item.name.strip())
        if key not in seen:
            seen[key] = item
            kept_ids.add(item.id)
        else:
            to_delete.append(item)

    if not to_delete:
        return {"message": "没有发现重复数据", "deleted_count": 0, "affected_categories": []}

    affected_categories = set()
    for dup in to_delete:
        kept = seen[(dup.category, dup.name.strip())]
        if dup.category == "department":
            db.query(Employee).filter(Employee.department_id == dup.id).update(
                {"department_id": kept.id}, synchronize_session=False
            )
        elif dup.category == "contract_company":
            db.query(Employee).filter(Employee.contract_company_id == dup.id).update(
                {"contract_company_id": kept.id}, synchronize_session=False
            )
        elif dup.category == "position":
            db.query(Employee).filter(Employee.position_id == dup.id).update(
                {"position_id": kept.id}, synchronize_session=False
            )
        elif dup.category == "employee_status":
            db.query(Employee).filter(Employee.status_id == dup.id).update(
                {"status_id": kept.id}, synchronize_session=False
            )
        affected_categories.add(dup.category)
        db.delete(dup)

    db.commit()
    write_log(
        db, "data_change", current_user.id, current_user.username, "system", "dedup",
        f"数据字典去重：删除 {len(to_delete)} 个重复项，涉及分类：{', '.join(affected_categories)}"
    )
    return {
        "message": f"去重完成，删除了 {len(to_delete)} 条重复记录",
        "deleted_count": len(to_delete),
        "affected_categories": list(affected_categories),
    }


@router.post("/dict/import")
async def import_dict(
    file: UploadFile = File(...),
    category: str = Form(...),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(require_permission("system:dict"))
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        contents = await file.read()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(contents), read_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel 文件：{str(e)}")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件为空或只有表头，没有数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    header_map = {}
    for i, h in enumerate(headers):
        if "编码" in h:
            header_map["code"] = i
        elif "名称" in h:
            header_map["name"] = i
        elif "排序" in h:
            header_map["sort_order"] = i
        elif "启用" in h:
            header_map["is_enabled"] = i

    if "code" not in header_map or "name" not in header_map:
        raise HTTPException(status_code=400, detail="Excel 表头缺少「编码」或「名称」列")

    created = 0
    updated = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            code = str(row[header_map["code"]]).strip() if row[header_map["code"]] else ""
            name = str(row[header_map["name"]]).strip() if row[header_map["name"]] else ""
            if not code or not name:
                errors.append(f"第{row_idx}行：编码或名称为空，跳过")
                continue

            sort_order = 0
            if "sort_order" in header_map and row[header_map["sort_order"]] is not None:
                try:
                    sort_order = int(row[header_map["sort_order"]])
                except (ValueError, TypeError):
                    sort_order = 0

            is_enabled = True
            if "is_enabled" in header_map and row[header_map["is_enabled"]] is not None:
                val = str(row[header_map["is_enabled"]]).strip().lower()
                is_enabled = val in ("是", "true", "1", "yes", "启用")

            existing = db.query(SysDictBase).filter(
                SysDictBase.category == category,
                SysDictBase.code == code
            ).first()

            if existing:
                existing.name = name
                existing.sort_order = sort_order
                existing.is_enabled = is_enabled
                updated += 1
            else:
                db_item = SysDictBase(
                    category=category, code=code, name=name,
                    sort_order=sort_order, is_enabled=is_enabled
                )
                db.add(db_item)
                created += 1
        except Exception as e:
            errors.append(f"第{row_idx}行：处理失败 - {str(e)}")

    db.commit()
    wb.close()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "import", f"导入字典数据：新增{created}条，更新{updated}条")

    return {
        "message": f"导入完成：新增 {created} 条，更新 {updated} 条",
        "created": created,
        "updated": updated,
        "errors": errors
    }


@router.get("/dict/export/{category}")
def export_dict(
    category: str,
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(require_permission("system:dict"))
):
    items = db.query(SysDictBase).filter(
        SysDictBase.category == category
    ).order_by(SysDictBase.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"数据字典_{category}"
    ws.append(["编码", "名称", "排序", "启用"])

    for item in items:
        ws.append([
            item.code,
            item.name,
            item.sort_order,
            "是" if item.is_enabled else "否"
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dict_{category}.xlsx"}
    )


@router.get("/permissions")
def get_permission_list(current_user: UserInfo = Depends(get_current_user)):
    """获取所有可用权限清单，用于角色权限配置界面"""
    if not (current_user.is_admin or current_user.has_permission("system:role")):
        raise HTTPException(status_code=403, detail="您没有权限查看权限清单")
    return PERMISSION_MODULES


@router.get("/roles", response_model=List[RoleOut])
def get_roles(db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not (current_user.is_admin or current_user.has_permission("system:role") or current_user.has_permission("system:user")):
        raise HTTPException(status_code=403, detail="您没有权限查看角色列表")
    return db.query(SysRole).all()


@router.post("/roles", response_model=RoleOut)
def create_role(role: RoleCreate, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:role"))):
    existing = db.query(SysRole).filter(SysRole.name == role.name).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"角色 [{role.name}] 已存在")
    db_role = SysRole(**role.model_dump())
    db.add(db_role)
    db.commit()
    db.refresh(db_role)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "create", f"新增角色 {role.name}")
    return db_role


@router.put("/roles/{role_id}", response_model=RoleOut)
def update_role(role_id: int, role: RoleUpdate, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:role"))):
    db_role = db.query(SysRole).filter(SysRole.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="角色不存在")
    update_data = role.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_role, key, value)
    db.commit()
    db.refresh(db_role)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "edit", f"编辑角色 (id={role_id})")
    return db_role


@router.delete("/roles/{role_id}")
def delete_role(role_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:role"))):
    db_role = db.query(SysRole).filter(SysRole.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="角色不存在")
    user_count = db.query(SysUserRole).filter(SysUserRole.role_id == role_id).count()
    if user_count > 0:
        raise HTTPException(status_code=400, detail=f"该角色下还有 {user_count} 个用户，请先解除用户关联后再删除")
    db.query(SysUserRole).filter(SysUserRole.role_id == role_id).delete()
    db.query(SysPermission).filter(SysPermission.role_id == role_id).delete()
    db.delete(db_role)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "delete", f"删除角色 {db_role.name}")
    return {"message": "删除成功"}


@router.post("/roles/batch-delete")
def batch_delete_roles(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:role"))):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的角色ID列表")
    roles = db.query(SysRole).filter(SysRole.id.in_(ids)).all()
    if not roles:
        raise HTTPException(status_code=404, detail="未找到指定的角色")
    used_roles = []
    for role in roles:
        user_count = db.query(SysUserRole).filter(SysUserRole.role_id == role.id).count()
        if user_count > 0:
            used_roles.append(f"{role.name}({user_count}人)")
    if used_roles:
        raise HTTPException(status_code=400, detail=f"以下角色下还有用户，请先解除关联：{', '.join(used_roles)}")
    for role in roles:
        db.query(SysUserRole).filter(SysUserRole.role_id == role.id).delete()
        db.query(SysPermission).filter(SysPermission.role_id == role.id).delete()
        db.delete(role)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "delete", f"批量删除 {len(roles)} 个角色")
    return {"message": f"成功删除 {len(roles)} 个角色", "deleted_count": len(roles)}


@router.get("/roles/{role_id}/permissions")
def get_role_permissions(role_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(get_current_user)):
    if not (current_user.is_admin or current_user.has_permission("system:role")):
        raise HTTPException(status_code=403, detail="您没有权限查看角色权限")
    perms = db.query(SysPermission).filter(SysPermission.role_id == role_id).all()
    return [{"module": p.module, "action": p.action} for p in perms]


@router.post("/roles/{role_id}/permissions")
def assign_permissions(role_id: int, data: PermissionAssign, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:role"))):
    db_role = db.query(SysRole).filter(SysRole.id == role_id).first()
    if not db_role:
        raise HTTPException(status_code=404, detail="角色不存在")
    db.query(SysPermission).filter(SysPermission.role_id == role_id).delete()
    for perm in data.permissions:
        db.add(SysPermission(role_id=role_id, module=perm["module"], action=perm["action"]))
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "edit", f"配置角色权限: {db_role.name}")
    return {"message": "权限配置成功"}


def _clear_user_references(db: Session, user_id: int):
    """清理用户在其他表中的外键引用（置为NULL，保留历史记录）"""
    db.query(SysUserRole).filter(SysUserRole.user_id == user_id).delete()
    db.query(SysLog).filter(SysLog.user_id == user_id).update({"user_id": None})
    db.query(EmployeeSalary).filter(EmployeeSalary.operator_id == user_id).update({"operator_id": None})
    db.query(PerformanceScore).filter(PerformanceScore.reviewer_id == user_id).update({"reviewer_id": None})
    db.query(CalculationLog).filter(CalculationLog.operator_id == user_id).update({"operator_id": None})
    db.query(ApprovalRecord).filter(ApprovalRecord.submitter_id == user_id).update({"submitter_id": None})
    db.query(ApprovalRecord).filter(ApprovalRecord.approver_id == user_id).update({"approver_id": None})
    db.query(ExportTemplate).filter(ExportTemplate.created_by == user_id).update({"created_by": None})
    db.query(SalaryPeriodStep).filter(SalaryPeriodStep.confirmed_by == user_id).update({"confirmed_by": None})


def _get_user_with_roles(db: Session, user: SysUser) -> dict:
    """获取用户信息及关联角色"""
    user_roles = db.query(SysRole).join(
        SysUserRole, SysUserRole.role_id == SysRole.id
    ).filter(SysUserRole.user_id == user.id).all()
    return {
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "is_admin": user.is_admin,
        "is_active": user.is_active,
        "role_ids": [r.id for r in user_roles],
        "role_names": [r.name for r in user_roles],
        "created_at": user.created_at.isoformat() if user.created_at else None,
    }


@router.get("/users", response_model=List[UserOut])
def get_users(db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:user"))):
    users = db.query(SysUser).all()
    return [_get_user_with_roles(db, u) for u in users]


@router.post("/users", response_model=UserOut)
def create_user(user: UserCreate, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:user"))):
    existing = db.query(SysUser).filter(SysUser.username == user.username).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"用户名 [{user.username}] 已存在")
    db_user = SysUser(
        username=user.username,
        password_hash=get_password_hash(user.password),
        display_name=user.display_name,
        is_admin=user.is_admin,
        is_active=user.is_active,
    )
    db.add(db_user)
    db.flush()
    for role_id in user.role_ids:
        db.add(SysUserRole(user_id=db_user.id, role_id=role_id))
    db.commit()
    db.refresh(db_user)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "create", f"新增用户 {user.username}")
    return _get_user_with_roles(db, db_user)


@router.put("/users/{user_id}", response_model=UserOut)
def update_user(user_id: int, user: UserUpdate, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:user"))):
    db_user = db.query(SysUser).filter(SysUser.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if db_user.username == "admin" and user.is_active is False:
        raise HTTPException(status_code=400, detail="超级管理员账号不可禁用")
    update_data = user.model_dump(exclude_unset=True)
    role_ids = update_data.pop("role_ids", None)
    for key, value in update_data.items():
        setattr(db_user, key, value)
    if role_ids is not None:
        db.query(SysUserRole).filter(SysUserRole.user_id == user_id).delete()
        for role_id in role_ids:
            db.add(SysUserRole(user_id=user_id, role_id=role_id))
    db.commit()
    db.refresh(db_user)
    write_log(db, "data_change", current_user.id, current_user.username, "system", "edit", f"编辑用户 (id={user_id})")
    return _get_user_with_roles(db, db_user)


@router.post("/users/{user_id}/reset-password")
def reset_user_password(user_id: int, data: PasswordReset, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:user"))):
    db_user = db.query(SysUser).filter(SysUser.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="新密码长度不能少于6位")
    db_user.password_hash = get_password_hash(data.new_password)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "edit", f"重置用户密码: {db_user.username}")
    return {"message": f"密码重置成功，新密码为：{data.new_password}"}


@router.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:user"))):
    db_user = db.query(SysUser).filter(SysUser.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if db_user.username == "admin":
        raise HTTPException(status_code=400, detail="超级管理员账号不可删除")
    if db_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="不能删除当前登录用户")
    _clear_user_references(db, user_id)
    db.delete(db_user)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "delete", f"删除用户 {db_user.username}")
    return {"message": "删除成功"}


@router.post("/users/batch-delete")
def batch_delete_users(ids: List[int], db: Session = Depends(get_db), current_user: UserInfo = Depends(require_permission("system:user"))):
    if not ids:
        raise HTTPException(status_code=400, detail="请提供要删除的用户ID列表")
    if current_user.id in ids:
        raise HTTPException(status_code=400, detail="不能删除当前登录用户")
    users = db.query(SysUser).filter(SysUser.id.in_(ids)).all()
    if not users:
        raise HTTPException(status_code=404, detail="未找到指定的用户")
    admin_names = [u.username for u in users if u.username == "admin"]
    if admin_names:
        raise HTTPException(status_code=400, detail=f"超级管理员账号不可删除")
    for user in users:
        _clear_user_references(db, user.id)
        db.delete(user)
    db.commit()
    write_log(db, "data_change", current_user.id, current_user.username, "system", "delete", f"批量删除 {len(users)} 个用户")
    return {"message": f"成功删除 {len(users)} 个用户", "deleted_count": len(users)}


@router.get("/logs", response_model=List[LogOut])
def get_logs(
    log_type: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(require_permission("system:log"))
):
    query = db.query(SysLog)
    if log_type:
        query = query.filter(SysLog.log_type == log_type)
    return query.order_by(SysLog.created_at.desc()).limit(limit).all()


# ========== 数据备份 ==========

def _parse_db_url() -> dict:
    """解析 DATABASE_URL，返回数据库类型和连接信息"""
    url = DATABASE_URL
    if url.startswith("sqlite:///"):
        rel_path = url.replace("sqlite:///", "")
        db_path = Path(__file__).parent.parent.parent / rel_path
        return {"type": "sqlite", "path": db_path.resolve()}
    if url.startswith("postgresql://") or url.startswith("postgres://"):
        parsed = urlparse(url)
        return {
            "type": "postgresql",
            "host": parsed.hostname or "localhost",
            "port": str(parsed.port or 5432),
            "user": parsed.username or "",
            "password": parsed.password or "",
            "database": parsed.path.lstrip("/") or "",
        }
    return {"type": "unknown"}


def _get_backup_dir() -> Path:
    """获取备份目录路径"""
    return Path(__file__).parent.parent.parent.parent / "backups"


@router.get("/backups")
def list_backups(current_user: UserInfo = Depends(require_permission("system:backup"))):
    """列出所有备份文件"""
    backup_dir = _get_backup_dir()
    if not backup_dir.exists():
        return {"backups": []}
    files = []
    for pattern in ["e9_salary_*.db", "e9_salary_*.sql"]:
        for f in sorted(backup_dir.glob(pattern), reverse=True):
            stat = f.stat()
            files.append({
                "filename": f.name,
                "size": stat.st_size,
                "size_display": f"{stat.st_size / 1024:.1f} KB",
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            })
    files.sort(key=lambda x: x["created_at"], reverse=True)
    return {"backups": files}


@router.post("/backup")
def create_backup(
    db: Session = Depends(get_db),
    current_user: UserInfo = Depends(require_permission("system:backup")),
):
    """手动创建数据库备份"""
    db_info = _parse_db_url()
    db_type = db_info["type"]

    backup_dir = _get_backup_dir()
    backup_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if db_type == "sqlite":
        db_path = db_info["path"]
        if not db_path.exists():
            raise HTTPException(status_code=404, detail=f"数据库文件不存在: {db_path}")
        backup_file = backup_dir / f"e9_salary_{timestamp}.db"
        shutil.copy2(db_path, backup_file)

    elif db_type == "postgresql":
        backup_file = backup_dir / f"e9_salary_{timestamp}.sql"
        env = {"PGPASSWORD": db_info["password"]}
        cmd = [
            "pg_dump",
            "-h", db_info["host"],
            "-p", db_info["port"],
            "-U", db_info["user"],
            "-d", db_info["database"],
            "-f", str(backup_file),
            "--no-owner",
            "--no-acl",
        ]
        try:
            result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=300)
            if result.returncode != 0:
                raise HTTPException(
                    status_code=500,
                    detail=f"数据库备份失败: {result.stderr.strip()}"
                )
        except FileNotFoundError:
            raise HTTPException(
                status_code=500,
                detail="pg_dump 命令不可用，请确认后端容器已安装 PostgreSQL 客户端工具"
            )
        except subprocess.TimeoutExpired:
            raise HTTPException(status_code=500, detail="数据库备份超时，请稍后重试")

    else:
        raise HTTPException(status_code=400, detail="当前数据库类型不支持自动备份，请手动导出")

    write_log(db, "data_change", current_user.id, current_user.username, "system", "backup",
              f"数据库备份完成: {backup_file.name}")

    return {
        "success": True,
        "message": "备份成功",
        "filename": backup_file.name,
        "size": backup_file.stat().st_size,
        "size_display": f"{backup_file.stat().st_size / 1024:.1f} KB",
    }