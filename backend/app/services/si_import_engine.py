"""
社保公积金智能导入引擎
支持：Excel(xlsx/xls) + PDF 多文件批量上传、多模板自动匹配、字段映射、
      员工姓名匹配、多层级数据校验、批量模板配置向导
"""
import re
import os
import uuid
import io
import json
import time
import shutil
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from decimal import Decimal, ROUND_HALF_EVEN

from sqlalchemy.orm import Session

from app.models.models import (
    SocialInsurance, Employee, SiImportTemplate, SiImportLog
)

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent.parent.parent
TEMP_DIR = PROJECT_ROOT / "temp" / "si_import"
TEMP_DIR.mkdir(parents=True, exist_ok=True)
BATCH_EXPIRE_SECONDS = 3600

# ── 所有可映射字段及其中文标签 ──────────────────────────────────────────
SI_FIELD_LABELS = {
    "employee_name": "员工姓名",
    "period": "缴费月份",
    "pension_personal_base": "养老保险个人基数",
    "pension_company_base": "养老保险单位基数",
    "unemployment_personal_base": "失业保险个人基数",
    "unemployment_company_base": "失业保险单位基数",
    "medical_personal_base": "医疗保险个人基数",
    "medical_company_base": "医疗保险单位基数",
    "injury_company_base": "工伤保险单位基数",
    "pension_personal": "养老保险个人金额",
    "pension_personal_rate": "养老保险个人比例",
    "pension_company": "养老保险单位金额",
    "pension_company_rate": "养老保险单位比例",
    "unemployment_personal": "失业保险个人金额",
    "unemployment_personal_rate": "失业保险个人比例",
    "unemployment_company": "失业保险单位金额",
    "unemployment_company_rate": "失业保险单位比例",
    "medical_personal": "医疗保险个人金额",
    "medical_personal_rate": "医疗保险个人比例",
    "medical_company": "医疗保险单位金额",
    "medical_company_rate": "医疗保险单位比例",
    "injury_company": "工伤保险单位金额",
    "injury_company_rate": "工伤保险单位比例",
    "si_personal": "社保个人合计",
    "si_company": "社保单位合计",
    "hf_base": "公积金缴存基数",
    "hf_personal": "公积金个人金额",
    "hf_personal_rate": "公积金个人比例",
    "hf_company": "公积金单位金额",
    "hf_company_rate": "公积金单位比例",
    "pension_total": "养老保险合计",
    "unemployment_total": "失业保险合计",
    "medical_total": "医疗保险合计",
    "injury_total": "工伤保险合计",
    "si_grand_total": "社保总合计",
    "hf_total": "公积金合计",
    "grand_total": "社保公积金总合计",
}

# ── 内部通用占位字段（用于文件名险种推断，不对用户显示）──────────────
# 当模板配置了这些通用字段名时，系统会从文件名推断具体险种，
# 自动将它们映射到对应的具体字段（如 pension_personal_base）
#
# 【完全通用字段】依赖文件名同时推断险种 + 个人/单位方向
#   amount_base → {ins_type}_{payer}_base  (如 pension_personal_base)
#   rate        → {ins_type}_{payer}_rate
#   amount      → {ins_type}_{payer}
#
# 【半通用字段】用户显式指定个人/单位，仅从文件名推断险种
#   personal_base / personal_amount / personal_rate
#   company_base  / company_amount  / company_rate
#
# 【公积金通用字段】
#   hf_amount_base → hf_base（公积金基数不分个人/单位）
GENERIC_INSURANCE_FIELDS = {
    "amount_base", "rate", "amount",
    "personal_base", "personal_amount", "personal_rate",
    "company_base", "company_amount", "company_rate",
}

# 所有系统可识别的字段（含内部通用字段）
ALL_SYSTEM_FIELDS = set(SI_FIELD_LABELS.keys()) | GENERIC_INSURANCE_FIELDS

# 金额类字段（需要做数值解析和校验）
AMOUNT_FIELDS = [
    "pension_personal_base", "pension_company_base",
    "unemployment_personal_base", "unemployment_company_base",
    "medical_personal_base", "medical_company_base",
    "injury_company_base",
    "pension_personal", "pension_personal_rate",
    "pension_company", "pension_company_rate",
    "unemployment_personal", "unemployment_personal_rate",
    "unemployment_company", "unemployment_company_rate",
    "medical_personal", "medical_personal_rate",
    "medical_company", "medical_company_rate",
    "injury_company", "injury_company_rate",
    "si_personal", "si_company",
    "hf_base", "hf_personal", "hf_personal_rate",
    "hf_company", "hf_company_rate",
    "pension_total", "unemployment_total", "medical_total", "injury_total",
    "si_grand_total", "hf_total", "grand_total",
    # 内部通用占位字段（数值类型）
    "amount_base", "rate", "amount",
    "personal_base", "personal_amount", "personal_rate",
    "company_base", "company_amount", "company_rate",
]

# 必须字段（导入时至少要有姓名）
REQUIRED_FIELDS = ["employee_name"]


class ImportResult:
    """导入结果"""
    def __init__(self):
        self.total_files = 0
        self.parsed_files = 0
        self.failed_files = 0
        self.total_rows = 0
        self.created = 0
        self.updated = 0
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []


def _parse_number(value, number_format: Optional[Dict] = None) -> Optional[Decimal]:
    """将单元格值解析为 Decimal"""
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        if isinstance(value, float):
            return Decimal(str(value))
        return Decimal(value)
    s = str(value).strip()
    if not s or s in ("-", "--", "——", "/"):
        return None
    # 根据模板配置去除不需要的字符，默认同时移除中英文逗号（千分位分隔符）和货币符号
    default_remove = [",", "，", "¥", "￥", "$"]
    custom_remove = (number_format or {}).get("remove_chars", [])
    remove_chars = list(set(default_remove + custom_remove))
    for ch in remove_chars:
        s = s.replace(ch, "")
    # 去除中文单位
    s = re.sub(r"[元圆]", "", s)
    s = s.strip()
    # 处理特殊格式：2%+3 → 只取百分比部分（2%），+3元作为固定附加额单独处理
    # 这种格式常见于北京医疗保险（个人2%+3元大病统筹）
    rate_plus_match = re.match(r"^([\d.]+)%\s*\+\s*[\d.]+", s)
    if rate_plus_match:
        try:
            return (Decimal(rate_plus_match.group(1)) / Decimal("100")).quantize(Decimal("0.0001"))
        except Exception:
            pass
    # 处理百分比：10% → 0.10，费率精度保留4位小数（如 0.5000% → 0.0050）
    if s.endswith("%"):
        try:
            return (Decimal(s[:-1]) / Decimal("100")).quantize(Decimal("0.0001"))
        except Exception:
            return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _get_excel_sheet_names(file_bytes: bytes, filename: str) -> List[str]:
    """获取 Excel 文件中所有工作表名称"""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    sheet_names = []
    if ext == "xls":
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            sheet_names = wb.sheet_names()
        except Exception:
            pass
    if not sheet_names:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            sheet_names = wb.sheetnames
        except Exception:
            pass
    return sheet_names


def _parse_excel(file_bytes: bytes, filename: str, sheet_name: Optional[str] = None, sheet_index: Optional[int] = None) -> List[List]:
    """解析 Excel 文件，返回二维列表（每行是一个列表），支持合并单元格填充
    
    Args:
        sheet_name: 按名称选择工作表（优先于sheet_index）
        sheet_index: 按索引选择工作表（默认0，即第一个工作表）
    """
    if filename.lower().endswith(".xls"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
            if sheet_name and sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
            elif sheet_index is not None and sheet_index < len(wb.sheets()):
                ws = wb.sheet_by_index(sheet_index)
            else:
                ws = wb.sheet_by_index(0)
            
            # 构建合并单元格映射：(row, col) -> value
            merged_values = {}
            for (rlo, rhi, clo, chi) in ws.merged_cells:
                parent_value = ws.cell(rlo, clo).value
                for r in range(rlo, rhi):
                    for c in range(clo, chi):
                        merged_values[(r, c)] = parent_value
            
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    if (r, c) in merged_values:
                        row.append(merged_values[(r, c)])
                    else:
                        cell = ws.cell(r, c)
                        row.append(cell.value)
                rows.append(row)
            return rows
        except Exception:
            pass
    # 默认用 openpyxl（支持 .xlsx 和伪装的 .xls）
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif sheet_index is not None and sheet_index < len(wb.worksheets):
        ws = wb.worksheets[sheet_index]
    else:
        ws = wb.active
    
    # 构建合并单元格映射
    merged_values = {}
    for merged_range in ws.merged_cells.ranges:
        min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1
        parent_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row - 1, col - 1)] = parent_value
    
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_data = []
        for col_idx, cell_value in enumerate(row):
            if (row_idx, col_idx) in merged_values:
                row_data.append(merged_values[(row_idx, col_idx)])
            else:
                row_data.append(cell_value)
        rows.append(row_data)
    return rows


def _parse_pdf(file_bytes: bytes) -> List[List]:
    """解析 PDF 文件，提取表格数据"""
    import pdfplumber
    all_rows = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if row and any(cell for cell in row):
                        all_rows.append([cell.strip() if cell else "" for cell in row])
    return all_rows


def _sheet_has_data(file_rows: List[List], min_data_rows: int = 2) -> bool:
    """
    判断工作表是否包含有效数据行（不是空表、说明页、目录页）。
    判定标准：
    1. 至少有 min_data_rows 行数据
    2. 数据区域有一定数量的非空单元格（排除全空行、标题行）
    3. 包含姓名相关关键词（"姓名"）或看起来像数据表格
    """
    if not file_rows or len(file_rows) < min_data_rows:
        return False

    non_empty_rows = 0
    has_name_column = False
    data_cells = 0

    for i, row in enumerate(file_rows):
        if not row:
            continue
        non_empty_cells = sum(1 for c in row if c is not None and str(c).strip() != "")
        if non_empty_cells >= 2:
            non_empty_rows += 1
            data_cells += non_empty_cells
            for c in row:
                if c and "姓名" in str(c):
                    has_name_column = True
                    break

    if has_name_column and non_empty_rows >= min_data_rows:
        return True

    if non_empty_rows >= 5 and data_cells >= 15:
        return True

    return False


def _extract_keywords_from_path(path: str) -> List[str]:
    """从文件路径中提取所有有意义的关键词（文件夹名、文件名）"""
    if not path:
        return []
    parts = re.split(r'[\\/]+', path)
    keywords = []
    for part in parts:
        if not part:
            continue
        part = os.path.splitext(part)[0]
        tokens = re.split(r'[_\-\s（）()\[\]【】\d]+', part)
        for t in tokens:
            t = t.strip()
            if len(t) >= 2:
                keywords.append(t.lower())
    return keywords


def _match_template(
    db: Session,
    file_rows: List[List],
    filename: str,
    sheet_name: Optional[str] = None,
    validate_columns: bool = True,
    context_keywords: Optional[List[str]] = None,
) -> Optional[SiImportTemplate]:
    """为文件匹配最合适的导入模板
    
    Args:
        sheet_name: 当前工作表名称，用于匹配模板的sheet_pattern
        validate_columns: 是否验证模板能建立有效的字段映射（默认True）
        context_keywords: 上下文关键词列表（从文件夹路径等提取，用于辅助匹配）
    """
    templates = (
        db.query(SiImportTemplate)
        .filter(SiImportTemplate.is_active == True)
        .order_by(SiImportTemplate.sort_order)
        .all()
    )
    if not templates:
        return None

    context_keywords = context_keywords or []
    context_keywords_lower = [kw.lower() for kw in context_keywords if kw]
    context_text = " ".join(context_keywords_lower)

    best_template = None
    best_score = 0

    for tpl in templates:
        score = 0
        has_filename_match = False
        has_header_match = False
        header_match_ratio = 0.0

        # 0. 工作表匹配（如果模板配置了sheet_pattern）
        if tpl.sheet_pattern and sheet_name:
            if tpl.sheet_pattern in sheet_name or sheet_name in tpl.sheet_pattern:
                score += 80  # 工作表名称匹配权重很高
        elif tpl.sheet_pattern and not sheet_name:
            score -= 30

        # 1. 文件名正则匹配
        if tpl.file_pattern:
            try:
                if re.search(tpl.file_pattern, filename, re.IGNORECASE):
                    score += 50
                    has_filename_match = True
            except re.error:
                pass

        # 1b. 文件名关键词匹配
        if tpl.file_keywords and isinstance(tpl.file_keywords, list) and len(tpl.file_keywords) > 0:
            kw_lower = [kw.lower() for kw in tpl.file_keywords if kw]
            fname_lower = filename.lower()
            if all(kw in fname_lower for kw in kw_lower):
                score += 50
                has_filename_match = True

        # 1c. 【新增】上下文关键词匹配（文件夹路径）
        # 从文件夹名中提取的关键词，如果匹配模板的城市、source_category或name中的关键词，加分
        context_match_score = 0
        tpl_name_lower = (tpl.name or "").lower()
        tpl_city_lower = (tpl.city or "").lower()
        tpl_category = (tpl.source_category or "").lower()
        
        category_keywords = {
            "social_insurance": ["社保", "社会保险", "五险"],
            "housing_fund": ["公积金", "住房公积金", "汇缴"],
        }
        
        for kw in context_keywords_lower:
            if not kw:
                continue
            if tpl_city_lower and kw in tpl_city_lower:
                context_match_score += 30
            if kw in tpl_name_lower:
                context_match_score += 25
            if tpl_category:
                cat_kws = category_keywords.get(tpl_category, [])
                for cat_kw in cat_kws:
                    if cat_kw in kw or kw in cat_kw:
                        context_match_score += 40
                        break
            if tpl.file_keywords and isinstance(tpl.file_keywords, list):
                for fkw in tpl.file_keywords:
                    if fkw and fkw.lower() in kw:
                        context_match_score += 20
                        break
        
        if context_match_score > 0:
            score += min(context_match_score, 60)

        # 2. 文件类型匹配
        ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if tpl.file_type == "excel" and ext in ("xlsx", "xls"):
            score += 10
        elif tpl.file_type == "pdf" and ext == "pdf":
            score += 10

        # 3. 表头列名匹配
        if tpl.header_rows and tpl.column_mappings:
            header_indices = tpl.header_rows if isinstance(tpl.header_rows, list) else [tpl.header_rows]
            for hr in header_indices:
                if hr < len(file_rows):
                    header_row = file_rows[hr]
                    header_texts = [str(c).strip() if c else "" for c in header_row]
                    mapping_keys = [k for k in tpl.column_mappings.keys() if k]
                    matched = sum(1 for k in mapping_keys if k in header_texts)
                    if mapping_keys:
                        ratio = matched / len(mapping_keys)
                        header_match_ratio = max(header_match_ratio, ratio)
                        score += int(ratio * 40)
                        if ratio >= 0.3:
                            has_header_match = True

        # 匹配有效性判断（上下文关键词匹配也算有效匹配）
        has_context_match = context_match_score >= 30
        is_valid = False
        if score >= 80 and (tpl.sheet_pattern and sheet_name):
            is_valid = True
        elif has_filename_match and score >= 50:
            is_valid = True
        elif has_context_match and score >= 40:
            is_valid = True
        elif header_match_ratio >= 0.5 and score >= 30:
            is_valid = True
        elif score >= 50:
            is_valid = True

        if validate_columns and is_valid:
            try:
                col_idx = _build_column_index(file_rows, tpl)
                if "employee_name" not in col_idx:
                    is_valid = False
            except Exception:
                is_valid = False

        if is_valid and score > best_score:
            best_score = score
            best_template = tpl

    return best_template


def _infer_insurance_from_filename(filename: str) -> Dict[str, Optional[str]]:
    """
    从文件名中推断险种类型和缴费方向。
    用于处理历史兼容格式：文件名含险种关键词（如"养老保险(单位缴纳)"），
    但文件表头中只有通用列名（如"缴费基数"、"费率"、"应缴费额"），不包含具体险种信息。

    返回 {"insurance_type": "pension", "payer": "company"} 或全部为 None。
    """
    result = {"insurance_type": None, "payer": None}

    # 险种关键词 → 系统字段前缀（按长度降序排列，避免短词误匹配）
    insurance_map = [
        ("养老保险", "pension"), ("失业保险", "unemployment"),
        ("医疗保险", "medical"), ("工伤保险", "injury"),
        ("生育保险", "maternity"), ("公积金", "hf"),
        ("养老", "pension"), ("失业", "unemployment"),
        ("医疗", "medical"), ("工伤", "injury"),
        ("生育", "maternity"), ("住房", "hf"),
    ]

    # 缴费方向关键词（按长度降序排列，避免"企业"误匹配"城镇企业职工"中的"企业"）
    payer_map = [
        ("个人缴纳", "personal"), ("单位缴纳", "company"),
        ("个人缴存", "personal"), ("单位缴存", "company"),
        ("个人部分", "personal"), ("单位部分", "company"),
        ("个人", "personal"), ("单位", "company"), ("企业", "company"),
    ]

    for kw, prefix in insurance_map:
        if kw in filename:
            result["insurance_type"] = prefix
            break

    for kw, payer in payer_map:
        if kw in filename:
            result["payer"] = payer
            break

    return result


def _build_column_index(
    file_rows: List[List],
    template: SiImportTemplate,
) -> Dict[str, int]:
    """根据模板的表头配置，建立「数据库字段名 → 文件列序号」的映射
    
    匹配策略（按得分从高到低）：
    1. 完全相等 → 100分
    2. 紧凑格式（去除空格、"-"、全角空格）完全相等 → 90分
    3. header_name 是 header_text 紧凑格式的后缀（如"缴费基数"匹配"XX - 缴费基数"）→ 80分
    4. 所有词组都在header_text中，且长度占比≥80% → 70分
    5. header_name紧凑格式是header_text紧凑格式的子串，且长度占比≥70% → 60分
    （删除了前两个字符匹配的危险兜底策略）
    """
    import re as _re
    
    column_index = {}
    used_columns = set()  # 记录已被映射的列，避免重复映射
    header_indices = template.header_rows if isinstance(template.header_rows, list) else [template.header_rows]

    # 合并多层表头为一行，用 " - " 分隔不同层级
    merged_header = []
    merged_header_compact = []  # 预计算紧凑格式，提高性能
    if header_indices:
        max_cols = max(len(file_rows[hr]) if hr < len(file_rows) else 0 for hr in header_indices)
        for col_idx in range(max_cols):
            parts = []
            for hr in header_indices:
                if hr < len(file_rows) and col_idx < len(file_rows[hr]):
                    val = str(file_rows[hr][col_idx]).strip() if file_rows[hr][col_idx] else ""
                    if val:
                        # 避免重复：如果和上一个 part 相同，则跳过
                        if not parts or val != parts[-1]:
                            parts.append(val)
            # 多层用 " - " 连接，单层直接用
            hdr_text = " - ".join(parts) if len(parts) > 1 else (parts[0] if parts else "")
            merged_header.append(hdr_text)
            merged_header_compact.append(hdr_text.replace(" - ", "").replace(" ", "").replace("\u3000", ""))

    def _calc_match_score(header_name: str, header_text: str, text_compact: str) -> int:
        """计算header_name与某个header_text的匹配得分，0表示不匹配"""
        if not header_name or not header_text:
            return 0
        
        name_compact = header_name.replace(" - ", "").replace(" ", "").replace("\u3000", "")
        name_len = len(name_compact)
        text_len = len(text_compact)
        
        # 策略1：完全相等
        if header_name == header_text:
            return 100
        
        # 策略2：紧凑格式完全相等
        if name_compact == text_compact:
            return 90
        
        # 策略3：header_name是header_text的后缀（要求header_name长度至少4个字符，避免短词误匹配）
        if name_len >= 4 and text_compact.endswith(name_compact):
            return 80
        
        # 策略4：所有词组都在header_text中（拆分"-"和空格）
        tokens = [t for t in _re.split(r"\s*[-]\s*|\s+|（|）|\(|\)", header_name) if t]
        tokens = [t for t in tokens if len(t) >= 2]  # 过滤掉单字token如"的"
        if tokens and all(t in header_text for t in tokens):
            # 要求长度占比至少80%，避免"单位"匹配到"工伤保险(单位缴纳)"这种情况
            if name_len >= text_len * 0.8:
                return 70
        
        # 策略5：header_name紧凑格式是header_text紧凑格式的子串
        if name_len >= 4 and name_compact in text_compact:
            # 要求长度占比至少70%
            if name_len >= text_len * 0.7:
                return 60
        
        return 0

    # 按 column_mappings 匹配，计算每个字段的最佳匹配列
    mappings = template.column_mappings or {}
    
    # 先收集所有(字段, 列, 得分)的候选
    candidates = []
    for header_name, db_field in mappings.items():
        if not db_field:  # 跳过未映射的列
            continue
        
        best_score = 0
        best_idx = None
        for idx, header_text in enumerate(merged_header):
            score = _calc_match_score(header_name, header_text, merged_header_compact[idx] if idx < len(merged_header_compact) else "")
            if score > best_score:
                best_score = score
                best_idx = idx
        
        if best_idx is not None and best_score >= 60:
            candidates.append((best_score, db_field, best_idx, header_name))
    
    # 按得分降序排列，得分高的优先选择列，避免被得分低的抢占
    candidates.sort(reverse=True, key=lambda x: x[0])
    
    for score, db_field, col_idx, header_name in candidates:
        if col_idx not in used_columns:
            column_index[db_field] = col_idx
            used_columns.add(col_idx)

    return column_index


def _extract_data(
    file_rows: List[List],
    template: SiImportTemplate,
    filename: str,
    period: str,
    logs: List[Dict],
    batch_id: str,
) -> List[Dict]:
    """根据模板从文件行中提取结构化数据"""
    column_index = _build_column_index(file_rows, template)

    # ── 文件名险种推断：修正无效字段映射 ──
    # 场景：历史邯郸社保文件，表头只有通用列名（缴费基数/费率/应缴费额），
    #       但文件名包含险种关键词（如"养老保险(单位缴纳)"），
    #       需要通过文件名推断将通用字段映射到具体险种字段。
    # 注意：GENERIC_INSURANCE_FIELDS（amount_base/rate/amount）是内部占位字段，
    #       不在 SI_FIELD_LABELS 中，但需要被识别并通过文件名推断转换。
    needs_inference = False
    for f in column_index:
        if f in GENERIC_INSURANCE_FIELDS:
            needs_inference = True
            break
        if f not in SI_FIELD_LABELS and not f.startswith("_"):
            needs_inference = True
            break

    if needs_inference:
        insurance_info = _infer_insurance_from_filename(filename)
        ins_type = insurance_info.get("insurance_type")
        payer = insurance_info.get("payer")
        if ins_type:
            # 定义通用占位字段 → 具体系统字段的修正规则
            # 社保险种：pension/unemployment/medical/injury
            # 公积金：hf（特殊处理，基数不分个人/单位）
            generic_to_specific = {}

            # 完全通用字段（依赖文件名的个人/单位关键词）
            if payer:
                if ins_type == "hf":
                    generic_to_specific.update({
                        "amount_base": "hf_base",
                        "rate": f"hf_{payer}_rate",
                        "amount": f"hf_{payer}",
                    })
                else:
                    generic_to_specific.update({
                        "amount_base": f"{ins_type}_{payer}_base",
                        "rate": f"{ins_type}_{payer}_rate",
                        "amount": f"{ins_type}_{payer}",
                    })

            # 半通用字段（用户显式指定个人/单位，只从文件名推断险种）
            if ins_type == "hf":
                # 公积金特殊处理
                generic_to_specific.update({
                    "personal_base": "hf_base",
                    "personal_amount": "hf_personal",
                    "personal_rate": "hf_personal_rate",
                    "company_base": "hf_base",
                    "company_amount": "hf_company",
                    "company_rate": "hf_company_rate",
                })
            else:
                generic_to_specific.update({
                    "personal_base": f"{ins_type}_personal_base",
                    "personal_amount": f"{ins_type}_personal",
                    "personal_rate": f"{ins_type}_personal_rate",
                    "company_base": f"{ins_type}_company_base",
                    "company_amount": f"{ins_type}_company",
                    "company_rate": f"{ins_type}_company_rate",
                })
            # 重建 column_index，替换无效字段为正确字段
            fixed_index = {}
            remap_count = 0
            for db_field, col_idx in column_index.items():
                if db_field in SI_FIELD_LABELS:
                    fixed_index[db_field] = col_idx
                elif db_field.startswith("_"):
                    # 跳过 _doc_type 等以下划线开头的占位字段
                    continue
                elif db_field in GENERIC_INSURANCE_FIELDS:
                    # 这是通用占位字段，需要根据文件名推断转换
                    new_field = generic_to_specific.get(db_field)
                    if new_field and new_field in SI_FIELD_LABELS:
                        fixed_index[new_field] = col_idx
                        remap_count += 1
                else:
                    new_field = generic_to_specific.get(db_field)
                    if new_field and new_field in SI_FIELD_LABELS:
                        fixed_index[new_field] = col_idx
                        remap_count += 1
                    # 无法推断的无效字段，直接跳过
            if remap_count > 0:
                logger.info(
                    f"文件名险种推断：从「{filename}」识别到险种={ins_type}、方向={payer}，"
                    f"已修正 {remap_count} 个字段映射"
                )
                column_index = fixed_index

    if "employee_name" not in column_index:
        logs.append({
            "file_name": filename, "row_number": None,
            "error_level": "error", "error_type": "template_error",
            "error_message": f"模板「{template.name}」未配置「员工姓名」字段映射，无法导入",
        })
        return []

    row_filters = template.row_filters or {}
    number_format = template.number_format or {}
    data_start = template.data_start_row
    skip_footer = template.skip_footer_rows or 0
    end_row = len(file_rows) - skip_footer if skip_footer > 0 else len(file_rows)

    results = []
    tpl_default_rates = template.default_rates or {}
    for row_idx in range(data_start, end_row):
        row = file_rows[row_idx] if row_idx < len(file_rows) else []
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        # 行过滤
        skip = False
        for filter_col, filter_val in row_filters.items():
            if filter_col not in column_index:
                continue
            ci = column_index[filter_col]
            cell_val = str(row[ci]).strip() if ci < len(row) and row[ci] else ""
            # 支持多种过滤操作符
            if isinstance(filter_val, dict):
                if "$not_in" in filter_val and cell_val in filter_val["$not_in"]:
                    skip = True; break
                if "$not_empty" in filter_val and filter_val["$not_empty"] and not cell_val:
                    skip = True; break
                if "$eq" in filter_val and cell_val != filter_val["$eq"]:
                    skip = True; break
            else:
                if cell_val != str(filter_val):
                    skip = True; break
        if skip:
            continue

        record = {}
        for db_field, col_idx in column_index.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if db_field in AMOUNT_FIELDS:
                val = _parse_number(raw_val, number_format)
                # 金额/比例字段：只有解析到有效数值（非None、非0）才放入record
                # 注意：val=0是有效值（比如工伤保险个人比例为0），也要保留
                if val is not None:
                    record[db_field] = val
            elif db_field == "employee_name":
                # 姓名字段特殊处理，始终保留（即使为空，后面会检查跳过）
                val = str(raw_val).strip() if raw_val else ""
                record[db_field] = val
            elif db_field == "period":
                if raw_val:
                    val = str(raw_val).strip()
                    if val:
                        record[db_field] = val
            else:
                # 普通字符串字段（如社保号）：非空才放入
                val = str(raw_val).strip() if raw_val else ""
                if val:
                    record[db_field] = val

        # 强制使用API传入的核算月份（即用户在页面选择的月份），不从文件中解析
        # 设计原则：在哪个月份页面导入，数据就归属到哪个月份
        record["period"] = period

        # 姓名检查
        name = record.get("employee_name", "")
        if not name:
            logs.append({
                "file_name": filename, "row_number": row_idx + 1,
                "error_level": "warning", "error_type": "empty_name",
                "error_message": f"第{row_idx + 1}行员工姓名为空，已跳过",
            })
            continue

        record["_source_file"] = filename
        record["_source_row"] = row_idx + 1
        record["_default_rates"] = tpl_default_rates
        results.append(record)

    return results


def _aggregate_by_name(all_records: List[Dict]) -> Dict[str, Dict]:
    """按员工姓名聚合多文件数据"""
    aggregated = {}
    for rec in all_records:
        name = rec.get("employee_name", "")
        if name not in aggregated:
            aggregated[name] = {
                "employee_name": name,
                "period": rec.get("period", ""),
                "_sources": [],
                "_default_rates": {},
            }
        # 合并 default_rates（多个模板的配置取并集，后者覆盖前者）
        rec_rates = rec.get("_default_rates", {})
        if rec_rates:
            for ins_type, rates in rec_rates.items():
                if ins_type not in aggregated[name]["_default_rates"]:
                    aggregated[name]["_default_rates"][ins_type] = {}
                aggregated[name]["_default_rates"][ins_type].update(rates)
        # 合并字段：_extract_data已保证rec中只有有效值，
        # 所以后导入的文件如果有同名字段且有值，则覆盖先导入的值
        for key, value in rec.items():
            if key.startswith("_"):
                if key == "_source_file":
                    aggregated[name]["_sources"].append({
                        "file": rec.get("_source_file", ""),
                        "row": rec.get("_source_row", ""),
                    })
                continue
            if key == "period":
                aggregated[name][key] = value
                continue
            # 对于业务字段：只有当新值有效时才更新（空/None不会进来）
            # 如果旧值不存在或为0，用新值；如果旧值已有非零值，新值非零则覆盖
            existing = aggregated[name].get(key)
            if existing is None or existing == "" or existing == 0:
                aggregated[name][key] = value
            elif value is not None and value != 0 and value != "":
                # 两个文件都有该字段且都有非零值，后导入覆盖先导入
                aggregated[name][key] = value
    return aggregated


def _auto_fill_shared_bases(aggregated: Dict[str, Dict]):
    """
    共用基数自动填充：
    横向格式文件（如北京社保）中，同一个险种的个人基数和单位基数通常是相同的，
    但模板可能只映射了其中一个。如果只有个人基数没有单位基数（或反之），
    自动复制填充，确保后续金额计算正确。
    """
    # 各险种的(个人基数字段, 单位基数字段)对
    base_pairs = [
        ("pension_personal_base", "pension_company_base"),
        ("unemployment_personal_base", "unemployment_company_base"),
        ("medical_personal_base", "medical_company_base"),
    ]
    # 工伤保险只有单位基数，无需配对

    for name, record in aggregated.items():
        for personal_base_field, company_base_field in base_pairs:
            p_base = record.get(personal_base_field)
            c_base = record.get(company_base_field)
            
            # 如果个人基数有值但单位基数没有，自动填充单位基数
            if p_base is not None and p_base != 0 and (c_base is None or c_base == 0):
                record[company_base_field] = p_base
            # 如果单位基数有值但个人基数没有，自动填充个人基数
            elif c_base is not None and c_base != 0 and (p_base is None or p_base == 0):
                record[personal_base_field] = c_base


def _format_source_files(sources: List[Dict]) -> str:
    """格式化来源文件列表，避免文件名过长导致数据库写入失败"""
    if not sources:
        return ""
    files = [s.get("file", "") for s in sources if s.get("file")]
    if len(files) == 1:
        return files[0]
    if len(files) <= 3:
        return ", ".join(files)
    return f"{files[0]}, {files[1]} 等{len(files)}个文件"


def _match_employees(
    db: Session,
    aggregated: Dict[str, Dict],
    logs: List[Dict],
    batch_id: str,
) -> Tuple[Dict[str, Dict], List[str]]:
    """按姓名匹配员工档案，返回 (有效数据, 跳过姓名列表)"""
    employees = db.query(Employee).all()
    # 按姓名建索引，注意同名情况
    name_index: Dict[str, List[Employee]] = {}
    for emp in employees:
        name_index.setdefault(emp.name, []).append(emp)

    valid_data = {}
    skipped = []

    for name, record in aggregated.items():
        matches = name_index.get(name, [])
        source_files = _format_source_files(record.get("_sources", []))
        if not matches:
            logs.append({
                "file_name": source_files,
                "row_number": None,
                "employee_name": name,
                "error_level": "warning",
                "error_type": "name_not_found",
                "error_message": f"员工「{name}」在系统中未找到，请确认姓名是否正确或该员工是否已录入档案",
            })
            skipped.append(name)
            continue

        if len(matches) > 1:
            logs.append({
                "file_name": source_files,
                "row_number": None,
                "employee_name": name,
                "error_level": "warning",
                "error_type": "duplicate_name",
                "error_message": f"系统中有{len(matches)}名同名员工「{name}」，已自动匹配第一条记录（工号：{matches[0].employee_no}），请确认是否正确",
            })

        record["employee_id"] = matches[0].id
        record["employee_no"] = matches[0].employee_no
        valid_data[name] = record

    return valid_data, skipped


def _format_rate(rate: Decimal) -> str:
    """将小数比例格式化为百分比显示，如 0.1 → 10%，0.005 → 0.50%"""
    if rate is None:
        return "无"
    return f"{(rate * Decimal('100')).quantize(Decimal('0.00'), rounding=ROUND_HALF_EVEN)}%"


def _validate_data(
    records: Dict[str, Dict],
    logs: List[Dict],
    batch_id: str,
):
    """多层级数据校验"""
    RATE_TOLERANCE = Decimal("0.001")

    for name, record in records.items():
        source_info = _format_source_files(record.get("_sources", []))
        default_rates = record.get("_default_rates", {})

        # 1. 缴费月份检查
        period = record.get("period", "")
        if not period or len(str(period)) < 6:
            logs.append({
                "file_name": source_info,
                "employee_name": name,
                "error_level": "error",
                "error_type": "missing_period",
                "error_message": f"员工「{name}」缴费月份缺失或格式错误",
            })

        # 2. 险种校验配置：(金额字段, 比例字段, 基数字段, 险种前缀, 缴费方)
        checks = [
            ("pension_personal", "pension_personal_rate", "pension_personal_base", "pension", "personal"),
            ("pension_company", "pension_company_rate", "pension_company_base", "pension", "company"),
            ("unemployment_personal", "unemployment_personal_rate", "unemployment_personal_base", "unemployment", "personal"),
            ("unemployment_company", "unemployment_company_rate", "unemployment_company_base", "unemployment", "company"),
            ("medical_personal", "medical_personal_rate", "medical_personal_base", "medical", "personal"),
            ("medical_company", "medical_company_rate", "medical_company_base", "medical", "company"),
            ("injury_company", "injury_company_rate", "injury_company_base", "injury", "company"),
            ("hf_personal", "hf_personal_rate", "hf_base", "hf", "personal"),
            ("hf_company", "hf_company_rate", "hf_base", "hf", "company"),
        ]
        for amount_field, rate_field, base_field, ins_prefix, payer in checks:
            amount = record.get(amount_field)
            rate = record.get(rate_field)
            base = record.get(base_field)
            field_label = SI_FIELD_LABELS.get(amount_field, amount_field)

            if (amount and isinstance(amount, Decimal) and amount > 0
                    and rate and isinstance(rate, Decimal) and rate > 0
                    and base and isinstance(base, Decimal) and base > 0):
                expected = (base * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_EVEN)
                diff = abs(amount - expected)
                if diff > Decimal("5.00"):
                    logs.append({
                        "file_name": source_info,
                        "employee_name": name,
                        "error_level": "warning",
                        "error_type": "amount_mismatch",
                        "error_message": (
                            f"员工「{name}」{field_label}金额不匹配："
                            f"基数{base} × 比例{_format_rate(rate)} = 期望{expected}，实际{amount}，差额{diff}"
                        ),
                    })

            template_rate = None
            ins_default = (default_rates or {}).get(ins_prefix, {})
            if ins_default:
                template_rate_val = ins_default.get(f"{payer}_rate")
                if template_rate_val is not None:
                    template_rate = Decimal(str(template_rate_val))

            if (rate and isinstance(rate, Decimal) and rate > 0
                    and template_rate is not None and template_rate > 0):
                rate_diff = abs(rate - template_rate)
                if rate_diff > RATE_TOLERANCE:
                    logs.append({
                        "file_name": source_info,
                        "employee_name": name,
                        "error_level": "warning",
                        "error_type": "rate_mismatch",
                        "error_message": (
                            f"员工「{name}」{field_label}比例与模板配置不符："
                            f"实际比例{_format_rate(rate)}，模板配置{_format_rate(template_rate)}，请确认"
                        ),
                    })


def _safe_add(*values) -> Optional[Decimal]:
    """安全相加多个可能为None的Decimal值
    - 如果所有值都是None → 返回None（保持空）
    - 否则把None当作0参与计算
    - 注意：0是有效值（如工伤保险个人部分为0），不会被当作None处理
    """
    has_value = False
    total = Decimal("0")
    for v in values:
        if v is not None:
            has_value = True
            total += Decimal(str(v))
    return total if has_value else None


def _calc_summaries(si_record):
    """自动计算所有合计/汇总字段
    核心逻辑：只有当子字段有实际数据时才计算合计；
    如果所有子字段都是NULL（未导入），合计字段也保持NULL，不填充0。
    """
    # 各险种合计（个人+单位）
    si_record.pension_total = _safe_add(si_record.pension_personal, si_record.pension_company)
    si_record.unemployment_total = _safe_add(si_record.unemployment_personal, si_record.unemployment_company)
    si_record.medical_total = _safe_add(si_record.medical_personal, si_record.medical_company)
    si_record.injury_total = _safe_add(si_record.injury_company)  # 工伤保险只有单位部分

    # 社保个人/单位合计
    si_record.si_personal = _safe_add(
        si_record.pension_personal, si_record.unemployment_personal, si_record.medical_personal
    )
    si_record.si_company = _safe_add(
        si_record.pension_company, si_record.unemployment_company,
        si_record.medical_company, si_record.injury_company
    )
    si_record.si_grand_total = _safe_add(si_record.si_personal, si_record.si_company)

    # 公积金合计
    # 如果个人和单位金额都为空/0，但 hf_total 已有值（如邯郸公积金仅含总发生额），保留原值
    hf_personal = si_record.hf_personal
    hf_company = si_record.hf_company
    has_hf_detail = hf_personal is not None or hf_company is not None
    if not has_hf_detail and si_record.hf_total is not None:
        pass  # 只有总计值，没有明细，保留原值
    elif has_hf_detail:
        si_record.hf_total = _safe_add(hf_personal, hf_company)
    else:
        si_record.hf_total = None  # 公积金无任何数据，保持空

    # 社保公积金总合计
    si_record.grand_total = _safe_add(si_record.si_grand_total, si_record.hf_total)


def _get_decimal(record: Dict, field: str) -> Optional[Decimal]:
    """安全获取 Decimal 值"""
    val = record.get(field)
    if val is None or val == "":
        return None
    if isinstance(val, Decimal):
        return val if val != 0 else None
    try:
        d = Decimal(str(val))
        return d if d != 0 else None
    except Exception:
        return None


def _smart_calculate_record(record: Dict, default_rates: Dict) -> Dict:
    """
    智能推算单条记录的缺失字段。优先级：实际导入数据 > 模板默认比例
    1. 已知金额+基数 → 反算实际比例（最高优先级，覆盖任何默认值）
    2. 缺失的比例 → 用模板配置的默认比例填充（仅当实际数据反算不出时）
    3. 根据已有数据推算缺失字段：基数+比例→金额、金额+比例→基数
    4. 公积金特殊处理：个人/单位比例联动、已知合计+比例→拆分金额/反推基数
    """
    TWO_PLACES = Decimal("0.01")
    FOUR_PLACES = Decimal("0.0001")

    insurance_configs = [
        ("pension", True),
        ("unemployment", True),
        ("medical", True),
        ("injury", False),
    ]

    # ── 1. 处理各社保险种 ──
    for prefix, has_personal in insurance_configs:
        ins_default = (default_rates or {}).get(prefix, {})
        default_personal_rate = ins_default.get("personal_rate")
        default_company_rate = ins_default.get("company_rate")

        personal_base = f"{prefix}_personal_base"
        company_base = f"{prefix}_company_base"
        personal_amount = f"{prefix}_personal"
        company_amount = f"{prefix}_company"
        personal_rate = f"{prefix}_personal_rate"
        company_rate = f"{prefix}_company_rate"

        p_base = _get_decimal(record, personal_base)
        c_base = _get_decimal(record, company_base)
        p_amt = _get_decimal(record, personal_amount) if has_personal else None
        c_amt = _get_decimal(record, company_amount)
        p_rate = _get_decimal(record, personal_rate) if has_personal else None
        c_rate = _get_decimal(record, company_rate)

        # 步骤1：如果Excel没有导入比例，但有金额和基数 → 从实际数据反算真实比例
        if has_personal and p_rate is None and p_amt is not None and p_base is not None and p_base > 0:
            p_rate = (p_amt / p_base).quantize(FOUR_PLACES, rounding=ROUND_HALF_EVEN)
            record[personal_rate] = p_rate
        if c_rate is None and c_amt is not None and c_base is not None and c_base > 0:
            c_rate = (c_amt / c_base).quantize(FOUR_PLACES, rounding=ROUND_HALF_EVEN)
            record[company_rate] = c_rate

        # 步骤2：实际数据反算不出的（缺金额或基数），才用模板默认比例填充
        if has_personal and p_rate is None and default_personal_rate is not None:
            p_rate = Decimal(str(default_personal_rate))
            record[personal_rate] = p_rate
        if c_rate is None and default_company_rate is not None:
            c_rate = Decimal(str(default_company_rate))
            record[company_rate] = c_rate

        # 步骤3：根据已有基数/比例/金额推算缺失字段
        if has_personal:
            if p_base is not None and p_rate is not None and p_rate > 0 and p_amt is None:
                p_amt = (p_base * p_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
                record[personal_amount] = p_amt
            elif p_amt is not None and p_rate is not None and p_rate > 0 and p_base is None:
                p_base = (p_amt / p_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
                record[personal_base] = p_base

        if c_base is not None and c_rate is not None and c_rate > 0 and c_amt is None:
            c_amt = (c_base * c_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
            record[company_amount] = c_amt
        elif c_amt is not None and c_rate is not None and c_rate > 0 and c_base is None:
            c_base = (c_amt / c_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
            record[company_base] = c_base

    # ── 2. 处理公积金 ──
    hf_default = (default_rates or {}).get("hf", {})
    default_hf_personal_rate = hf_default.get("personal_rate")
    default_hf_company_rate = hf_default.get("company_rate")
    split_equal = hf_default.get("split_equal", True)

    hf_base = _get_decimal(record, "hf_base")
    hf_personal = _get_decimal(record, "hf_personal")
    hf_company = _get_decimal(record, "hf_company")
    hf_personal_rate = _get_decimal(record, "hf_personal_rate")
    hf_company_rate = _get_decimal(record, "hf_company_rate")
    hf_total = _get_decimal(record, "hf_total")

    # 步骤1：如果Excel没有导入比例，但有金额和基数 → 从实际数据反算真实比例
    if hf_personal_rate is None and hf_personal is not None and hf_base is not None and hf_base > 0:
        hf_personal_rate = (hf_personal / hf_base).quantize(FOUR_PLACES, rounding=ROUND_HALF_EVEN)
        record["hf_personal_rate"] = hf_personal_rate
    if hf_company_rate is None and hf_company is not None and hf_base is not None and hf_base > 0:
        hf_company_rate = (hf_company / hf_base).quantize(FOUR_PLACES, rounding=ROUND_HALF_EVEN)
        record["hf_company_rate"] = hf_company_rate

    # 步骤2：实际数据反算不出的，才用模板默认比例填充
    if hf_personal_rate is None and default_hf_personal_rate is not None:
        hf_personal_rate = Decimal(str(default_hf_personal_rate))
        record["hf_personal_rate"] = hf_personal_rate
    if hf_company_rate is None and default_hf_company_rate is not None:
        hf_company_rate = Decimal(str(default_hf_company_rate))
        record["hf_company_rate"] = hf_company_rate

    # 步骤3：split_equal处理：如果勾选了个人单位比例相同，补全缺失的那个
    if split_equal:
        if hf_personal_rate is not None and hf_company_rate is None:
            hf_company_rate = hf_personal_rate
            record["hf_company_rate"] = hf_company_rate
        if hf_company_rate is not None and hf_personal_rate is None:
            hf_personal_rate = hf_company_rate
            record["hf_personal_rate"] = hf_personal_rate

    # 步骤4：已知合计(hf_total)且有比例 → 拆分个人/单位金额，反推基数
    if hf_total is not None and hf_total > 0:
        total_rate = Decimal("0")
        if hf_personal_rate is not None:
            total_rate += hf_personal_rate
        if hf_company_rate is not None:
            total_rate += hf_company_rate

        if total_rate > 0:
            if hf_base is None:
                hf_base = (hf_total / total_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
                record["hf_base"] = hf_base

            if hf_personal is None and hf_personal_rate is not None and hf_base is not None:
                hf_personal = (hf_base * hf_personal_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
                record["hf_personal"] = hf_personal
            if hf_company is None and hf_company_rate is not None and hf_base is not None:
                hf_company = (hf_base * hf_company_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
                record["hf_company"] = hf_company

    # 步骤5：已知基数 + 比例 → 计算缺失的金额
    if hf_base is not None and hf_personal_rate is not None and hf_personal_rate > 0 and hf_personal is None:
        hf_personal = (hf_base * hf_personal_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
        record["hf_personal"] = hf_personal
    if hf_base is not None and hf_company_rate is not None and hf_company_rate > 0 and hf_company is None:
        hf_company = (hf_base * hf_company_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
        record["hf_company"] = hf_company

    # 步骤6：已知金额 + 比例 → 反推基数
    if hf_personal is not None and hf_personal_rate is not None and hf_personal_rate > 0 and hf_base is None:
        hf_base = (hf_personal / hf_personal_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
        record["hf_base"] = hf_base
    if hf_company is not None and hf_company_rate is not None and hf_company_rate > 0 and hf_base is None:
        hf_base = (hf_company / hf_company_rate).quantize(TWO_PLACES, rounding=ROUND_HALF_EVEN)
        record["hf_base"] = hf_base

    # 个人/单位金额都有了但合计缺失 → 计算合计
    if hf_total is None and hf_personal is not None and hf_company is not None:
        hf_total = hf_personal + hf_company
        record["hf_total"] = hf_total

    return record


def _has_value(val) -> bool:
    """判断字段是否有有效值（非None、非空字符串、非0的Decimal/数值）"""
    if val is None:
        return False
    if isinstance(val, str):
        return val.strip() != ""
    if isinstance(val, Decimal):
        return val != 0
    if isinstance(val, (int, float)):
        return val != 0
    return True


def _save_records(
    db: Session,
    period: str,
    records: Dict[str, Dict],
    user_id: int,
    username: str,
    logs: List[Dict],
    batch_id: str,
) -> Tuple[int, int]:
    """保存数据到数据库（新增或更新）
    
    关键设计：更新现有记录时，只覆盖本次导入中有实际值的字段，
    不会将本次导入缺失的字段设置为0，避免不同文件分次导入时互相覆盖数据。
    """
    created = 0
    updated = 0

    # 获取现有记录
    existing_map = {}
    existing_records = db.query(SocialInsurance).filter(
        SocialInsurance.period == period
    ).all()
    for r in existing_records:
        existing_map[r.employee_id] = r

    for name, record in records.items():
        emp_id = record.get("employee_id")
        if not emp_id:
            continue

        # 检查同一周期是否有重复
        existing = existing_map.get(emp_id)
        if existing:
            existing_map.pop(emp_id, None)  # 已处理则移除

        # 所有可写入的业务字段（合计字段最后统一计算，不在此处设置）
        business_fields = [
            # 社保 — 各险种单独基数
            "pension_personal_base", "pension_company_base",
            "unemployment_personal_base", "unemployment_company_base",
            "medical_personal_base", "medical_company_base",
            "injury_company_base",
            # 社保 — 个人/单位 金额和比例
            "pension_personal", "pension_personal_rate",
            "pension_company", "pension_company_rate",
            "unemployment_personal", "unemployment_personal_rate",
            "unemployment_company", "unemployment_company_rate",
            "medical_personal", "medical_personal_rate",
            "medical_company", "medical_company_rate",
            "injury_company", "injury_company_rate",
            # 公积金
            "hf_base", "hf_personal", "hf_personal_rate",
            "hf_company", "hf_company_rate",
            "hf_total",
        ]

        if existing:
            # ── 更新模式：只更新本次导入中实际提取到的字段 ──
            # _extract_data保证了只有有效值（包括0）才会在record中，
            # 不在record中的字段说明文件中没有该列，保留数据库原有值（NULL或之前的值）
            for f in business_fields:
                if f in record:
                    setattr(existing, f, record[f])
            # 自动计算合计字段
            _calc_summaries(existing)
            updated += 1
        else:
            # ── 新增模式：只设置本次导入中有值的字段，无关字段保持NULL ──
            # 不再自动填充0，让页面清晰展示哪些字段有数据、哪些没有
            si_data = {"period": period, "employee_id": emp_id}
            for f in business_fields:
                val = record.get(f)
                if isinstance(val, Decimal):
                    si_data[f] = val
                elif isinstance(val, (int, float)):
                    si_data[f] = val
                elif isinstance(val, str) and val.strip():
                    si_data[f] = val
                else:
                    si_data[f] = None
            si = SocialInsurance(**si_data)
            db.add(si)
            # flush让si获取id，然后计算合计字段
            db.flush()
            _calc_summaries(si)
            created += 1

    db.commit()
    return created, updated


def run_smart_import(
    db: Session,
    period: str,
    files: List,
    user_id: int,
    username: str,
    file_paths: Optional[List[str]] = None,
) -> ImportResult:
    """
    智能导入主流程：
    1. 逐个解析文件（Excel/PDF），Excel自动遍历所有工作表
    2. 每个工作表匹配模板并提取字段
    3. 按姓名聚合多文件/多工作表数据
    4. 姓名匹配员工
    5. 数据校验
    6. 保存入库
    
    Args:
        file_paths: 每个文件的相对路径（可选，用于文件夹导入时提供上下文关键词）
    """
    result = ImportResult()
    batch_id = str(uuid.uuid4())
    all_records: List[Dict] = []
    logs: List[Dict] = []

    result.total_files = len(files)

    for file_idx, file in enumerate(files):
        raw_filename = file.filename or "unknown"
        # 提取纯文件名（去掉可能的路径分隔符）
        filename = os.path.basename(raw_filename.replace('\\', '/'))
        relative_path = file_paths[file_idx] if file_paths and file_idx < len(file_paths) else raw_filename.replace('\\', '/')
        context_keywords = _extract_keywords_from_path(relative_path)
        try:
            file_bytes = file.file.read()
            file.file.seek(0)

            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

            file_matched = False
            sheets_processed = 0

            if ext in ("xlsx", "xls"):
                # Excel文件：遍历所有工作表，每个工作表尝试匹配模板
                sheet_names = _get_excel_sheet_names(file_bytes, filename)
                if not sheet_names:
                    # 无法获取工作表列表，回退到默认方式（第一个工作表）
                    sheet_names = [None]  # None表示用默认sheet
                
                data_sheets_without_template = 0
                for sheet_name in sheet_names:
                    try:
                        file_rows = _parse_excel(file_bytes, filename, sheet_name=sheet_name)
                    except Exception as e:
                        logger.warning(f"解析工作表 {sheet_name} 失败: {e}")
                        continue

                    if not file_rows or len(file_rows) < 2:
                        continue

                    # 判断是否是有效数据工作表
                    if not _sheet_has_data(file_rows):
                        continue

                    template = _match_template(db, file_rows, filename, sheet_name=sheet_name, context_keywords=context_keywords)
                    if not template:
                        # 有数据但无匹配模板，记录警告
                        sheet_display = sheet_name if sheet_name else "(默认工作表)"
                        data_sheets_without_template += 1
                        logs.append({
                            "file_name": f"{filename}（工作表：{sheet_display}）",
                            "error_level": "warning",
                            "error_type": "no_template",
                            "error_message": f"工作表「{sheet_display}」包含数据但未匹配到模板，该工作表数据已跳过",
                        })
                        continue

                    # 提取数据
                    display_name = f"{filename}（工作表：{sheet_name}）" if sheet_name else filename
                    records = _extract_data(file_rows, template, display_name, period, logs, batch_id)
                    
                    if records:
                        all_records.extend(records)
                        file_matched = True
                        sheets_processed += 1
                        logger.info(f"成功从工作表「{sheet_name}」提取 {len(records)} 条记录（模板：{template.name}）")

            elif ext == "pdf":
                file_rows = _parse_pdf(file_bytes)
                
                if not file_rows or len(file_rows) < 2:
                    logs.append({
                        "file_name": filename,
                        "error_level": "error",
                        "error_type": "empty_file",
                        "error_message": "文件为空或数据行数不足",
                    })
                    result.failed_files += 1
                    continue

                template = _match_template(db, file_rows, filename, context_keywords=context_keywords)
                if not template:
                    logs.append({
                        "file_name": filename,
                        "error_level": "error",
                        "error_type": "no_template",
                        "error_message": f"未匹配到适合文件「{filename}」的导入模板，请在「社保公积金导入模板配置」页面中为该文件创建对应的解析模板后再导入",
                    })
                    result.failed_files += 1
                    continue

                records = _extract_data(file_rows, template, filename, period, logs, batch_id)
                if records:
                    all_records.extend(records)
                    file_matched = True
                    sheets_processed = 1
            else:
                logs.append({
                    "file_name": filename,
                    "error_level": "error",
                    "error_type": "unsupported_format",
                    "error_message": f"不支持的文件格式「.{ext}」，仅支持 Excel(.xlsx/.xls) 和 PDF(.pdf)",
                })
                result.failed_files += 1
                continue

            if file_matched:
                result.parsed_files += 1
            else:
                # 所有工作表都没匹配到模板或没提取到数据
                logs.append({
                    "file_name": filename,
                    "error_level": "error",
                    "error_type": "no_template",
                    "error_message": f"未匹配到适合文件「{filename}」的导入模板（已检查所有工作表），请在「社保公积金导入模板配置」页面中为该文件创建对应的解析模板后再导入",
                })
                result.failed_files += 1

        except Exception as e:
            logger.exception(f"解析文件失败: {filename}")
            logs.append({
                "file_name": filename,
                "error_level": "error",
                "error_type": "file_error",
                "error_message": f"文件解析失败：{str(e)}",
            })
            result.failed_files += 1

    result.total_rows = len(all_records)

    if not all_records:
        result.errors = logs
        _save_logs(db, logs, period, batch_id)
        return result

    # ── 第2步：按姓名聚合 ──
    aggregated = _aggregate_by_name(all_records)

    # ── 第2.5步：共用基数自动填充 ──
    # 如果只有个人基数没有单位基数（或反之），自动复制（绝大多数情况下基数相同）
    _auto_fill_shared_bases(aggregated)

    # ── 第3步：姓名匹配员工 ──
    valid_data, skipped = _match_employees(db, aggregated, logs, batch_id)

    # ── 第3.5步：智能推算缺失字段（应用默认比例、反算比例/金额、拆分公积金等）──
    for name, record in valid_data.items():
        default_rates = record.get("_default_rates", {})
        _smart_calculate_record(record, default_rates)

    # ── 第4步：数据校验 ──
    _validate_data(valid_data, logs, batch_id)

    # ── 第5步：保存 ──
    created, updated = _save_records(db, period, valid_data, user_id, username, logs, batch_id)
    result.created = created
    result.updated = updated

    # ── 第6步：分类日志 ──
    for log_entry in logs:
        if log_entry.get("error_level") == "error":
            result.errors.append(log_entry)
        else:
            result.warnings.append(log_entry)

    _save_logs(db, logs, period, batch_id)

    return result


def _save_logs(db: Session, logs: List[Dict], period: str, batch_id: str):
    """将导入日志写入数据库（日志保存失败不影响主流程）"""
    if not logs:
        return
    try:
        for log_entry in logs:
            file_name = log_entry.get("file_name") or ""
            error_msg = log_entry.get("error_message", "") or ""
            emp_name = log_entry.get("employee_name") or ""
            db_log = SiImportLog(
                period=period,
                batch_id=batch_id,
                file_name=file_name[:1000] if len(file_name) > 1000 else file_name,
                row_number=log_entry.get("row_number"),
                employee_name=emp_name[:100] if len(emp_name) > 100 else emp_name,
                error_level=log_entry.get("error_level", "warning"),
                error_type=log_entry.get("error_type", "unknown")[:50],
                error_message=error_msg[:2000] if len(error_msg) > 2000 else error_msg,
            )
            db.add(db_log)
        db.commit()
    except Exception as e:
        logger.error(f"保存导入日志失败（不影响数据导入结果）: {e}")
        db.rollback()



# ── 模板自动识别已移除 ──
# 智能导入时，若未匹配到已配置的模板，将直接返回错误提示，
# 引导用户先在「导入模板配置」中手动创建模板。


def _keyword_match_columns(column_names: List[str], source_category: str) -> Dict[str, Optional[str]]:
    """
    基于关键词规则的兜底匹配（AI 不可用时使用）。
    按优先级从高到低逐字段匹配，已匹配的系统字段不再重复匹配。
    【重要】宁可不匹配，也不要错配。只使用高置信度的精确关键词。
    """
    # 关键词匹配规则：(关键词列表, 系统字段名) 按优先级排序
    # 注意：不要使用过于宽泛的词（如单独的"基数"、"金额"、"合计"），避免误匹配
    RULES = [
        # ── 姓名 ──
                (["姓名", "员工姓名", "人员姓名", "名字"], "employee_name"),
        # ── 养老保险 ──
                (["养老个人基数", "养老保险个人缴费基数", "基本养老保险个人缴费基数"], "pension_personal_base"),
                (["养老单位基数", "养老保险单位缴费基数", "基本养老保险单位缴费基数"], "pension_company_base"),
                (["养老个人金额", "养老个人缴费", "养老保险个人缴纳", "基本养老保险个人缴纳金额"], "pension_personal"),
                (["养老个人比例", "养老个人费率", "养老保险个人比例", "基本养老保险个人比例"], "pension_personal_rate"),
                (["养老单位金额", "养老单位缴费", "养老保险单位缴纳", "基本养老保险单位缴纳金额", "单位缴纳基本养老保险"], "pension_company"),
                (["养老单位比例", "养老单位费率", "养老保险单位比例", "基本养老保险单位比例"], "pension_company_rate"),
        # ── 失业保险 ──
                (["失业个人基数", "失业保险个人缴费基数"], "unemployment_personal_base"),
                (["失业单位基数", "失业保险单位缴费基数"], "unemployment_company_base"),
                (["失业个人金额", "失业个人缴费", "失业保险个人缴纳", "失业保险个人缴纳金额"], "unemployment_personal"),
                (["失业个人比例", "失业个人费率", "失业保险个人比例"], "unemployment_personal_rate"),
                (["失业单位金额", "失业单位缴费", "失业保险单位缴纳", "失业保险单位缴纳金额", "单位缴纳失业保险"], "unemployment_company"),
                (["失业单位比例", "失业单位费率", "失业保险单位比例"], "unemployment_company_rate"),
        # ── 医疗保险 ──
                (["医疗个人基数", "医疗保险个人缴费基数", "基本医疗保险个人缴费基数"], "medical_personal_base"),
                (["医疗单位基数", "医疗保险单位缴费基数", "基本医疗保险单位缴费基数"], "medical_company_base"),
                (["医疗个人金额", "医疗个人缴费", "医疗保险个人缴纳", "医疗保险个人缴纳金额"], "medical_personal"),
                (["医疗个人比例", "医疗个人费率", "医疗保险个人比例"], "medical_personal_rate"),
                (["医疗单位金额", "医疗单位缴费", "医疗保险单位缴纳", "医疗保险单位缴纳金额", "单位缴纳基本医疗保险"], "medical_company"),
                (["医疗单位比例", "医疗单位费率", "医疗保险单位比例"], "medical_company_rate"),
        # ── 工伤保险 ──
                (["工伤单位基数", "工伤保险单位缴费基数", "工伤保险缴费基数"], "injury_company_base"),
                (["工伤单位金额", "工伤单位缴费", "工伤保险单位缴纳", "工伤保险单位缴纳金额", "单位缴纳工伤保险"], "injury_company"),
                (["工伤单位比例", "工伤单位费率", "工伤保险单位比例", "工伤保险比例"], "injury_company_rate"),
        # ── 生育保险 ──
                (["生育单位基数", "生育保险单位缴费基数", "生育保险缴费基数"], "maternity_company_base"),
                (["生育单位金额", "生育单位缴费", "生育保险单位缴纳", "生育保险单位缴纳金额", "单位缴纳生育保险"], "maternity_company"),
                (["生育单位比例", "生育单位费率", "生育保险单位比例"], "maternity_company_rate"),
        # ── 公积金 ──
                (["公积金缴存基数", "公积金缴费基数", "住房公积金缴存基数", "住房公积金缴费基数"], "hf_base"),
                (["公积金个人金额", "公积金个人缴费", "公积金个人缴存", "公积金个人部分", "个人缴存额", "住房公积金个人缴纳", "住房公积金个人金额"], "hf_personal"),
                (["公积金个人比例", "公积金个人费率", "公积金个人缴存比例", "个人缴存比例"], "hf_personal_rate"),
                (["公积金单位金额", "公积金单位缴费", "公积金单位缴存", "公积金单位部分", "单位缴存额", "住房公积金单位缴纳", "住房公积金单位金额"], "hf_company"),
                (["公积金单位比例", "公积金单位费率", "公积金单位缴存比例", "单位缴存比例"], "hf_company_rate"),
        # ── 社保汇总 ──
                (["社保个人合计", "社保个人小计", "个人社保合计", "社保个人总额"], "si_personal"),
                (["社保单位合计", "社保单位小计", "单位社保合计", "社保单位总额"], "si_company"),
        # ── 合计 ──
                (["养老保险合计", "养老小计", "养老保险小计"], "pension_total"),
                (["失业保险合计", "失业小计", "失业保险小计"], "unemployment_total"),
                (["医疗保险合计", "医疗小计", "医疗保险小计"], "medical_total"),
                (["工伤保险合计", "工伤小计", "工伤保险小计"], "injury_total"),
                (["社保总合计", "社保合计", "社保汇总", "社保总额"], "si_grand_total"),
                (["公积金合计", "公积金汇总", "公积金总额", "住房公积金合计"], "hf_total"),
                (["总合计", "应缴合计", "社保公积金总合计", "总计", "缴费合计"], "grand_total"),
    ]

    # 对列名做标准化处理（去空格、去括号、统一全角半角）
    def _normalize(s: str) -> str:
        return s.replace(" ", "").replace("\u3000", "").replace("（", "(").replace("）", ")")

    mappings: Dict[str, Optional[str]] = {}
    used_fields: set = set()

    for col in column_names:
        norm_col = _normalize(col)
        # 收集所有可能的匹配：(关键词, 系统字段, 匹配得分, 关键词长度)
        # 得分规则：精确匹配=3, 关键词完整包含于列名=2（要求关键词长度≥4，避免短词误匹配）
        candidates = []
        for keywords, field in RULES:
            if field in used_fields:
                continue
            for kw in keywords:
                norm_kw = _normalize(kw)
                kw_len = len(norm_kw)
                score = 0
                if norm_col == norm_kw:
                    score = 3
                elif kw_len >= 4 and norm_kw in norm_col:
                    # 对于包含匹配，要求关键词长度至少4个字符，避免"基数"、"金额"等短词误匹配
                    score = 2
                if score > 0:
                    candidates.append((kw, field, score, kw_len))

        # 按得分降序，同分按关键词长度降序（越长越精准）
        if candidates:
            candidates.sort(key=lambda x: (x[2], x[3]), reverse=True)
            best = candidates[0]
            # 只有得分≥2才匹配，得分=1（列名包含于关键词但关键词不包含列名）不再使用
            if best[2] >= 2:
                mappings[col] = best[1]
                used_fields.add(best[1])
            else:
                mappings[col] = None
        else:
            mappings[col] = None

    return mappings


def _ai_match_columns(column_names: List[str], source_category: str, city: Optional[str] = None) -> Dict[str, Optional[str]]:
    """
    使用 AI（百炼）智能匹配文件列名到系统字段。
    AI 不可用时返回空字典，由调用方降级到关键词匹配。
    """
    from app.core.ai_service import is_available, chat_json

    if not is_available():
        logger.info("AI 服务未配置，将使用关键词规则匹配字段")
        return {}

    if not column_names:
        return {}

    # 构建系统字段说明
    field_desc_lines = []
    for field_key, field_label in SI_FIELD_LABELS.items():
        field_desc_lines.append(f"  - {field_key}: {field_label}")

    system_prompt = (
        "你是一个社保公积金数据字段匹配专家，要求极度严谨、精准。\n"
        "你的任务是根据文件中的列名，判断它对应哪个系统数据库字段。\n\n"
        "【核心原则：宁可不匹配，也不要错配！】\n"
        "- 只有当你100%确定列名对应某个系统字段时才填写，否则一律填null\n"
        "- 不要因为列名包含某个词就强行映射，避免过度匹配\n"
        "- 如果列名已经明确包含险种名称（如'基本养老保险'、'失业保险'），必须映射到对应险种的具体字段，不要使用通用字段\n\n"
        "规则：\n"
        "1. 如果列名明确对应某个系统字段，填入字段名\n"
        "2. 如果不确定、模糊或有歧义，填入null\n"
        "3. 严格区分「个人」和「单位」、「基数」和「金额」和「比例」\n"
        "4. 严格区分不同险种：养老、失业、医疗、工伤、生育、公积金\n"
        "5. 每个系统字段只能匹配一次，不要重复匹配\n"
        f"6. 数据来源类别是「{'社保' if source_category == 'social_insurance' else '公积金'}」，"
        f"请优先匹配{'社保' if source_category == 'social_insurance' else '公积金'}相关字段\n"
        "7. 对于包含明确险种+方向的列名（如'基本养老保险(单位缴纳) - 缴费基数'），必须映射到对应的具体字段（如pension_company_base），不要映射到通用字段\n"
        "8. 只有当列名确实是纯通用名称（不包含任何具体险种名）时，才考虑映射，但如果不确定还是填null\n\n"
        "请直接返回 JSON，格式为：\n"
        '{"mappings": {"文件列名A": "系统字段名", "文件列名B": null, ...}}\n'
        "确保每个文件列名都出现在 mappings 中。"
    )

    user_message = (
        f"文件中的列名列表：\n{json.dumps(column_names, ensure_ascii=False, indent=2)}\n\n"
        f"系统可用字段（字段名: 中文标签）：\n"
        + "\n".join(field_desc_lines)
        + f"\n\n{'城市：' + city if city else ''}"
    )

    try:
        result = chat_json(system_prompt, user_message, temperature=0.1)
        raw_mappings = result.get("mappings", {})
        if not isinstance(raw_mappings, dict):
            logger.warning(f"AI 返回的 mappings 格式异常: {type(raw_mappings)}")
            return {}

        # 验证返回的字段名是否都在 SI_FIELD_LABELS 中
        mappings: Dict[str, Optional[str]] = {}
        for col in column_names:
            db_field = raw_mappings.get(col)
            if db_field and db_field in SI_FIELD_LABELS:
                mappings[col] = db_field
            elif db_field and db_field not in SI_FIELD_LABELS:
                logger.warning(f"AI 返回了无效的系统字段名: {db_field}（列名: {col}），已忽略")
                mappings[col] = None
            else:
                mappings[col] = None
        return mappings

    except Exception as e:
        logger.warning(f"AI 字段匹配失败，将使用关键词规则兜底: {e}")
        return {}


def auto_detect_template(
    file_bytes: bytes,
    filename: str,
    sheet_name: Optional[str] = None,
) -> Dict:
    """
    文件结构自动检测（含 AI 智能字段匹配）。
    提取文件中的列名，调用 AI 自动匹配系统字段映射，
    AI 不可用时降级为关键词规则匹配。
    用户在前端核对确认后保存即可。
    
    Args:
        sheet_name: Excel工作表名称，指定解析哪个工作表；None则使用默认工作表
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in ("xlsx", "xls"):
        file_rows = _parse_excel(file_bytes, filename, sheet_name=sheet_name)
        file_type = "excel"
    elif ext == "pdf":
        file_rows = _parse_pdf(file_bytes)
        file_type = "pdf"
    else:
        raise ValueError(f"不支持的文件格式「.{ext}」")

    if not file_rows or len(file_rows) < 2:
        raise ValueError("文件为空或数据行数不足")

    total_rows = len(file_rows)

    # 检测表头行：扫描前30行，找包含"姓名"或列数最多的行
    header_row = 0
    max_cols = 0
    for i in range(min(30, total_rows)):
        row = file_rows[i]
        non_empty = [c for c in row if c and str(c).strip()]
        if len(non_empty) > max_cols:
            max_cols = len(non_empty)
            header_row = i
        # 如果包含"姓名"列，优先选这一行
        name_found = any("姓名" in str(c) for c in row if c)
        if name_found and len(non_empty) >= 2:
            header_row = i
            break

    header = file_rows[header_row]
    header_rows_list = [header_row]

    # ── 二级表头检测：检查下一行是否是子列表头 ──
    # 场景：广州社保文件格式
    #   行6: 基本养老保险(单位缴纳) | 基本养老保险(单位缴纳) | ...  (一级：险种名)
    #   行7: 缴费基数 | 应缴金额 | 缴费基数 | 应缴金额 | ...  (二级：子列名)
    # 判断标准：一级表头中有重复列名（同一险种占多列），下一行都是中文关键词
    next_row_idx = header_row + 1
    if next_row_idx < total_rows:
        header_texts = [str(c).strip() if c else "" for c in header]
        next_row = file_rows[next_row_idx]
        next_texts = [str(c).strip() if c else "" for c in next_row]
        # 检查一级表头是否有大量重复列名（排除开头固定列如序号、姓名等）
        meaningful_headers = [h for h in header_texts if h and h not in ("序号", "姓名", "证件号码", "证件类型", "个人社保号", "费款所属期起", "费款所属期止", "单位", "个人")]
        has_duplicates = len(meaningful_headers) > len(set(meaningful_headers)) * 1.2 if meaningful_headers else False
        # 检查下一行是否像子表头行：包含典型的子列关键词
        sub_header_keywords = ["缴费基数", "应缴金额", "费率", "比例", "基数", "金额", "缴费工资",
                                "个人缴纳", "单位缴纳", "个人缴存", "单位缴存", "缴存额", "应缴费额",
                                "个人部分", "单位部分", "合计", "小计"]
        next_non_empty = [t for t in next_texts if t]
        sub_keyword_count = sum(1 for t in next_non_empty for kw in sub_header_keywords if kw in t)
        # 子表头行特征：基本都是中文（不含数字），且包含子列关键词
        has_numbers = sum(1 for t in next_non_empty if any(ch.isdigit() for ch in t))
        is_text_row = has_numbers <= len(next_non_empty) * 0.2  # 数字比例低于20%
        if (has_duplicates or sub_keyword_count >= 3) and is_text_row and len(next_non_empty) >= 3:
            header_rows_list.append(next_row_idx)

    header = file_rows[header_row]
    column_names = [str(c).strip() if c else "" for c in header]

    # ── 多级表头合并：生成完整的二级列名 ──
    # 如果有二级表头（header_rows_list 有2+行），将多层合并为完整列名
    # 例如：行6 "基本养老保险(单位缴纳)" + 行7 "缴费基数" → "基本养老保险(单位缴纳) - 缴费基数"
    if len(header_rows_list) > 1:
        max_cols = max(len(file_rows[hr]) if hr < len(file_rows) else 0 for hr in header_rows_list)
        merged_names = []
        for col_idx in range(max_cols):
            parts = []
            for hr in header_rows_list:
                if hr < len(file_rows) and col_idx < len(file_rows[hr]):
                    val = str(file_rows[hr][col_idx]).strip() if file_rows[hr][col_idx] else ""
                    if val and (not parts or val != parts[-1]):
                        parts.append(val)
            merged_names.append(" - ".join(parts) if len(parts) > 1 else (parts[0] if parts else ""))
        column_names = merged_names

    # 数据起始行 = 最后一行表头 + 1
    data_start_row = header_rows_list[-1] + 1

    # 跳过末尾的空行
    skip_footer_rows = 0
    for i in range(total_rows - 1, data_start_row, -1):
        row = file_rows[i]
        if row and any(c for c in row if c and str(c).strip()):
            break
        skip_footer_rows += 1

    # 检测数据类别（简单规则）
    all_text = " ".join(str(c) for row in file_rows[:min(10, total_rows)] for c in row if c)
    if any(kw in all_text for kw in ["公积金", "缴存额", "缴存比例", "公积金账号", "缴存名单", "汇缴"]):
        source_category = "housing_fund"
    else:
        source_category = "social_insurance"

    # 检测城市
    city = None
    city_keywords = {"广州": "广州", "邯郸": "邯郸", "深圳": "深圳", "北京": "北京", "上海": "上海"}
    for kw, city_name in city_keywords.items():
        if kw in all_text or kw in filename:
            city = city_name
            break

    # ── 智能字段匹配：AI 优先，关键词兜底 ──
    valid_column_names = [name for name in column_names if name]

    # 先尝试 AI 匹配
    ai_mappings = _ai_match_columns(valid_column_names, source_category, city)

    if ai_mappings:
        # AI 匹配成功
        column_mappings = ai_mappings
        matched_count = sum(1 for v in ai_mappings.values() if v is not None)
        description = f"AI 已自动匹配 {matched_count}/{len(valid_column_names)} 个字段，请核对后保存"
        logger.info(f"AI 字段匹配完成: {matched_count}/{len(valid_column_names)} 个字段已匹配（文件: {filename}）")
    else:
        # AI 不可用，降级为关键词规则匹配
        keyword_mappings = _keyword_match_columns(valid_column_names, source_category)
        column_mappings = keyword_mappings
        matched_count = sum(1 for v in keyword_mappings.values() if v is not None)
        description = f"已自动匹配 {matched_count}/{len(valid_column_names)} 个字段，请核对后保存"
        logger.info(f"关键词规则匹配完成: {matched_count}/{len(valid_column_names)} 个字段已匹配（文件: {filename}）")

    basename = filename.rsplit(".", 1)[0] if "." in filename else filename

    # ── 险种感知的列名解析与修正 ──
    # 核心改进：
    # 1. 如果列名已包含明确险种关键词（养老/失业/医疗/工伤/生育/公积金），优先解析为具体险种字段
    # 2. 只有当列名是纯通用名称（不包含任何险种名）时，才使用通用字段映射
    # 3. 不再覆盖AI/关键词已经正确匹配的具体险种字段

    # 险种关键词（按长度降序，避免短词误匹配）
    INSURANCE_KEYWORDS = [
        ("基本养老保险", "pension"), ("养老保险", "pension"), ("养老", "pension"),
        ("失业保险", "unemployment"), ("失业", "unemployment"),
        ("基本医疗保险", "medical"), ("医疗保险", "medical"), ("医疗", "medical"),
        ("工伤保险", "injury"), ("工伤", "injury"),
        ("生育保险", "maternity"), ("生育", "maternity"),
        ("住房公积金", "hf"), ("公积金", "hf"),
    ]

    # 缴费方向关键词（按长度降序）
    PAYER_KEYWORDS = [
        ("个人缴纳", "personal"), ("单位缴纳", "company"),
        ("个人缴存", "personal"), ("单位缴存", "company"),
        ("个人部分", "personal"), ("单位部分", "company"),
        ("(个人)", "personal"), ("(单位)", "company"),
        ("（个人）", "personal"), ("（单位）", "company"),
        ("个人", "personal"), ("单位", "company"), ("企业", "company"),
    ]

    # 字段类型关键词
    TYPE_KEYWORDS = {
        "base": ["缴费基数", "缴费工资", "缴存基数", "基数"],
        "rate": ["费率", "缴费比例", "缴存比例", "比例"],
        "amount": ["应缴金额", "应缴费额", "缴费金额", "应缴", "缴存额", "金额"],
    }

    def _parse_column_insurance(col_name: str) -> Optional[Tuple[str, str, str]]:
        """
        从列名中解析 (险种, 缴费方向, 字段类型)。
        如果列名不包含明确险种信息，返回None。
        特殊处理：
        1. 工伤保险、生育保险只有单位部分，列名可能不标注(单位缴纳)
        2. 横向格式共用基数（如"缴费基数_养老保险"）未标注个人/单位，默认personal_base（后续自动补全company_base）
        3. 公积金基数不分个人/单位，统一返回hf_base
        """
        col_lower = col_name
        ins_type = None
        payer = None
        ftype = None

        # 1. 识别险种
        for kw, itype in INSURANCE_KEYWORDS:
            if kw in col_lower:
                ins_type = itype
                break

        # 如果没有识别到险种，返回None（说明是纯通用列名）
        if not ins_type:
            # 额外检查：合计类字段
            if "合计" in col_lower or "总计" in col_lower:
                # 先尝试识别具体险种合计
                for kw, itype in INSURANCE_KEYWORDS:
                    if kw in col_lower:
                        if itype == "pension":
                            return ("pension", "total", "total")
                        elif itype == "unemployment":
                            return ("unemployment", "total", "total")
                        elif itype == "medical":
                            return ("medical", "total", "total")
                        elif itype == "injury":
                            return ("injury", "total", "total")
                # 社保合计/公积金合计/总合计
                if "社保" in col_lower and "公积金" not in col_lower:
                    return ("si", "total", "total")
                if "公积金" in col_lower:
                    return ("hf", "total", "total")
                if "总合计" in col_lower or "应缴合计" in col_lower:
                    return ("grand", "total", "total")
            return None

        # 2. 识别缴费方向
        for kw, ptype in PAYER_KEYWORDS:
            if kw in col_lower:
                payer = ptype
                break

        # 3. 识别字段类型
        for ftype_name, kw_list in TYPE_KEYWORDS.items():
            for kw in kw_list:
                if kw in col_lower:
                    ftype = ftype_name
                    break
            if ftype:
                break

        # 特殊处理1：工伤保险、生育保险只有单位部分，如果列名没标注方向，默认是单位
        if not payer and ins_type in ("injury", "maternity"):
            payer = "company"
        
        # 特殊处理2：公积金基数不分个人/单位，直接返回base（_build_specific_field会处理为hf_base）
        if ins_type == "hf" and ftype == "base":
            payer = "personal"  # 任意值，hf_base不区分
        
        # 特殊处理3：共用基数（识别到险种和base类型，但没有明确个人/单位方向）
        # 例如"缴费基数_养老保险"这种横向格式，默认按personal_base映射，
        # 后续由_auto_fill_shared_bases自动填充company_base
        if not payer and ftype == "base" and ins_type not in ("hf",):
            payer = "personal"

        if ins_type and payer and ftype:
            return (ins_type, payer, ftype)
        # 合计类特殊处理
        if "合计" in col_lower or "小计" in col_lower:
            return (ins_type, "total", "total")
        return None

    def _build_specific_field(ins_type: str, payer: str, ftype: str) -> Optional[str]:
        """根据险种、方向、类型构建具体系统字段名"""
        # 合计字段
        if ftype == "total":
            total_map = {
                "pension": "pension_total",
                "unemployment": "unemployment_total",
                "medical": "medical_total",
                "injury": "injury_total",
                "si": "si_grand_total",
                "hf": "hf_total",
                "grand": "grand_total",
            }
            return total_map.get(ins_type)

        if ins_type == "hf":
            # 公积金特殊处理
            if ftype == "base":
                return "hf_base"
            elif ftype == "rate":
                return f"hf_{payer}_rate"
            elif ftype == "amount":
                return f"hf_{payer}"
        else:
            # 社保险种
            if ftype == "base":
                return f"{ins_type}_{payer}_base"
            elif ftype == "rate":
                return f"{ins_type}_{payer}_rate"
            elif ftype == "amount":
                return f"{ins_type}_{payer}"
        return None

    # ── 第一步：基于列名内容，尝试解析为具体险种字段 ──
    # 这是高置信度匹配，优先于通用字段处理
    parsed_mappings = {}
    parsed_count = 0
    for col_name in column_mappings.keys():
        parsed = _parse_column_insurance(col_name)
        if parsed:
            ins_type, payer, ftype = parsed
            specific_field = _build_specific_field(ins_type, payer, ftype)
            if specific_field and specific_field in SI_FIELD_LABELS:
                parsed_mappings[col_name] = specific_field
                parsed_count += 1

    # ── 第二步：处理纯通用列名（列名中不包含任何险种信息）──
    generic_col_patterns = {
        "缴费基数": "amount_base",
        "缴费工资": "amount_base",
        "费率": "rate",
        "缴费比例": "rate",
        "应缴费额": "amount",
        "缴费金额": "amount",
        "应缴金额": "amount",
        "单位应缴": "company_amount",
        "个人应缴": "personal_amount",
    }
    personal_col_keywords = ["个人", "个人缴纳", "个人部分", "个人应缴"]
    company_col_keywords = ["单位", "企业", "公司", "单位缴纳", "单位部分", "单位应缴"]

    # 合计类通用列名（根据数据来源类别判断）
    total_col_patterns = {}
    if source_category == "social_insurance":
        total_col_patterns = {
            "单位部分合计": "si_company",
            "个人部分合计": "si_personal",
            "应缴金额合计": "si_grand_total",
            "单位合计": "si_company",
            "个人合计": "si_personal",
            "社保合计": "si_grand_total",
        }
    elif source_category == "housing_fund":
        total_col_patterns = {
            "单位部分合计": "hf_company",
            "个人部分合计": "hf_personal",
            "应缴金额合计": "hf_total",
            "单位合计": "hf_company",
            "个人合计": "hf_personal",
            "公积金合计": "hf_total",
        }

    fixed_mappings = {}
    generic_fix_count = 0

    for col_name, db_field in column_mappings.items():
        # 如果列名能解析为具体险种字段，优先使用解析结果
        if col_name in parsed_mappings:
            fixed_mappings[col_name] = parsed_mappings[col_name]
            continue

        # 检查列名是否包含险种关键词
        has_insurance = any(kw in col_name for kw, _ in INSURANCE_KEYWORDS)

        # 如果列名包含险种关键词，但我们没能成功解析（缺少方向或类型），
        # 保留原有匹配结果（可能AI已经正确匹配），不要强制改成通用字段
        if has_insurance:
            fixed_mappings[col_name] = db_field
            continue

        # 列名不包含险种关键词，才检查是否是纯通用列名
        matched_generic = None
        matched_total = None

        # 优先匹配合计类列名
        for pattern, total_field in total_col_patterns.items():
            if pattern in col_name:
                matched_total = total_field
                break

        if matched_total:
            if matched_total != db_field:
                fixed_mappings[col_name] = matched_total
                generic_fix_count += 1
            else:
                fixed_mappings[col_name] = db_field
            continue

        # 再检查普通通用列名
        for pattern, generic_field in generic_col_patterns.items():
            if pattern in col_name:
                matched_generic = generic_field
                break

        if matched_generic:
            # 根据列名内容判断是个人/单位/完全通用
            col_personal = any(kw in col_name for kw in personal_col_keywords)
            col_company = any(kw in col_name for kw in company_col_keywords)

            if matched_generic == "personal_amount" or (col_personal and matched_generic == "amount"):
                target_field = "personal_amount"
            elif matched_generic == "company_amount" or (col_company and matched_generic == "amount"):
                target_field = "company_amount"
            elif col_personal and matched_generic == "amount_base":
                target_field = "personal_base"
            elif col_company and matched_generic == "amount_base":
                target_field = "company_base"
            elif col_personal and matched_generic == "rate":
                target_field = "personal_rate"
            elif col_company and matched_generic == "rate":
                target_field = "company_rate"
            else:
                # 列名未明确个人/单位，使用完全通用字段（导入时从文件名推断）
                target_field = matched_generic

            if target_field != db_field:
                fixed_mappings[col_name] = target_field
                generic_fix_count += 1
            else:
                fixed_mappings[col_name] = db_field
        else:
            fixed_mappings[col_name] = db_field

    column_mappings = fixed_mappings
    matched_count = sum(1 for v in column_mappings.values() if v is not None)

    # 生成描述信息
    if parsed_count > 0 and generic_fix_count > 0:
        description = (
            f"已自动识别 {parsed_count} 个含具体险种的字段，{generic_fix_count} 个通用字段，"
            f"共匹配 {matched_count}/{len(valid_column_names)} 个字段，请核对后保存"
        )
    elif parsed_count > 0:
        description = (
            f"AI 已自动匹配 {matched_count}/{len(valid_column_names)} 个字段，"
            f"其中自动识别了{parsed_count}个含具体险种名的列，请核对后保存"
        )
    elif generic_fix_count > 0:
        insurance_info = _infer_insurance_from_filename(filename)
        ins_type = insurance_info.get("insurance_type")
        payer = insurance_info.get("payer")
        detected_info = ""
        if ins_type or payer:
            ins_name = {"pension": "养老", "unemployment": "失业", "medical": "医疗",
                        "injury": "工伤", "maternity": "生育", "hf": "公积金"}.get(ins_type, "")
            payer_name = {"personal": "个人", "company": "单位"}.get(payer, "")
            if ins_name or payer_name:
                detected_info = f"从文件名识别到「{ins_name}{payer_name}」，"
        description = (
            f"{detected_info}检测到通用表头格式（列名不含具体险种），已自动配置为通用字段映射"
            f"（共匹配 {matched_count}/{len(valid_column_names)} 个字段）。"
            f"系统将根据导入文件的文件名自动识别险种。如需限定为单一险种，请手动调整映射字段。"
        )
    else:
        description = f"已自动匹配 {matched_count}/{len(valid_column_names)} 个字段，请核对后保存"

    # 从文件名提取关键词（用于关键词匹配模式）
    # 过滤掉具体险种名和个人/单位方向词，保留城市、"社保"等通用词，
    # 确保通用模板能匹配所有同格式的险种文件
    basename_no_ext = basename[:50]
    import re as re2
    filter_keywords = {
        "养老", "养老保险", "失业", "失业保险", "医疗", "医疗保险", "医保",
        "工伤", "工伤保险", "生育", "生育保险", "公积金", "住房公积金",
        "个人", "单位", "企业", "公司", "缴纳", "缴存", "部分", "应缴",
        "缴费", "明细", "表", "汇总", "清册",
        # 拼音/英文常见
        "pension", "unemployment", "medical", "injury", "hf",
        "personal", "company",
    }
    keywords_candidates = [w for w in re2.split(r'[_\-（）()\s\u3000,，\.]+', basename_no_ext) if len(w) >= 2 and not w.isdigit()]
    # 过滤掉险种/方向关键词，只保留通用词（如城市名、"社保"等）
    keywords = []
    for kw in keywords_candidates:
        skip = False
        for fk in filter_keywords:
            if fk in kw:
                skip = True
                break
        if not skip:
            keywords.append(kw)
    keywords = keywords[:5]

    # 获取所有工作表名称（供前端选择）
    all_sheet_names = None
    if ext in ("xlsx", "xls"):
        try:
            all_sheet_names = _get_excel_sheet_names(file_bytes, filename)
        except Exception:
            all_sheet_names = None

    # 如果指定了sheet_name，建议将其作为sheet_pattern
    suggested_sheet_pattern = sheet_name if sheet_name else None

    return {
        "name": f"{city or '通用'}{'社保' if source_category == 'social_insurance' else '公积金'}-{basename[:20]}",
        "source_category": source_category,
        "file_type": file_type,
        "city": city,
        "description": description,
        "file_pattern": None,
        "file_keywords": keywords if keywords else None,
        "sheet_pattern": suggested_sheet_pattern,
        "header_rows": header_rows_list,
        "data_start_row": data_start_row,
        "skip_footer_rows": skip_footer_rows,
        "column_mappings": column_mappings,
        "row_filters": None,
        "number_format": {"remove_chars": [",", "，"], "decimal_separator": "."},
        "sheet_name": sheet_name,
        "all_sheet_names": all_sheet_names,
    }


# ── 批量模板配置向导 ──────────────────────────────────────────────────

class BatchFileInfo:
    """批量上传文件信息"""
    def __init__(self, filename: str, original_filename: str, file_path: Path):
        self.filename = filename
        self.original_filename = original_filename
        self.file_path = file_path


def _cleanup_expired_batches():
    """清理过期的临时批次文件"""
    now = time.time()
    if not TEMP_DIR.exists():
        return
    for batch_dir in TEMP_DIR.iterdir():
        if not batch_dir.is_dir():
            continue
        try:
            meta_path = batch_dir / "meta.json"
            if not meta_path.exists():
                if now - batch_dir.stat().st_mtime > BATCH_EXPIRE_SECONDS:
                    shutil.rmtree(batch_dir, ignore_errors=True)
                continue
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            created_at = meta.get("created_at", 0)
            if now - created_at > BATCH_EXPIRE_SECONDS:
                shutil.rmtree(batch_dir, ignore_errors=True)
        except Exception:
            pass


def save_batch_files(period: str, files: List, file_paths: Optional[List[str]] = None) -> Dict:
    """
    保存上传的文件到临时目录，返回batch信息。
    Args:
        file_paths: 每个文件的相对路径（用于文件夹导入时保留目录结构），与files一一对应
    返回: {batch_id, period, files: [{index, filename, relative_path, size}]}
    """
    _cleanup_expired_batches()

    batch_id = str(uuid.uuid4())
    batch_dir = TEMP_DIR / batch_id
    batch_dir.mkdir(parents=True, exist_ok=True)

    file_infos = []
    for idx, file in enumerate(files):
        raw_filename = file.filename or f"unknown_{idx}"
        # 从raw_filename中提取纯文件名（去掉可能的路径分隔符，浏览器文件夹上传时filename可能包含路径）
        pure_filename = os.path.basename(raw_filename.replace('\\', '/'))
        # 相对路径优先使用前端传来的file_paths，否则使用raw_filename（包含路径）
        if file_paths and idx < len(file_paths) and file_paths[idx]:
            relative_path = file_paths[idx].replace('\\', '/')
        else:
            relative_path = raw_filename.replace('\\', '/')
        
        safe_name = f"{idx}_{pure_filename}"
        file_path = batch_dir / safe_name
        file_bytes = file.file.read()
        file.file.seek(0)
        with open(file_path, "wb") as f:
            f.write(file_bytes)
        file_infos.append({
            "index": idx,
            "filename": pure_filename,
            "relative_path": relative_path,
            "size": len(file_bytes),
        })

    meta = {
        "batch_id": batch_id,
        "period": period,
        "created_at": time.time(),
        "files": file_infos,
    }
    with open(batch_dir / "meta.json", "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    return meta


def precheck_batch(db: Session, batch_id: str) -> Dict:
    """
    预检查批次中的文件，判断哪些能匹配模板、哪些不能。
    对于Excel多工作表文件，会检查所有包含数据的工作表。
    每个未匹配的（文件, 工作表）组合生成一个独立的待配置条目。
    
    返回: {batch_id, matched_files: [...], unmatched_files: [...], all_files: [...]}
    unmatched_files 中的每个条目可能包含 sheet_name 字段，标识需要配置模板的具体工作表
    """
    batch_dir = TEMP_DIR / batch_id
    meta_path = batch_dir / "meta.json"
    if not meta_path.exists():
        raise ValueError(f"批次 {batch_id} 不存在或已过期")

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    matched_files = []
    unmatched_files = []
    file_results = []

    for finfo in meta.get("files", []):
        idx = finfo["index"]
        filename = finfo["filename"]
        relative_path = finfo.get("relative_path", filename)
        context_keywords = _extract_keywords_from_path(relative_path)
        
        safe_name = f"{idx}_{filename}"
        file_path = batch_dir / safe_name

        if not file_path.exists():
            unmatched_files.append({**finfo, "reason": "文件不存在", "sheet_name": None})
            file_results.append({**finfo, "matched": False, "reason": "文件不存在"})
            continue

        try:
            with open(file_path, "rb") as f:
                file_bytes = f.read()

            ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
            file_fully_matched = True
            matched_template_info = None
            all_matched_sheets = []
            unmatched_sheet_entries = []

            if ext in ("xlsx", "xls"):
                sheet_names = _get_excel_sheet_names(file_bytes, filename)
                if not sheet_names:
                    sheet_names = [None]
                
                data_sheet_count = 0
                for sheet_name in sheet_names:
                    try:
                        file_rows = _parse_excel(file_bytes, filename, sheet_name=sheet_name)
                    except Exception:
                        continue
                    
                    if not file_rows or len(file_rows) < 2:
                        continue

                    if not _sheet_has_data(file_rows):
                        continue

                    data_sheet_count += 1
                    sheet_display = sheet_name if sheet_name else f"工作表{data_sheet_count}"

                    template = _match_template(db, file_rows, filename, sheet_name=sheet_name, context_keywords=context_keywords)
                    if template:
                        all_matched_sheets.append(sheet_display)
                        if not matched_template_info:
                            matched_template_info = {"name": template.name, "id": template.id}
                    else:
                        file_fully_matched = False
                        unmatched_entry = {
                            **finfo,
                            "sheet_name": sheet_name,
                            "sheet_display": sheet_display,
                            "reason": f"工作表「{sheet_display}」未匹配到模板",
                            "display_name": f"{filename}（工作表：{sheet_display}）" if sheet_name else filename,
                        }
                        unmatched_sheet_entries.append(unmatched_entry)

                if data_sheet_count == 0:
                    reason = "文件中未找到有效数据工作表"
                    unmatched_files.append({**finfo, "reason": reason, "sheet_name": None})
                    file_results.append({**finfo, "matched": False, "reason": reason, "total_sheets": len(sheet_names)})
                    continue
                elif not file_fully_matched:
                    for entry in unmatched_sheet_entries:
                        unmatched_files.append(entry)
                    total_sheets = data_sheet_count
                    file_results.append({
                        **finfo,
                        "matched": False,
                        "reason": f"{len(unmatched_sheet_entries)}/{total_sheets}个工作表未匹配模板",
                        "matched_sheets": all_matched_sheets,
                        "unmatched_sheets": [e["sheet_display"] for e in unmatched_sheet_entries],
                    })
                    continue
                else:
                    pass
                            
            elif ext == "pdf":
                file_rows = _parse_pdf(file_bytes)
                if not file_rows or len(file_rows) < 2:
                    unmatched_files.append({**finfo, "reason": "PDF文件为空或无法解析", "sheet_name": None})
                    file_results.append({**finfo, "matched": False, "reason": "PDF文件为空或无法解析"})
                    continue
                if not _sheet_has_data(file_rows, min_data_rows=1):
                    unmatched_files.append({**finfo, "reason": "PDF文件中未找到有效数据表格", "sheet_name": None})
                    file_results.append({**finfo, "matched": False, "reason": "PDF文件中未找到有效数据表格"})
                    continue
                template = _match_template(db, file_rows, filename, context_keywords=context_keywords)
                if template:
                    matched_template_info = {"name": template.name, "id": template.id}
                    all_matched_sheets = ["(PDF)"]
                else:
                    file_fully_matched = False
                    unmatched_files.append({
                        **finfo,
                        "sheet_name": None,
                        "sheet_display": None,
                        "reason": "未匹配到模板",
                        "display_name": filename,
                    })
                    file_results.append({**finfo, "matched": False, "reason": "未匹配到模板"})
                    continue
            else:
                unmatched_files.append({**finfo, "reason": "不支持的文件格式", "sheet_name": None})
                file_results.append({**finfo, "matched": False, "reason": "不支持的文件格式"})
                continue

            if file_fully_matched and matched_template_info:
                matched_files.append({
                    **finfo, 
                    "template_name": matched_template_info["name"], 
                    "template_id": matched_template_info["id"],
                    "matched_sheets": all_matched_sheets,
                })
                file_results.append({
                    **finfo, 
                    "matched": True, 
                    "template_name": matched_template_info["name"], 
                    "template_id": matched_template_info["id"],
                    "matched_sheets": all_matched_sheets,
                })
            else:
                unmatched_files.append({**finfo, "reason": "未匹配到模板（已检查所有工作表）", "sheet_name": None})
                file_results.append({**finfo, "matched": False, "reason": "未匹配到模板（已检查所有工作表）"})
        except Exception as e:
            logger.exception(f"预检查文件失败: {filename}")
            unmatched_files.append({**finfo, "reason": f"文件解析失败: {str(e)}", "sheet_name": None})
            file_results.append({**finfo, "matched": False, "reason": f"文件解析失败: {str(e)}"})

    return {
        "batch_id": batch_id,
        "period": meta.get("period", ""),
        "total_files": len(meta.get("files", [])),
        "matched_count": len(matched_files),
        "unmatched_count": len(unmatched_files),
        "has_unmatched": len(unmatched_files) > 0,
        "matched_files": matched_files,
        "unmatched_files": unmatched_files,
        "all_files": file_results,
    }


def auto_detect_from_batch(
    batch_id: str,
    file_index: int,
    sheet_name: Optional[str] = None,
) -> Dict:
    """
    对批次中指定索引的文件（可指定工作表）执行自动模板识别。
    返回auto_detect_template的结果。
    
    Args:
        batch_id: 批次ID
        file_index: 文件索引
        sheet_name: Excel工作表名称，None则使用默认工作表
    """
    batch_dir = TEMP_DIR / batch_id
    meta_path = batch_dir / "meta.json"
    if not meta_path.exists():
        raise ValueError(f"批次 {batch_id} 不存在或已过期")

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    finfo = None
    for f in meta.get("files", []):
        if f["index"] == file_index:
            finfo = f
            break
    if not finfo:
        raise ValueError(f"批次中未找到索引为 {file_index} 的文件")

    filename = finfo["filename"]
    safe_name = f"{file_index}_{filename}"
    file_path = batch_dir / safe_name
    if not file_path.exists():
        raise ValueError(f"文件不存在: {filename}")

    with open(file_path, "rb") as f:
        file_bytes = f.read()

    return auto_detect_template(file_bytes, filename, sheet_name=sheet_name)


def run_smart_import_from_batch(
    db: Session,
    batch_id: str,
    user_id: int,
    username: str,
) -> ImportResult:
    """从临时批次目录读取文件并执行智能导入"""
    batch_dir = TEMP_DIR / batch_id
    meta_path = batch_dir / "meta.json"
    if not meta_path.exists():
        raise ValueError(f"批次 {batch_id} 不存在或已过期")

    with open(meta_path, "r", encoding="utf-8") as f:
        meta = json.load(f)

    period = meta.get("period", "")

    class TempFileWrapper:
        """模拟UploadFile的简单包装器"""
        def __init__(self, filename: str, file_path: Path):
            self.filename = filename
            self._file_path = file_path
            self._data = None

        @property
        def file(self):
            if self._data is None:
                with open(self._file_path, "rb") as f:
                    self._data = io.BytesIO(f.read())
            self._data.seek(0)
            return self._data

    temp_files = []
    file_paths = []
    for finfo in meta.get("files", []):
        idx = finfo["index"]
        filename = finfo["filename"]
        relative_path = finfo.get("relative_path", filename)
        safe_name = f"{idx}_{filename}"
        file_path = batch_dir / safe_name
        if file_path.exists():
            temp_files.append(TempFileWrapper(filename, file_path))
            file_paths.append(relative_path)

    result = run_smart_import(db, period, temp_files, user_id, username, file_paths=file_paths)
    # 收集未匹配模板的文件名（仅error级别的no_template错误，即整个文件未匹配到模板）
    # warning级别的是单个工作表未匹配（预检查阶段应已拦截，此处仅作兜底）
    result.no_template_filenames = list({
        e["file_name"] for e in result.errors 
        if e.get("error_type") == "no_template" and e.get("error_level") == "error"
    })
    # 不自动清理批次目录，由调用方决定是否清理（若需要配置模板则先复制失败文件）
    return result


def cleanup_batch(batch_id: str):
    """清理批次临时文件"""
    batch_dir = TEMP_DIR / batch_id
    if batch_dir.exists():
        shutil.rmtree(batch_dir, ignore_errors=True)


def create_batch_from_unmatched(original_batch_id: str, failed_filenames: List[str]) -> Dict:
    """
    从原批次中提取未匹配模板的文件，创建一个新的批次用于模板配置。
    返回新批次的 meta 信息（含 batch_id、period、files 列表）。
    """
    _cleanup_expired_batches()

    src_dir = TEMP_DIR / original_batch_id
    src_meta_path = src_dir / "meta.json"
    if not src_meta_path.exists():
        raise ValueError(f"原批次 {original_batch_id} 不存在或已过期")

    with open(src_meta_path, "r", encoding="utf-8") as f:
        src_meta = json.load(f)

    period = src_meta.get("period", "")
    failed_set = set(failed_filenames)

    new_batch_id = str(uuid.uuid4())
    new_dir = TEMP_DIR / new_batch_id
    new_dir.mkdir(parents=True, exist_ok=True)

    new_file_infos = []
    new_idx = 0
    for finfo in src_meta.get("files", []):
        if finfo["filename"] not in failed_set:
            continue
        old_safe_name = f"{finfo['index']}_{finfo['filename']}"
        src_file = src_dir / old_safe_name
        if not src_file.exists():
            continue
        new_safe_name = f"{new_idx}_{finfo['filename']}"
        dst_file = new_dir / new_safe_name
        shutil.copy2(str(src_file), str(dst_file))
        new_file_infos.append({
            "index": new_idx,
            "filename": finfo["filename"],
            "relative_path": finfo.get("relative_path", finfo["filename"]),
            "size": finfo.get("size", 0),
        })
        new_idx += 1

    if not new_file_infos:
        shutil.rmtree(new_dir, ignore_errors=True)
        raise ValueError("没有需要配置模板的文件")

    new_meta = {
        "batch_id": new_batch_id,
        "period": period,
        "created_at": time.time(),
        "files": new_file_infos,
    }
    with open(new_dir / "meta.json", "w", encoding="utf-8") as f:
        json.dump(new_meta, f, ensure_ascii=False, indent=2)

    return new_meta
